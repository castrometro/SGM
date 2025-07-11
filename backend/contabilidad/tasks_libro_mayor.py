# tasks_libro_mayor.py

import os
import re
import hashlib
import datetime
from decimal import Decimal

from django.utils import timezone
from django.core.cache import cache
from django.core.files.storage import default_storage
from django.db import transaction
from django.db.models import Q
from celery import shared_task, chain
from celery.utils.log import get_task_logger
from openpyxl import load_workbook
from django.contrib.auth import get_user_model

from .models import (
    UploadLog,
    LibroMayorArchivo,
    CuentaContable,
    AperturaCuenta,
    MovimientoContable,
    TipoDocumento,
    NombreIngles,
    # ClasificacionCuentaArchivo,  # OBSOLETO - ELIMINADO EN REDISEÑO
    ClasificacionSet,
    ClasificacionOption,
    AccountClassification,
    Incidencia,
    ExcepcionValidacion,
    ExcepcionClasificacionSet,
)

logger = get_task_logger(__name__)


def crear_chain_libro_mayor(upload_log_id, user_correo_bdo):
    """
    Crea un chain de tasks para procesar el libro mayor.
    Usamos .s() para el primer task y .si() para los demás para mantener
    los mismos argumentos en cada task del chain.
    """
    return chain(
        validar_nombre_archivo_libro_mayor.s(upload_log_id, user_correo_bdo),
        verificar_archivo_libro_mayor.si(upload_log_id, user_correo_bdo),
        validar_contenido_libro_mayor.si(upload_log_id, user_correo_bdo),
        procesar_libro_mayor_raw.si(upload_log_id, user_correo_bdo),
        generar_incidencias_libro_mayor.si(upload_log_id, user_correo_bdo),
        finalizar_procesamiento_libro_mayor.si(upload_log_id, user_correo_bdo),
    )
# ─── Utilidades ───────────────────────────────────────────────────────────────

def _parse_fecha(value, fecha_cierre=None):
    """
    Parsea una fecha desde el Excel.
    Si no se puede parsear, usa fecha_cierre como fallback, o fecha actual.
    """
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, str) and value.strip():
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    
    # Si no se pudo parsear y hay fecha_cierre, usarla
    if fecha_cierre:
        return fecha_cierre
    
    # Fallback a fecha actual
    return timezone.now().date()


def _clean_header(h):
    if not isinstance(h, str):
        return ""
    hh = h.strip().upper()
    for a, b in [("Á","A"),("É","E"),("Í","I"),("Ó","O"),("Ú","U"),("Ñ","N")]:
        hh = hh.replace(a, b)
    return re.sub(r"[^A-Z0-9]", "", hh)


# ─── Tasks 1–3: Validaciones previas ──────────────────────────────────────────

@shared_task
def validar_nombre_archivo_libro_mayor(upload_log_id, user_correo_bdo):
    try:
        upload = UploadLog.objects.get(pk=upload_log_id)
        nombre = upload.nombre_archivo_original
        rut = upload.cliente.rut or str(upload.cliente.id)
        rut_clean = rut.replace(".", "").replace("-", "")
        pattern = rf"^{rut_clean}_(LibroMayor|Mayor)_(\d{{6}})\.xlsx?$"
        match = re.match(pattern, nombre, flags=re.IGNORECASE)
        if not match:
            msg = f"Nombre inválido: {nombre}"
            upload.estado  = "error"
            upload.errores = msg
            upload.save()
            raise ValueError(msg)
        periodo = match.group(2)
        resumen = upload.resumen or {}
        resumen["validacion_nombre"] = {"periodo": periodo}
        upload.resumen = resumen
        upload.estado = "procesando"
        upload.save(update_fields=["resumen","estado"])
        return upload_log_id
    except UploadLog.DoesNotExist:
        logger.error(f"UploadLog con id {upload_log_id} no encontrado")
        raise
    except Exception as e:
        logger.error(f"Error validando nombre archivo libro mayor {upload_log_id}: {e}")
        try:
            upload = UploadLog.objects.get(pk=upload_log_id)
            upload.estado = "error"
            upload.errores = f"Error validando nombre: {str(e)}"
            upload.save()
        except:
            pass
        raise


@shared_task
def verificar_archivo_libro_mayor(upload_log_id, user_correo_bdo):
    upload = UploadLog.objects.get(pk=upload_log_id)
    ruta_rel = upload.ruta_archivo
    if not ruta_rel:
        raise ValueError("Ruta de archivo no definida")
    full = default_storage.path(ruta_rel)
    if not os.path.exists(full):
        raise FileNotFoundError(full)
    with open(full, "rb") as f:
        data = f.read()
    upload.hash_archivo = hashlib.sha256(data).hexdigest()
    resumen = upload.resumen or {}
    resumen["verificacion_archivo"] = {"bytes": len(data)}
    upload.resumen = resumen
    upload.save(update_fields=["hash_archivo","resumen"])
    return upload_log_id


@shared_task
def validar_contenido_libro_mayor(upload_log_id, user_correo_bdo):
    upload = UploadLog.objects.get(pk=upload_log_id)
    full = default_storage.path(upload.ruta_archivo)
    wb = load_workbook(full, read_only=True, data_only=True)
    ws = wb.active
    headers = next(ws.iter_rows(min_row=9, max_row=9, values_only=True))
    cleaned = [_clean_header(h) for h in headers]
    required = {"CUENTA","FECHA","DEBE","HABER","DESCRIPCION"}
    if not required.issubset(cleaned):
        faltan = required - set(cleaned)
        msg = f"Headers faltantes: {faltan}"
        upload.estado  = "error"
        upload.errores = msg
        upload.save()
        raise ValueError(msg)
    # chequear datos en fila 11
    sample = list(ws.iter_rows(min_row=11, max_row=11, values_only=True))[0]
    if all(cell is None for cell in sample):
        msg = "Sin datos en fila 11"
        upload.estado  = "error"
        upload.errores = msg
        upload.save()
        raise ValueError(msg)
    resumen = upload.resumen or {}
    resumen["validacion_contenido"] = {"ok": True}
    upload.resumen = resumen
    upload.save(update_fields=["resumen"])
    return upload_log_id


# ─── Task 4: Procesar aperturas y movimientos ─────────────────────────────────

@shared_task
def procesar_libro_mayor_raw(upload_log_id, user_correo_bdo):
    upload = UploadLog.objects.get(pk=upload_log_id)
    inicio = timezone.now()
    full = default_storage.path(upload.ruta_archivo)
    wb = load_workbook(full, read_only=True, data_only=True)
    ws = wb.active

    # -- índices de columnas --
    headers = next(ws.iter_rows(min_row=9, max_row=9, values_only=True))
    idx = {_clean_header(h): i for i,h in enumerate(headers) if isinstance(h, str)}
    C, F, D, H, DS = idx["CUENTA"], idx["FECHA"], idx["DEBE"], idx["HABER"], idx["DESCRIPCION"]
    S = idx.get("SALDO")  # Column for SALDO
    TD = idx.get("TIPODOC") or idx.get("TIPODOC.")
    
    # Columnas adicionales que pueden estar presentes
    ND = idx.get("NUMERODOCUMENTO") or idx.get("NUMDOCUMENTO") or idx.get("NUMDOC") or idx.get("NUMERODOC")
    TP = idx.get("TIPO")
    NC = idx.get("NUMEROCOMPROBANTE") or idx.get("NUMCOMPROBANTE") or idx.get("COMPROBANTE") or idx.get("NCOMPROBANTE")
    NI = idx.get("NUMEROINTERNO") or idx.get("NUMINTERNO") or idx.get("INTERNO") or idx.get("NINTERNO")
    CC = idx.get("CENTROCOSTO") or idx.get("CENTRO_COSTO") or idx.get("CCOSTO") or idx.get("CENTRODECOSTO")
    AUX = idx.get("AUXILIAR")
    DG = idx.get("DETALLEGASTO") or idx.get("DETALLE_GASTO") or idx.get("GASTO") or idx.get("DETDEGASTOINSTFINANCIERO")
    
    logger.info(f"Columnas detectadas: {list(idx.keys())}")
    logger.info(f"Columnas adicionales encontradas: ND={ND}, TP={TP}, NC={NC}, NI={NI}, CC={CC}, AUX={AUX}, DG={DG}")

    cliente = upload.cliente
    cliente_id = cliente.id
    cierre = upload.cierre

    # 1. CARGA INICIAL COMPLETA - Una sola vez al inicio
    
    # Mapas de datos principales
    nombres_ingles_map = {
        ni.cuenta_codigo: ni.nombre_ingles
        for ni in NombreIngles.objects.filter(cliente=cliente)
    }
    # Mapeo de clasificaciones existentes (fuente única de verdad)
    # Usamos AccountClassification en lugar del modelo obsoleto
    clasif_existentes = {}
    for ac in AccountClassification.objects.filter(cliente=cliente).select_related('set_clas', 'opcion', 'cuenta'):
        codigo_cuenta = ac.codigo_cuenta_display  # Usa el property que maneja tanto FK como código temporal
        # Log si es temporal o tiene FK
        if ac.cuenta:
            logger.debug(f"Clasificación con FK: cuenta {ac.cuenta.codigo} - set {ac.set_clas.nombre}")
        else:
            logger.debug(f"Clasificación temporal: cuenta_codigo {ac.cuenta_codigo} - set {ac.set_clas.nombre}")
        if codigo_cuenta not in clasif_existentes:
            clasif_existentes[codigo_cuenta] = {}
        clasif_existentes[codigo_cuenta][ac.set_clas.nombre] = ac.opcion.valor
    
    # Sets de clasificación con sus opciones precargadas
    sets_clasificacion = list(ClasificacionSet.objects.filter(cliente=cliente).prefetch_related('opciones'))
    sets_clasificacion_dict = {s.id: s for s in sets_clasificacion}
    
    # Excepciones por set de clasificación
    excepciones_por_set = {}
    total_excepciones = 0
    for exc in ExcepcionClasificacionSet.objects.filter(cliente=cliente, activa=True).select_related('set_clasificacion'):
        set_id = exc.set_clasificacion.id
        if set_id not in excepciones_por_set:
            excepciones_por_set[set_id] = set()
        excepciones_por_set[set_id].add(exc.cuenta_codigo)
        total_excepciones += 1
    
    logger.info(f"Cargadas {total_excepciones} excepciones de clasificación activas para {len(excepciones_por_set)} sets")
    
    # Tipos de documento precargados (con caché optimizado)
    tipos_documento_map = {
        td.codigo: td
        for td in TipoDocumento.objects.filter(cliente=cliente)
    }
    
    # Excepciones de validación por tipo
    excepciones_nombres_ingles = set(
        ExcepcionValidacion.objects.filter(
            cliente=cliente,
            tipo_excepcion='cuentas_sin_nombre_ingles',
            activa=True
        ).values_list('codigo_cuenta', flat=True)
    )
    
    excepciones_tipo_doc_null = set(
        ExcepcionValidacion.objects.filter(
            cliente=cliente,
            tipo_excepcion='movimientos_tipodoc_nulo',
            activa=True
        ).values_list('codigo_cuenta', flat=True)
    )
    
    excepciones_tipo_doc_no_reconocido = set(
        ExcepcionValidacion.objects.filter(
            cliente=cliente,
            tipo_excepcion='tipos_doc_no_reconocidos',
            activa=True
        ).values_list('codigo_cuenta', flat=True)
    )
    
    # Arrays de trabajo
    processed_accounts = {}
    aperturas = []
    movimientos = []
    incidencias_pendientes = []  # Array para recopilar incidencias
    
    # Sets para agrupar incidencias por cuenta (evitar duplicados por cuenta)
    cuentas_sin_clasificacion_por_set = {}  # {set_id: {cuenta_codigo: set_nombre}}
    cuentas_sin_nombre_ingles = set()
    cuentas_con_tipo_doc_null = set()
    cuentas_con_tipo_doc_no_reconocido = {}  # dict para guardar también el código no reconocido
    
    # NUEVO: Contador de movimientos por cuenta para debugging
    movimientos_por_cuenta = {}  # {cuenta_codigo: cantidad_movimientos}
    
    # NUEVO: Totales ESF/ERI calculados durante el procesamiento + CATEGORÍA PARA CUENTAS NO CLASIFICADAS
    totales_esf_eri = {
        'ESF': {'saldo_ant': Decimal('0'), 'debe': Decimal('0'), 'haber': Decimal('0')},
        'ERI': {'saldo_ant': Decimal('0'), 'debe': Decimal('0'), 'haber': Decimal('0')},
        'NO_CLASIFICADAS': {'saldo_ant': Decimal('0'), 'debe': Decimal('0'), 'haber': Decimal('0')},
    }
    
    # NUEVO: Contadores para debugging
    contadores_clasificacion = {
        'ESF': 0,
        'ERI': 0,
        'NO_CLASIFICADAS': 0,
        'Sin_clasificacion': 0  # Mantenemos para retrocompatibilidad
    }
    
    stats = {"cuentas_nuevas": 0, "aperturas": 0, "movimientos": 0, "incidencias_detectadas": 0}
    
    logger.info(f"CARGA INICIAL COMPLETADA:")
    logger.info(f"  - {len(nombres_ingles_map)} nombres en inglés")
    logger.info(f"  - {len(clasif_existentes)} cuentas con clasificaciones")
    logger.info(f"  - {len(sets_clasificacion)} sets de clasificación")
    logger.info(f"  - {len(excepciones_por_set)} sets con excepciones")
    logger.info(f"  - {len(tipos_documento_map)} tipos de documento")
    logger.info(f"  - {len(excepciones_nombres_ingles)} excepciones nombres inglés")
    logger.info(f"  - {len(excepciones_tipo_doc_null)} excepciones tipo doc null")
    logger.info(f"  - {len(excepciones_tipo_doc_no_reconocido)} excepciones tipo doc no reconocido")

    # -- helper optimizado para tipos de documento --
    def get_tipo_doc(codigo_td, cuenta_codigo):
        """
        Obtiene tipo de documento optimizado con validación de excepciones en línea
        Retorna: (tipo_documento_obj, generar_incidencia_null, generar_incidencia_no_reconocido)
        """
        if not codigo_td:
            # Sin código de tipo documento
            generar_incidencia = cuenta_codigo not in excepciones_tipo_doc_null
            return None, generar_incidencia, False
        
        # Buscar tipo documento en el mapa precargado
        tipo_doc = tipos_documento_map.get(codigo_td.strip())
        
        if tipo_doc:
            # Tipo documento encontrado
            return tipo_doc, False, False
        else:
            # Tipo documento no reconocido
            generar_incidencia = cuenta_codigo not in excepciones_tipo_doc_no_reconocido
            return None, False, generar_incidencia

    # -- helper para identificar clasificación ESF/ERI --
    def identificar_clasificacion_esf_eri(cuenta_obj):
        """
        Identifica si una cuenta es ESF o ERI basándose en sus clasificaciones
        Retorna: 'ESF', 'ERI' o None
        """
        try:
            # Obtener todas las clasificaciones de la cuenta
            clasificaciones = AccountClassification.objects.filter(
                cuenta=cuenta_obj
            ).select_related('set_clas', 'opcion')
            
            if not clasificaciones.exists():
                logger.debug(f"Cuenta {cuenta_obj.codigo}: Sin clasificaciones")
                return None
            
            for clasificacion in clasificaciones:
                if not clasificacion.opcion:
                    continue
                    
                set_nombre = clasificacion.set_clas.nombre.upper()
                valor = clasificacion.opcion.valor.upper()
                
                logger.debug(f"Cuenta {cuenta_obj.codigo}: Set='{set_nombre}' | Valor='{valor}'")
                
                # Identificar ESF (Estado de Situación Financiera)
                if ('ESTADO' in set_nombre and 'SITUACION' in set_nombre and 'FINANCIERA' in set_nombre) or \
                   ('BALANCE' in set_nombre) or \
                   (valor in ['ACTIVO CORRIENTE', 'ACTIVO NO CORRIENTE', 'PASIVO CORRIENTE', 'PASIVO NO CORRIENTE', 'PATRIMONIO']) or \
                   (valor in ['CASH AND CASH EQUIVALENT', 'OTHER CURRENT FINANCIAL ASSETS', 'INVENTORIES',
                             'COMMERCIAL DEBTORS AND OTHER RECEIVABLES, CURRENT', 'CURRENT TAX RECEIVABLE',
                             'OTHER CURRENT ASSETS', 'PROPERTIES, FACILITY AND EQUIPMENT', 'DEFERRED TAX ASSETS',
                             'OTHER NON-CURRENT ASSETS', 'OTHER FINANCIAL, NON-CURRENT ASSETS',
                             'COMMERCIAL ACCOUNTS AND OTHER ACCOUNTS PAYABLE, CURRENT', 'LIABILITY FOR CURRENT TAXES',
                             'OTHER LIABILITY, CURRENT', 'PROVISIONS', 'ACCOUNTS PAYABLE TO RELATED ENTITIES, NO CURRENT',
                             'DEFERRED TAX LIABILITY', 'PAID-IN CAPITAL', 'OTHER RESERVES']):
                    # LOGGING ESPECIAL PARA PATRIMONIO Y ESF
                    if 'PATRIMONIO' in valor or 'CAPITAL' in valor or 'RESERVES' in valor:
                        logger.info(f"🏛️  CUENTA PATRIMONIO ESF: {cuenta_obj.codigo} ({cuenta_obj.nombre}) | Set: '{set_nombre}' | Valor: '{valor}'")
                    else:
                        logger.debug(f"Cuenta {cuenta_obj.codigo}: Identificada como ESF por Set='{set_nombre}' | Valor='{valor}'")
                    return 'ESF'
                
                # Identificar ERI (Estado de Resultados Integrales)
                elif ('RESULTADO' in set_nombre) or ('INCOME' in set_nombre) or \
                     (valor in ['INGRESOS', 'GASTOS', 'COSTOS', 'OTROS INGRESOS', 'OTROS GASTOS']) or \
                     (valor in ['ADMINISTRATION EXPENSES', 'COST SALE', 'FINANCIAL EXPENSES', 'FINANCIAL INCOME', 
                               'OTHER EXPENSES', 'INCOME', 'GANANCIAS', 'PERDIDAS', 'GAINS (LOSSES) ACCUMULATED',
                               'GANANCIAS (ANTES DE IMPUESTOS)', 'GANANCIAS (PERDIDAS)', 'GANANCIAS BRUTAS',
                               'INCOME / (LOSS) OF THE OPERATION', 'OTHER EXPENSES, BY FUNCTION', 'DIFFERENCE IN CHANGES']):
                    # LOGGING ESPECIAL para detectar si alguna cuenta patrimonio está clasificada como ERI
                    if 'PATRIMONIO' in cuenta_obj.nombre.upper() or 'CAPITAL' in cuenta_obj.nombre.upper():
                        logger.warning(f"⚠️  POSIBLE ERROR: Cuenta aparentemente de PATRIMONIO clasificada como ERI: {cuenta_obj.codigo} ({cuenta_obj.nombre}) | Set: '{set_nombre}' | Valor: '{valor}'")
                    else:
                        logger.debug(f"Cuenta {cuenta_obj.codigo}: Identificada como ERI por Set='{set_nombre}' | Valor='{valor}'")
                    return 'ERI'
                
                # Búsqueda por valores específicos que identifiquen ESF/ERI
                elif valor in ['ESF', 'ERI']:
                    logger.debug(f"Cuenta {cuenta_obj.codigo}: Identificada como {valor} por valor directo")
                    return valor
            
            logger.debug(f"Cuenta {cuenta_obj.codigo}: No identificada como ESF/ERI")
            return None
        except Exception as e:
            logger.debug(f"Error identificando clasificación ESF/ERI para cuenta {cuenta_obj.codigo}: {e}")
            return None

    # 2. FUNCIONES AUXILIARES
    def procesar_saldo_anterior(row, cierre, code, aperturas, stats, totales_esf_eri, identificar_clasificacion_esf_eri):
        """Procesa un saldo anterior y lo añade a la lista de aperturas"""
        try:
            # Obtener el saldo exclusivamente de la columna SALDO
            if S is not None:
                saldo = Decimal(row[S] or 0)
                origen_saldo = f"columna SALDO = {row[S]}"
            else:
                # Si no existe la columna SALDO, usar 0
                saldo = Decimal(0)
                origen_saldo = "columna SALDO no encontrada - usando saldo 0"
            
            logger.info(f"📊 PROCESANDO SALDO ANTERIOR: Cuenta {code} | {origen_saldo} | Saldo final: ${saldo:,.2f}")
            cuenta_obj = processed_accounts[code]
            apertura = AperturaCuenta(
                cierre=cierre,
                cuenta=cuenta_obj,
                saldo_anterior=saldo
            )
            aperturas.append(apertura)
            stats["aperturas"] += 1
            
            # NUEVO: Acumular en totales ESF/ERI si corresponde, o en NO_CLASIFICADAS
            clasificacion_esf_eri = identificar_clasificacion_esf_eri(cuenta_obj)
            if clasificacion_esf_eri in ['ESF', 'ERI']:
                totales_esf_eri[clasificacion_esf_eri]['saldo_ant'] += saldo
                contadores_clasificacion[clasificacion_esf_eri] += 1
                logger.info(f"🔍 SALDO ANTERIOR {clasificacion_esf_eri}: Cuenta {code} = ${saldo:,.2f} | Total acumulado: ${totales_esf_eri[clasificacion_esf_eri]['saldo_ant']:,.2f}")
            else:
                # Incluir cuentas no clasificadas en la categoría NO_CLASIFICADAS
                totales_esf_eri['NO_CLASIFICADAS']['saldo_ant'] += saldo
                contadores_clasificacion['NO_CLASIFICADAS'] += 1
                contadores_clasificacion['Sin_clasificacion'] += 1  # Mantener para retrocompatibilidad
                logger.warning(f"⚠️  CUENTA NO CLASIFICADA (incluida en totales): {code} = ${saldo:,.2f} | Clasificación: {clasificacion_esf_eri} | Total NO_CLASIFICADAS: ${totales_esf_eri['NO_CLASIFICADAS']['saldo_ant']:,.2f}")
            
            logger.debug(f"Saldo anterior procesado para cuenta {code}: {saldo}")
        except Exception as e:
            logger.error(f"Error procesando saldo anterior para cuenta {code}: {e}")

    def procesar_movimiento(row, cierre, cuenta_obj, get_tipo_doc, movimientos, stats, indices_cols, totales_esf_eri, identificar_clasificacion_esf_eri):
        """Procesa un movimiento contable con validación completa en línea"""
        try:
            # Desempaquetar índices de columnas
            ND, TP, NC, NI, CC, AUX, DG, DS, TD, D, H, F = indices_cols
            
            # Extraer código de tipo documento
            codigo_td = (str(row[TD]).strip()[:10] if (TD is not None and row[TD]) else "")
            
            # Extraer valores debe/haber
            debe = Decimal(row[D] or 0)
            haber = Decimal(row[H] or 0)
            
            # NUEVO: Contador de movimientos por cuenta
            if cuenta_obj.codigo not in movimientos_por_cuenta:
                movimientos_por_cuenta[cuenta_obj.codigo] = 0
            movimientos_por_cuenta[cuenta_obj.codigo] += 1
            
            logger.debug(f"📝 PROCESANDO MOVIMIENTO #{movimientos_por_cuenta[cuenta_obj.codigo]} para cuenta {cuenta_obj.codigo}: Debe=${debe:,.2f} | Haber=${haber:,.2f}")
            
            # Validar tipo documento con el helper optimizado
            tipo_doc_obj, generar_inc_null, generar_inc_no_reconocido = get_tipo_doc(codigo_td, cuenta_obj.codigo)
            
            # Generar incidencias de tipo documento si es necesario (AGRUPADAS POR CUENTA)
            if generar_inc_null:
                cuentas_con_tipo_doc_null.add(cuenta_obj.codigo)
                logger.debug(f"Cuenta {cuenta_obj.codigo} marcada para incidencia tipo doc null")
            
            if generar_inc_no_reconocido:
                cuentas_con_tipo_doc_no_reconocido[cuenta_obj.codigo] = codigo_td
                logger.debug(f"Cuenta {cuenta_obj.codigo} marcada para incidencia tipo doc no reconocido: '{codigo_td}'")
            
            # NUEVO: Acumular en totales ESF/ERI si corresponde, o en NO_CLASIFICADAS
            clasificacion_esf_eri = identificar_clasificacion_esf_eri(cuenta_obj)
            if clasificacion_esf_eri in ['ESF', 'ERI']:
                totales_esf_eri[clasificacion_esf_eri]['debe'] += debe
                totales_esf_eri[clasificacion_esf_eri]['haber'] += haber
                logger.debug(f"Movimiento {clasificacion_esf_eri} acumulado para cuenta {cuenta_obj.codigo}: Debe={debe}, Haber={haber}")
            else:
                # Incluir movimientos de cuentas no clasificadas en la categoría NO_CLASIFICADAS
                totales_esf_eri['NO_CLASIFICADAS']['debe'] += debe
                totales_esf_eri['NO_CLASIFICADAS']['haber'] += haber
                logger.debug(f"Movimiento NO_CLASIFICADO acumulado para cuenta {cuenta_obj.codigo}: Debe={debe}, Haber={haber} | Total acumulado: Debe=${totales_esf_eri['NO_CLASIFICADAS']['debe']:,.2f}, Haber=${totales_esf_eri['NO_CLASIFICADAS']['haber']:,.2f}")
            
            # Crear el movimiento contable con TODOS los datos disponibles
            mov = MovimientoContable(
                cierre=cierre,
                cuenta=cuenta_obj,
                fecha=_parse_fecha(row[F], cierre.fecha_cierre),
                debe=debe,
                haber=haber,
                descripcion=str(row[DS] or "")[:500] if (DS is not None and row[DS]) else "",
                tipo_doc_codigo=codigo_td,
                tipo_documento=tipo_doc_obj,  # Asignar directamente si existe
                # Campos adicionales si están disponibles
                numero_documento=str(row[ND])[:50] if (ND is not None and row[ND]) else "",
                tipo=str(row[TP])[:50] if (TP is not None and row[TP]) else "",
                numero_comprobante=str(row[NC])[:50] if (NC is not None and row[NC]) else "",
                numero_interno=str(row[NI])[:50] if (NI is not None and row[NI]) else "",
                detalle_gasto=str(row[DG]) if (DG is not None and row[DG]) else "",
            )
            
            # TODO: Manejar centro_costo y auxiliar si hay columnas disponibles
            # Estas requieren lookup a otras tablas, se puede implementar después
            if CC is not None and row[CC]:
                logger.debug(f"Centro de costo detectado pero no procesado: {row[CC]}")
            if AUX is not None and row[AUX]:
                logger.debug(f"Auxiliar detectado pero no procesado: {row[AUX]}")
            
            movimientos.append(mov)
            stats["movimientos"] += 1
            
        except Exception as e:
            logger.error(f"Error procesando movimiento: {e}")
            # Crear movimiento con campos mínimos para no perder el registro
            try:
                mov_minimo = MovimientoContable(
                    cierre=cierre,
                    cuenta=cuenta_obj,
                    fecha=_parse_fecha(row[F], cierre.fecha_cierre),
                    debe=Decimal(row[D] or 0),
                    haber=Decimal(row[H] or 0),
                    descripcion="Error procesando campos adicionales",
                    flag_incompleto=True
                )
                movimientos.append(mov_minimo)
                stats["movimientos"] += 1
            except Exception as e2:
                logger.error(f"Error creando movimiento mínimo: {e2}")

    # 3. BUCLE PRINCIPAL
    current_code = None
    fila_numero = 10  # Empezamos en fila 11 del Excel (10 + 1)
    
    with transaction.atomic():
        for row in ws.iter_rows(min_row=11, values_only=True):
            fila_numero += 1
            cell = row[C]
            
            # SALDO ANTERIOR → nuevo bloque de cuenta
            if isinstance(cell, str) and cell.startswith("SALDO ANTERIOR"):
                # Extrae code y nombre real de la cuenta
                # Formato: "SALDO ANTERIOR DE LA CUENTA: 5-04-004-002-0002  Comisiones y gastos bancarios"
                texto_completo = cell.split(":", 1)[1].strip()  # "5-04-004-002-0002  Comisiones y gastos bancarios"
                
                # Separar código y nombre
                partes = texto_completo.split(" ", 1)  # ["5-04-004-002-0002", "Comisiones y gastos bancarios"]
                code = partes[0].strip()
                nombre_real = partes[1].strip() if len(partes) > 1 else f"Cuenta {code}"
                
                current_code = code
                logger.info(f"🆕 NUEVA CUENTA DETECTADA en fila {fila_numero}: {code} - {nombre_real}")
                
                # Reiniciar contador para esta cuenta
                movimientos_por_cuenta[code] = 0
                
                # Si code no está en processed_accounts
                if code not in processed_accounts:
                    # get_or_create la CuentaContable CON EL NOMBRE REAL
                    cuenta, created = CuentaContable.objects.get_or_create(
                        cliente=cliente,
                        codigo=code[:50],
                        defaults={"nombre": nombre_real[:50]}  # Usar el nombre real del Excel
                    )
                    # Si la cuenta ya existía pero no tenía nombre, actualizarlo
                    if not created and (not cuenta.nombre or cuenta.nombre.startswith("Cuenta ")):
                        cuenta.nombre = nombre_real[:50]
                        cuenta.save(update_fields=["nombre"])
                        logger.debug(f"Actualizado nombre de cuenta existente {cuenta.codigo}: {nombre_real}")
                    
                    # Incrementar stats["cuentas_nuevas"] si es nueva
                    if created:
                        stats["cuentas_nuevas"] += 1
                        logger.debug(f"Cuenta nueva creada: {code} - {nombre_real}")
                    
                    # Aplicar nombre en inglés UNA SOLA VEZ usando nombres_ingles_map
                    ing = nombres_ingles_map.get(cuenta.codigo)
                    if ing and not cuenta.nombre_en:
                        cuenta.nombre_en = ing
                        cuenta.save(update_fields=["nombre_en"])
                        logger.debug(f"Aplicado nombre en inglés a cuenta {cuenta.codigo}: {ing}")
                    
                    # NUEVO: Sincronizar clasificaciones temporales con la cuenta real
                    if created:
                        clasificaciones_temporales = AccountClassification.objects.filter(
                            cliente=cliente,
                            cuenta_codigo=code,
                            cuenta__isnull=True  # Solo las temporales
                        )
                        
                        if clasificaciones_temporales.exists():
                            num_actualizadas = clasificaciones_temporales.update(cuenta=cuenta)
                            logger.info(f"✅ Sincronizadas {num_actualizadas} clasificaciones temporales con FK para cuenta {code}")
                    
                    # Aplicar clasificaciones existentes (si las hay)
                    clasificaciones_cuenta = clasif_existentes.get(cuenta.codigo)
                    if clasificaciones_cuenta:
                        logger.debug(f"Encontradas clasificaciones existentes para cuenta {cuenta.codigo}: {list(clasificaciones_cuenta.keys())}")
                        # Las clasificaciones ya están aplicadas en AccountClassification
                        # Solo registramos que esta cuenta ya tiene clasificaciones
                    
                    # VALIDACIÓN DE CLASIFICACIONES EN LÍNEA - Solo si hay sets configurados
                    if sets_clasificacion:
                        cuentas_exceptuadas = 0
                        for set_clas in sets_clasificacion:
                            # Verificar si la cuenta tiene excepción para este set
                            excepciones_set = excepciones_por_set.get(set_clas.id, set())
                            if cuenta.codigo in excepciones_set:
                                cuentas_exceptuadas += 1
                                logger.debug(f"Cuenta {cuenta.codigo} tiene excepción para set {set_clas.nombre} - NO se validará clasificación")
                                continue  # Esta cuenta está excenta de clasificación en este set
                            
                            # CORREGIDO: Verificar clasificación usando TANTO FK como código temporal
                            # Esto incluye clasificaciones que pueden estar guardadas solo con cuenta_codigo (temporales)
                            tiene_clasificacion = AccountClassification.objects.filter(
                                Q(cuenta=cuenta) | Q(cuenta_codigo=cuenta.codigo, cuenta__isnull=True),
                                cliente=cliente,
                                set_clas=set_clas
                            ).exists()
                            if not tiene_clasificacion:
                                # Agregar cuenta y set específico al diccionario de incidencias
                                if set_clas.id not in cuentas_sin_clasificacion_por_set:
                                    cuentas_sin_clasificacion_por_set[set_clas.id] = {}
                                cuentas_sin_clasificacion_por_set[set_clas.id][cuenta.codigo] = set_clas.nombre
                                logger.debug(f"Cuenta {cuenta.codigo} marcada para incidencia de clasificación en set {set_clas.nombre}")
                            else:
                                logger.debug(f"Cuenta {cuenta.codigo} SÍ tiene clasificación en set {set_clas.nombre}")
                        
                        if cuentas_exceptuadas > 0:
                            logger.debug(f"Cuenta {cuenta.codigo} exceptuada de {cuentas_exceptuadas} sets de clasificación")
                    
                    # VALIDACIÓN DE NOMBRES EN INGLÉS EN LÍNEA - Solo si el cliente es bilingüe
                    if cliente.bilingue and not cuenta.nombre_en:
                        # Solo crear incidencia si NO existe un nombre en inglés disponible 
                        # Y la cuenta no está marcada como excepción
                        if cuenta.codigo not in nombres_ingles_map and cuenta.codigo not in excepciones_nombres_ingles:
                            cuentas_sin_nombre_ingles.add(cuenta.codigo)
                            logger.debug(f"Cuenta {cuenta.codigo} marcada para incidencia de nombre en inglés")
                            stats["incidencias_detectadas"] += 1
                            logger.debug(f"Incidencia detectada: Cuenta {cuenta.codigo} sin nombre en inglés")
                    
                    # Guardar la cuenta en processed_accounts
                    processed_accounts[code] = cuenta
                
                # Llamar a procesar_saldo_anterior para añadir la apertura
                procesar_saldo_anterior(row, cierre, code, aperturas, stats, totales_esf_eri, identificar_clasificacion_esf_eri)
                continue
            
            # Si current_code definido y hay datos de movimiento (debe o haber)
            if current_code and (row[D] is not None or row[H] is not None):
                # Verificar si hay al menos debe o haber con valores > 0
                debe_val = Decimal(row[D] or 0) if row[D] is not None else Decimal('0')
                haber_val = Decimal(row[H] or 0) if row[H] is not None else Decimal('0')
                
                logger.debug(f"🔍 EVALUANDO FILA {fila_numero} para cuenta {current_code}: Debe={debe_val}, Haber={haber_val}")
                
                # Solo procesar si hay movimiento real (debe > 0 o haber > 0)
                if debe_val > 0 or haber_val > 0:
                    # Recuperar cuenta_obj = processed_accounts[current_code]
                    cuenta_obj = processed_accounts[current_code]
                    
                    # LOGGING para detectar movimientos sin fecha
                    if row[F] is None or row[F] == "":
                        logger.warning(f"⚠️  MOVIMIENTO SIN FECHA en fila {fila_numero}: Cuenta {current_code} | Debe=${debe_val:,.2f} | Haber=${haber_val:,.2f} | Usando fecha de cierre como default")
                    
                    # Llamar a procesar_movimiento con validación completa
                    indices_cols = (ND, TP, NC, NI, CC, AUX, DG, DS, TD, D, H, F)
                    procesar_movimiento(row, cierre, cuenta_obj, get_tipo_doc, movimientos, stats, indices_cols, totales_esf_eri, identificar_clasificacion_esf_eri)
                    logger.debug(f"✅ MOVIMIENTO PROCESADO en fila {fila_numero} para cuenta {current_code}")
                else:
                    logger.debug(f"⏭️  FILA {fila_numero} IGNORADA para cuenta {current_code}: debe={debe_val}, haber={haber_val} (sin movimiento real)")
                continue
            
            # Loggear filas que se ignoran completamente
            elif current_code:
                logger.debug(f"⏭️  FILA {fila_numero} IGNORADA para cuenta {current_code}: sin datos de debe/haber válidos")
            else:
                logger.debug(f"⏭️  FILA {fila_numero} IGNORADA: sin cuenta activa definida")
            
            # Ignorar el resto de filas

        # 4. BULK-CREATE Y RESUMEN
        # Al final, bulk-create de aperturas y movimientos
        if aperturas:
            AperturaCuenta.objects.bulk_create(aperturas, batch_size=500)
        if movimientos:
            MovimientoContable.objects.bulk_create(movimientos, batch_size=500)
        
        # NUEVO: Resumen de movimientos por cuenta
        logger.info("="*80)
        logger.info("📊 RESUMEN DE MOVIMIENTOS PROCESADOS POR CUENTA")
        logger.info("="*80)
        
        for cuenta_codigo, cantidad in sorted(movimientos_por_cuenta.items()):
            cuenta_obj = processed_accounts.get(cuenta_codigo)
            nombre_cuenta = cuenta_obj.nombre if cuenta_obj else "N/A"
            logger.info(f"  {cuenta_codigo} | {nombre_cuenta} | {cantidad} movimientos procesados")
        
        logger.info(f"TOTAL: {len(movimientos_por_cuenta)} cuentas con movimientos | {sum(movimientos_por_cuenta.values())} movimientos totales")
        logger.info("="*80)

    # 5. CONVERTIR SETS AGRUPADOS A INCIDENCIAS INDIVIDUALES
    # Convertir cuentas sin clasificación por set específico
    for set_id, cuentas_por_set in cuentas_sin_clasificacion_por_set.items():
        for cuenta_codigo, set_nombre in cuentas_por_set.items():
            incidencias_pendientes.append({
                'tipo': 'cuenta_no_clasificada',
                'cuenta_codigo': cuenta_codigo,
                'set_clasificacion_id': set_id,
                'set_clasificacion_nombre': set_nombre,
                'descripcion': f"Cuenta {cuenta_codigo} sin clasificación en set '{set_nombre}'"
            })
    
    # Convertir cuentas sin nombre en inglés
    for cuenta_codigo in cuentas_sin_nombre_ingles:
        incidencias_pendientes.append({
            'tipo': 'cuenta_sin_ingles',
            'cuenta_codigo': cuenta_codigo,
            'descripcion': f"Cuenta {cuenta_codigo} sin nombre en inglés"
        })
    
    # Convertir cuentas con tipo documento null
    for cuenta_codigo in cuentas_con_tipo_doc_null:
        incidencias_pendientes.append({
            'tipo': 'movimiento_tipo_doc_null',
            'cuenta_codigo': cuenta_codigo,
            'descripcion': f"Cuenta {cuenta_codigo} tiene movimientos sin código de tipo documento"
        })
    
    # Convertir cuentas con tipo documento no reconocido
    for cuenta_codigo, tipo_doc_codigo in cuentas_con_tipo_doc_no_reconocido.items():
        incidencias_pendientes.append({
            'tipo': 'movimiento_tipo_doc_no_reconocido',
            'cuenta_codigo': cuenta_codigo,
            'tipo_doc_codigo': tipo_doc_codigo,
            'descripcion': f"Cuenta {cuenta_codigo} tiene tipo documento '{tipo_doc_codigo}' no reconocido"
        })
    
    # Actualizar estadísticas
    stats["incidencias_detectadas"] = len(incidencias_pendientes)
    stats["excepciones_aplicadas"] = {
        "clasificacion_sets": sum(len(excepciones) for excepciones in excepciones_por_set.values()),
        "nombres_ingles": len(excepciones_nombres_ingles),
        "tipo_doc_null": len(excepciones_tipo_doc_null),
        "tipo_doc_no_reconocido": len(excepciones_tipo_doc_no_reconocido)
    }

    # Calcular balances ESF/ERI + NO_CLASIFICADAS
    balance_esf = float(totales_esf_eri['ESF']['saldo_ant'] + totales_esf_eri['ESF']['debe'] - totales_esf_eri['ESF']['haber'])
    balance_eri = float(totales_esf_eri['ERI']['saldo_ant'] + totales_esf_eri['ERI']['debe'] - totales_esf_eri['ERI']['haber'])
    balance_no_clasificadas = float(totales_esf_eri['NO_CLASIFICADAS']['saldo_ant'] + totales_esf_eri['NO_CLASIFICADAS']['debe'] - totales_esf_eri['NO_CLASIFICADAS']['haber'])
    balance_total = balance_esf + balance_eri + balance_no_clasificadas

    # LOGGING DETALLADO DE TOTALES FINALES
    logger.info("="*80)
    logger.info("📊 RESUMEN FINAL DE TOTALES ESF/ERI + NO_CLASIFICADAS")
    logger.info("="*80)
    logger.info(f"CONTADORES DE CUENTAS:")
    logger.info(f"  ESF: {contadores_clasificacion['ESF']} cuentas")
    logger.info(f"  ERI: {contadores_clasificacion['ERI']} cuentas")
    logger.info(f"  NO_CLASIFICADAS: {contadores_clasificacion['NO_CLASIFICADAS']} cuentas")
    logger.info(f"  Sin clasificación (legacy): {contadores_clasificacion['Sin_clasificacion']} cuentas")
    logger.info("")
    logger.info(f"ESF TOTALES:")
    logger.info(f"  Saldo Anterior: ${totales_esf_eri['ESF']['saldo_ant']:,.2f}")
    logger.info(f"  Debe:           ${totales_esf_eri['ESF']['debe']:,.2f}")
    logger.info(f"  Haber:          ${totales_esf_eri['ESF']['haber']:,.2f}")
    logger.info(f"  Balance:        ${balance_esf:,.2f}")
    logger.info("")
    logger.info(f"ERI TOTALES:")
    logger.info(f"  Saldo Anterior: ${totales_esf_eri['ERI']['saldo_ant']:,.2f}")
    logger.info(f"  Debe:           ${totales_esf_eri['ERI']['debe']:,.2f}")
    logger.info(f"  Haber:          ${totales_esf_eri['ERI']['haber']:,.2f}")
    logger.info(f"  Balance:        ${balance_eri:,.2f}")
    logger.info("")
    logger.info(f"NO_CLASIFICADAS TOTALES:")
    logger.info(f"  Saldo Anterior: ${totales_esf_eri['NO_CLASIFICADAS']['saldo_ant']:,.2f}")
    logger.info(f"  Debe:           ${totales_esf_eri['NO_CLASIFICADAS']['debe']:,.2f}")
    logger.info(f"  Haber:          ${totales_esf_eri['NO_CLASIFICADAS']['haber']:,.2f}")
    logger.info(f"  Balance:        ${balance_no_clasificadas:,.2f}")
    logger.info("")
    logger.info(f"BALANCE TOTAL (ESF + ERI + NO_CLASIFICADAS): ${balance_total:,.2f}")
    logger.info("="*80)
    
    # NUEVO: DETALLAR TODAS LAS CUENTAS ESF Y ERI PROCESADAS
    logger.info("🏛️  DETALLE DE CUENTAS ESF (Estado de Situación Financiera):")
    logger.info("="*80)
    cuentas_esf_procesadas = []
    cuentas_eri_procesadas = []
    
    # Recorrer todas las cuentas procesadas y clasificarlas
    for codigo, cuenta_obj in processed_accounts.items():
        clasificacion = identificar_clasificacion_esf_eri(cuenta_obj)
        if clasificacion == 'ESF':
            # Buscar el saldo anterior de esta cuenta
            for apertura in aperturas:
                if apertura.cuenta.codigo == codigo:
                    cuentas_esf_procesadas.append({
                        'codigo': codigo,
                        'nombre': cuenta_obj.nombre,
                        'saldo_anterior': apertura.saldo_anterior
                    })
                    logger.info(f"  {codigo} | {cuenta_obj.nombre} | Saldo: ${apertura.saldo_anterior:,.2f}")
                    break
        elif clasificacion == 'ERI':
            # Buscar el saldo anterior de esta cuenta
            for apertura in aperturas:
                if apertura.cuenta.codigo == codigo:
                    cuentas_eri_procesadas.append({
                        'codigo': codigo,
                        'nombre': cuenta_obj.nombre,
                        'saldo_anterior': apertura.saldo_anterior
                    })
                    break
    
    logger.info(f"TOTAL CUENTAS ESF PROCESADAS: {len(cuentas_esf_procesadas)}")
    logger.info("="*80)
    
    logger.info("💼 DETALLE DE CUENTAS ERI (Estado de Resultados Integrales):")
    logger.info("="*80)
    for cuenta in cuentas_eri_procesadas:
        logger.info(f"  {cuenta['codigo']} | {cuenta['nombre']} | Saldo: ${cuenta['saldo_anterior']:,.2f}")
    
    logger.info(f"TOTAL CUENTAS ERI PROCESADAS: {len(cuentas_eri_procesadas)}")
    logger.info("="*80)

    # Guardar stats en upload.resumen
    resumen = upload.resumen or {}
    resumen["procesamiento"] = stats
    resumen["incidencias_pendientes"] = incidencias_pendientes  # Guardar para el siguiente task
    
    # NUEVO: Guardar totales ESF/ERI + NO_CLASIFICADAS calculados durante el procesamiento
    resumen['totales_esf_eri'] = {
        'totales': {
            'ESF': {
                'saldo_ant': float(totales_esf_eri['ESF']['saldo_ant']),
                'debe': float(totales_esf_eri['ESF']['debe']),
                'haber': float(totales_esf_eri['ESF']['haber'])
            },
            'ERI': {
                'saldo_ant': float(totales_esf_eri['ERI']['saldo_ant']),
                'debe': float(totales_esf_eri['ERI']['debe']),
                'haber': float(totales_esf_eri['ERI']['haber'])
            },
            'NO_CLASIFICADAS': {
                'saldo_ant': float(totales_esf_eri['NO_CLASIFICADAS']['saldo_ant']),
                'debe': float(totales_esf_eri['NO_CLASIFICADAS']['debe']),
                'haber': float(totales_esf_eri['NO_CLASIFICADAS']['haber'])
            }
        },
        'balance_esf': balance_esf,
        'balance_eri': balance_eri,
        'balance_no_clasificadas': balance_no_clasificadas,
        'balance_total': balance_total,
        'balance_validado': abs(balance_total) <= 1000.00,
        'diferencia_balance': abs(balance_total),
        'explicacion_balance': "Balance total debe ser cercano a $0.00 (contabilidad balanceada)"
    }
    
    upload.resumen = resumen
    upload.tiempo_procesamiento = timezone.now() - inicio
    upload.save(update_fields=["resumen", "tiempo_procesamiento"])

    logger.info(f"Procesamiento completado: {stats}")
    logger.info(f"Incidencias detectadas en línea: {len(incidencias_pendientes)}")
    logger.info(f"Totales calculados - ESF: {balance_esf:,.2f}, ERI: {balance_eri:,.2f}, NO_CLASIFICADAS: {balance_no_clasificadas:,.2f}, Total: {balance_total:,.2f}")
    logger.info(f"Excepciones aplicadas - Sets clasificación: {stats['excepciones_aplicadas']['clasificacion_sets']}, "
               f"Nombres inglés: {stats['excepciones_aplicadas']['nombres_ingles']}, "
               f"Tipo doc null: {stats['excepciones_aplicadas']['tipo_doc_null']}, "
               f"Tipo doc no reconocido: {stats['excepciones_aplicadas']['tipo_doc_no_reconocido']}")
    return upload_log_id

# ─── Task 5: Generar incidencias ───────────────────────────────────────────────

@shared_task
def generar_incidencias_libro_mayor(upload_log_id, user_correo_bdo):
    User = get_user_model()
    try:
        creador = User.objects.get(correo_bdo=user_correo_bdo)
    except User.DoesNotExist:
        creador = None

    upload_log = UploadLog.objects.get(id=upload_log_id)
    cierre = upload_log.cierre

    print("\n" + "🔄 INICIANDO GENERACIÓN DE INCIDENCIAS LIBRO MAYOR")
    print("="*80)
    print(f"📄 Upload Log ID: {upload_log_id}")
    print(f"🏢 Cliente: {upload_log.cliente}")
    print(f"📅 Cierre: {cierre}")
    print("="*80)

    # Obtener totales ESF/ERI ya calculados durante el procesamiento
    totales_esf_eri = upload_log.resumen.get('totales_esf_eri', {})
    
    if not totales_esf_eri:
        print("⚠️  WARNING: No se encontraron totales ESF/ERI precalculados")
        logger.warning("No se encontraron totales ESF/ERI en el resumen del upload_log")
        # Crear estructura vacía para evitar errores
        totales_esf_eri = {
            'totales': {
                'ESF': {'saldo_ant': 0, 'debe': 0, 'haber': 0},
                'ERI': {'saldo_ant': 0, 'debe': 0, 'haber': 0}
            },
            'balance_esf': 0,
            'balance_eri': 0,
            'balance_total': 0,
            'balance_validado': True
        }
    else:
        print(f"✅ Totales ESF/ERI obtenidos del procesamiento:")
        print(f"   ESF: Saldo=${totales_esf_eri['totales']['ESF']['saldo_ant']:,.2f} | "
              f"Debe=${totales_esf_eri['totales']['ESF']['debe']:,.2f} | "
              f"Haber=${totales_esf_eri['totales']['ESF']['haber']:,.2f}")
        print(f"   ERI: Saldo=${totales_esf_eri['totales']['ERI']['saldo_ant']:,.2f} | "
              f"Debe=${totales_esf_eri['totales']['ERI']['debe']:,.2f} | "
              f"Haber=${totales_esf_eri['totales']['ERI']['haber']:,.2f}")
        print(f"   Balance total: ${totales_esf_eri['balance_total']:,.2f}")

    # Limpiar incidencias anteriores
    Incidencia.objects.filter(cierre=cierre).delete()
    creadas = existentes = 0

    # 1. CREAR INCIDENCIAS DETECTADAS EN EL PROCESAMIENTO (Task 4)
    incidencias_pendientes = upload_log.resumen.get("incidencias_pendientes", [])
    logger.info(f"Creando {len(incidencias_pendientes)} incidencias detectadas durante el procesamiento")
    
    incidencias_a_crear = []
    for inc_data in incidencias_pendientes:
        if inc_data['tipo'] == 'cuenta_no_clasificada':
            incidencias_a_crear.append(Incidencia(
                cierre=cierre,
                tipo=Incidencia.CUENTA_NO_CLASIFICADA,
                cuenta_codigo=inc_data['cuenta_codigo'],
                set_clasificacion_id=inc_data.get('set_clasificacion_id'),
                set_clasificacion_nombre=inc_data.get('set_clasificacion_nombre'),
                descripcion=inc_data['descripcion'],
                creada_por=creador
            ))
        elif inc_data['tipo'] == 'cuenta_sin_ingles':
            incidencias_a_crear.append(Incidencia(
                cierre=cierre,
                tipo=Incidencia.CUENTA_SIN_INGLES,
                cuenta_codigo=inc_data['cuenta_codigo'],
                descripcion=inc_data['descripcion'],
                creada_por=creador
            ))
        elif inc_data['tipo'] == 'movimiento_tipo_doc_null':
            incidencias_a_crear.append(Incidencia(
                cierre=cierre,
                tipo=Incidencia.DOC_NULL,
                cuenta_codigo=inc_data['cuenta_codigo'],
                descripcion=inc_data['descripcion'],
                creada_por=creador
            ))
        elif inc_data['tipo'] == 'movimiento_tipo_doc_no_reconocido':
            incidencias_a_crear.append(Incidencia(
                cierre=cierre,
                tipo=Incidencia.DOC_NO_RECONOCIDO,
                cuenta_codigo=inc_data['cuenta_codigo'],
                tipo_doc_codigo=inc_data.get('tipo_doc_codigo'),
                descripcion=inc_data['descripcion'],
                creada_por=creador
            ))
    
    # Bulk create de TODAS las incidencias detectadas en procesamiento
    if incidencias_a_crear:
        Incidencia.objects.bulk_create(incidencias_a_crear, batch_size=500)
        creadas += len(incidencias_a_crear)
        logger.info(f"Creadas {len(incidencias_a_crear)} incidencias desde procesamiento en línea")

    # 2. SINCRONIZACIÓN FINAL DE CLASIFICACIONES TEMPORALES
    try:
        # Verificar si hay clasificaciones temporales (sin FK a cuenta) para el cliente
        clasificaciones_temporales = AccountClassification.objects.filter(
            cliente=upload_log.cliente,
            cuenta__isnull=True  # Solo las temporales sin FK
        )
        
        temporales_count = clasificaciones_temporales.count()
        logger.info(f"Clasificaciones temporales disponibles para sincronización: {temporales_count} para cliente {upload_log.cliente.id}")
        
        if temporales_count > 0:
            # Sincronizar clasificaciones temporales con cuentas reales creadas
            sincronizadas = 0
            for clasif_temp in clasificaciones_temporales:
                try:
                    cuenta_real = CuentaContable.objects.get(
                        cliente=upload_log.cliente,
                        codigo=clasif_temp.cuenta_codigo
                    )
                    clasif_temp.cuenta = cuenta_real
                    clasif_temp.save(update_fields=['cuenta'])
                    sincronizadas += 1
                except CuentaContable.DoesNotExist:
                    logger.debug(f"Cuenta {clasif_temp.cuenta_codigo} no existe aún para sincronizar clasificación temporal")
                    continue
            
            logger.info(f"Sincronizadas {sincronizadas} clasificaciones temporales con cuentas reales")
        
        # Verificar estado final de clasificaciones
        total_clasificaciones = AccountClassification.objects.filter(
            cliente=upload_log.cliente
        ).count()
        clasificaciones_con_fk = AccountClassification.objects.filter(
            cliente=upload_log.cliente,
            cuenta__isnull=False
        ).count()
        clasificaciones_temporales_restantes = total_clasificaciones - clasificaciones_con_fk
        
        logger.info(f"Estado final de clasificaciones - Total: {total_clasificaciones}, Con FK: {clasificaciones_con_fk}, Temporales: {clasificaciones_temporales_restantes}")
        
    except Exception as e:
        logger.error(f"Error en sincronización final de clasificaciones: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")

    # 3. VALIDAR BALANCE ESF/ERI - Usar totales ya calculados
    try:
        print("\n" + "="*80)
        print("🧮 VALIDACIÓN DE BALANCE ESF/ERI")
        print("="*80)
        
        # Extraer valores de balance ya calculados
        balance_esf = totales_esf_eri['balance_esf']
        balance_eri = totales_esf_eri['balance_eri']
        balance_total = totales_esf_eri['balance_total']
        
        # Mostrar totales detallados
        print(f"\n� TOTALES ACUMULADOS (calculados durante procesamiento):")
        print(f"   ESF (Estado de Situación Financiera):")
        print(f"      Saldo Anterior: ${totales_esf_eri['totales']['ESF']['saldo_ant']:,.2f}")
        print(f"      Debe:           ${totales_esf_eri['totales']['ESF']['debe']:,.2f}")
        print(f"      Haber:          ${totales_esf_eri['totales']['ESF']['haber']:,.2f}")
        
        print(f"\n   ERI (Estado de Resultados Integrales):")
        print(f"      Saldo Anterior: ${totales_esf_eri['totales']['ERI']['saldo_ant']:,.2f}")
        print(f"      Debe:           ${totales_esf_eri['totales']['ERI']['debe']:,.2f}")
        print(f"      Haber:          ${totales_esf_eri['totales']['ERI']['haber']:,.2f}")
        
        print(f"\n🔢 CÁLCULOS DE BALANCE:")
        print(f"   ESF = Saldo_Ant + Debe - Haber")
        print(f"   ESF = {totales_esf_eri['totales']['ESF']['saldo_ant']:,.2f} + {totales_esf_eri['totales']['ESF']['debe']:,.2f} - {totales_esf_eri['totales']['ESF']['haber']:,.2f}")
        print(f"   ESF = ${balance_esf:,.2f}")
        
        print(f"\n   ERI = Saldo_Ant + Debe - Haber")
        print(f"   ERI = {totales_esf_eri['totales']['ERI']['saldo_ant']:,.2f} + {totales_esf_eri['totales']['ERI']['debe']:,.2f} - {totales_esf_eri['totales']['ERI']['haber']:,.2f}")
        print(f"   ERI = ${balance_eri:,.2f}")
        
        print(f"\n⚖️  BALANCE TOTAL:")
        print(f"   Total = ESF + ERI")
        print(f"   Total = {balance_esf:,.2f} + {balance_eri:,.2f}")
        print(f"   Total = ${balance_total:,.2f}")
        
        # VALIDACIÓN CONCEPTUAL MEJORADA
        print(f"\n✅ ANÁLISIS DE BALANCE:")
        print(f"   Balance Total calculado: ${balance_total:,.2f}")
        print(f"   Saldo Anterior ESF: ${totales_esf_eri['totales']['ESF']['saldo_ant']:,.2f}")
        
        # VALIDACIÓN SIMPLIFICADA: Balance total debe ser cercano a 0 (contabilidad balanceada)
        diferencia_balance = abs(balance_total)
        
        # 🚧 BYPASS TEMPORAL: Descomenta la siguiente línea para bypasear validación de balance
        # diferencia_balance = 0  # BYPASS: Simula balance correcto para desarrollo
        
        if diferencia_balance <= 1000.00:  # Tolerancia más amplia para balances reales
            print(f"   ✅ BALANCE CORRECTO: Balance total = ${balance_total:,.2f} (contabilidad balanceada)")
            print(f"   📊 Interpretación: ESF + ERI ≈ 0 indica que los movimientos están balanceados")
            print(f"   🔍 Validación: Diferencia ${diferencia_balance:,.2f} dentro de tolerancia (±$1,000.00)")
            logger.info("✓ Balance ESF/ERI validado correctamente - contabilidad balanceada")
        else:
            print(f"   ❌ BALANCE DESCUADRADO")
            print(f"   📊 Balance total: ${balance_total:,.2f}")
            print(f"   📊 ESF: ${balance_esf:,.2f} | ERI: ${balance_eri:,.2f}")
            print(f"   📊 Diferencia absoluta: ${diferencia_balance:,.2f}")
            print(f"   📊 El balance debería ser cercano a $0.00 para una contabilidad balanceada")
            print(f"   📝 Creando incidencia de balance descuadrado...")
            
            # Crear descripción detallada con todos los datos del balance
            descripcion_detallada = f"""BALANCE DESCUADRADO - Análisis Completo:

📊 RESUMEN DEL BALANCE:
• Balance Total: ${balance_total:,.2f}
• Diferencia Absoluta: ${diferencia_balance:,.2f}
• Tolerancia Permitida: ±$1,000.00
• Estado: DESCUADRADO (diferencia excede tolerancia)

🏛️ DETALLE ESF (Estado de Situación Financiera):
• Saldo Anterior: ${totales_esf_eri['totales']['ESF']['saldo_ant']:,.2f}
• Debe Total: ${totales_esf_eri['totales']['ESF']['debe']:,.2f}
• Haber Total: ${totales_esf_eri['totales']['ESF']['haber']:,.2f}
• Balance ESF: ${balance_esf:,.2f}

💼 DETALLE ERI (Estado de Resultados Integrales):
• Saldo Anterior: ${totales_esf_eri['totales']['ERI']['saldo_ant']:,.2f}
• Debe Total: ${totales_esf_eri['totales']['ERI']['debe']:,.2f}
• Haber Total: ${totales_esf_eri['totales']['ERI']['haber']:,.2f}
• Balance ERI: ${balance_eri:,.2f}

🔢 CÁLCULO DEL BALANCE:
• ESF = Saldo_Ant + Debe - Haber = {totales_esf_eri['totales']['ESF']['saldo_ant']:,.2f} + {totales_esf_eri['totales']['ESF']['debe']:,.2f} - {totales_esf_eri['totales']['ESF']['haber']:,.2f} = ${balance_esf:,.2f}
• ERI = Saldo_Ant + Debe - Haber = {totales_esf_eri['totales']['ERI']['saldo_ant']:,.2f} + {totales_esf_eri['totales']['ERI']['debe']:,.2f} - {totales_esf_eri['totales']['ERI']['haber']:,.2f} = ${balance_eri:,.2f}
• Total = ESF + ERI = {balance_esf:,.2f} + {balance_eri:,.2f} = ${balance_total:,.2f}

⚠️ PROBLEMA DETECTADO:
El balance total debería ser cercano a $0.00 para indicar una contabilidad balanceada. La diferencia de ${diferencia_balance:,.2f} excede la tolerancia permitida.

🔧 ACCIONES SUGERIDAS:
1. Revisar las clasificaciones de cuentas ESF vs ERI
2. Verificar que todos los movimientos estén correctamente registrados
3. Confirmar que los saldos anteriores sean correctos
4. Revisar posibles errores en la carga del archivo Excel"""
            
            incidencia_balance = Incidencia(
                cierre=cierre,
                tipo=Incidencia.BALANCE_DESCUADRADO,
                descripcion=descripcion_detallada,
                creada_por=creador
            )
            incidencia_balance.save()
            creadas += 1
            logger.warning(f"BALANCE DESCUADRADO - ESF: {balance_esf}, ERI: {balance_eri}, Total: {balance_total}, Diferencia: {diferencia_balance}")
        
        logger.info(f"Balance ESF: {balance_esf}, Balance ERI: {balance_eri}, Balance total: {balance_total}")
        
        print("="*80)
            
    except Exception as e:
        print(f"\n❌ ERROR validando balance ESF/ERI: {e}")
        logger.error(f"Error validando balance ESF/ERI: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        # Usar valores por defecto en caso de error
        balance_esf = 0
        balance_eri = 0
        balance_total = 0
    
    # Guardar resumen (los totales ESF/ERI ya están calculados y guardados desde el procesamiento)
    resumen = upload_log.resumen or {}
    resumen['incidencias'] = {
        'creadas': creadas,
        'existentes': existentes,
        'total_bd': Incidencia.objects.filter(cierre=cierre).count(),
        'desde_procesamiento_en_linea': len(incidencias_pendientes),
        'clasificaciones': len([i for i in incidencias_pendientes if i['tipo'] == 'cuenta_no_clasificada']),
        'nombres_ingles': len([i for i in incidencias_pendientes if i['tipo'] == 'cuenta_sin_ingles']),
        'tipos_documento_null': len([i for i in incidencias_pendientes if i['tipo'] == 'movimiento_tipo_doc_null']),
        'tipos_documento_no_reconocido': len([i for i in incidencias_pendientes if i['tipo'] == 'movimiento_tipo_doc_no_reconocido']),
        'balance_descuadrado': 1 if abs(balance_total) > 1000.00 else 0
    }
    
    # Actualizar solo el estado de validación del balance en los totales existentes
    if 'totales_esf_eri' in resumen:
        # Validar solo si el balance total está cerca de 0 (contabilidad balanceada)
        diferencia_balance = abs(balance_total)
        balance_validado = diferencia_balance <= 1000.00
        resumen['totales_esf_eri']['balance_validado'] = balance_validado
        resumen['totales_esf_eri']['diferencia_balance'] = float(diferencia_balance)
        resumen['totales_esf_eri']['explicacion_balance'] = "Balance total debe ser cercano a $0.00 (contabilidad balanceada)"
    
    upload_log.resumen = resumen
    upload_log.save(update_fields=['resumen'])

    print("\n" + "📋 RESUMEN FINAL DE INCIDENCIAS")
    print("="*80)
    print(f"✅ Total incidencias creadas: {creadas}")
    print(f"📊 Desglose por tipo:")
    print(f"   - Clasificaciones: {resumen['incidencias']['clasificaciones']}")
    print(f"   - Nombres inglés: {resumen['incidencias']['nombres_ingles']}")
    print(f"   - Tipos doc null: {resumen['incidencias']['tipos_documento_null']}")
    print(f"   - Tipos doc no reconocido: {resumen['incidencias']['tipos_documento_no_reconocido']}")
    print(f"   - Balance descuadrado: {resumen['incidencias']['balance_descuadrado']}")
    
    print(f"\n💰 BALANCE ESF/ERI FINAL:")
    print(f"   - Balance total: ${balance_total:,.2f}")
    diferencia_balance = abs(balance_total)
    print(f"   - Diferencia absoluta: ${diferencia_balance:,.2f}")
    print(f"   - Balance validado: {'✅ SÍ' if diferencia_balance <= 1000.00 else '❌ NO'}")
    print("="*80)

    logger.info(f"Generación de incidencias completada - Total creadas: {creadas}")
    logger.info(f"Desglose: Clasificaciones: {resumen['incidencias']['clasificaciones']}, " +
               f"Nombres inglés: {resumen['incidencias']['nombres_ingles']}, " +
               f"Tipos doc null: {resumen['incidencias']['tipos_documento_null']}, " +
               f"Tipos doc no reconocido: {resumen['incidencias']['tipos_documento_no_reconocido']}, " +
               f"Balance descuadrado: {resumen['incidencias']['balance_descuadrado']}")
    logger.info(f"Balance ESF/ERI - Total: {balance_total}, Diferencia absoluta: {abs(balance_total)}, " +
               f"Validado: {abs(balance_total) <= 1000.00}")
    return upload_log_id

# ─── Task 6: Finalizar y limpiar ───────────────────────────────────────────────

@shared_task
def finalizar_procesamiento_libro_mayor(upload_log_id, user_correo_bdo):
    upload = UploadLog.objects.get(pk=upload_log_id)
    cierre = upload.cierre

    # NUEVO: Auto-incrementar iteración para libro mayor
    if upload.tipo_upload == 'libro_mayor' and upload.cierre:
        iteraciones_anteriores = UploadLog.objects.filter(
            cierre=upload.cierre,
            tipo_upload='libro_mayor',
            estado='completado'
        ).exclude(id=upload.id).count()
        
        upload.iteracion = iteraciones_anteriores + 1
        logger.info(f"Upload {upload.id} asignado iteración {upload.iteracion}")

    # marcar completado
    upload.estado = "completado"
    upload.save(update_fields=["estado", "iteracion"])

    # Crear el registro de LibroMayorArchivo para que el frontend pueda encontrarlo
    try:
        crear_registro_libro_mayor_archivo(upload_log_id)
        logger.info(f"Registro LibroMayorArchivo creado para upload_log {upload_log_id}")
    except Exception as e:
        logger.error(f"Error creando registro LibroMayorArchivo: {e}")

    # NUEVO: Crear snapshot de incidencias consolidadas
    try:
        crear_snapshot_incidencias_consolidadas(upload_log_id)
        logger.info(f"Snapshot de incidencias creado para upload_log {upload_log_id}")
    except Exception as e:
        logger.error(f"Error creando snapshot de incidencias: {e}")

    # post-procesos asíncronos - clasificaciones ya se procesaron en generar_incidencias
    # (se movió el mapeo de clasificaciones para que ocurra antes de generar incidencias)

    # Aplicar nombres en inglés pendientes
    try:
        aplicar_nombres_ingles_pendientes(upload_log_id)
        logger.info(f"Aplicados nombres en inglés pendientes para upload_log {upload_log_id}")
    except Exception as e:
        logger.error(f"Error aplicando nombres en inglés: {e}")

    # limpiar archivo temp
    try:
        path = default_storage.path(upload.ruta_archivo)
        if os.path.exists(path):
            os.remove(path)
    except Exception as e:
        logger.warning(f"No se eliminó temp: {e}")

    stats_proc = upload.resumen.get("procesamiento", {})
    stats_inc  = upload.resumen.get("incidencias", {})
    msg = (
        f"Movimientos: {stats_proc.get('movimientos',0)}, "
        f"Aperturas: {stats_proc.get('aperturas',0)}, "
        f"Incidencias: {stats_inc.get('creadas',0)}"
    )
    return msg


def aplicar_nombres_ingles_pendientes(upload_log_id):
    """
    Aplica nombres en inglés pendientes a las cuentas que no los tienen
    """
    try:
        upload_log = UploadLog.objects.get(id=upload_log_id)
        cliente = upload_log.cliente
        
        # Obtener mapeo de nombres en inglés
        nombres_ingles_map = {
            ni.cuenta_codigo: ni.nombre_ingles
            for ni in NombreIngles.objects.filter(cliente=cliente)
        }
        
        # Aplicar a cuentas que no tienen nombre en inglés
        cuentas_sin_nombre = CuentaContable.objects.filter(
            cliente=cliente,
            nombre_en__isnull=True,
            codigo__in=nombres_ingles_map.keys()
        )
        
        aplicados = 0
        for cuenta in cuentas_sin_nombre:
            if cuenta.codigo in nombres_ingles_map:
                cuenta.nombre_en = nombres_ingles_map[cuenta.codigo]
                cuenta.save(update_fields=['nombre_en'])
                aplicados += 1
        
        logger.info(f"Aplicados {aplicados} nombres en inglés pendientes para cliente {cliente.id}")
        
    except Exception as e:
        logger.error(f"Error en aplicar_nombres_ingles_pendientes: {e}")

@shared_task
def mapear_clasificaciones_desde_sets_existentes(cliente_id, cierre_id=None):
    """
    Intenta mapear cuentas a clasificaciones usando sets y opciones existentes
    cuando no hay archivos RAW disponibles. Útil para casos donde se tienen
    sets configurados pero se necesita aplicar a un nuevo conjunto de cuentas.
    """
    try:
        from .models import Cliente, CierreContabilidad, CuentaContable
        
        cliente = Cliente.objects.get(id=cliente_id)
        logger.info(f"Iniciando mapeo desde sets existentes para cliente {cliente_id}")
        
        # Buscar cierre
        if cierre_id:
            cierre = CierreContabilidad.objects.filter(id=cierre_id, cliente=cliente).first()
        else:
            cierre = CierreContabilidad.objects.filter(
                cliente=cliente, 
                estado__in=['pendiente', 'procesando', 'clasificacion', 'incidencias', 'en_revision']
            ).order_by('-fecha_creacion').first()
        
        if not cierre:
            return {'error': f'No hay cierre activo para cliente {cliente_id}'}
        
        # Obtener sets de clasificación
        sets_clasificacion = ClasificacionSet.objects.filter(cliente=cliente)
        if not sets_clasificacion.exists():
            return {'error': f'No hay sets de clasificación para cliente {cliente_id}'}
        
        # Obtener cuentas del cierre actual
        cuentas = CuentaContable.objects.filter(
            movimientocontable__cierre=cierre
        ).distinct()
        
        logger.info(f"Encontradas {cuentas.count()} cuentas en el cierre {cierre.id}")
        logger.info(f"Encontrados {sets_clasificacion.count()} sets de clasificación")
        
        # Aquí podrías implementar lógica de mapeo inteligente
        # Por ejemplo, mapear por patrones de códigos de cuenta, nombres, etc.
        # Por ahora, solo retornamos información para que el usuario tome decisiones
        
        return {
            'info': f'Análisis completado para cliente {cliente_id}',
            'total_cuentas': cuentas.count(),
            'total_sets': sets_clasificacion.count(),
            'sets_disponibles': list(sets_clasificacion.values_list('nombre', flat=True)),
            'mensaje': 'Para aplicar clasificaciones, suba un archivo de clasificaciones que mapee las cuentas actuales con los sets existentes.'
        }
        
    except Exception as e:
        logger.error(f"Error en mapeo desde sets existentes: {e}")
        return {'error': str(e)}


def crear_registro_libro_mayor_archivo(upload_log_id):
    """
    Crea el registro LibroMayorArchivo necesario para que el frontend
    pueda mostrar el estado correcto del libro mayor
    """
    from django.core.files.base import ContentFile
    
    upload_log = UploadLog.objects.get(id=upload_log_id)
    
    # Extraer el período del nombre del archivo
    rut_limpio = upload_log.cliente.rut.replace(".", "").replace("-", "") if upload_log.cliente.rut else str(upload_log.cliente.id)
    nombre_sin_ext = re.sub(r"\.(xlsx|xls)$", "", upload_log.nombre_archivo_original, flags=re.IGNORECASE)
    patron_periodo = rf"^{rut_limpio}_(LibroMayor|Mayor)_(\d{{6}})$"
    match = re.match(patron_periodo, nombre_sin_ext)
    periodo = match.group(2) if match else "000000"
    
    # Leer el archivo original para guardarlo en LibroMayorArchivo
    ruta_completa = default_storage.path(upload_log.ruta_archivo)
    if os.path.exists(ruta_completa):
        with open(ruta_completa, "rb") as f:
            contenido = f.read()
        
        nombre_final = f"libro_mayor_{upload_log.cliente.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        archivo_obj, created = LibroMayorArchivo.objects.get_or_create(
            cliente=upload_log.cliente,
            cierre=upload_log.cierre,
            defaults={
                "archivo": ContentFile(contenido, name=nombre_final),
                "estado": "completado",
                "procesado": True,
                "upload_log": upload_log,
                "periodo": periodo,
                "errores": "",
            },
        )
        
        if not created:
            # Si ya existe, actualizar el estado
            archivo_obj.estado = "completado"
            archivo_obj.procesado = True
            archivo_obj.upload_log = upload_log
            archivo_obj.errores = ""
            archivo_obj.save(update_fields=["estado", "procesado", "upload_log", "errores"])
        
        logger.info(f"LibroMayorArchivo {'creado' if created else 'actualizado'} para upload_log {upload_log_id}")
    else:
        logger.warning(f"Archivo temporal no encontrado para upload_log {upload_log_id}")

# ─── NUEVO: Snapshot de Incidencias ──────────────────────────────────────────

def crear_snapshot_incidencias_consolidadas(upload_log_id):
    """
    Crea un snapshot de las incidencias consolidadas y lo guarda en el resumen del UploadLog
    """
    import json
    from collections import defaultdict
    from django.utils import timezone
    
    try:
        upload_log = UploadLog.objects.get(id=upload_log_id)
        
        if not upload_log.cierre:
            logger.warning(f"UploadLog {upload_log_id} no tiene cierre asociado")
            return
        
        # Obtener incidencias del cierre
        incidencias_query = Incidencia.objects.filter(cierre_id=upload_log.cierre.id)
        
        if not incidencias_query.exists():
            # No hay incidencias - crear snapshot vacío
            snapshot = {
                'iteracion': upload_log.iteracion,
                'timestamp': timezone.now().isoformat(),
                'incidencias_detectadas': [],
                'estadisticas': {
                    'total_incidencias': 0,
                    'por_tipo': {}
                },
                'total_elementos_afectados': 0,
                'comparacion_anterior': None,
                'sin_incidencias': True,  # Flag para indicar que no hay incidencias
                'mensaje_sin_incidencias': 'Procesamiento completado sin incidencias detectadas'
            }
        else:
            # Agrupar incidencias por tipo (igual que en views/incidencias.py)
            incidencias_agrupadas = defaultdict(list)
            
            for inc in incidencias_query:
                incidencias_agrupadas[inc.tipo].append(inc)
            
            # Transformar a formato consolidado
            incidencias_consolidadas = []
            
            for tipo, incidencias_list in incidencias_agrupadas.items():
                cantidad_total = len(incidencias_list)
                
                # Determinar severidad
                if cantidad_total >= 50:
                    severidad = 'critica'
                elif cantidad_total >= 20:
                    severidad = 'alta'
                elif cantidad_total >= 5:
                    severidad = 'media'
                else:
                    severidad = 'baja'
                
                # Crear mensaje descriptivo
                tipo_display = dict(Incidencia.TIPO_CHOICES).get(tipo, tipo)
                
                if tipo == Incidencia.CUENTA_NO_CLASIFICADA:
                    mensaje_usuario = f"Se encontraron {cantidad_total} cuentas sin clasificación"
                    accion_sugerida = "Revisar y clasificar las cuentas pendientes en el sistema de clasificaciones"
                elif tipo == Incidencia.CUENTA_SIN_INGLES:
                    mensaje_usuario = f"Se encontraron {cantidad_total} cuentas sin traducción al inglés"
                    accion_sugerida = "Completar las traducciones faltantes en la sección de Nombres en Inglés"
                elif tipo == Incidencia.DOC_NO_RECONOCIDO:
                    mensaje_usuario = f"Se encontraron {cantidad_total} documentos con tipos no reconocidos"
                    accion_sugerida = "Revisar y agregar los tipos de documento faltantes"
                elif tipo == Incidencia.DOC_NULL:
                    mensaje_usuario = f"Se encontraron {cantidad_total} movimientos sin tipo de documento"
                    accion_sugerida = "Corregir los movimientos que no tienen tipo de documento asignado"
                else:
                    mensaje_usuario = f"Se encontraron {cantidad_total} incidencias de tipo {tipo_display}"
                    accion_sugerida = "Revisar y corregir las incidencias detectadas"
                
                # Recopilar elementos afectados (todos para el modal)
                elementos_afectados = []
                for inc in incidencias_list:  # Incluir TODOS los elementos
                    if inc.cuenta_codigo:
                        elemento = {
                            'tipo': 'cuenta',
                            'codigo': inc.cuenta_codigo,
                            'descripcion': inc.descripcion or ''
                        }
                        # Para incidencias de clasificación, incluir información del set específico
                        if tipo == Incidencia.CUENTA_NO_CLASIFICADA and inc.set_clasificacion_id:
                            elemento['set_id'] = inc.set_clasificacion_id
                            elemento['set_nombre'] = inc.set_clasificacion_nombre or f'Set ID {inc.set_clasificacion_id}'
                        
                        elementos_afectados.append(elemento)
                    if inc.tipo_doc_codigo:
                        elementos_afectados.append({
                            'tipo': 'documento',
                            'codigo': inc.tipo_doc_codigo,
                            'descripcion': inc.descripcion or ''
                        })
                
                incidencias_consolidadas.append({
                    'id': f"consolidada_{tipo}_{upload_log.cierre.id}",
                    'tipo_incidencia': tipo,
                    'tipo_codigo': tipo,
                    'codigo_problema': tipo,
                    'cantidad_afectada': cantidad_total,
                    'severidad': severidad,
                    'severidad_codigo': severidad,
                    'estado': 'activa',
                    'estado_codigo': 'activa',
                    'mensaje_usuario': mensaje_usuario,
                    'accion_sugerida': accion_sugerida,
                    'elementos_afectados': elementos_afectados,
                    'detalle_muestra': {
                        'primeros_ejemplos': [
                            {
                                'cuenta_codigo': inc.cuenta_codigo,
                                'tipo_doc_codigo': inc.tipo_doc_codigo,
                                'descripcion': inc.descripcion
                            } for inc in incidencias_list[:5]
                        ]
                    },
                    'fecha_deteccion': min(inc.fecha_creacion for inc in incidencias_list).isoformat(),
                    'fecha_resolucion': None,
                    'resuelto_por': None,
                })
            
            # Ordenar por severidad y cantidad
            severidad_order = {'critica': 0, 'alta': 1, 'media': 2, 'baja': 3}
            incidencias_consolidadas.sort(
                key=lambda x: (severidad_order.get(x['severidad'], 4), -x['cantidad_afectada'])
            )
            
            # Estadísticas
            estadisticas = {
                'total_incidencias': len(incidencias_consolidadas),
                'por_tipo': {}
            }
            
            for inc in incidencias_consolidadas:
                tipo = inc['tipo_incidencia']
                estadisticas['por_tipo'][tipo] = estadisticas['por_tipo'].get(tipo, 0) + 1
            
            # Comparar con iteración anterior si existe
            comparacion_anterior = comparar_con_iteracion_anterior(upload_log, incidencias_consolidadas)
            
            snapshot = {
                'iteracion': upload_log.iteracion,
                'timestamp': timezone.now().isoformat(),
                'incidencias_detectadas': incidencias_consolidadas,
                'estadisticas': estadisticas,
                'total_elementos_afectados': sum(inc['cantidad_afectada'] for inc in incidencias_consolidadas),
                'comparacion_anterior': comparacion_anterior,
                'sin_incidencias': False,  # Flag para indicar que SÍ hay incidencias
                'mensaje_sin_incidencias': None
            }
        
        # Guardar snapshot en el resumen del UploadLog
        if not upload_log.resumen:
            upload_log.resumen = {}
        
        upload_log.resumen['incidencias_snapshot'] = snapshot
        upload_log.save(update_fields=['resumen'])
        
        logger.info(f"Snapshot de incidencias creado para iteración {upload_log.iteracion} "
                   f"con {len(snapshot['incidencias_detectadas'])} tipos de incidencias")
        
        return snapshot
        
    except Exception as e:
        logger.error(f"Error creando snapshot de incidencias para upload_log {upload_log_id}: {e}")
        raise


def comparar_con_iteracion_anterior(upload_log_actual, incidencias_actuales):
    """
    Compara las incidencias actuales con la iteración anterior
    """
    try:
        if upload_log_actual.iteracion <= 1:
            return {'es_primera_iteracion': True}
        
        upload_anterior = UploadLog.objects.filter(
            cierre=upload_log_actual.cierre,
            tipo_upload='libro_mayor',
            iteracion=upload_log_actual.iteracion - 1,
            estado='completado'
        ).first()
        
        if not upload_anterior or not upload_anterior.resumen:
            return {'anterior_no_encontrada': True}
        
        snapshot_anterior = upload_anterior.resumen.get('incidencias_snapshot')
        if not snapshot_anterior:
            return {'snapshot_anterior_no_encontrado': True}
        
        incidencias_anteriores = snapshot_anterior.get('incidencias_detectadas', [])
        
        # Comparación simple por tipo
        tipos_anteriores = {inc['tipo_codigo']: inc['cantidad_afectada'] for inc in incidencias_anteriores}
        tipos_actuales = {inc['tipo_codigo']: inc['cantidad_afectada'] for inc in incidencias_actuales}
        
        cambios = {
            'resueltas': [],
            'nuevas': [],
            'empeoradas': [],
            'mejoradas': []
        }
        
        # Tipos que mejoraron o empeoraron
        for tipo, cantidad_actual in tipos_actuales.items():
            cantidad_anterior = tipos_anteriores.get(tipo, 0)
            
            if cantidad_anterior > cantidad_actual:
                cambios['mejoradas'].append({
                    'tipo': tipo,
                    'anterior': cantidad_anterior,
                    'actual': cantidad_actual,
                    'diferencia': cantidad_anterior - cantidad_actual
                })
            elif cantidad_anterior < cantidad_actual:
                cambios['empeoradas'].append({
                    'tipo': tipo,
                    'anterior': cantidad_anterior,
                    'actual': cantidad_actual,
                    'diferencia': cantidad_actual - cantidad_anterior
                })
        
        # Tipos que aparecieron por primera vez
        for tipo, cantidad_actual in tipos_actuales.items():
            if tipo not in tipos_anteriores:
                cambios['nuevas'].append({
                    'tipo': tipo,
                    'cantidad': cantidad_actual
                })
        
        # Tipos que desaparecieron completamente
        for tipo, cantidad_anterior in tipos_anteriores.items():
            if tipo not in tipos_actuales:
                cambios['resueltas'].append({
                    'tipo': tipo,
                    'cantidad_resuelta': cantidad_anterior
                })
        
        return cambios
        
    except Exception as e:
        logger.error(f"Error comparando con iteración anterior: {e}")
        return {'error_comparacion': str(e)}
