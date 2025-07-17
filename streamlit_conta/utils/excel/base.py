"""
Clase base para templates Excel con estilos y métodos comunes
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import io
from datetime import datetime
import logging

from .translations import get_text, get_account_name

logger = logging.getLogger(__name__)


class BaseExcelTemplate:
    """Clase base para templates Excel con estilos y métodos comunes"""
    
    def __init__(self):
        # Estilos predefinidos
        self.styles = {
            'title': Font(name='Arial', size=16, bold=True, color='FFFFFF'),
            'header': Font(name='Arial', size=12, bold=True, color='FFFFFF'),
            'subheader': Font(name='Arial', size=11, bold=True, color='000000'),
            'data': Font(name='Arial', size=10, color='000000'),
            'total': Font(name='Arial', size=11, bold=True, color='FFFFFF'),
            'metadata': Font(name='Arial', size=9, italic=True, color='666666')
        }
        
        self.fills = {
            'title': PatternFill(start_color='0A58CA', end_color='0A58CA', fill_type='solid'),
            'header': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
            'subheader': PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid'),
            'total': PatternFill(start_color='2F5233', end_color='2F5233', fill_type='solid'),
            'alternate': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
            'initial_balance': PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid'),  # Verde claro para saldo inicial
            'final_balance': PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')     # Verde oscuro para saldo final
        }
        
        self.borders = {
            'thin': Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin')
            ),
            'thick': Border(
                left=Side(border_style='thick'),
                right=Side(border_style='thick'),
                top=Side(border_style='thick'),
                bottom=Side(border_style='thick')
            )
        }

    def _get_text(self, key, language='es', metadata=None):
        """Obtener texto traducido según el idioma, con soporte para textos dinámicos"""
        return get_text(key, language, metadata)

    def _get_account_name(self, cuenta, language='es'):
        """Obtener nombre de cuenta según el idioma"""
        return get_account_name(cuenta, language)

    def _apply_header_style(self, worksheet, start_row, start_col, end_col, title=""):
        """Aplicar estilo de encabezado a un rango de celdas"""
        if title:
            cell = worksheet.cell(row=start_row, column=start_col)
            cell.value = title
            cell.font = self.styles['title']
            cell.fill = self.fills['title']
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Combinar celdas para el título
            if end_col > start_col:
                worksheet.merge_cells(
                    start_row=start_row, start_column=start_col,
                    end_row=start_row, end_column=end_col
                )

    def _apply_data_formatting(self, worksheet, start_row, end_row, start_col, end_col):
        """Aplicar formato a los datos"""
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = self.borders['thin']
                cell.font = self.styles['data']
                
                # Alternar colores de fila
                if row % 2 == 0:
                    cell.fill = self.fills['alternate']

    def _add_metadata_sheet(self, workbook, metadata, language='es'):
        """Agregar hoja con metadatos del informe"""
        ws = workbook.create_sheet(self._get_text('info_sheet', language, metadata))
        
        # Título
        title_text = self._get_text('report_info', language, metadata)
        ws.cell(row=1, column=1, value=title_text)
        ws.cell(row=1, column=1).font = self.styles['title']
        ws.cell(row=1, column=1).fill = self.fills['title']
        ws.merge_cells('A1:B1')
        
        # Metadatos
        info_data = [
            (f"{self._get_text('client', language)}:", metadata.get('cliente_nombre', 'N/A')),
            (f"{self._get_text('period', language)}:", metadata.get('periodo', 'N/A')),
            (f"{self._get_text('currency', language)}:", metadata.get('moneda', 'CLP')),
            (f"{self._get_text('language', language)}:", metadata.get('idioma', self._get_text('spanish', language))),
            (f"{self._get_text('generation_date', language)}:", datetime.now().strftime("%d/%m/%Y %H:%M")),
            (f"{self._get_text('system', language)}:", "SGM Dashboard Contable"),
            (f"{self._get_text('version', language)}:", "v1.0")
        ]
        
        for i, (label, value) in enumerate(info_data, start=3):
            ws.cell(row=i, column=1, value=label).font = self.styles['subheader']
            ws.cell(row=i, column=2, value=value).font = self.styles['data']
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30

    def _add_period_info(self, ws, current_row, metadata, language='es'):
        """Agregar información del período al worksheet"""
        ws.cell(row=current_row, column=1, value=f"{self._get_text('client', language)}: {metadata.get('cliente_nombre', 'N/A')}")
        ws.cell(row=current_row, column=3, value=f"{self._get_text('period', language)}: {metadata.get('periodo', 'N/A')}")
        current_row += 1
        ws.cell(row=current_row, column=1, value=f"{self._get_text('currency', language)}: {metadata.get('moneda', 'CLP')}")
        ws.cell(row=current_row, column=3, value=f"{self._get_text('date', language)}: {datetime.now().strftime('%d/%m/%Y')}")
        return current_row + 2

    def _format_amount(self, amount, moneda="CLP"):
        """Formatear monto según la moneda (sin sufijo de moneda)"""
        if pd.isna(amount) or amount is None:
            return 0
        
        try:
            amount = float(amount)
        except (ValueError, TypeError):
            return 0
        
        if moneda == "USD":
            return f"${amount:,.2f}"
        elif moneda == "EUR":
            return f"€{amount:,.2f}"
        else:  # CLP por defecto
            return f"${amount:,.0f}"

    def _debug_data_structure(self, data, section_name=""):
        """Debug helper para verificar estructura de datos"""
        logger.info(f"=== DEBUG {section_name} ===")
        if isinstance(data, dict):
            for key, value in data.items():
                if isinstance(value, dict):
                    logger.info(f"{key}: dict con {len(value)} elementos")
                    if "total" in value:
                        logger.info(f"  - total: {value['total']}")
                    if "grupos" in value:
                        logger.info(f"  - grupos: {len(value['grupos'])} grupos")
                        for grupo_name in value['grupos'].keys():
                            logger.info(f"    * {grupo_name}")
                    if "cuentas" in value:
                        logger.info(f"  - cuentas: {len(value['cuentas'])} cuentas")
                elif isinstance(value, list):
                    logger.info(f"{key}: lista con {len(value)} elementos")
                else:
                    logger.info(f"{key}: {type(value)} = {value}")
        else:
            logger.info(f"Data type: {type(data)} = {data}")
        logger.info(f"=== FIN DEBUG {section_name} ===")

    def workbook_to_bytes(self, workbook):
        """Convertir workbook a bytes para descarga"""
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _create_group_collapsible(self, ws, grupo_start_row, current_row):
        """Crear grupo colapsable en Excel"""
        if current_row > grupo_start_row:
            try:
                # Crear grupo colapsable desde grupo_start_row hasta current_row-1
                ws.row_dimensions.group(grupo_start_row, current_row - 1, outline_level=1)
                # Por defecto, dejar el grupo colapsado
                for row_num in range(grupo_start_row, current_row):
                    ws.row_dimensions[row_num].hidden = True
            except Exception as e:
                # Si falla la agrupación, continuar sin ella
                logger.warning(f"No se pudo crear grupo colapsable: {e}")
                pass
