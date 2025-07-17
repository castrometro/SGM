"""
Template Excel específico para Estado de Cambios en el Patrimonio (ECP)
"""

import openpyxl
import logging
from datetime import datetime
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from .base import BaseExcelTemplate
from .translations import detect_language_improved

logger = logging.getLogger(__name__)


class ECPTemplate(BaseExcelTemplate):
    """Template Excel para Estado de Cambios en el Patrimonio"""

    def generate(self, data_ecp, metadata, data_eri=None):
        """Generar template Excel para Estado de Cambios en el Patrimonio"""
        # Obtener idioma usando detección mejorada
        language = detect_language_improved(metadata)
        
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = self._get_text('title_ecp', language, metadata)
        
        # Título principal
        title_text = self._get_text('title_ecp', language, metadata)
        self._apply_header_style(ws, 1, 1, 7, title_text)
        
        # Información del período
        current_row = 3
        ws.cell(row=current_row, column=1, value=f"{self._get_text('client', language, metadata)}: {metadata.get('cliente_nombre', 'N/A')}")
        ws.cell(row=current_row, column=5, value=f"{self._get_text('period', language, metadata)}: {metadata.get('periodo', 'N/A')}")
        current_row += 1
        ws.cell(row=current_row, column=1, value=f"{self._get_text('currency', language, metadata)}: {metadata.get('moneda', 'CLP')}")
        ws.cell(row=current_row, column=5, value=f"{self._get_text('date', language, metadata)}: {datetime.now().strftime('%d/%m/%Y')}")
        current_row += 2
        
        # Encabezados de columnas
        headers = [
            self._get_text('concept', language, metadata),
            self._get_text('capital', language, metadata),
            self._get_text('other_reserves', language, metadata),
            self._get_text('retained_earnings', language, metadata),
            self._get_text('attributable_capital', language, metadata),
            self._get_text('non_controlling_interests', language, metadata),
            self._get_text('total', language, metadata)
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = self.styles['header']
            cell.fill = self.fills['header']
            cell.alignment = Alignment(horizontal='center')
        
        current_row += 1
        
        if data_ecp:
            current_row = self._add_ecp_data(ws, current_row, data_ecp, data_eri, metadata, language)
        
        # Ajustar anchos de columna
        for col_idx in range(1, 8):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 18
        
        # Agregar hoja de metadatos
        self._add_metadata_sheet(workbook, metadata, language)
        
        return workbook

    def _add_ecp_data(self, ws, current_row, data_ecp, data_eri, metadata, language):
        """Agregar datos del ECP al worksheet"""
        # DEBUG: Verificar estructura de datos
        logger.info(f"=== DEBUG ECP DATA ===")
        logger.info(f"data_ecp: {data_ecp}")
        logger.info(f"data_eri: {data_eri}")
        
        # Procesar datos del ECP - manejar diferentes estructuras posibles (igual que excel_templates)
        patrimonio_data = data_ecp.get("patrimonio", data_ecp)  # Puede estar directamente en data_ecp
        
        # Extraer datos de Capital
        capital_data = patrimonio_data.get("capital", {})
        if isinstance(capital_data, dict):
            capital_inicial = capital_data.get("saldo_inicial", capital_data.get("saldo_anterior", 0))
            capital_cambios = capital_data.get("cambios", capital_data.get("movimientos", 0))
            capital_final = capital_data.get("saldo_final", capital_inicial + capital_cambios)
        else:
            capital_inicial = capital_cambios = capital_final = 0
        
        # Extraer datos de Otras Reservas
        otras_reservas_data = patrimonio_data.get("otras_reservas", patrimonio_data.get("reservas", {}))
        if isinstance(otras_reservas_data, dict):
            otras_reservas_inicial = otras_reservas_data.get("saldo_inicial", otras_reservas_data.get("saldo_anterior", 0))
            otras_reservas_cambios = otras_reservas_data.get("cambios", otras_reservas_data.get("movimientos", 0))
            otras_reservas_final = otras_reservas_data.get("saldo_final", otras_reservas_inicial + otras_reservas_cambios)
        else:
            otras_reservas_inicial = otras_reservas_cambios = otras_reservas_final = 0
        
        # Extraer datos de Resultados Acumulados
        resultados_data = patrimonio_data.get("resultados_acumulados", patrimonio_data.get("utilidades_retenidas", {}))
        if isinstance(resultados_data, dict):
            resultados_inicial = resultados_data.get("saldo_inicial", resultados_data.get("saldo_anterior", 0))
            resultados_cambios = resultados_data.get("cambios", resultados_data.get("movimientos", 0))
            resultados_final = resultados_data.get("saldo_final", resultados_inicial + resultados_cambios)
        else:
            resultados_inicial = resultados_cambios = resultados_final = 0
        
        # Resultado del ejercicio del ERI - IGUAL QUE excel_templates
        total_eri = 0
        if data_eri:
            # Calcular total del ERI sumando todos los bloques
            for bloque_key in ["ganancias_brutas", "ganancia_perdida", "ganancia_perdida_antes_impuestos"]:
                bloque_data = data_eri.get(bloque_key)
                if bloque_data and "total" in bloque_data:
                    total_eri += bloque_data["total"]
        
        # DEBUG: Verificar valores calculados
        logger.info(f"=== VALORES CALCULADOS ===")
        logger.info(f"capital_inicial: {capital_inicial}")
        logger.info(f"capital_cambios: {capital_cambios}")
        logger.info(f"otras_reservas_inicial: {otras_reservas_inicial}")
        logger.info(f"otras_reservas_cambios: {otras_reservas_cambios}")
        logger.info(f"resultados_inicial: {resultados_inicial}")
        logger.info(f"resultados_cambios: {resultados_cambios}")
        logger.info(f"total_eri (ERI): {total_eri}")
        
        # Datos de las filas del ECP (formato original)
        filas_ecp = [
            # 1. BALANCE INICIAL
            {
                self._get_text('concept', language, metadata): self._get_text('initial_balance', language, metadata),
                self._get_text('capital', language, metadata): capital_inicial,  # Capital: Balance inicial
                self._get_text('other_reserves', language, metadata): otras_reservas_inicial,  # Otras Reservas: Balance inicial
                self._get_text('retained_earnings', language, metadata): resultados_inicial,  # Resultado acumulado: Balance inicial
                self._get_text('attributable_capital', language, metadata): capital_inicial + otras_reservas_inicial + resultados_inicial,
                self._get_text('non_controlling_interests', language, metadata): 0,
                self._get_text('total', language, metadata): capital_inicial + otras_reservas_inicial + resultados_inicial
            },
            # 2. RESULTADO DEL EJERCICIO
            {
                self._get_text('concept', language, metadata): self._get_text('period_result_ecp', language, metadata),  # "Resultado del ejercicio" / "Result of the Exercise"
                self._get_text('capital', language, metadata): 0,  # El resultado NO va en capital
                self._get_text('other_reserves', language, metadata): 0,  # El resultado NO va en otras reservas
                self._get_text('retained_earnings', language, metadata): total_eri,  # Resultado acumulado: ERI total
                self._get_text('attributable_capital', language, metadata): total_eri,  # Solo el resultado
                self._get_text('non_controlling_interests', language, metadata): 0,
                self._get_text('total', language, metadata): total_eri  # Solo el resultado
            },
            # 3. OTROS CAMBIOS
            {
                self._get_text('concept', language, metadata): self._get_text('other_changes', language, metadata),
                self._get_text('capital', language, metadata): capital_cambios,
                self._get_text('other_reserves', language, metadata): otras_reservas_cambios,
                self._get_text('retained_earnings', language, metadata): resultados_cambios,  # Otros cambios en resultados
                self._get_text('attributable_capital', language, metadata): capital_cambios + otras_reservas_cambios + resultados_cambios,
                self._get_text('non_controlling_interests', language, metadata): 0,
                self._get_text('total', language, metadata): capital_cambios + otras_reservas_cambios + resultados_cambios
            },
            # 4. SALDO FINAL
            {
                self._get_text('concept', language, metadata): self._get_text('final_balance', language, metadata),
                self._get_text('capital', language, metadata): capital_inicial + capital_cambios,
                self._get_text('other_reserves', language, metadata): otras_reservas_inicial + otras_reservas_cambios,
                self._get_text('retained_earnings', language, metadata): resultados_inicial + resultados_cambios + total_eri,
                self._get_text('attributable_capital', language, metadata): capital_inicial + capital_cambios + otras_reservas_inicial + otras_reservas_cambios + resultados_inicial + resultados_cambios + total_eri,
                self._get_text('non_controlling_interests', language, metadata): 0,
                self._get_text('total', language, metadata): capital_inicial + capital_cambios + otras_reservas_inicial + otras_reservas_cambios + resultados_inicial + resultados_cambios + total_eri
            }
        ]
        
        # Agregar filas de datos
        for fila in filas_ecp:
            for col_idx, (key, value) in enumerate(fila.items(), 1):
                cell = ws.cell(row=current_row, column=col_idx)
                if col_idx == 1:  # Concepto
                    cell.value = value
                    cell.font = self.styles['subheader']
                else:  # Valores numéricos
                    cell.value = self._format_amount(value, metadata.get("moneda", "CLP"))
                    cell.font = self.styles['data']
                
                cell.border = self.borders['thin']
                
                # Resaltar filas especiales con colores
                concept_key = self._get_text('concept', language, metadata)
                initial_balance_text = self._get_text('initial_balance', language, metadata)
                final_balance_text = self._get_text('final_balance', language, metadata)
                
                if fila[concept_key] == initial_balance_text:
                    cell.fill = self.fills['initial_balance']   # Verde claro para saldo inicial
                    cell.font = self.styles['total']
                elif fila[concept_key] == final_balance_text:
                    cell.fill = self.fills['final_balance']     # Verde oscuro para saldo final
                    cell.font = self.styles['total']
            
            current_row += 1
        
        return current_row
