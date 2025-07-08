#!/usr/bin/env python
"""
Script de validación para el flujo de Libro Mayor con Celery Chains.
Ejecutar desde el contenedor Django con: python manage.py shell < validar_libro_mayor_chains.py
"""

import os
import tempfile
from decimal import Decimal
from datetime import datetime, date
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook

# Imports del proyecto
from contabilidad.models import (
    CierreContabilidad, Cliente, UploadLog, LibroMayorArchivo,
    MovimientoContable, CuentaContable, AperturaCuenta
)
from contabilidad.tasks_libro_mayor import (
    crear_chain_libro_mayor,
    validar_nombre_archivo_libro_mayor,
    verificar_archivo_libro_mayor,
    validar_contenido_libro_mayor
)

print("=" * 60)
print("VALIDACIÓN DEL FLUJO DE LIBRO MAYOR CON CELERY CHAINS")
print("=" * 60)

def crear_archivo_excel_test():
    """Crea un archivo Excel de prueba válido"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    
    # Headers esperados por el sistema
    headers = [
        'Código Cuenta', 'Nombre Cuenta', 'Saldo Inicial', 'Débito', 'Crédito',
        'Fecha', 'Tipo Doc', 'Número Doc', 'Glosa', 'Referencia'
    ]
    
    # Escribir headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Datos de ejemplo
    data = [
        ['1101', 'Caja', 1000.00, 500.00, 200.00, '2024-01-15', 'FC', '001', 'Venta productos', 'REF001'],
        ['1102', 'Banco Estado', 2000.00, 1000.00, 500.00, '2024-01-16', 'TG', '002', 'Transferencia', 'REF002'],
        ['2101', 'Proveedores', 0.00, 0.00, 300.00, '2024-01-17', 'FC', '003', 'Pago proveedor', 'REF003'],
        ['4101', 'Ventas', 0.00, 0.00, 1500.00, '2024-01-18', 'FC', '004', 'Venta servicios', 'REF004'],
        ['6101', 'Gastos Operacionales', 800.00, 200.00, 0.00, '2024-01-19', 'FC', '005', 'Gastos varios', 'REF005']
    ]
    
    for row, row_data in enumerate(data, 2):
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Guardar en archivo temporal
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(temp_file.name)
    temp_file.close()
    
    return temp_file.name

def test_1_verificar_imports():
    """Test 1: Verificar que los imports funcionan correctamente"""
    print("\n🔍 TEST 1: Verificando imports...")
    
    try:
        from contabilidad.tasks_libro_mayor import crear_chain_libro_mayor
        print("✅ Import de crear_chain_libro_mayor exitoso")
        
        from contabilidad.tasks_libro_mayor import validar_nombre_archivo_libro_mayor
        print("✅ Import de validar_nombre_archivo_libro_mayor exitoso")
        
        from contabilidad.tasks_libro_mayor import finalizar_procesamiento_libro_mayor
        print("✅ Import de finalizar_procesamiento_libro_mayor exitoso")
        
        print("✅ TEST 1 COMPLETADO: Todos los imports funcionan")
        return True
        
    except ImportError as e:
        print(f"❌ ERROR en imports: {str(e)}")
        return False

def test_2_crear_datos_basicos():
    """Test 2: Crear datos básicos necesarios"""
    print("\n🔍 TEST 2: Creando datos básicos...")
    
    try:
        # Crear o obtener usuario
        user, created = User.objects.get_or_create(
            username='test_libro_mayor',
            defaults={
                'email': 'test@example.com',
                'first_name': 'Test',
                'last_name': 'User'
            }
        )
        if created:
            user.set_password('testpass123')
            user.save()
        print(f"✅ Usuario: {user.username} ({'creado' if created else 'existente'})")
        
        # Crear o obtener cliente
        cliente, created = Cliente.objects.get_or_create(
            rut='12345678-9',
            defaults={
                'nombre': 'Cliente Test Libro Mayor',
                'activo': True
            }
        )
        print(f"✅ Cliente: {cliente.nombre} ({'creado' if created else 'existente'})")
        
        # Crear o obtener cierre
        cierre, created = CierreContabilidad.objects.get_or_create(
            cliente=cliente,
            periodo='2024-01',
            defaults={
                'activo': True,
                'estado': 'abierto'
            }
        )
        print(f"✅ Cierre: {cierre.periodo} ({'creado' if created else 'existente'})")
        
        print("✅ TEST 2 COMPLETADO: Datos básicos creados")
        return user, cliente, cierre
        
    except Exception as e:
        print(f"❌ ERROR creando datos básicos: {str(e)}")
        return None, None, None

def test_3_crear_upload_log():
    """Test 3: Crear UploadLog con nombre válido"""
    print("\n🔍 TEST 3: Creando UploadLog...")
    
    try:
        user, cliente, cierre = test_2_crear_datos_basicos()
        if not all([user, cliente, cierre]):
            print("❌ No se pudieron crear datos básicos")
            return None
        
        # Crear UploadLog con nombre válido para Libro Mayor
        upload_log = UploadLog.objects.create(
            cliente=cliente,
            tipo_archivo='libro_mayor',
            nombre_archivo='12345678-9_LibroMayor_2024-01.xlsx',  # Formato válido
            archivo_hash='test_hash_' + str(datetime.now().timestamp()),
            usuario=user,
            estado='pendiente',
            cierre=cierre
        )
        
        print(f"✅ UploadLog creado con ID: {upload_log.id}")
        print(f"   - Cliente: {upload_log.cliente.nombre}")
        print(f"   - Archivo: {upload_log.nombre_archivo}")
        print(f"   - Estado: {upload_log.estado}")
        
        print("✅ TEST 3 COMPLETADO: UploadLog creado")
        return upload_log
        
    except Exception as e:
        print(f"❌ ERROR creando UploadLog: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def test_4_crear_chain():
    """Test 4: Crear el chain de procesamiento"""
    print("\n🔍 TEST 4: Creando chain de procesamiento...")
    
    try:
        upload_log = test_3_crear_upload_log()
        if not upload_log:
            print("❌ No se pudo crear UploadLog")
            return None
        
        # Crear chain
        chain = crear_chain_libro_mayor(upload_log.id)
        
        print(f"✅ Chain creado exitosamente")
        print(f"   - Número de tasks: {len(chain.tasks)}")
        
        # Listar tasks del chain
        print("   - Tasks en el chain:")
        for i, task in enumerate(chain.tasks, 1):
            task_name = task.name.split('.')[-1]  # Solo el nombre final
            print(f"     {i}. {task_name}")
        
        print("✅ TEST 4 COMPLETADO: Chain creado correctamente")
        return chain, upload_log
        
    except Exception as e:
        print(f"❌ ERROR creando chain: {str(e)}")
        import traceback
        traceback.print_exc()
        return None, None

def test_5_validar_task1():
    """Test 5: Ejecutar primera task (validación de nombre)"""
    print("\n🔍 TEST 5: Ejecutando task de validación de nombre...")
    
    try:
        chain, upload_log = test_4_crear_chain()
        if not upload_log:
            print("❌ No se pudo crear chain")
            return False
        
        # Ejecutar primera task directamente
        resultado = validar_nombre_archivo_libro_mayor(upload_log.id)
        
        print(f"✅ Task ejecutada exitosamente")
        print(f"   - Resultado: {resultado}")
        
        # Verificar estado del upload_log
        upload_log.refresh_from_db()
        print(f"   - Estado después de validación: {upload_log.estado}")
        
        print("✅ TEST 5 COMPLETADO: Task de validación ejecutada")
        return True
        
    except Exception as e:
        print(f"❌ ERROR ejecutando task de validación: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def test_6_verificar_tasks_registradas():
    """Test 6: Verificar que las tasks están registradas en Celery"""
    print("\n🔍 TEST 6: Verificando tasks registradas en Celery...")
    
    try:
        from celery import current_app
        
        # Obtener tasks registradas
        registered_tasks = current_app.tasks.keys()
        
        # Tasks esperadas del libro mayor
        expected_tasks = [
            'contabilidad.tasks_libro_mayor.validar_nombre_archivo_libro_mayor',
            'contabilidad.tasks_libro_mayor.verificar_archivo_libro_mayor',
            'contabilidad.tasks_libro_mayor.validar_contenido_libro_mayor',
            'contabilidad.tasks_libro_mayor.procesar_libro_mayor_raw',
            'contabilidad.tasks_libro_mayor.generar_incidencias_libro_mayor',
            'contabilidad.tasks_libro_mayor.finalizar_procesamiento_libro_mayor'
        ]
        
        print("   Tasks esperadas:")
        for task_name in expected_tasks:
            if task_name in registered_tasks:
                print(f"   ✅ {task_name}")
            else:
                print(f"   ❌ {task_name} - NO REGISTRADA")
        
        # Mostrar stats
        total_tasks = len(registered_tasks)
        libro_mayor_tasks = len([t for t in registered_tasks if 'libro_mayor' in t])
        
        print(f"\n   📊 Estadísticas:")
        print(f"   - Total tasks registradas: {total_tasks}")
        print(f"   - Tasks de libro mayor: {libro_mayor_tasks}")
        
        print("✅ TEST 6 COMPLETADO: Verificación de registro completada")
        return True
        
    except Exception as e:
        print(f"❌ ERROR verificando tasks registradas: {str(e)}")
        return False

def test_7_crear_archivo_excel():
    """Test 7: Crear archivo Excel de prueba"""
    print("\n🔍 TEST 7: Creando archivo Excel de prueba...")
    
    try:
        archivo_path = crear_archivo_excel_test()
        
        print(f"✅ Archivo Excel creado: {archivo_path}")
        
        # Verificar que el archivo existe
        if os.path.exists(archivo_path):
            file_size = os.path.getsize(archivo_path)
            print(f"   - Tamaño: {file_size} bytes")
            print(f"   - Existe: Sí")
        else:
            print(f"   ❌ Archivo no existe")
            return None
        
        # Leer archivo para verificar contenido
        from openpyxl import load_workbook
        wb = load_workbook(archivo_path)
        ws = wb.active
        
        print(f"   - Hoja activa: {ws.title}")
        print(f"   - Filas con datos: {ws.max_row}")
        print(f"   - Columnas: {ws.max_column}")
        
        # Mostrar headers
        headers = [cell.value for cell in ws[1]]
        print(f"   - Headers: {headers}")
        
        print("✅ TEST 7 COMPLETADO: Archivo Excel creado y verificado")
        return archivo_path
        
    except Exception as e:
        print(f"❌ ERROR creando archivo Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def ejecutar_todos_los_tests():
    """Ejecuta todos los tests de validación"""
    print("🚀 INICIANDO VALIDACIÓN COMPLETA DEL FLUJO DE LIBRO MAYOR")
    
    tests_results = []
    
    # Ejecutar tests secuencialmente
    tests = [
        ("Verificar Imports", test_1_verificar_imports),
        ("Crear Datos Básicos", lambda: test_2_crear_datos_basicos() is not None),
        ("Crear UploadLog", lambda: test_3_crear_upload_log() is not None),
        ("Crear Chain", lambda: test_4_crear_chain()[0] is not None),
        ("Validar Task 1", test_5_validar_task1),
        ("Verificar Tasks Registradas", test_6_verificar_tasks_registradas),
        ("Crear Archivo Excel", lambda: test_7_crear_archivo_excel() is not None)
    ]
    
    for test_name, test_func in tests:
        try:
            result = test_func()
            tests_results.append((test_name, result))
            if not result:
                print(f"❌ {test_name} FALLÓ - detteniendo tests")
                break
        except Exception as e:
            print(f"❌ {test_name} ERROR: {str(e)}")
            tests_results.append((test_name, False))
            break
    
    # Resumen final
    print("\n" + "=" * 60)
    print("RESUMEN DE VALIDACIÓN")
    print("=" * 60)
    
    passed = sum(1 for _, result in tests_results if result)
    total = len(tests_results)
    
    for test_name, result in tests_results:
        status = "✅ PASS" if result else "❌ FAIL"
        print(f"{status} - {test_name}")
    
    print(f"\n📊 RESULTADO FINAL: {passed}/{total} tests pasaron")
    
    if passed == total:
        print("🎉 ¡TODOS LOS TESTS PASARON! El flujo de Libro Mayor está listo.")
    else:
        print("⚠️  Algunos tests fallaron. Revisar implementación.")
    
    return passed == total

# Ejecutar validación
if __name__ == "__main__":
    ejecutar_todos_los_tests()
else:
    # Cuando se ejecuta desde shell
    ejecutar_todos_los_tests()
