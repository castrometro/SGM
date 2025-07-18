"""
Utilidades para crear y manejar templates de Excel para los informes contables.
DEPRECATED: Este archivo se está migrando al módulo excel/ modular.
Mantiene compatibilidad hacia atrás mientras se completa la migración.
"""

# Importar el nuevo sistema modular
try:
    from .excel import excel_generator
    print("✅ Usando sistema modular de Excel templates")
    
    # Re-exportar para compatibilidad
    __all__ = ['excel_generator']
    
except ImportError as e:
    # Fallback al sistema antiguo si hay problemas
    print(f"⚠️  Fallback al sistema antiguo de Excel templates: {e}")
    
    # Importar todo el código original como fallback

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import pandas as pd
import io
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class ExcelTemplateGenerator:
    """Generador de templates Excel para informes contables"""
    
    def __init__(self):
        # Traducciones para títulos y etiquetas
        self.translations = {
            'es': {
                'title_esf': 'ESTADO DE SITUACIÓN FINANCIERA',
                'title_eri': 'ESTADO DE RESULTADO INTEGRAL',
                'title_ecp': 'ESTADO DE CAMBIOS EN EL PATRIMONIO',
                'title_movimientos': 'MOVIMIENTOS CONTABLES',
                'assets': 'ACTIVOS',
                'current_assets': 'Activos Corrientes',
                'non_current_assets': 'Activos No Corrientes',
                'total_assets': 'TOTAL ACTIVOS',
                'liabilities_equity': 'PASIVOS Y PATRIMONIO',
                'current_liabilities': 'Pasivos Corrientes',
                'non_current_liabilities': 'Pasivos No Corrientes',
                'total_liabilities': 'TOTAL PASIVOS',
                'equity': 'PATRIMONIO',
                'total_equity': 'TOTAL PATRIMONIO',
                'total_liabilities_equity': 'TOTAL PASIVOS Y PATRIMONIO',
                'period_result': 'Resultado del Ejercicio',
                'profit_loss': 'Ganancia del Ejercicio',
                'loss': 'Pérdida del Ejercicio',
                'total_period_result': 'TOTAL RESULTADO DEL EJERCICIO',
                'client': 'Cliente',
                'period': 'Período',
                'currency': 'Moneda',
                'date': 'Fecha',
                'no_movements': '(Sin movimientos)',
                'no_data': '(Sin datos disponibles)',
                'totals': 'TOTALES:',
                'total_general': 'TOTAL GENERAL (Ganancia/Pérdida)',
                'info_sheet': 'Información del Reporte',
                'report_info': 'INFORMACIÓN DEL REPORTE',
                'language': 'Idioma',
                'spanish': 'Español',
                'generation_date': 'Fecha de generación',
                'system': 'Sistema',
                'version': 'Versión',
                # Bloques ERI
                'ganancias_brutas': 'Ganancias Brutas',
                'ganancia_perdida': 'Ganancia (Pérdida)',
                'ganancia_perdida_antes_impuestos': 'Ganancia (Pérdida) Antes de Impuestos',
                # ECP
                'initial_balance': 'Saldo Inicial al 1 de Enero',
                'period_result_ecp': 'Resultado del ejercicio',
                'other_changes': 'Otros cambios',
                'final_balance': 'Saldo Final',
                'concept': 'Concepto',
                'capital': 'Capital',
                'other_reserves': 'Otras Reservas',
                'accumulated_results': 'Resultados Acumulados',
                'attributable_capital': 'Capital Atribuible',
                'non_controlling_interests': 'Participaciones no Controladoras'
            },
            'en': {
                'title_esf': 'STATUS OF FINANCIAL SITUATION',
                'title_eri': 'STATUS OF COMPREHENSIVE INCOME, BY FUNCTION',
                'title_ecp': 'STATUS OF CHANGES OF PATRIMONY',
                'title_movimientos': 'ACCOUNTING TRANSACTIONS',
                'assets': 'ASSETS',
                'current_assets': 'Current Assets',
                'non_current_assets': 'Non-Current Assets',
                'total_assets': 'TOTAL ASSETS',
                'liabilities_equity': 'LIABILITIES AND PATRIMONY',
                'current_liabilities': 'Current Liabilities',
                'non_current_liabilities': 'Non-Current Liabilities',
                'total_liabilities': 'TOTAL LIABILITIES',
                'equity': 'PATRIMONY',
                'total_equity': 'TOTAL PATRIMONY',
                'total_liabilities_equity': 'TOTAL LIABILITIES AND PATRIMONY',
                'period_result': 'Period Result',
                'profit_loss': 'Profit for the Period',
                'loss': 'Loss for the Period',
                'total_period_result': 'TOTAL PERIOD RESULT',
                'client': 'Client',
                'period': 'Period',
                'currency': 'Currency',
                'date': 'Date',
                'no_movements': '(No movements)',
                'no_data': '(No data available)',
                'totals': 'TOTAL:',
                'total_general': 'TOTAL GENERAL (Profit/Loss)',
                'info_sheet': 'Report Information',
                'report_info': 'REPORT INFORMATION',
                'language': 'Language',
                'spanish': 'Spanish',
                'english': 'English',
                'generation_date': 'Generation Date',
                'system': 'System',
                'version': 'Version',
                # Bloques ERI
                'ganancias_brutas': 'Gross Earnings',
                'ganancia_perdida': 'Earnings (Loss)',
                'ganancia_perdida_antes_impuestos': 'Earnings (Loss) Before Taxes',
                # ECP
                'initial_balance': 'Initial Balance as of January 1',
                'period_result_ecp': 'Result of the Exercise',
                'other_changes': 'Other Settings',
                'final_balance': 'Final Balance',
                'concept': 'Concept',
                'capital': 'Capital',
                'other_reserves': 'Other Reserves',
                'accumulated_results': 'R. Accumulated',
                'attributable_capital': 'Capital Attributable to the owners of the controller\'s instruments',
                'non_controlling_interests': 'Uncontrolled participations'
            }
        }
        
        # Estilos predefinidos
        self.styles = {
            'title': Font(name='Arial', size=16, bold=True, color='FFFFFF'),
            'header': Font(name='Arial', size=12, bold=True, color='FFFFFF'),
            'subheader': Font(name='Arial', size=11, bold=True, color='000000'),
            'data': Font(name='Arial', size=10, color='000000'),
            'total': Font(name='Arial', size=11, bold=True, color='FFFFFF'),
            'metadata': Font(name='Arial', size=9, italic=True, color='666666')
        }
        
        self.fills = {
            'title': PatternFill(start_color='0A58CA', end_color='0A58CA', fill_type='solid'),
            'header': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
            'subheader': PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid'),
            'total': PatternFill(start_color='2F5233', end_color='2F5233', fill_type='solid'),
            'alternate': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
            'initial_balance': PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid'),  # Verde claro para saldo inicial
            'final_balance': PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')     # Verde oscuro para saldo final
        }
        
        self.borders = {
            'thin': Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin')
            ),
            'thick': Border(
                left=Side(border_style='thick'),
                right=Side(border_style='thick'),
                top=Side(border_style='thick'),
                bottom=Side(border_style='thick')
            )
        }

    def _get_text(self, key, language='es'):
        """Obtener texto traducido según el idioma"""
        lang = 'en' if language.lower() in ['en', 'english', 'inglés'] else 'es'
        return self.translations.get(lang, {}).get(key, key)

    def _get_account_name(self, cuenta, language='es'):
        """Obtener nombre de cuenta según el idioma"""
        if language.lower() in ['en', 'english', 'inglés']:
            return cuenta.get("nombre_en", cuenta.get("nombre_es", ""))
        else:
            return cuenta.get("nombre_es", cuenta.get("nombre_en", ""))

    def _apply_header_style(self, worksheet, start_row, start_col, end_col, title=""):
        """Aplicar estilo de encabezado a un rango de celdas"""
        if title:
            cell = worksheet.cell(row=start_row, column=start_col)
            cell.value = title
            cell.font = self.styles['title']
            cell.fill = self.fills['title']
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Combinar celdas para el título
            if end_col > start_col:
                worksheet.merge_cells(
                    start_row=start_row, start_column=start_col,
                    end_row=start_row, end_column=end_col
                )

    def _apply_data_formatting(self, worksheet, start_row, end_row, start_col, end_col):
        """Aplicar formato a los datos"""
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = self.borders['thin']
                cell.font = self.styles['data']
                
                # Alternar colores de fila
                if row % 2 == 0:
                    cell.fill = self.fills['alternate']

    def _add_metadata_sheet(self, workbook, metadata, language='es'):
        """Agregar hoja con metadatos del informe"""
        ws = workbook.create_sheet(self._get_text('info_sheet', language))
        
        # Título
        title_text = self._get_text('report_info', language)
        ws.cell(row=1, column=1, value=title_text)
        ws.cell(row=1, column=1).font = self.styles['title']
        ws.cell(row=1, column=1).fill = self.fills['title']
        ws.merge_cells('A1:B1')
        
        # Metadatos
        info_data = [
            (f"{self._get_text('client', language)}:", metadata.get('cliente_nombre', 'N/A')),
            (f"{self._get_text('period', language)}:", metadata.get('periodo', 'N/A')),
            (f"{self._get_text('currency', language)}:", metadata.get('moneda', 'CLP')),
            (f"{self._get_text('language', language)}:", metadata.get('idioma', self._get_text('spanish', language))),
            (f"{self._get_text('generation_date', language)}:", datetime.now().strftime("%d/%m/%Y %H:%M")),
            (f"{self._get_text('system', language)}:", "SGM Dashboard Contable"),
            (f"{self._get_text('version', language)}:", "v1.0")
        ]
        
        for i, (label, value) in enumerate(info_data, start=3):
            ws.cell(row=i, column=1, value=label).font = self.styles['subheader']
            ws.cell(row=i, column=2, value=value).font = self.styles['data']
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30

    def _calcular_total_eri(self, data_eri):
        """Calcular el total de ganancia/pérdida del ERI"""
        if not data_eri:
            return 0
        
        total_eri = 0
        # Sumar todos los bloques del ERI
        bloques_eri = [
            "ganancias_brutas",
            "ganancia_perdida", 
            "ganancia_perdida_antes_impuestos"
        ]
        
        for bloque_key in bloques_eri:
            bloque_data = data_eri.get(bloque_key, {})
            if isinstance(bloque_data, dict) and "total" in bloque_data:
                total_eri += bloque_data["total"]
        
        return total_eri

    def _debug_data_structure(self, data, section_name=""):
        """Debug helper para verificar estructura de datos"""
        logger.info(f"=== DEBUG {section_name} ===")
        #if isinstance(data, dict):
           # for key, value in data.items():
               ##if isinstance(value, dict):
                   # logger.info(f"{key}: dict con {len(value)} elementos")
                   # if "total" in value:
                    #   logger.info(f"  - total: {value['total']}")
                   # if "grupos" in value:
                    #    logger.info(f"  - grupos: {len(value['grupos'])} grupos")
                     #      logger.info(f"    * {grupo_name}")
                   #if "cuentas" in value:
                     #   logger.info(f"  - cuentas: {len(value['cuentas'])} cuentas")
               # elif isinstance(value, list):
                #    logger.info(f"{key}: lista con {len(value)} elementos")
               # else:
                 #   logger.info(f"{key}: {type(value)} = {value}")
       # else:
           # logger.info(f"Data type: {type(data)} = {data}")
        logger.info(f"=== FIN DEBUG {section_name} ===")

    def generate_esf_template(self, data_esf, metadata, data_eri=None):
        """Generar template Excel para Estado de Situación Financiera"""
        # Obtener idioma de los metadatos
        language = metadata.get('idioma', 'es')
        
        # Debug de la estructura de datos
        if data_esf and logger.isEnabledFor(logging.INFO):
            self._debug_data_structure(data_esf, "DATA_ESF_COMPLETA")
            if "patrimonio" in data_esf:
                self._debug_data_structure(data_esf["patrimonio"], "PATRIMONIO")
        
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = self._get_text('title_esf', language)
        
        # Título principal
        title_text = self._get_text('title_esf', language)
        ws.cell(row=1, column=1, value=title_text)
        self._apply_header_style(ws, 1, 1, 4, title_text)
        
        # Información del período
        current_row = 3
        ws.cell(row=current_row, column=1, value=f"{self._get_text('client', language)}: {metadata.get('cliente_nombre', 'N/A')}")
        ws.cell(row=current_row, column=3, value=f"{self._get_text('period', language)}: {metadata.get('periodo', 'N/A')}")
        current_row += 1
        ws.cell(row=current_row, column=1, value=f"{self._get_text('currency', language)}: {metadata.get('moneda', 'CLP')}")
        ws.cell(row=current_row, column=3, value=f"{self._get_text('date', language)}: {datetime.now().strftime('%d/%m/%Y')}")
        current_row += 2
        
        # Extraer y procesar datos
        if data_esf:
            # ACTIVOS
            ws.cell(row=current_row, column=1, value=self._get_text('assets', language)).font = self.styles['header']
            ws.cell(row=current_row, column=1).fill = self.fills['header']
            current_row += 1
            
            # Activos Corrientes
            activos_corrientes = data_esf.get("activos", {}).get("corrientes", {})
            current_row = self._add_esf_section(ws, current_row, self._get_text('current_assets', language), activos_corrientes, metadata.get("moneda", "CLP"), language)
            
            # Activos No Corrientes
            activos_no_corrientes = data_esf.get("activos", {}).get("no_corrientes", {})
            current_row = self._add_esf_section(ws, current_row, self._get_text('non_current_assets', language), activos_no_corrientes, metadata.get("moneda", "CLP"), language)
            
            # Total Activos
            total_activos = data_esf.get("activos", {}).get("total_activos", 0)
            if total_activos != 0:
                current_row += 1
                ws.cell(row=current_row, column=1, value=self._get_text('total_assets', language)).font = self.styles['total']
                ws.cell(row=current_row, column=2, value=self._format_amount(total_activos, metadata.get("moneda", "CLP"))).font = self.styles['total']
                ws.cell(row=current_row, column=1).fill = self.fills['total']
                ws.cell(row=current_row, column=2).fill = self.fills['total']
            
            current_row += 1
            
            # PASIVOS Y PATRIMONIO
            ws.cell(row=current_row, column=1, value=self._get_text('liabilities_equity', language)).font = self.styles['header']
            ws.cell(row=current_row, column=1).fill = self.fills['header']
            current_row += 1
            
            # Pasivos Corrientes
            pasivos_corrientes = data_esf.get("pasivos", {}).get("corrientes", {})
            current_row = self._add_esf_section(ws, current_row, self._get_text('current_liabilities', language), pasivos_corrientes, metadata.get("moneda", "CLP"), language)
            
            # Pasivos No Corrientes
            pasivos_no_corrientes = data_esf.get("pasivos", {}).get("no_corrientes", {})
            current_row = self._add_esf_section(ws, current_row, self._get_text('non_current_liabilities', language), pasivos_no_corrientes, metadata.get("moneda", "CLP"), language)
            
            # Total Pasivos
            total_pasivos = data_esf.get("pasivos", {}).get("total_pasivos", 0)
            if total_pasivos != 0:
                current_row += 1
                ws.cell(row=current_row, column=1, value=self._get_text('total_liabilities', language)).font = self.styles['total']
                ws.cell(row=current_row, column=2, value=self._format_amount(total_pasivos, metadata.get("moneda", "CLP"))).font = self.styles['total']
                ws.cell(row=current_row, column=1).fill = self.fills['total']
                ws.cell(row=current_row, column=2).fill = self.fills['total']
                current_row += 2
            
            # Patrimonio - manejar estructura anidada
            patrimonio = data_esf.get("patrimonio", {})
            if patrimonio:
                # Título principal de Patrimonio
                ws.cell(row=current_row, column=1, value=self._get_text('equity', language)).font = self.styles['header']
                ws.cell(row=current_row, column=1).fill = self.fills['header']
                current_row += 1
                
                # Procesar cada subcategoría de patrimonio (ej: capital)
                total_patrimonio = 0
                for subcategoria_key, subcategoria_data in patrimonio.items():
                    if isinstance(subcategoria_data, dict) and subcategoria_key not in ['total_patrimonio']:
                        # Usar nombre en idioma apropiado si está disponible
                        if language.lower() in ['en', 'english', 'inglés']:
                            section_name = subcategoria_data.get('nombre_en', subcategoria_data.get('nombre_es', subcategoria_key))
                        else:
                            section_name = subcategoria_data.get('nombre_es', subcategoria_data.get('nombre_en', subcategoria_key))
                        current_row = self._add_esf_section(ws, current_row, section_name, subcategoria_data, metadata.get("moneda", "CLP"), language)
                        total_patrimonio += subcategoria_data.get('total', 0)
                
                # Agregar Ganancia/(Pérdida) del Ejercicio del ERI
                total_eri = self._calcular_total_eri(data_eri)
                if total_eri != 0:
                    # Título de la subcategoría
                    ws.cell(row=current_row, column=1, value=self._get_text('period_result', language)).font = self.styles['subheader']
                    ws.cell(row=current_row, column=1).fill = self.fills['subheader']
                    current_row += 1
                    
                    # Línea del resultado
                    ganancia_perdida_texto = self._get_text('profit_loss', language) if total_eri > 0 else self._get_text('loss', language)
                    ws.cell(row=current_row, column=1, value=f"  {ganancia_perdida_texto} (Del ERI)")
                    ws.cell(row=current_row, column=2, value=self._format_amount(total_eri, metadata.get("moneda", "CLP")))
                    current_row += 1
                    
                    # Total de resultado del ejercicio
                    ws.cell(row=current_row, column=1, value=self._get_text('total_period_result', language)).font = self.styles['total']
                    ws.cell(row=current_row, column=2, value=self._format_amount(total_eri, metadata.get("moneda", "CLP"))).font = self.styles['total']
                    ws.cell(row=current_row, column=1).fill = self.fills['total']
                    ws.cell(row=current_row, column=2).fill = self.fills['total']
                    current_row += 2
                    
                    # Agregar al total del patrimonio
                    total_patrimonio += total_eri
                
                # Total patrimonio
                if total_patrimonio != 0:
                    current_row += 1
                    ws.cell(row=current_row, column=1, value=self._get_text('total_equity', language)).font = self.styles['total']
                    ws.cell(row=current_row, column=2, value=self._format_amount(total_patrimonio, metadata.get("moneda", "CLP"))).font = self.styles['total']
                    ws.cell(row=current_row, column=1).fill = self.fills['total']
                    ws.cell(row=current_row, column=2).fill = self.fills['total']
                    current_row += 2
            
            # TOTAL FINAL: PASIVOS Y PATRIMONIO
            current_row += 1
            total_pasivos = data_esf.get("pasivos", {}).get("total_pasivos", 0)
            # Usar el total de patrimonio ya calculado que incluye el ERI
            total_pasivos_patrimonio = total_pasivos + total_patrimonio
            
            ws.cell(row=current_row, column=1, value=self._get_text('total_liabilities_equity', language)).font = self.styles['total']
            ws.cell(row=current_row, column=2, value=self._format_amount(total_pasivos_patrimonio, metadata.get("moneda", "CLP"))).font = self.styles['total']
            ws.cell(row=current_row, column=1).fill = self.fills['total']
            ws.cell(row=current_row, column=2).fill = self.fills['total']
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        
        # Agregar hoja de metadatos
        self._add_metadata_sheet(workbook, metadata, language)
        
        return workbook

    def _add_esf_section(self, ws, start_row, section_title, section_data, moneda, language='es'):
        """Agregar una sección del ESF al worksheet con grupos colapsables"""
        current_row = start_row
        
        # Título de la sección
        ws.cell(row=current_row, column=1, value=section_title).font = self.styles['subheader']
        ws.cell(row=current_row, column=1).fill = self.fills['subheader']
        current_row += 1
        
        total_section = 0
        
        if isinstance(section_data, dict):
            if "grupos" in section_data and section_data["grupos"]:
                # Formato con grupos - crear grupos colapsables
                for grupo_nombre, grupo_data in section_data.get("grupos", {}).items():
                    if isinstance(grupo_data, dict):
                        # Título del grupo - usar nombre según idioma
                        if language.lower() in ['en', 'english', 'inglés']:
                            grupo_display_name = grupo_data.get('nombre_en', grupo_data.get('nombre_es', grupo_nombre))
                        else:
                            grupo_display_name = grupo_data.get('nombre_es', grupo_data.get('nombre_en', grupo_nombre))
                        
                        grupo_row = current_row
                        ws.cell(row=current_row, column=1, value=f"  {grupo_display_name}").font = self.styles['data']
                        ws.cell(row=current_row, column=2, value=self._format_amount(grupo_data.get('total', 0), moneda)).font = self.styles['data']
                        current_row += 1
                        
                        # Filas de inicio y fin para el grupo
                        grupo_start_row = current_row
                        
                        # Agregar cuentas del grupo
                        for cuenta in grupo_data.get("cuentas", []):
                            codigo = cuenta.get("codigo", "")
                            nombre = self._get_account_name(cuenta, language)
                            saldo = cuenta.get("saldo_final", 0)
                            
                            ws.cell(row=current_row, column=1, value=f"    {codigo} - {nombre}")
                            ws.cell(row=current_row, column=2, value=self._format_amount(saldo, moneda))
                            current_row += 1
                            total_section += saldo
                        
                        # Crear grupo colapsable solo si hay cuentas
                        if grupo_data.get("cuentas") and current_row > grupo_start_row:
                            try:
                                # Crear grupo colapsable desde grupo_start_row hasta current_row-1
                                ws.row_dimensions.group(grupo_start_row, current_row - 1, outline_level=1)
                                # Por defecto, dejar el grupo colapsado
                                for row_num in range(grupo_start_row, current_row):
                                    ws.row_dimensions[row_num].hidden = True
                            except Exception as e:
                                # Si falla la agrupación, continuar sin ella
                                pass
                        
                        # Agregar espacio después del grupo si tiene cuentas
                        if grupo_data.get("cuentas"):
                            current_row += 1
                            
            elif "cuentas" in section_data and section_data["cuentas"]:
                # Formato con cuentas directas
                for cuenta in section_data.get("cuentas", []):
                    codigo = cuenta.get("codigo", "")
                    nombre = self._get_account_name(cuenta, language)
                    saldo = cuenta.get("saldo_final", 0)
                    
                    ws.cell(row=current_row, column=1, value=f"  {codigo} - {nombre}")
                    ws.cell(row=current_row, column=2, value=self._format_amount(saldo, moneda))
                    current_row += 1
                    total_section += saldo
            else:
                # Si no hay grupos ni cuentas, mostrar mensaje
                ws.cell(row=current_row, column=1, value=f"  {self._get_text('no_movements', language)}").font = self.styles['metadata']
                current_row += 1
            
            # Usar el total precalculado si existe, sino usar el calculado
            if "total" in section_data:
                total_section = section_data["total"]
            
            # Solo mostrar total si hay datos
            if total_section != 0:
                ws.cell(row=current_row, column=1, value=f"TOTAL {section_title.upper()}").font = self.styles['total']
                ws.cell(row=current_row, column=2, value=self._format_amount(total_section, moneda)).font = self.styles['total']
                ws.cell(row=current_row, column=1).fill = self.fills['total']
                ws.cell(row=current_row, column=2).fill = self.fills['total']
                current_row += 2
        else:
            # Si section_data no es un diccionario válido
            ws.cell(row=current_row, column=1, value=f"  {self._get_text('no_data', language)}").font = self.styles['metadata']
            current_row += 2
        
        return current_row

    def generate_eri_template(self, data_eri, metadata):
        """Generar template Excel para Estado de Resultado Integral"""
        # Obtener idioma de los metadatos
        language = metadata.get('idioma', 'es')
        
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = self._get_text('title_eri', language)
        
        # Título principal
        title_text = self._get_text('title_eri', language)
        self._apply_header_style(ws, 1, 1, 4, title_text)
        
        # Información del período
        current_row = 3
        ws.cell(row=current_row, column=1, value=f"{self._get_text('client', language)}: {metadata.get('cliente_nombre', 'N/A')}")
        ws.cell(row=current_row, column=3, value=f"{self._get_text('period', language)}: {metadata.get('periodo', 'N/A')}")
        current_row += 1
        ws.cell(row=current_row, column=1, value=f"{self._get_text('currency', language)}: {metadata.get('moneda', 'CLP')}")
        ws.cell(row=current_row, column=3, value=f"{self._get_text('date', language)}: {datetime.now().strftime('%d/%m/%Y')}")
        current_row += 2
        
        if data_eri:
            # Definir bloques ERI a procesar
            bloques_eri = [
                "ganancias_brutas",
                "ganancia_perdida", 
                "ganancia_perdida_antes_impuestos"
            ]
            
            total_general = 0
            
            for bloque_key in bloques_eri:
                bloque_data = data_eri.get(bloque_key)
                if bloque_data:
                    current_row, total_bloque = self._add_eri_section(ws, current_row, bloque_key, bloque_data, metadata.get("moneda", "CLP"), language)
                    total_general += total_bloque
            
            # Total General
            current_row += 1
            ws.cell(row=current_row, column=1, value=self._get_text('total_general', language)).font = self.styles['total']
            ws.cell(row=current_row, column=2, value=self._format_amount(total_general, metadata.get("moneda", "CLP"))).font = self.styles['total']
            ws.cell(row=current_row, column=1).fill = self.fills['total']
            ws.cell(row=current_row, column=2).fill = self.fills['total']
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        # Agregar hoja de metadatos
        self._add_metadata_sheet(workbook, metadata, language)
        
        return workbook

    def _add_eri_section(self, ws, start_row, bloque_key, bloque_data, moneda, language='es'):
        """Agregar una sección del ERI al worksheet con grupos colapsables"""
        current_row = start_row
        
        # Título del bloque - usar traducción si está disponible
        titulo_bloque = self._get_text(bloque_key, language)
        ws.cell(row=current_row, column=1, value=titulo_bloque).font = self.styles['subheader']
        ws.cell(row=current_row, column=1).fill = self.fills['subheader']
        current_row += 1
        
        total_bloque = 0
        
        if isinstance(bloque_data, dict):
            if "grupos" in bloque_data:
                # Procesar grupos con agrupación colapsable
                for grupo_nombre, grupo_data in bloque_data.get("grupos", {}).items():
                    if isinstance(grupo_data, dict):
                        # Título del grupo - usar nombre según idioma
                        if language.lower() in ['en', 'english', 'inglés']:
                            grupo_display_name = grupo_data.get('nombre_en', grupo_data.get('nombre_es', grupo_nombre))
                        else:
                            grupo_display_name = grupo_data.get('nombre_es', grupo_data.get('nombre_en', grupo_nombre))
                        
                        # Mostrar total del grupo
                        ws.cell(row=current_row, column=1, value=f"  {grupo_display_name}").font = self.styles['data']
                        ws.cell(row=current_row, column=2, value=self._format_amount(grupo_data.get('total', 0), moneda)).font = self.styles['data']
                        current_row += 1
                        
                        # Filas de inicio y fin para el grupo
                        grupo_start_row = current_row
                        
                        # Agregar cuentas del grupo
                        for cuenta in grupo_data.get("cuentas", []):
                            codigo = cuenta.get("codigo", "")
                            nombre = self._get_account_name(cuenta, language)
                            saldo = cuenta.get("saldo_final", 0)
                            
                            ws.cell(row=current_row, column=1, value=f"    {codigo} - {nombre}")
                            ws.cell(row=current_row, column=2, value=self._format_amount(saldo, moneda))
                            current_row += 1
                            total_bloque += saldo
                        
                        # Crear grupo colapsable solo si hay cuentas
                        if grupo_data.get("cuentas") and current_row > grupo_start_row:
                            try:
                                # Crear grupo colapsable desde grupo_start_row hasta current_row-1
                                ws.row_dimensions.group(grupo_start_row, current_row - 1, outline_level=1)
                                # Por defecto, dejar el grupo colapsado
                                for row_num in range(grupo_start_row, current_row):
                                    ws.row_dimensions[row_num].hidden = True
                            except Exception as e:
                                # Si falla la agrupación, continuar sin ella
                                pass
                        
                        # Agregar espacio después del grupo si tiene cuentas
                        if grupo_data.get("cuentas"):
                            current_row += 1
            
            # Total del bloque
            if "total" in bloque_data:
                total_bloque = bloque_data["total"]
            
            if total_bloque != 0:
                ws.cell(row=current_row, column=1, value=f"TOTAL {titulo_bloque.upper()}").font = self.styles['total']
                ws.cell(row=current_row, column=2, value=self._format_amount(total_bloque, moneda)).font = self.styles['total']
                ws.cell(row=current_row, column=1).fill = self.fills['total']
                ws.cell(row=current_row, column=2).fill = self.fills['total']
                current_row += 2
        
        return current_row, total_bloque

    def generate_movimientos_template(self, df_movimientos, metadata, tipo_vista="Todos los movimientos"):
        """Generar template Excel para movimientos contables"""
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "Movimientos Contables"
        
        # Título principal
        self._apply_header_style(ws, 1, 1, len(df_movimientos.columns), f"MOVIMIENTOS CONTABLES - {tipo_vista.upper()}")
        
        # Información del período
        current_row = 3
        ws.cell(row=current_row, column=1, value=f"Cliente: {metadata.get('cliente_nombre', 'N/A')}")
        ws.cell(row=current_row, column=4, value=f"Período: {metadata.get('periodo', 'N/A')}")
        current_row += 1
        ws.cell(row=current_row, column=1, value=f"Total movimientos: {len(df_movimientos)}")
        ws.cell(row=current_row, column=4, value=f"Fecha: {datetime.now().strftime('%d/%m/%Y')}")
        current_row += 2
        
        # Encabezados de columna
        for col_idx, column_name in enumerate(df_movimientos.columns, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=column_name)
            cell.font = self.styles['header']
            cell.fill = self.fills['header']
            cell.alignment = Alignment(horizontal='center')
        
        current_row += 1
        
        # Datos
        for row_idx, row in df_movimientos.iterrows():
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font = self.styles['data']
                cell.border = self.borders['thin']
                
                # Alternar colores de fila
                if current_row % 2 == 0:
                    cell.fill = self.fills['alternate']
            
            current_row += 1
        
        # Totales si aplica
        if "Debe" in df_movimientos.columns and "Haber" in df_movimientos.columns:
            current_row += 1
            debe_col = list(df_movimientos.columns).index("Debe") + 1
            haber_col = list(df_movimientos.columns).index("Haber") + 1
            
            ws.cell(row=current_row, column=debe_col - 1, value="TOTALES:").font = self.styles['total']
            ws.cell(row=current_row, column=debe_col, value=self._format_amount(df_movimientos["Debe"].sum(), metadata.get("moneda", "CLP"))).font = self.styles['total']
            ws.cell(row=current_row, column=haber_col, value=self._format_amount(df_movimientos["Haber"].sum(), metadata.get("moneda", "CLP"))).font = self.styles['total']
        
        # Ajustar anchos de columna
        for col_idx in range(1, len(df_movimientos.columns) + 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 15
        
        # Agregar hoja de metadatos
        self._add_metadata_sheet(workbook, metadata)
        
        return workbook

    def generate_ecp_template(self, data_ecp, metadata, data_eri=None):
        """Generar template Excel para Estado de Cambios en el Patrimonio"""
        # Obtener idioma de los metadatos - usar lógica más robusta
        language_raw = metadata.get('idioma', 'es')
        
        # Normalizar el idioma - ser más permisivo con la detección
        if isinstance(language_raw, str):
            language_lower = language_raw.lower()
            is_english = any(keyword in language_lower for keyword in ['en', 'english', 'inglés', 'ing'])
        else:
            is_english = False
        
        language = 'en' if is_english else 'es'
        
        # Función para obtener el período de cierre formateado
        def obtener_periodo_cierre(periodo_str):
            """
            Convierte una cadena de período en mes y año para el balance final.
            También extrae el año para el saldo inicial (siempre 1 de enero).
            Esperado: formato como "2024-12" o "Diciembre 2024"
            """
            if not periodo_str or periodo_str == "Periodo desconocido":
                return "Diciembre 2024", "December 2024", "2024"
            
            # Mapeos de meses
            meses_es = {
                "01": "Enero", "02": "Febrero", "03": "Marzo", "04": "Abril",
                "05": "Mayo", "06": "Junio", "07": "Julio", "08": "Agosto",
                "09": "Septiembre", "10": "Octubre", "11": "Noviembre", "12": "Diciembre"
            }
            
            meses_en = {
                "01": "January", "02": "February", "03": "March", "04": "April",
                "05": "May", "06": "June", "07": "July", "08": "August",
                "09": "September", "10": "October", "11": "November", "12": "December"
            }
            
            # Intentar diferentes formatos
            try:
                # Formato YYYY-MM
                if "-" in periodo_str and len(periodo_str.split("-")) == 2:
                    año, mes = periodo_str.split("-")
                    mes_es = meses_es.get(mes.zfill(2), "Diciembre")
                    mes_en = meses_en.get(mes.zfill(2), "December")
                    return f"{mes_es} {año}", f"{mes_en} {año}", año
                
                # Formato "Mes YYYY" (ya en español)
                elif any(mes in periodo_str for mes in meses_es.values()):
                    # Buscar el mes en español
                    for num, mes_es in meses_es.items():
                        if mes_es in periodo_str:
                            año = ''.join(filter(str.isdigit, periodo_str))
                            mes_en = meses_en[num]
                            return f"{mes_es} {año}", f"{mes_en} {año}", año
                
                # Formato "Month YYYY" (ya en inglés)
                elif any(mes in periodo_str for mes in meses_en.values()):
                    # Buscar el mes en inglés
                    for num, mes_en in meses_en.items():
                        if mes_en in periodo_str:
                            año = ''.join(filter(str.isdigit, periodo_str))
                            mes_es = meses_es[num]
                            return f"{mes_es} {año}", f"{mes_en} {año}", año
                
            except:
                pass
            
            # Fallback
            return "Diciembre 2024", "December 2024", "2024"
        
        # Obtener período de cierre
        periodo_raw = metadata.get("periodo", "Periodo desconocido")
        periodo_es, periodo_en, año = obtener_periodo_cierre(periodo_raw)
        
        # Crear textos dinámicos para saldos
        if is_english:
            initial_balance_text = f"Initial Balance as of January 1, {año}"
            final_balance_text = f"Final Balance as of {periodo_en}"
        else:
            initial_balance_text = f"Saldo Inicial al 1 de Enero {año}"
            final_balance_text = f"Saldo Final a {periodo_es}"
        
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = self._get_text('title_ecp', language)
        
        # Título principal
        title_text = self._get_text('title_ecp', language)
        self._apply_header_style(ws, 1, 1, 7, title_text)
        
        # Información del período
        current_row = 3
        ws.cell(row=current_row, column=1, value=f"{self._get_text('client', language)}: {metadata.get('cliente_nombre', 'N/A')}")
        ws.cell(row=current_row, column=5, value=f"{self._get_text('period', language)}: {metadata.get('periodo', 'N/A')}")
        current_row += 1
        ws.cell(row=current_row, column=1, value=f"{self._get_text('currency', language)}: {metadata.get('moneda', 'CLP')}")
        ws.cell(row=current_row, column=5, value=f"{self._get_text('date', language)}: {datetime.now().strftime('%d/%m/%Y')}")
        current_row += 2
        
        # Encabezados de columnas
        headers = [
            self._get_text('concept', language),
            self._get_text('capital', language),
            self._get_text('other_reserves', language),
            self._get_text('accumulated_results', language),
            self._get_text('attributable_capital', language),
            self._get_text('non_controlling_interests', language),
            self._get_text('totals', language).replace(':', '')
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = self.styles['header']
            cell.fill = self.fills['header']
            cell.alignment = Alignment(horizontal='center')
        
        current_row += 1
        
        if data_ecp:
            # Debug de la estructura de datos ECP
            if logger.isEnabledFor(logging.INFO):
                self._debug_data_structure(data_ecp, "DATA_ECP_COMPLETA")
            
            # Procesar datos del ECP - manejar diferentes estructuras posibles
            patrimonio_data = data_ecp.get("patrimonio", data_ecp)  # Puede estar directamente en data_ecp
            
            # Extraer datos de Capital
            capital_data = patrimonio_data.get("capital", {})
            if isinstance(capital_data, dict):
                capital_inicial = capital_data.get("saldo_inicial", capital_data.get("saldo_anterior", 0))
                capital_cambios = capital_data.get("cambios", capital_data.get("movimientos", 0))
                capital_final = capital_data.get("saldo_final", capital_inicial + capital_cambios)
            else:
                capital_inicial = capital_cambios = capital_final = 0
            
            # Extraer datos de Otras Reservas
            otras_reservas_data = patrimonio_data.get("otras_reservas", patrimonio_data.get("reservas", {}))
            if isinstance(otras_reservas_data, dict):
                otras_reservas_inicial = otras_reservas_data.get("saldo_inicial", otras_reservas_data.get("saldo_anterior", 0))
                otras_reservas_cambios = otras_reservas_data.get("cambios", otras_reservas_data.get("movimientos", 0))
                otras_reservas_final = otras_reservas_data.get("saldo_final", otras_reservas_inicial + otras_reservas_cambios)
            else:
                otras_reservas_inicial = otras_reservas_cambios = otras_reservas_final = 0
            
            # Extraer datos de Resultados Acumulados
            resultados_data = patrimonio_data.get("resultados_acumulados", patrimonio_data.get("utilidades_retenidas", {}))
            if isinstance(resultados_data, dict):
                resultados_inicial = resultados_data.get("saldo_inicial", resultados_data.get("saldo_anterior", 0))
                resultados_cambios = resultados_data.get("cambios", resultados_data.get("movimientos", 0))
                resultados_final = resultados_data.get("saldo_final", resultados_inicial + resultados_cambios)
            else:
                resultados_inicial = resultados_cambios = resultados_final = 0
            
            # Resultado del ejercicio del ERI
            total_eri = 0
            if data_eri:
                # Calcular total del ERI
                for bloque_key in ["ganancias_brutas", "ganancia_perdida", "ganancia_perdida_antes_impuestos"]:
                    bloque_data = data_eri.get(bloque_key)
                    if bloque_data and "total" in bloque_data:
                        total_eri += bloque_data["total"]
            
            # Datos de las filas
            filas_ecp = [
                {
                    "Concepto": initial_balance_text,
                    "Capital": capital_inicial,
                    "Otras Reservas": otras_reservas_inicial,
                    "Resultados Acumulados": resultados_inicial,
                    "Capital Atribuible": capital_inicial + otras_reservas_inicial + resultados_inicial,
                    "Participaciones no Controladoras": 0,
                    "Total": capital_inicial + otras_reservas_inicial + resultados_inicial
                },
                {
                    "Concepto": self._get_text('period_result_ecp', language),
                    "Capital": 0,
                    "Otras Reservas": 0,
                    "Resultados Acumulados": total_eri,
                    "Capital Atribuible": total_eri,
                    "Participaciones no Controladoras": 0,
                    "Total": total_eri
                },
                {
                    "Concepto": self._get_text('other_changes', language),
                    "Capital": capital_cambios,
                    "Otras Reservas": otras_reservas_cambios,
                    "Resultados Acumulados": resultados_cambios,
                    "Capital Atribuible": capital_cambios + otras_reservas_cambios + resultados_cambios,
                    "Participaciones no Controladoras": 0,
                    "Total": capital_cambios + otras_reservas_cambios + resultados_cambios
                },
                {
                    "Concepto": final_balance_text,
                    "Capital": capital_inicial + capital_cambios,
                    "Otras Reservas": otras_reservas_inicial + otras_reservas_cambios,
                    "Resultados Acumulados": resultados_inicial + resultados_cambios + total_eri,
                    "Capital Atribuible": capital_inicial + capital_cambios + otras_reservas_inicial + otras_reservas_cambios + resultados_inicial + resultados_cambios + total_eri,
                    "Participaciones no Controladoras": 0,
                    "Total": capital_inicial + capital_cambios + otras_reservas_inicial + otras_reservas_cambios + resultados_inicial + resultados_cambios + total_eri
                }
            ]
            
            # Agregar filas de datos
            for fila in filas_ecp:
                for col_idx, (key, value) in enumerate(fila.items(), 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    if col_idx == 1:  # Concepto
                        cell.value = value
                        cell.font = self.styles['subheader']
                    else:  # Valores numéricos
                        cell.value = self._format_amount(value, metadata.get("moneda", "CLP"))
                        cell.font = self.styles['data']
                    
                    cell.border = self.borders['thin']
                    
                    # Resaltar solo la fila de saldo final con color verde oscuro
                    final_balance_concepts = [
                        final_balance_text,  # Usar el texto dinámico
                        "Saldo Final"  # Compatibilidad
                    ]
                    
                    if fila["Concepto"] in final_balance_concepts:
                        cell.fill = self.fills['final_balance']    # Verde oscuro solo para saldo final
                        cell.font = self.styles['total']
                
                current_row += 1
        
        # Ajustar anchos de columna
        for col_idx in range(1, 8):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 18
        
        # Agregar hoja de metadatos
        self._add_metadata_sheet(workbook, metadata, language)
        
        return workbook

    def _format_amount(self, amount, moneda="CLP"):
        """Formatear monto según la moneda (sin sufijo de moneda)"""
        if pd.isna(amount) or amount is None:
            return 0
        
        try:
            amount = float(amount)
        except (ValueError, TypeError):
            return 0
        
        if moneda == "USD":
            return f"${amount:,.2f}"
        elif moneda == "EUR":
            return f"€{amount:,.2f}"
        else:  # CLP por defecto
            return f"${amount:,.0f}"

    def workbook_to_bytes(self, workbook):
        """Convertir workbook a bytes para descarga"""
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


# Instancia global del generador
excel_generator = ExcelTemplateGenerator()
