"""
Template Excel específico para Estado de Situación Financiera (ESF)
"""

import openpyxl
import logging
from .base import BaseExcelTemplate
from .translations import detect_language_improved

logger = logging.getLogger(__name__)


class ESFTemplate(BaseExcelTemplate):
    """Template Excel para Estado de Situación Financiera"""

    def generate(self, data_esf, metadata, data_eri=None):
        """Generar template Excel para Estado de Situación Financiera"""
        # Obtener idioma usando detección mejorada
        language = detect_language_improved(metadata)
        
        # Debug de la estructura de datos
        if data_esf and logger.isEnabledFor(logging.INFO):
            self._debug_data_structure(data_esf, "DATA_ESF_COMPLETA")
            if "patrimonio" in data_esf:
                self._debug_data_structure(data_esf["patrimonio"], "PATRIMONIO")
        
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = self._get_text('title_esf', language)
        
        # Título principal
        title_text = self._get_text('title_esf', language)
        ws.cell(row=1, column=1, value=title_text)
        self._apply_header_style(ws, 1, 1, 4, title_text)
        
        # Información del período
        current_row = self._add_period_info(ws, 3, metadata, language)
        
        # Extraer y procesar datos
        if data_esf:
            current_row = self._add_assets_section(ws, current_row, data_esf, metadata, language)
            current_row = self._add_liabilities_equity_section(ws, current_row, data_esf, metadata, language, data_eri)
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        
        # Agregar hoja de metadatos
        self._add_metadata_sheet(workbook, metadata, language)
        
        return workbook

    def _add_assets_section(self, ws, current_row, data_esf, metadata, language):
        """Agregar sección de ACTIVOS"""
        # ACTIVOS
        ws.cell(row=current_row, column=1, value=self._get_text('assets', language)).font = self.styles['header']
        ws.cell(row=current_row, column=1).fill = self.fills['header']
        current_row += 1
        
        # Activos Corrientes
        activos_corrientes = data_esf.get("activos", {}).get("corrientes", {})
        current_row = self._add_esf_section(ws, current_row, self._get_text('current_assets', language), 
                                          activos_corrientes, metadata.get("moneda", "CLP"), language)
        
        # Activos No Corrientes
        activos_no_corrientes = data_esf.get("activos", {}).get("no_corrientes", {})
        current_row = self._add_esf_section(ws, current_row, self._get_text('non_current_assets', language), 
                                          activos_no_corrientes, metadata.get("moneda", "CLP"), language)
        
        # Total Activos
        total_activos = data_esf.get("activos", {}).get("total_activos", 0)
        if total_activos != 0:
            current_row += 1
            ws.cell(row=current_row, column=1, value=self._get_text('total_assets', language)).font = self.styles['total']
            ws.cell(row=current_row, column=2, value=self._format_amount(total_activos, metadata.get("moneda", "CLP"))).font = self.styles['total']
            ws.cell(row=current_row, column=1).fill = self.fills['total']
            ws.cell(row=current_row, column=2).fill = self.fills['total']
        
        return current_row + 1

    def _add_liabilities_equity_section(self, ws, current_row, data_esf, metadata, language, data_eri=None):
        """Agregar sección de PASIVOS Y PATRIMONIO"""
        # PASIVOS Y PATRIMONIO
        ws.cell(row=current_row, column=1, value=self._get_text('liabilities_equity', language)).font = self.styles['header']
        ws.cell(row=current_row, column=1).fill = self.fills['header']
        current_row += 1
        
        # Pasivos Corrientes
        pasivos_corrientes = data_esf.get("pasivos", {}).get("corrientes", {})
        current_row = self._add_esf_section(ws, current_row, self._get_text('current_liabilities', language), 
                                          pasivos_corrientes, metadata.get("moneda", "CLP"), language)
        
        # Pasivos No Corrientes
        pasivos_no_corrientes = data_esf.get("pasivos", {}).get("no_corrientes", {})
        current_row = self._add_esf_section(ws, current_row, self._get_text('non_current_liabilities', language), 
                                          pasivos_no_corrientes, metadata.get("moneda", "CLP"), language)
        
        # Total Pasivos
        total_pasivos = data_esf.get("pasivos", {}).get("total_pasivos", 0)
        if total_pasivos != 0:
            current_row += 1
            ws.cell(row=current_row, column=1, value=self._get_text('total_liabilities', language)).font = self.styles['total']
            ws.cell(row=current_row, column=2, value=self._format_amount(total_pasivos, metadata.get("moneda", "CLP"))).font = self.styles['total']
            ws.cell(row=current_row, column=1).fill = self.fills['total']
            ws.cell(row=current_row, column=2).fill = self.fills['total']
            current_row += 2
        
        # Patrimonio
        current_row = self._add_equity_section(ws, current_row, data_esf, metadata, language, data_eri)
        
        # TOTAL FINAL: PASIVOS Y PATRIMONIO
        total_pasivos = data_esf.get("pasivos", {}).get("total_pasivos", 0)
        total_patrimonio = self._calcular_total_patrimonio(data_esf, data_eri)
        total_pasivos_patrimonio = total_pasivos + total_patrimonio
        
        current_row += 1
        ws.cell(row=current_row, column=1, value=self._get_text('total_liabilities_equity', language)).font = self.styles['total']
        ws.cell(row=current_row, column=2, value=self._format_amount(total_pasivos_patrimonio, metadata.get("moneda", "CLP"))).font = self.styles['total']
        ws.cell(row=current_row, column=1).fill = self.fills['total']
        ws.cell(row=current_row, column=2).fill = self.fills['total']
        
        return current_row

    def _add_equity_section(self, ws, current_row, data_esf, metadata, language, data_eri=None):
        """Agregar sección de PATRIMONIO"""
        patrimonio = data_esf.get("patrimonio", {})
        if not patrimonio:
            return current_row
        
        # Título principal de Patrimonio
        ws.cell(row=current_row, column=1, value=self._get_text('equity', language)).font = self.styles['header']
        ws.cell(row=current_row, column=1).fill = self.fills['header']
        current_row += 1
        
        # Procesar cada subcategoría de patrimonio (ej: capital)
        total_patrimonio = 0
        for subcategoria_key, subcategoria_data in patrimonio.items():
            if isinstance(subcategoria_data, dict) and subcategoria_key not in ['total_patrimonio']:
                # Usar nombre en idioma apropiado si está disponible
                if language.lower() in ['en', 'english', 'inglés']:
                    section_name = subcategoria_data.get('nombre_en', subcategoria_data.get('nombre_es', subcategoria_key))
                else:
                    section_name = subcategoria_data.get('nombre_es', subcategoria_data.get('nombre_en', subcategoria_key))
                current_row = self._add_esf_section(ws, current_row, section_name, subcategoria_data, metadata.get("moneda", "CLP"), language)
                total_patrimonio += subcategoria_data.get('total', 0)
        
        # Agregar Ganancia/(Pérdida) del Ejercicio del ERI
        total_eri = self._calcular_total_eri(data_eri)
        if total_eri != 0:
            current_row = self._add_eri_result_section(ws, current_row, total_eri, metadata, language)
            total_patrimonio += total_eri
        
        # Total patrimonio
        if total_patrimonio != 0:
            current_row += 1
            ws.cell(row=current_row, column=1, value=self._get_text('total_equity', language)).font = self.styles['total']
            ws.cell(row=current_row, column=2, value=self._format_amount(total_patrimonio, metadata.get("moneda", "CLP"))).font = self.styles['total']
            ws.cell(row=current_row, column=1).fill = self.fills['total']
            ws.cell(row=current_row, column=2).fill = self.fills['total']
            current_row += 2
        
        return current_row

    def _add_eri_result_section(self, ws, current_row, total_eri, metadata, language):
        """Agregar sección de resultado del ERI"""
        # Título de la subcategoría
        ws.cell(row=current_row, column=1, value=self._get_text('period_result', language)).font = self.styles['subheader']
        ws.cell(row=current_row, column=1).fill = self.fills['subheader']
        current_row += 1
        
        # Línea del resultado
        ganancia_perdida_texto = self._get_text('profit_loss', language) if total_eri > 0 else self._get_text('loss', language)
        ws.cell(row=current_row, column=1, value=f"  {ganancia_perdida_texto} (Del ERI)")
        ws.cell(row=current_row, column=2, value=self._format_amount(total_eri, metadata.get("moneda", "CLP")))
        current_row += 1
        
        # Total de resultado del ejercicio
        ws.cell(row=current_row, column=1, value=self._get_text('total_period_result', language)).font = self.styles['total']
        ws.cell(row=current_row, column=2, value=self._format_amount(total_eri, metadata.get("moneda", "CLP"))).font = self.styles['total']
        ws.cell(row=current_row, column=1).fill = self.fills['total']
        ws.cell(row=current_row, column=2).fill = self.fills['total']
        
        return current_row + 2

    def _add_esf_section(self, ws, start_row, section_title, section_data, moneda, language='es'):
        """Agregar una sección del ESF al worksheet con grupos colapsables"""
        current_row = start_row
        
        # Título de la sección
        ws.cell(row=current_row, column=1, value=section_title).font = self.styles['subheader']
        ws.cell(row=current_row, column=1).fill = self.fills['subheader']
        current_row += 1
        
        total_section = 0
        
        if isinstance(section_data, dict):
            if "grupos" in section_data and section_data["grupos"]:
                current_row = self._add_groups_section(ws, current_row, section_data, moneda, language)
                total_section = section_data.get("total", 0)
            elif "cuentas" in section_data and section_data["cuentas"]:
                current_row = self._add_accounts_section(ws, current_row, section_data, moneda, language)
                total_section = sum(cuenta.get("saldo_final", 0) for cuenta in section_data.get("cuentas", []))
            else:
                # Si no hay grupos ni cuentas, mostrar mensaje
                ws.cell(row=current_row, column=1, value=f"  {self._get_text('no_movements', language)}").font = self.styles['metadata']
                current_row += 1
            
            # Solo mostrar total si hay datos
            if total_section != 0:
                ws.cell(row=current_row, column=1, value=f"TOTAL {section_title.upper()}").font = self.styles['total']
                ws.cell(row=current_row, column=2, value=self._format_amount(total_section, moneda)).font = self.styles['total']
                ws.cell(row=current_row, column=1).fill = self.fills['total']
                ws.cell(row=current_row, column=2).fill = self.fills['total']
                current_row += 2
        else:
            # Si section_data no es un diccionario válido
            ws.cell(row=current_row, column=1, value=f"  {self._get_text('no_data', language)}").font = self.styles['metadata']
            current_row += 2
        
        return current_row

    def _add_groups_section(self, ws, current_row, section_data, moneda, language):
        """Agregar sección con grupos colapsables"""
        for grupo_nombre, grupo_data in section_data.get("grupos", {}).items():
            if isinstance(grupo_data, dict):
                # Título del grupo - usar nombre según idioma
                if language.lower() in ['en', 'english', 'inglés']:
                    grupo_display_name = grupo_data.get('nombre_en', grupo_data.get('nombre_es', grupo_nombre))
                else:
                    grupo_display_name = grupo_data.get('nombre_es', grupo_data.get('nombre_en', grupo_nombre))
                
                ws.cell(row=current_row, column=1, value=f"  {grupo_display_name}").font = self.styles['data']
                ws.cell(row=current_row, column=2, value=self._format_amount(grupo_data.get('total', 0), moneda)).font = self.styles['data']
                current_row += 1
                
                # Filas de inicio y fin para el grupo
                grupo_start_row = current_row
                
                # Agregar cuentas del grupo
                for cuenta in grupo_data.get("cuentas", []):
                    codigo = cuenta.get("codigo", "")
                    nombre = self._get_account_name(cuenta, language)
                    saldo = cuenta.get("saldo_final", 0)
                    
                    ws.cell(row=current_row, column=1, value=f"    {codigo} - {nombre}")
                    ws.cell(row=current_row, column=2, value=self._format_amount(saldo, moneda))
                    current_row += 1
                
                # Crear grupo colapsable solo si hay cuentas
                if grupo_data.get("cuentas"):
                    self._create_group_collapsible(ws, grupo_start_row, current_row)
                    current_row += 1
        
        return current_row

    def _add_accounts_section(self, ws, current_row, section_data, moneda, language):
        """Agregar sección con cuentas directas"""
        for cuenta in section_data.get("cuentas", []):
            codigo = cuenta.get("codigo", "")
            nombre = self._get_account_name(cuenta, language)
            saldo = cuenta.get("saldo_final", 0)
            
            ws.cell(row=current_row, column=1, value=f"  {codigo} - {nombre}")
            ws.cell(row=current_row, column=2, value=self._format_amount(saldo, moneda))
            current_row += 1
        
        return current_row

    def _calcular_total_eri(self, data_eri):
        """Calcular el total de ganancia/pérdida del ERI"""
        if not data_eri:
            return 0
        
        total_eri = 0
        # Sumar todos los bloques del ERI
        bloques_eri = [
            "ganancias_brutas",
            "ganancia_perdida", 
            "ganancia_perdida_antes_impuestos"
        ]
        
        for bloque_key in bloques_eri:
            bloque_data = data_eri.get(bloque_key, {})
            if isinstance(bloque_data, dict) and "total" in bloque_data:
                total_eri += bloque_data["total"]
        
        return total_eri

    def _calcular_total_patrimonio(self, data_esf, data_eri=None):
        """Calcular el total del patrimonio incluyendo ERI"""
        total_patrimonio = 0
        patrimonio = data_esf.get("patrimonio", {})
        
        for subcategoria_key, subcategoria_data in patrimonio.items():
            if isinstance(subcategoria_data, dict) and subcategoria_key not in ['total_patrimonio']:
                total_patrimonio += subcategoria_data.get('total', 0)
        
        # Agregar resultado del ERI
        total_patrimonio += self._calcular_total_eri(data_eri)
        
        return total_patrimonio
