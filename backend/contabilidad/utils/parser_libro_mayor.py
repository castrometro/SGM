from openpyxl import load_workbook
import re
from django.utils import timezone
import datetime

from contabilidad.models import (
    CuentaContable,
    CentroCosto,
    Auxiliar,
    TipoDocumento,
    AperturaCuenta,
    MovimientoContable,
)

def parsear_libro_mayor(ruta_excel, upload):
    from contabilidad.models import (
        CuentaContable, CentroCosto, Auxiliar, TipoDocumento,
        AperturaCuenta, MovimientoContable
    )
    wb = load_workbook(ruta_excel, read_only=True, data_only=True)
    ws = wb.active
    # Al inicio de la función:
    cuentas_existentes = set(
        CuentaContable.objects.filter(cliente=upload.cierre.cliente)
        .values_list('codigo', flat=True)
    )
    cuentas_nuevas = set()
    total_cuentas = set()
    # Función para normalizar encabezados
    def clean(h):
        h = h.strip().upper()
        for a, b in [("Á","A"),("É","E"),("Í","I"),("Ó","O"),("Ú","U"),("Ñ","N")]:
            h = h.replace(a, b)
        return re.sub(r"[^A-Z0-9]", "", h)

    # Leer fila 9 y construir índice de columnas
    raw = next(ws.iter_rows(min_row=9, max_row=9, values_only=True))
    idx = { clean(h): i for i, h in enumerate(raw) if isinstance(h, str) }

    # Índices requeridos según el Excel
    C  = idx["CUENTA"]
    F  = idx["FECHA"]
    NC = idx.get("NCOMPROBANTE")
    T  = idx.get("TIPO")
    NI = idx.get("NINTERNO")
    CC = idx.get("CENTRODECOSTO")
    AX = idx.get("AUXILIAR")
    TD = idx.get("TIPODOC")
    ND = idx.get("NUMERODOC")
    DG = idx.get("DETDEGASTOINSTFINANCIERO")
    D  = idx["DEBE"]
    H  = idx["HABER"]
    S  = idx["SALDO"]
    DS = idx["DESCRIPCION"]

    errores = []
    current_code = None
    fechas_movimientos = []  # <- lista de todas las fechas válidas de movimientos

    for row_idx, row in enumerate(ws.iter_rows(min_row=11, values_only=True), start=11):
        cell = row[C]

        # 1) Apertura de cuenta
        if isinstance(cell, str) and cell.startswith("SALDO ANTERIOR"):
            try:
                resto = cell.split(":", 1)[1].strip()
                code, name = resto.split(" ", 1)
                saldo_ant = row[S] or 0

                cuenta_obj, created = CuentaContable.objects.get_or_create(
                    cliente=upload.cierre.cliente,
                    codigo=code,
                    defaults={"nombre": name}
                )
                total_cuentas.add(code)
                if created:
                    cuentas_nuevas.add(code)
                AperturaCuenta.objects.update_or_create(
                    cierre=upload.cierre,
                    cuenta=cuenta_obj,
                    defaults={"saldo_anterior": saldo_ant}
                )
                current_code = code
            except Exception as e:
                errores.append(f"F{row_idx} Apertura: {e}")
            continue

        # 2) Movimiento
        if current_code and row[F] is not None:
            # Parseo de fecha robusto
            raw_fecha = row[F]
            try:
                if isinstance(raw_fecha, str):
                    s = raw_fecha.strip().strip('"').strip("'").replace('“','').replace('”','')
                    try:
                        fecha = datetime.datetime.strptime(s, "%d/%m/%Y").date()
                    except ValueError:
                        fecha = datetime.datetime.strptime(s, "%Y-%m-%d").date()
                elif isinstance(raw_fecha, datetime.datetime):
                    fecha = raw_fecha.date()
                elif isinstance(raw_fecha, datetime.date):
                    fecha = raw_fecha
                else:
                    raise ValueError(f"Formato de fecha no soportado: {raw_fecha!r}")
                fechas_movimientos.append(fecha)  # <--- Guardar todas las fechas
            except Exception as e:
                errores.append(f"F{row_idx} Fecha: {e}")
                continue

            try:
                cuenta_obj = CuentaContable.objects.get(
                    cliente=upload.cierre.cliente,
                    codigo=current_code
                )

                # CentroCosto por nombre
                centro_obj = None
                if row[CC]:
                    nombre_cc = str(row[CC]).strip()
                    centro_obj, _ = CentroCosto.objects.get_or_create(
                        cliente=upload.cierre.cliente,
                        nombre=nombre_cc
                    )

                # Auxiliar por rut_auxiliar
                aux_obj = None
                if row[AX]:
                    rut = str(row[AX]).strip()
                    aux_obj, _ = Auxiliar.objects.get_or_create(
                        rut_auxiliar=rut,
                        defaults={
                            "nombre": "",
                            "fecha_creacion": timezone.now()
                        }
                    )

                # TipoDocumento por cliente + código
                td_obj = None
                if row[TD]:
                    td_obj, _ = TipoDocumento.objects.get_or_create(
                        cliente=upload.cierre.cliente,
                        codigo=str(row[TD]).strip(),
                        defaults={"descripcion": ""}
                    )

                MovimientoContable.objects.create(                   
                    cierre=upload.cierre,
                    cuenta=cuenta_obj,
                    fecha=fecha,
                    tipo_documento=td_obj,
                    numero_documento=str(row[ND] or ""),
                    tipo=str(row[T] or ""),
                    numero_comprobante=str(row[NC] or ""),
                    numero_interno=str(row[NI] or ""),
                    centro_costo=centro_obj,
                    auxiliar=aux_obj,
                    detalle_gasto=str(row[DG] or ""),
                    debe=row[D] or 0,
                    haber=row[H] or 0,
                    descripcion=str(row[DS] or ""),
                )
            except Exception as e:
                errores.append(f"F{row_idx} Movimiento: {e}")

    # Detectar fechas mínima y máxima de movimientos
    fecha_inicio_libro = min(fechas_movimientos) if fechas_movimientos else None
    fecha_fin_libro = max(fechas_movimientos) if fechas_movimientos else None
    resumen = {
        "total_cuentas": len(total_cuentas),
        "cuentas_nuevas": len(cuentas_nuevas),
        "codigos_cuentas_nuevas": list(cuentas_nuevas)
    }
    return errores, fecha_inicio_libro, fecha_fin_libro, resumen
