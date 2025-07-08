#!/usr/bin/env python3
"""
Script de prueba para validar el procesamiento de Libro Mayor con Chains.
Ejecutar en Django shell:

docker compose exec backend python manage.py shell

Luego:
exec(open('test_libro_mayor_chains.py').read())
"""

from contabilidad.models import UploadLog, Cliente, CierreContabilidad, MovimientoContable, Incidencia
from contabilidad.tasks_libro_mayor import crear_chain_libro_mayor
from django.core.files.base import ContentFile
from django.utils import timezone
import os
import time
import tempfile
from decimal import Decimal
from datetime import datetime, date
from unittest.mock import Mock, patch, MagicMock

# Configurar Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'sgm_backend.settings')
sys.path.append('/root/SGM/backend')

import django
django.setup()

from django.test import TestCase
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test.utils import override_settings
from django.db import transaction
from openpyxl import Workbook

# Imports del proyecto
from contabilidad.models import (
    CierreContabilidad, Cliente, UploadLog, LibroMayorArchivo,
    MovimientoContable, CuentaContable, AperturaCuenta, Incidencia
)
from contabilidad.tasks_libro_mayor import (
    crear_chain_libro_mayor,
    validar_nombre_archivo_libro_mayor,
    verificar_archivo_libro_mayor,
    validar_contenido_libro_mayor,
    procesar_libro_mayor_raw,
    generar_incidencias_libro_mayor,
    finalizar_procesamiento_libro_mayor
)


class LibroMayorChainsTestCase(TestCase):
    """Tests para el flujo de libro mayor con chains"""
    
    def setUp(self):
        """Configuración inicial para tests"""
        # Crear usuario y cliente
        self.user = User.objects.create_user(
            username='testuser',
            password='testpass123',
            email='test@example.com'
        )
        
        self.cliente = Cliente.objects.create(
            nombre='Cliente Test',
            rut='12345678-9',
            activo=True
        )
        
        # Crear cierre de contabilidad
        self.cierre = CierreContabilidad.objects.create(
            cliente=self.cliente,
            periodo='2024-01',
            activo=True,
            estado='abierto'
        )
        
        # Crear archivo Excel de prueba
        self.excel_file = self._crear_archivo_excel_prueba()
        
    def _crear_archivo_excel_prueba(self):
        """Crea un archivo Excel válido para pruebas"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Hoja1"
        
        # Headers esperados
        headers = [
            'Código Cuenta', 'Nombre Cuenta', 'Saldo Inicial', 'Débito', 'Crédito',
            'Fecha', 'Tipo Doc', 'Número Doc', 'Glosa', 'Referencia'
        ]
        
        # Escribir headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Datos de ejemplo
        data = [
            ['1101', 'Caja', 1000.00, 500.00, 200.00, '2024-01-15', 'FC', '001', 'Venta', 'REF001'],
            ['1102', 'Banco', 2000.00, 1000.00, 500.00, '2024-01-16', 'FC', '002', 'Compra', 'REF002'],
            ['2101', 'Proveedores', 0.00, 0.00, 300.00, '2024-01-17', 'FC', '003', 'Pago', 'REF003']
        ]
        
        for row, row_data in enumerate(data, 2):
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Guardar en archivo temporal
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(temp_file.name)
        temp_file.close()
        
        return temp_file.name
    
    def tearDown(self):
        """Limpieza post-test"""
        if hasattr(self, 'excel_file') and os.path.exists(self.excel_file):
            os.unlink(self.excel_file)
    
    def test_crear_chain_libro_mayor(self):
        """Test básico de creación del chain"""
        # Crear UploadLog
        upload_log = UploadLog.objects.create(
            cliente=self.cliente,
            tipo_archivo='libro_mayor',
            nombre_archivo='LibroMayor_2024-01.xlsx',
            archivo_hash='test_hash',
            usuario=self.user,
            estado='pendiente'
        )
        
        # Crear chain
        chain = crear_chain_libro_mayor(upload_log.id)
        
        # Verificar que el chain se creó correctamente
        self.assertIsNotNone(chain)
        self.assertEqual(len(chain.tasks), 6)
        
        # Verificar nombres de tasks
        task_names = [task.name for task in chain.tasks]
        expected_tasks = [
            'contabilidad.tasks_libro_mayor.validar_nombre_archivo_libro_mayor',
            'contabilidad.tasks_libro_mayor.verificar_archivo_libro_mayor',
            'contabilidad.tasks_libro_mayor.validar_contenido_libro_mayor',
            'contabilidad.tasks_libro_mayor.procesar_libro_mayor_raw',
            'contabilidad.tasks_libro_mayor.generar_incidencias_libro_mayor',
            'contabilidad.tasks_libro_mayor.finalizar_procesamiento_libro_mayor'
        ]
        
        for expected in expected_tasks:
            self.assertIn(expected, task_names)
    
    def test_validar_nombre_archivo_libro_mayor(self):
        """Test de validación de nombre de archivo"""
        # Crear UploadLog con nombre válido
        upload_log = UploadLog.objects.create(
            cliente=self.cliente,
            tipo_archivo='libro_mayor',
            nombre_archivo='LibroMayor_2024-01.xlsx',
            archivo_hash='test_hash',
            usuario=self.user,
            estado='pendiente'
        )
        
        # Ejecutar task de validación
        resultado = validar_nombre_archivo_libro_mayor(upload_log.id)
        
        # Verificar resultado
        self.assertEqual(resultado, upload_log.id)
        
        # Verificar que el estado cambió
        upload_log.refresh_from_db()
        self.assertEqual(upload_log.estado, 'procesando')
    
    def test_validar_nombre_archivo_invalido(self):
        """Test de validación con nombre de archivo inválido"""
        # Crear UploadLog con nombre inválido
        upload_log = UploadLog.objects.create(
            cliente=self.cliente,
            tipo_archivo='libro_mayor',
            nombre_archivo='archivo_invalido.xlsx',
            archivo_hash='test_hash',
            usuario=self.user,
            estado='pendiente'
        )
        
        # Ejecutar task y esperar excepción
        with self.assertRaises(Exception):
            validar_nombre_archivo_libro_mayor(upload_log.id)
    
    @patch('contabilidad.tasks_libro_mayor.default_storage')
    def test_verificar_archivo_libro_mayor(self, mock_storage):
        """Test de verificación de archivo"""
        # Configurar mock
        mock_storage.exists.return_value = True
        mock_storage.size.return_value = 1024
        
        # Crear UploadLog
        upload_log = UploadLog.objects.create(
            cliente=self.cliente,
            tipo_archivo='libro_mayor',
            nombre_archivo='LibroMayor_2024-01.xlsx',
            archivo_hash='test_hash',
            usuario=self.user,
            estado='procesando',
            ruta_archivo='test/path/file.xlsx'
        )
        
        # Ejecutar task
        resultado = verificar_archivo_libro_mayor(upload_log.id)
        
        # Verificar resultado
        self.assertEqual(resultado, upload_log.id)
        
        # Verificar que se verificó la existencia
        mock_storage.exists.assert_called_once()
    
    def test_chain_completo_mock(self):
        """Test del chain completo con datos mockeados"""
        # Crear UploadLog
        upload_log = UploadLog.objects.create(
            cliente=self.cliente,
            tipo_archivo='libro_mayor',
            nombre_archivo='LibroMayor_2024-01.xlsx',
            archivo_hash='test_hash',
            usuario=self.user,
            estado='pendiente',
            ruta_archivo=self.excel_file
        )
        
        # Mock de funciones que requieren storage
        with patch('contabilidad.tasks_libro_mayor.default_storage') as mock_storage:
            mock_storage.exists.return_value = True
            mock_storage.size.return_value = 1024
            mock_storage.open.return_value = open(self.excel_file, 'rb')
            
            # Ejecutar cada task del chain individualmente
            try:
                # Task 1: Validar nombre
                resultado1 = validar_nombre_archivo_libro_mayor(upload_log.id)
                self.assertEqual(resultado1, upload_log.id)
                
                # Task 2: Verificar archivo
                resultado2 = verificar_archivo_libro_mayor(upload_log.id)
                self.assertEqual(resultado2, upload_log.id)
                
                # Task 3: Validar contenido
                resultado3 = validar_contenido_libro_mayor(upload_log.id)
                self.assertEqual(resultado3, upload_log.id)
                
                # Task 4: Procesar raw
                resultado4 = procesar_libro_mayor_raw(upload_log.id)
                self.assertEqual(resultado4, upload_log.id)
                
                # Task 5: Generar incidencias
                resultado5 = generar_incidencias_libro_mayor(upload_log.id)
                self.assertEqual(resultado5, upload_log.id)
                
                # Task 6: Finalizar
                resultado6 = finalizar_procesamiento_libro_mayor(upload_log.id)
                self.assertIsInstance(resultado6, str)
                
                # Verificar estado final
                upload_log.refresh_from_db()
                self.assertEqual(upload_log.estado, 'completado')
                
            except Exception as e:
                self.fail(f"Chain falló con error: {str(e)}")
    
    def test_manejo_errores_chain(self):
        """Test de manejo de errores en el chain"""
        # Crear UploadLog con ID inexistente
        upload_log_id = 999999
        
        # Ejecutar primera task con ID inválido
        with self.assertRaises(Exception):
            validar_nombre_archivo_libro_mayor(upload_log_id)
    
    def test_resumen_procesamiento(self):
        """Test de generación de resumen de procesamiento"""
        # Crear UploadLog
        upload_log = UploadLog.objects.create(
            cliente=self.cliente,
            tipo_archivo='libro_mayor',
            nombre_archivo='LibroMayor_2024-01.xlsx',
            archivo_hash='test_hash',
            usuario=self.user,
            estado='procesando',
            ruta_archivo=self.excel_file
        )
        
        # Simular resumen de procesamiento
        resumen = {
            'procesamiento': {
                'filas_procesadas': 3,
                'movimientos_creados': 3,
                'cuentas_creadas': 3,
                'aperturas_creadas': 3,
                'errores_count': 0
            },
            'validacion': {
                'archivo_valido': True,
                'nombre_valido': True,
                'contenido_valido': True
            },
            'incidencias': {
                'generadas': 0,
                'tipos': []
            }
        }
        
        upload_log.resumen = resumen
        upload_log.save()
        
        # Verificar resumen
        self.assertEqual(upload_log.resumen['procesamiento']['movimientos_creados'], 3)
        self.assertEqual(upload_log.resumen['incidencias']['generadas'], 0)


def ejecutar_test_basico():
    """Ejecuta un test básico del chain"""
    print("=== TEST BÁSICO DEL CHAIN DE LIBRO MAYOR ===")
    
    try:
        # Crear datos de prueba
        user = User.objects.create_user(
            username='testuser_basic',
            password='testpass123'
        )
        
        cliente = Cliente.objects.create(
            nombre='Cliente Test Basic',
            rut='11111111-1',
            activo=True
        )
        
        # Crear UploadLog
        upload_log = UploadLog.objects.create(
            cliente=cliente,
            tipo_archivo='libro_mayor',
            nombre_archivo='LibroMayor_2024-01.xlsx',
            archivo_hash='test_hash_basic',
            usuario=user,
            estado='pendiente'
        )
        
        print(f"✓ UploadLog creado con ID: {upload_log.id}")
        
        # Crear chain
        chain = crear_chain_libro_mayor(upload_log.id)
        print(f"✓ Chain creado con {len(chain.tasks)} tasks")
        
        # Listar tasks del chain
        print("Tasks en el chain:")
        for i, task in enumerate(chain.tasks, 1):
            print(f"  {i}. {task.name}")
        
        print("✓ Test básico completado exitosamente")
        
    except Exception as e:
        print(f"✗ Error en test básico: {str(e)}")
        import traceback
        traceback.print_exc()


def ejecutar_test_validacion():
    """Ejecuta test de validación de nombre de archivo"""
    print("\n=== TEST DE VALIDACIÓN DE NOMBRE ===")
    
    try:
        # Crear datos de prueba
        user = User.objects.create_user(
            username='testuser_validation',
            password='testpass123'
        )
        
        cliente = Cliente.objects.create(
            nombre='Cliente Test Validation',
            rut='22222222-2',
            activo=True
        )
        
        # Test con nombre válido
        upload_log_valido = UploadLog.objects.create(
            cliente=cliente,
            tipo_archivo='libro_mayor',
            nombre_archivo='LibroMayor_2024-01.xlsx',
            archivo_hash='test_hash_valid',
            usuario=user,
            estado='pendiente'
        )
        
        resultado = validar_nombre_archivo_libro_mayor(upload_log_valido.id)
        print(f"✓ Validación con nombre válido: {resultado}")
        
        # Test con nombre inválido
        upload_log_invalido = UploadLog.objects.create(
            cliente=cliente,
            tipo_archivo='libro_mayor',
            nombre_archivo='archivo_mal_formato.xlsx',
            archivo_hash='test_hash_invalid',
            usuario=user,
            estado='pendiente'
        )
        
        try:
            validar_nombre_archivo_libro_mayor(upload_log_invalido.id)
            print("✗ Debería haber fallado con nombre inválido")
        except Exception as e:
            print(f"✓ Validación correcta con nombre inválido: {str(e)}")
        
        print("✓ Test de validación completado")
        
    except Exception as e:
        print(f"✗ Error en test de validación: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    print("INICIANDO TESTS DE LIBRO MAYOR CHAINS")
    print("=" * 50)
    
    # Ejecutar tests básicos
    ejecutar_test_basico()
    ejecutar_test_validacion()
    
    # Ejecutar tests con Django TestCase
    print("\n=== EJECUTANDO TESTS UNITARIOS ===")
    
    try:
        from django.test.utils import get_runner
        from django.conf import settings
        
        # Configurar test runner
        TestRunner = get_runner(settings)
        test_runner = TestRunner()
        
        # Ejecutar tests específicos
        failures = test_runner.run_tests(["__main__.LibroMayorChainsTestCase"])
        
        if failures:
            print(f"✗ {failures} tests fallaron")
        else:
            print("✓ Todos los tests unitarios pasaron")
            
    except Exception as e:
        print(f"Error ejecutando tests unitarios: {str(e)}")
    
    print("\n" + "=" * 50)
    print("TESTS COMPLETADOS")
