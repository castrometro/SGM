from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from openpyxl import load_workbook, Workbook
from io import BytesIO
import unicodedata
from django.http import HttpResponse
import json
from contabilidad.tasks import (
    get_headers_salida_contabilidad,
    get_redis_client_db1,
    get_redis_client_db1_binary,
)
## Endpoint sincrónico eliminado: se fuerza uso de Celery
from contabilidad.task_rindegastos import rg_procesar_step1_task


def _normalize(text):
    if text is None:
        return ""
    if not isinstance(text, str):
        text = str(text)
    text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii')
    return text.strip().lower()


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def leer_headers_excel_rindegastos(request):
    """
    Contabilidad: Endpoint exclusivo RindeGastos para leer headers y detectar CC
    como el rango entre la última columna 'Nombre cuenta' y la columna 'Fecha aprobacion'.
    """
    try:
        if 'archivo' not in request.FILES:
            return Response({'error': 'No se encontró archivo en la petición'}, status=400)

        archivo = request.FILES['archivo']

        if not archivo.name.lower().endswith(('.xlsx', '.xls')):
            return Response({'error': 'El archivo debe ser un Excel (.xlsx o .xls)'}, status=400)

        contenido = archivo.read()
        wb = load_workbook(BytesIO(contenido), read_only=True)
        ws = wb.active

        primera_fila = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [(v if v is not None else '') for v in primera_fila]

        # Última columna que contiene "Nombre cuenta"
        last_nombre_idx = -1
        for i, h in enumerate(headers):
            if 'nombre cuenta' in _normalize(h):
                last_nombre_idx = i

        # Columna de "Fecha aprobacion"
        fecha_ap_idx = None
        for i, h in enumerate(headers):
            hn = _normalize(h)
            if 'fecha' in hn and 'aprobacion' in hn:
                fecha_ap_idx = i
                break

        centros_costo = {}
        if last_nombre_idx != -1 and fecha_ap_idx is not None and fecha_ap_idx - last_nombre_idx > 1:
            for pos in range(last_nombre_idx + 1, fecha_ap_idx):
                nombre = headers[pos]
                if nombre and str(nombre).strip() != '':
                    centros_costo[str(nombre)] = {"posicion": pos, "nombre": str(nombre)}

        # Fallback por nombres conocidos
        if not centros_costo:
            conocidos = ['PyC', 'PS', 'EB', 'CO', 'RE', 'TR', 'CF', 'LRC']
            for i, h in enumerate(headers):
                hs = str(h).strip() if h is not None else ''
                if hs in conocidos:
                    centros_costo[hs] = {"posicion": i, "nombre": hs}

        wb.close()

        return Response({
            'headers': [str(h) if h is not None else '' for h in headers],
            'total_columnas': len(headers),
            'centros_costo': centros_costo,
            'mensaje': 'Headers leídos exitosamente (RindeGastos/Contabilidad)'
        })

    except Exception as e:
        return Response({'error': f'Error leyendo headers del Excel: {str(e)}'}, status=500)


def _find_cc_range(headers):
    """Encuentra el rango (start_exclusive, end_inclusive_exclusive) para CC: entre última 'Nombre cuenta' y 'Fecha aprobacion'.
    Retorna (start_idx, end_idx). Si no hay coincidencia válida, retorna (None, None)."""
    last_nombre_idx = -1
    for i, h in enumerate(headers):
        if 'nombre cuenta' in _normalize(h):
            last_nombre_idx = i
    fecha_ap_idx = None
    for i, h in enumerate(headers):
        hn = _normalize(h)
        if 'fecha' in hn and 'aprobacion' in hn:
            fecha_ap_idx = i
            break
    if last_nombre_idx != -1 and fecha_ap_idx is not None and fecha_ap_idx - last_nombre_idx > 1:
        return last_nombre_idx + 1, fecha_ap_idx
    return None, None


def _parse_numeric(value):
    """Intenta parsear a float. Acepta strings con % o espacios. Retorna None si no parseable."""
    if value is None:
        return None
    try:
        if isinstance(value, str):
            s = value.replace('%', '').replace(',', '.').strip()
            if s == '':
                return None
            return float(s)
        if isinstance(value, (int, float)):
            return float(value)
    except Exception:
        return None
    return None


def _sanitize_sheet_name(name: str) -> str:
    s = str(name).replace(':', '-').replace('/', '-').replace('\\', '-')
    return s[:31] if len(s) > 31 else s


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def procesar_step1_rindegastos(request):
    """Inicia Step1 (asíncrono) exigiendo parametros_contables (JSON o campos sueltos)."""
    try:
        if 'archivo' not in request.FILES:
            return Response({'error': 'No se encontró archivo en la petición'}, status=400)
        archivo = request.FILES['archivo']
        if not archivo.name.lower().endswith(('.xlsx', '.xls')):
            return Response({'error': 'El archivo debe ser un Excel (.xlsx o .xls)'}, status=400)

        # Parse parametros contables obligatorios
        raw_param = request.data.get('parametros_contables')
        
        parametros_contables = None
        if raw_param:
            try:
                parametros_contables = json.loads(raw_param)
            except Exception as e:
                return Response({'error': 'parametros_contables no es JSON válido'}, status=400)
        else:
            # Fallback: construir desde campos sueltos
            cuenta_iva = request.data.get('cuentaIva') or request.data.get('cuenta_iva')
            cuenta_prov = request.data.get('cuentaProveedores') or request.data.get('cuenta_proveedores')
            cuenta_gasto = request.data.get('cuentaGasto') or request.data.get('cuenta_gasto')
            # CC: prefijo cc_*
            mapeo_cc = {}
            for k, v in request.data.items():
                if k.startswith('cc_') and v:
                    nombre = k[3:]
                    mapeo_cc[nombre] = v
            if any([cuenta_iva, cuenta_prov, cuenta_gasto]) or mapeo_cc:
                parametros_contables = {
                    'cuentasGlobales': {
                        'iva': cuenta_iva,
                        'proveedores': cuenta_prov,
                        'gasto_default': cuenta_gasto,
                    },
                    'mapeoCC': mapeo_cc
                }
        if not parametros_contables:
            return Response({'error': 'Se requieren parametros_contables (JSON) o campos individuales de cuentas y CC.'}, status=400)
        cg = parametros_contables.get('cuentasGlobales') or {}
        requeridas = ['iva', 'proveedores', 'gasto_default']
        faltantes = [r for r in requeridas if not cg.get(r)]
        if faltantes:
            return Response({'error': f'Faltan cuentasGlobales requeridas: {", ".join(faltantes)}'}, status=400)

        contenido = archivo.read()
        task = rg_procesar_step1_task.delay(contenido, archivo.name, request.user.id, parametros_contables)
        return Response({
            'task_id': task.id,
            'estado': 'procesando',
            'archivo_nombre': archivo.name,
            'mensaje': 'Archivo enviado para Step1 (RG) con parametros contables'
        }, status=202)
    except Exception as e:
        return Response({'error': f'Error iniciando Step1: {str(e)}'}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def estado_step1_rindegastos(request, task_id):
    try:
        r = get_redis_client_db1()
        meta_raw = r.get(f"rg_step1_meta:{request.user.id}:{task_id}")
        if not meta_raw:
            return Response({'error': 'No se encontró información de la tarea'}, status=404)
        return Response(json.loads(meta_raw))
    except Exception as e:
        return Response({'error': f'Error consultando estado: {str(e)}'}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def descargar_step1_rindegastos(request, task_id):
    try:
        r = get_redis_client_db1()
        meta_raw = r.get(f"rg_step1_meta:{request.user.id}:{task_id}")
        if not meta_raw:
            return Response({'error': 'No se encontró información de la tarea'}, status=404)
        meta = json.loads(meta_raw)
        if meta.get('estado') != 'completado':
            return Response({'error': 'La tarea aún no ha sido completada'}, status=400)

        r_bin = get_redis_client_db1_binary()
        excel_content = r_bin.get(f"rg_step1_excel:{request.user.id}:{task_id}")
        if not excel_content:
            return Response({'error': 'El archivo procesado no está disponible'}, status=404)

        resp = HttpResponse(
            excel_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = f'attachment; filename="rg_step1_{task_id}.xlsx"'
        return resp
    except Exception as e:
        return Response({'error': f'Error descargando archivo: {str(e)}'}, status=500)
