"""
Template Excel específico para Estado de Resultado Integral (ERI)
"""

import openpyxl
import logging
from .base import BaseExcelTemplate
from .translations import detect_language_improved

logger = logging.getLogger(__name__)


class ERITemplate(BaseExcelTemplate):
    """Template Excel para Estado de Resultado Integral"""

    def generate(self, data_eri, metadata):
        """Generar template Excel para Estado de Resultado Integral"""
        # Obtener idioma usando detección mejorada
        language = detect_language_improved(metadata)
        
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = self._get_text('title_eri', language)
        
        # Título principal
        title_text = self._get_text('title_eri', language)
        self._apply_header_style(ws, 1, 1, 4, title_text)
        
        # Información del período
        current_row = self._add_period_info(ws, 3, metadata, language)
        
        if data_eri:
            current_row = self._add_eri_sections(ws, current_row, data_eri, metadata, language)
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        # Agregar hoja de metadatos
        self._add_metadata_sheet(workbook, metadata, language)
        
        return workbook

    def _add_eri_sections(self, ws, current_row, data_eri, metadata, language):
        """Agregar todas las secciones del ERI"""
        # Definir bloques ERI a procesar con sus traducciones
        bloques_eri = [
            ("ganancias_brutas", "ganancias_brutas"),
            ("ganancia_perdida", "ganancia_perdida"), 
            ("ganancia_perdida_antes_impuestos", "ganancia_perdida_antes_impuestos")
        ]
        
        total_general = 0
        
        for bloque_key, translation_key in bloques_eri:
            bloque_data = data_eri.get(bloque_key)
            if bloque_data:
                current_row, total_bloque = self._add_eri_section(
                    ws, current_row, bloque_key, bloque_data, 
                    metadata.get("moneda", "CLP"), language, translation_key
                )
                total_general += total_bloque
        
        # Total General
        current_row += 1
        ws.cell(row=current_row, column=1, value=self._get_text('total_general', language)).font = self.styles['total']
        ws.cell(row=current_row, column=2, value=self._format_amount(total_general, metadata.get("moneda", "CLP"))).font = self.styles['total']
        ws.cell(row=current_row, column=1).fill = self.fills['total']
        ws.cell(row=current_row, column=2).fill = self.fills['total']
        
        return current_row

    def _add_eri_section(self, ws, start_row, bloque_key, bloque_data, moneda, language, translation_key=None):
        """Agregar una sección del ERI al worksheet"""
        current_row = start_row
        
        # Título del bloque - usar traducción si está disponible
        if translation_key:
            titulo_bloque = self._get_text(translation_key, language)
        else:
            titulo_bloque = bloque_key.replace("_", " ").title()
        
        ws.cell(row=current_row, column=1, value=titulo_bloque).font = self.styles['subheader']
        ws.cell(row=current_row, column=1).fill = self.fills['subheader']
        current_row += 1
        
        total_bloque = 0
        
        if isinstance(bloque_data, dict):
            if "grupos" in bloque_data:
                current_row = self._add_eri_groups(ws, current_row, bloque_data, moneda, language)
            
            # Total del bloque
            if "total" in bloque_data:
                total_bloque = bloque_data["total"]
            
            if total_bloque != 0:
                ws.cell(row=current_row, column=1, value=f"TOTAL {titulo_bloque.upper()}").font = self.styles['total']
                ws.cell(row=current_row, column=2, value=self._format_amount(total_bloque, moneda)).font = self.styles['total']
                ws.cell(row=current_row, column=1).fill = self.fills['total']
                ws.cell(row=current_row, column=2).fill = self.fills['total']
                current_row += 2
        
        return current_row, total_bloque

    def _add_eri_groups(self, ws, current_row, bloque_data, moneda, language):
        """Agregar grupos del ERI con agrupación colapsable"""
        for grupo_nombre, grupo_data in bloque_data.get("grupos", {}).items():
            # Título del grupo - usar nombre según idioma
            if language.lower() in ['en', 'english', 'inglés']:
                grupo_display_name = grupo_data.get('nombre_en', grupo_data.get('nombre_es', grupo_nombre))
            else:
                grupo_display_name = grupo_data.get('nombre_es', grupo_data.get('nombre_en', grupo_nombre))
            
            ws.cell(row=current_row, column=1, value=f"  {grupo_display_name}").font = self.styles['data']
            ws.cell(row=current_row, column=2, value=self._format_amount(grupo_data.get('total', 0), moneda)).font = self.styles['data']
            current_row += 1
            
            # Filas de inicio y fin para el grupo
            grupo_start_row = current_row
            
            # Agregar cuentas del grupo
            for cuenta in grupo_data.get("cuentas", []):
                codigo = cuenta.get("codigo", "")
                nombre = self._get_account_name(cuenta, language)
                saldo = cuenta.get("saldo_final", 0)
                
                ws.cell(row=current_row, column=1, value=f"    {codigo} - {nombre}")
                ws.cell(row=current_row, column=2, value=self._format_amount(saldo, moneda))
                current_row += 1
            
            # Crear grupo colapsable solo si hay cuentas
            if grupo_data.get("cuentas"):
                self._create_group_collapsible(ws, grupo_start_row, current_row)
                current_row += 1
        
        return current_row
