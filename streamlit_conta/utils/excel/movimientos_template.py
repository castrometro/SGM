"""
Template Excel específico para Movimientos Contables
"""

import openpyxl
import logging
from datetime import datetime
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from .base import BaseExcelTemplate

logger = logging.getLogger(__name__)


class MovimientosTemplate(BaseExcelTemplate):
    """Template Excel para Movimientos Contables"""

    def generate(self, df_movimientos, metadata, tipo_vista="Todos los movimientos"):
        """Generar template Excel para movimientos contables"""
        # Obtener idioma de los metadatos
        language = metadata.get('idioma', 'es')
        
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = self._get_text('title_movimientos', language)
        
        # Título principal
        title_text = f"{self._get_text('title_movimientos', language)} - {tipo_vista.upper()}"
        self._apply_header_style(ws, 1, 1, len(df_movimientos.columns), title_text)
        
        # Información del período
        current_row = 3
        ws.cell(row=current_row, column=1, value=f"{self._get_text('client', language)}: {metadata.get('cliente_nombre', 'N/A')}")
        ws.cell(row=current_row, column=4, value=f"{self._get_text('period', language)}: {metadata.get('periodo', 'N/A')}")
        current_row += 1
        ws.cell(row=current_row, column=1, value=f"{self._get_text('total_movements', language)}: {len(df_movimientos)}")
        ws.cell(row=current_row, column=4, value=f"{self._get_text('date', language)}: {datetime.now().strftime('%d/%m/%Y')}")
        current_row += 2
        
        # Encabezados de columna
        current_row = self._add_column_headers(ws, current_row, df_movimientos)
        
        # Datos
        current_row = self._add_movement_data(ws, current_row, df_movimientos)
        
        # Totales si aplica
        current_row = self._add_totals(ws, current_row, df_movimientos, metadata)
        
        # Ajustar anchos de columna
        self._adjust_column_widths(ws, df_movimientos)
        
        # Agregar hoja de metadatos
        self._add_metadata_sheet(workbook, metadata, language)
        
        return workbook

    def _add_column_headers(self, ws, current_row, df_movimientos):
        """Agregar encabezados de columna"""
        for col_idx, column_name in enumerate(df_movimientos.columns, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=column_name)
            cell.font = self.styles['header']
            cell.fill = self.fills['header']
            cell.alignment = Alignment(horizontal='center')
        
        return current_row + 1

    def _add_movement_data(self, ws, current_row, df_movimientos):
        """Agregar datos de movimientos"""
        for row_idx, row in df_movimientos.iterrows():
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font = self.styles['data']
                cell.border = self.borders['thin']
                
                # Alternar colores de fila
                if current_row % 2 == 0:
                    cell.fill = self.fills['alternate']
            
            current_row += 1
        
        return current_row

    def _add_totals(self, ws, current_row, df_movimientos, metadata):
        """Agregar totales si aplica"""
        if "Debe" in df_movimientos.columns and "Haber" in df_movimientos.columns:
            current_row += 1
            debe_col = list(df_movimientos.columns).index("Debe") + 1
            haber_col = list(df_movimientos.columns).index("Haber") + 1
            
            language = metadata.get('idioma', 'es')
            
            ws.cell(row=current_row, column=debe_col - 1, value=self._get_text('totals', language)).font = self.styles['total']
            ws.cell(row=current_row, column=debe_col, value=self._format_amount(df_movimientos["Debe"].sum(), metadata.get("moneda", "CLP"))).font = self.styles['total']
            ws.cell(row=current_row, column=haber_col, value=self._format_amount(df_movimientos["Haber"].sum(), metadata.get("moneda", "CLP"))).font = self.styles['total']
        
        return current_row

    def _adjust_column_widths(self, ws, df_movimientos):
        """Ajustar anchos de columna"""
        for col_idx in range(1, len(df_movimientos.columns) + 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 15
