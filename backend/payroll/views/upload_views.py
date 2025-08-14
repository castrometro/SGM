# ============================================================================
#                           UPLOAD VIEWS
# ============================================================================
# Views para carga de archivos y importación de datos

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.http import JsonResponse
from django.views.generic import TemplateView
from django.core.files.storage import default_storage
from django.conf import settings
import os
import pandas as pd
from openpyxl import load_workbook

from ..models import CierrePayroll, Empleados_Cierre, Item_Cierre, Logs_Actividad
from api.models import Usuario  # Cambiado de Empleado a Usuario


class UploadExcelView(LoginRequiredMixin, TemplateView):
    """
    Vista para cargar archivos Excel
    """
    template_name = 'payroll/upload/excel.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cierre_id = kwargs.get('cierre_id')
        context['cierre'] = get_object_or_404(CierrePayroll, pk=cierre_id)
        return context


@login_required
def upload_excel_ajax(request, cierre_id):
    """
    Vista AJAX para procesar carga de Excel
    """
    cierre = get_object_or_404(CierrePayroll, pk=cierre_id)
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            
            # Validar archivo
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                return JsonResponse({
                    'success': False,
                    'error': 'Formato de archivo no válido. Use .xlsx o .xls'
                })
            
            # Guardar archivo temporalmente
            file_path = default_storage.save(
                f'temp/payroll_{cierre.id}_{excel_file.name}',
                excel_file
            )
            
            # Procesar archivo
            resultado = procesar_excel_payroll(file_path, cierre, request.user)
            
            # Limpiar archivo temporal
            default_storage.delete(file_path)
            
            if resultado['success']:
                # Guardar archivo original en el cierre
                cierre.archivo_excel_original = excel_file
                cierre.save()
                
                # Log de éxito
                Logs_Actividad.objects.create(
                    cierre_payroll=cierre,
                    usuario=request.user,
                    accion='carga_archivo',
                    descripcion=f'Excel cargado exitosamente: {excel_file.name}',
                    resultado='exitoso',
                    detalles=f"Empleados procesados: {resultado['empleados_procesados']}, Items: {resultado['items_procesados']}"
                )
            
            return JsonResponse(resultado)
            
        except Exception as e:
            # Log de error
            Logs_Actividad.objects.create(
                cierre_payroll=cierre,
                usuario=request.user,
                accion='carga_archivo',
                descripcion=f'Error al cargar Excel: {str(e)}',
                resultado='error',
                detalles=str(e)
            )
            
            return JsonResponse({
                'success': False,
                'error': f'Error al procesar archivo: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'No se recibió archivo válido'
    })


def procesar_excel_payroll(file_path, cierre, usuario):
    """
    Procesa archivo Excel de payroll y crea registros
    """
    try:
        # Leer archivo Excel
        full_path = os.path.join(settings.MEDIA_ROOT, file_path)
        workbook = load_workbook(full_path, data_only=True)
        
        resultado = {
            'success': True,
            'empleados_procesados': 0,
            'items_procesados': 0,
            'errores': [],
            'advertencias': []
        }
        
        # Procesar hoja de empleados (asumiendo formato estándar)
        if 'Empleados' in workbook.sheetnames:
            sheet_empleados = workbook['Empleados']
            resultado.update(procesar_hoja_empleados(sheet_empleados, cierre))
        
        # Procesar hoja de items
        if 'Items' in workbook.sheetnames:
            sheet_items = workbook['Items']
            resultado.update(procesar_hoja_items(sheet_items, cierre))
        
        # Procesar hoja de liquidaciones
        if 'Liquidaciones' in workbook.sheetnames:
            sheet_liquidaciones = workbook['Liquidaciones']
            resultado.update(procesar_hoja_liquidaciones(sheet_liquidaciones, cierre))
        
        return resultado
        
    except Exception as e:
        return {
            'success': False,
            'error': f'Error al procesar Excel: {str(e)}'
        }


def procesar_hoja_empleados(sheet, cierre):
    """
    Procesa hoja de empleados del Excel
    """
    empleados_procesados = 0
    errores = []
    
    # Asumir formato: RUT | Nombre | Apellido | Dias | Horas_Extras
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Si no hay RUT, saltar fila
            continue
            
        try:
            rut = str(row[0]).strip()
            nombre = str(row[1]).strip() if row[1] else ''
            apellido = str(row[2]).strip() if row[2] else ''
            dias_trabajados = int(row[3]) if row[3] else 30
            horas_extras = float(row[4]) if row[4] else 0
            
            # Buscar o crear empleado (usuario)
            empleado, created = Usuario.objects.get_or_create(
                username=rut,  # Usar RUT como username
                defaults={
                    'first_name': nombre,
                    'last_name': apellido,
                    'is_active': True
                }
            )
            
            # Crear empleado_cierre
            empleado_cierre, created = Empleados_Cierre.objects.get_or_create(
                cierre_payroll=cierre,
                empleado=empleado,
                defaults={
                    'dias_trabajados': dias_trabajados,
                    'horas_extras': horas_extras,
                    'estado_procesamiento': 'pendiente'
                }
            )
            
            if created:
                empleados_procesados += 1
            
        except Exception as e:
            errores.append(f'Error en fila {row}: {str(e)}')
    
    return {
        'empleados_procesados': empleados_procesados,
        'errores_empleados': errores
    }


def procesar_hoja_items(sheet, cierre):
    """
    Procesa hoja de items del Excel
    """
    items_procesados = 0
    errores = []
    
    # Formato: Codigo | Nombre | Tipo | Imponible | Variable
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
            
        try:
            codigo = str(row[0]).strip()
            nombre = str(row[1]).strip()
            tipo = str(row[2]).lower().strip()
            es_imponible = str(row[3]).lower() in ['si', 'sí', 'true', '1'] if row[3] else False
            es_variable = str(row[4]).lower() in ['si', 'sí', 'true', '1'] if row[4] else False
            
            # Validar tipo
            if tipo not in ['haberes', 'descuentos']:
                tipo = 'haberes'  # Default
            
            # Crear item
            item, created = Item_Cierre.objects.get_or_create(
                cierre_payroll=cierre,
                codigo_item=codigo,
                defaults={
                    'nombre_item': nombre,
                    'tipo_item': tipo,
                    'es_imponible': es_imponible,
                    'es_variable': es_variable
                }
            )
            
            if created:
                items_procesados += 1
                
        except Exception as e:
            errores.append(f'Error en item {row}: {str(e)}')
    
    return {
        'items_procesados': items_procesados,
        'errores_items': errores
    }


def procesar_hoja_liquidaciones(sheet, cierre):
    """
    Procesa hoja de liquidaciones (RUT + Items con montos)
    """
    liquidaciones_procesadas = 0
    errores = []
    
    # Este es más complejo, requiere formato específico
    # Por ahora, implementación básica
    
    return {
        'liquidaciones_procesadas': liquidaciones_procesadas,
        'errores_liquidaciones': errores
    }


@login_required
def upload_pdf_view(request, cierre_id):
    """
    Vista para cargar PDF de comparación
    """
    cierre = get_object_or_404(CierrePayroll, pk=cierre_id)
    
    if request.method == 'POST' and request.FILES.get('pdf_file'):
        try:
            pdf_file = request.FILES['pdf_file']
            
            if not pdf_file.name.endswith('.pdf'):
                messages.error(request, 'Solo se permiten archivos PDF')
                return redirect('payroll:cierre_detail', pk=cierre_id)
            
            # Guardar PDF
            cierre.archivo_pdf_comparacion = pdf_file
            cierre.save()
            
            # Log
            Logs_Actividad.objects.create(
                cierre_payroll=cierre,
                usuario=request.user,
                accion='carga_archivo',
                descripcion=f'PDF de comparación cargado: {pdf_file.name}',
                resultado='exitoso'
            )
            
            messages.success(request, 'PDF cargado exitosamente')
            
        except Exception as e:
            messages.error(request, f'Error al cargar PDF: {str(e)}')
    
    return redirect('payroll:cierre_detail', pk=cierre_id)


@login_required
def importar_empleados_masivo(request):
    """
    Vista para importar empleados de forma masiva desde Excel
    """
    if request.method == 'POST' and request.FILES.get('excel_empleados'):
        try:
            excel_file = request.FILES['excel_empleados']
            
            # Procesar archivo
            df = pd.read_excel(excel_file)
            
            empleados_creados = 0
            empleados_actualizados = 0
            errores = []
            
            for index, row in df.iterrows():
                try:
                    rut = str(row['RUT']).strip()
                    nombre = str(row['Nombre']).strip()
                    apellido = str(row['Apellido']).strip()
                    email = str(row.get('Email', '')).strip()
                    departamento = str(row.get('Departamento', '')).strip()
                    
                    empleado, created = Usuario.objects.update_or_create(
                        username=rut,  # Usar RUT como username
                        defaults={
                            'first_name': nombre,
                            'last_name': apellido,
                            'email': email,
                            'is_active': True
                        }
                    )
                    
                    if created:
                        empleados_creados += 1
                    else:
                        empleados_actualizados += 1
                        
                except Exception as e:
                    errores.append(f'Fila {index + 2}: {str(e)}')
            
            # Crear log
            Logs_Actividad.objects.create(
                cierre_payroll=None,
                usuario=request.user,
                accion='importacion_masiva',
                descripcion=f'Importación masiva de empleados completada',
                resultado='exitoso' if not errores else 'advertencia',
                detalles=f'Creados: {empleados_creados}, Actualizados: {empleados_actualizados}, Errores: {len(errores)}'
            )
            
            return JsonResponse({
                'success': True,
                'empleados_creados': empleados_creados,
                'empleados_actualizados': empleados_actualizados,
                'errores': errores
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Archivo no válido'
    })


@login_required
def validar_formato_excel(request):
    """
    Vista AJAX para validar formato de archivo Excel antes de procesar
    """
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            
            # Guardar temporalmente
            file_path = default_storage.save(f'temp/validation_{excel_file.name}', excel_file)
            full_path = os.path.join(settings.MEDIA_ROOT, file_path)
            
            # Validar estructura
            workbook = load_workbook(full_path, data_only=True)
            
            validacion = {
                'success': True,
                'hojas_encontradas': list(workbook.sheetnames),
                'hojas_requeridas': ['Empleados', 'Items'],
                'formato_valido': True,
                'errores': [],
                'advertencias': []
            }
            
            # Validar hojas requeridas
            for hoja_requerida in validacion['hojas_requeridas']:
                if hoja_requerida not in workbook.sheetnames:
                    validacion['errores'].append(f'Falta la hoja: {hoja_requerida}')
                    validacion['formato_valido'] = False
            
            # Validar estructura de hojas
            if 'Empleados' in workbook.sheetnames:
                sheet = workbook['Empleados']
                headers = [cell.value for cell in sheet[1]]
                headers_requeridos = ['RUT', 'Nombre', 'Apellido']
                
                for header in headers_requeridos:
                    if header not in headers:
                        validacion['errores'].append(f'Falta columna en Empleados: {header}')
                        validacion['formato_valido'] = False
            
            # Limpiar archivo temporal
            default_storage.delete(file_path)
            
            return JsonResponse(validacion)
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al validar archivo: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'No se recibió archivo'
    })
