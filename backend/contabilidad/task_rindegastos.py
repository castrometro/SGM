from celery import shared_task
from io import BytesIO
from openpyxl import load_workbook, Workbook
from django.utils import timezone
import json

from contabilidad.tasks import (
    get_redis_client_db1,
    get_redis_client_db1_binary,
    get_headers_salida_contabilidad,
)


def _normalize(text):
    if text is None:
        return ""
    return str(text).strip()


def _parse_numeric(value):
    if value is None:
        return None
    try:
        if isinstance(value, str):
            s = value.replace('%', '').replace(',', '.').strip()
            if s == '':
                return None
            return float(s)
        if isinstance(value, (int, float)):
            return float(value)
    except Exception:
        return None
    return None


def _find_cc_range(headers):
    last_nombre_idx = -1
    for i, h in enumerate(headers):
        if 'nombre cuenta' in str(h).lower():
            last_nombre_idx = i
    fecha_ap_idx = None
    for i, h in enumerate(headers):
        hn = str(h).lower()
        if 'fecha' in hn and 'aprobacion' in hn:
            fecha_ap_idx = i
            break
    if last_nombre_idx != -1 and fecha_ap_idx is not None and fecha_ap_idx - last_nombre_idx > 1:
        return last_nombre_idx + 1, fecha_ap_idx
    return None, None


@shared_task(bind=True)
def rg_procesar_archivo_task(self, archivo_content, archivo_nombre, usuario_id, mapeo_cc=None, parametros_contables=None):
    """
    Tarea placeholder exclusiva de RindeGastos. Por ahora no procesa;
    se implementará en fases siguientes.
    """
    # Simplemente devuelve un estado simulado
    return {
        'task_id': self.request.id,
        'estado': 'no-implementado',
        'archivo_nombre': archivo_nombre,
        'mensaje': 'Placeholder RindeGastos: procesamiento no implementado aún'
    }


@shared_task(bind=True)
def rg_procesar_step1_task(self, archivo_content, archivo_nombre, usuario_id, parametros_contables=None):
    """
    Genera Excel con hojas por grupo (Tipo Doc + cantidad de CC > 0) y guarda en Redis.
    Guarda metadatos en rg_step1_meta:{usuario_id}:{task_id} y el archivo en rg_step1_excel:{usuario_id}:{task_id}
    """
    task_id = self.request.id
    redis_client = get_redis_client_db1()

    # Validar parametros contables obligatorios
    if not parametros_contables:
        raise ValueError("parametros_contables es obligatorio (cuentasGlobales y mapeoCC)")
    cuentas_globales = parametros_contables.get('cuentasGlobales') or {}
    mapeo_cc_param = parametros_contables.get('mapeoCC') or {}
    requeridas = ['iva', 'proveedores', 'gasto_default']
    faltantes = [r for r in requeridas if not cuentas_globales.get(r)]
    if faltantes:
        raise ValueError(f"Faltan cuentasGlobales requeridas: {', '.join(faltantes)}")

    # Meta inicial
    metadata = {
        'task_id': task_id,
        'usuario_id': usuario_id,
        'archivo_nombre': archivo_nombre,
        'inicio': timezone.now().isoformat(),
        'estado': 'procesando',
        'grupos': [],
        'archivo_excel_disponible': False,
        'cuentas_globales_usadas': list(cuentas_globales.keys()),
    }
    redis_client.setex(
        f"rg_step1_meta:{usuario_id}:{task_id}", 300, json.dumps(metadata, ensure_ascii=False)
    )

    # Leer archivo y agrupar
    wb_in = load_workbook(BytesIO(archivo_content), read_only=True)
    ws_in = wb_in.active
    headers = [(v if v is not None else '') for v in next(ws_in.iter_rows(min_row=1, max_row=1, values_only=True))]

    # Índice Tipo Doc (robusto: sinónimos y case-insensitive)
    posibles_nombres = {'tipo doc', 'tipodoc', 'tipo_documento', 'tipo documento', 'tipo_doc'}
    idx_tipo_doc = None
    for i, h in enumerate(headers):
        nombre_norm = str(h).strip().lower()
        if nombre_norm in posibles_nombres:
            idx_tipo_doc = i
            break
    if idx_tipo_doc is None:
        raise ValueError("No se encontró la columna de Tipo de Documento (buscó: Tipo Doc / tipodoc / tipo_documento)")

    cc_start, cc_end = _find_cc_range(headers)
    conocidos = ['PyC', 'PS', 'EB', 'CO', 'RE', 'TR', 'CF', 'LRC']
    cc_indices_conocidos = {str(h).strip(): i for i, h in enumerate(headers) if str(h).strip() in conocidos}

    grupos = {}
    grupos_filas = {}  # clave -> lista de (tipo_doc, cc_count, fila_index)
    total_filas = 0
    debug_filas = []
    # Guardar filas originales para cálculos posteriores
    filas_originales = {}
    for row_idx, row in enumerate(ws_in.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue
        total_filas += 1
        tipo_doc = row[idx_tipo_doc] if idx_tipo_doc < len(row) else None
        tipo_doc = str(tipo_doc) if tipo_doc is not None else 'Sin Tipo'

        # Contar CC válidos
        cc_count = 0
        if cc_start is not None and cc_end is not None:
            for col in range(cc_start, cc_end):
                val = row[col] if col < len(row) else None
                num = _parse_numeric(val)
                # Contar si valor numérico distinto de cero o si es texto no vacío significativo
                if (num is not None and abs(num) > 0) or (isinstance(val, str) and val.strip() not in ['', '-', '0']):
                    cc_count += 1
        else:
            vistos = set()
            for nombre, col in cc_indices_conocidos.items():
                if nombre == 'EB' and 'PS' in cc_indices_conocidos:
                    continue
                val = row[col] if col < len(row) else None
                num = _parse_numeric(val)
                if ((num is not None and abs(num) > 0) or (isinstance(val, str) and val and val.strip() not in ['', '-', '0'])) and nombre not in vistos:
                    cc_count += 1
                    vistos.add(nombre)

        clave = f"{tipo_doc} con {cc_count}CC"
        grupos[clave] = grupos.get(clave, 0) + 1
        grupos_filas.setdefault(clave, []).append((tipo_doc, cc_count, row_idx))
        filas_originales[row_idx] = row
        debug_filas.append({
            'fila_excel': row_idx,
            'tipo_doc': tipo_doc,
            'cc_count': cc_count,
            'clave': clave
        })

    wb_in.close()

    # Crear Excel de salida
    wb_out = Workbook()
    default_sheet = wb_out.active
    wb_out.remove(default_sheet)
    headers_salida = get_headers_salida_contabilidad()

    # Asegurar incorporación de la nueva columna requerida si aún no existe.
    # Se intenta colocar inmediatamente después de 'Monto 3 Detalle Libro' para mantener coherencia.
    if 'Monto Suma Detalle Libro' not in headers_salida:
        if 'Monto 3 Detalle Libro' in headers_salida:
            idx_m3 = headers_salida.index('Monto 3 Detalle Libro')
            headers_salida.insert(idx_m3 + 1, 'Monto Suma Detalle Libro')
        else:
            headers_salida.append('Monto Suma Detalle Libro')

    def sanitize(name: str) -> str:
        s = str(name).replace(':', '-').replace('/', '-').replace('\\', '-')
        return s[:31] if len(s) > 31 else s

    for clave in sorted(grupos.keys()):
        ws = wb_out.create_sheet(title=sanitize(clave))
        for col_idx, h in enumerate(headers_salida, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        row_cursor = 2

        filas_grupo = grupos_filas.get(clave, [])
        # Mapeo header salida -> índice para escribir
        header_to_col = {h: i + 1 for i, h in enumerate(headers_salida)}

        # Detectar índices de columnas relevantes en entrada (heurística case-insensitive)
        headers_lower = [str(h).strip().lower() for h in headers]
        def find_idx(posibles):
            for nombre in headers_lower:
                if nombre in posibles:
                    return headers_lower.index(nombre)
            return None

        idx_monto_neto = find_idx({'monto neto', 'neto', 'monto_neto'})
        idx_monto_iva_rec = find_idx({'monto iva recuperable', 'iva recuperable', 'monto iva', 'iva'})
        idx_monto_total = find_idx({'monto total', 'total', 'monto_total'})
        # Funciones auxiliares
        def trunc(v):
            try:
                return int(float(v))
            except Exception:
                return 0

        for idx_fila, (tipo_doc_str, cc_count, fila_original_idx) in enumerate(filas_grupo, start=1):
            row_in = filas_originales.get(fila_original_idx, [])
            monto_neto = _parse_numeric(row_in[idx_monto_neto]) if idx_monto_neto is not None and idx_monto_neto < len(row_in) else 0.0
            if monto_neto is None:
                monto_neto = 0.0
            monto_total_input = _parse_numeric(row_in[idx_monto_total]) if idx_monto_total is not None and idx_monto_total < len(row_in) else None
            monto_iva_rec_input = _parse_numeric(row_in[idx_monto_iva_rec]) if idx_monto_iva_rec is not None and idx_monto_iva_rec < len(row_in) else None
            # IVA: si no existe columna o valor, calcular 0.19 * neto (truncado)
            iva_monto = monto_iva_rec_input if (monto_iva_rec_input is not None) else trunc(monto_neto * 0.19)
            # Total: si no existe usar neto + iva
            if monto_total_input is None:
                monto_total = (monto_neto + (iva_monto if monto_iva_rec_input is None else iva_monto))
            else:
                monto_total = monto_total_input

            # Recolectar montos por CC (cada fila de gasto corresponde a 1 CC)
            gastos_rows = []  # lista de (descripcion, debe, codigo_cc)
            if cc_count > 0:
                if cc_start is not None and cc_end is not None:
                    for col in range(cc_start, cc_end):
                        if col >= len(row_in):
                            continue
                        val = row_in[col]
                        perc = _parse_numeric(val)
                        if perc is None:
                            continue
                        if abs(perc) > 0:
                            debe = (perc / 100.0) * monto_neto
                            codigo_cc = headers[col] if col < len(headers) else f'CC{col}'
                            gastos_rows.append((f'Gasto {codigo_cc}', debe, codigo_cc))
                else:
                    for nombre, col in cc_indices_conocidos.items():
                        if col >= len(row_in):
                            continue
                        val = row_in[col]
                        perc = _parse_numeric(val)
                        if perc is None:
                            continue
                        if abs(perc) > 0:
                            debe = (perc / 100.0) * monto_neto
                            gastos_rows.append((f'Gasto {nombre}', debe, nombre))

            suma_debe_gastos = sum(g[1] for g in gastos_rows)

            def _truncate_number(v):
                """Trunca (corta decimales) hacia cero cualquier número convertible a float."""
                if v is None:
                    return None
                try:
                    return int(float(v))
                except Exception:
                    return v

            def write_row(descripcion, debe=None, haber=None, extra=None):
                nonlocal row_cursor
                if debe is not None:
                    ws.cell(row=row_cursor, column=header_to_col.get('Monto al Debe Moneda Base', 3), value=_truncate_number(debe))
                if haber is not None:
                    ws.cell(row=row_cursor, column=header_to_col.get('Monto al Haber Moneda Base', 4), value=_truncate_number(haber))
                ws.cell(row=row_cursor, column=header_to_col.get('Descripción Movimiento', 5), value=descripcion)
                if extra:
                    for hname, val in extra.items():
                        col_idx_h = header_to_col.get(hname)
                        if col_idx_h:
                            ws.cell(row=row_cursor, column=col_idx_h, value=_truncate_number(val))
                row_cursor += 1

            # Tipos 33 / 64: IVA + Proveedores + Gastos
            if tipo_doc_str in ['33', '64']:
                # Fila IVA
                write_row(
                    descripcion=f'IVA Doc {fila_original_idx}',
                    debe=None,
                    haber=iva_monto,
                    extra={
                        'Código Plan de Cuenta': cuentas_globales.get('iva')
                    }
                )
                # Fila Proveedores (usa IVA y suma gastos)
                monto1 = suma_debe_gastos
                monto3 = iva_monto
                write_row(
                    descripcion=f'Proveedor Doc {fila_original_idx}',
                    debe=None,
                    haber=monto_total,
                    extra={
                        'Monto 1 Detalle Libro': monto1,
                        'Monto 3 Detalle Libro': monto3,
                        'Monto Suma Detalle Libro': (monto1 if monto1 is not None else 0) + (monto3 if monto3 is not None else 0),
                        'Código Plan de Cuenta': cuentas_globales.get('proveedores')
                    }
                )
                # Filas de Gasto
                for desc_gasto, debe_val, codigo_cc in gastos_rows:
                    codigo_cc_final = mapeo_cc_param.get(codigo_cc, codigo_cc)
                    write_row(
                        descripcion=desc_gasto,
                        debe=debe_val,
                        haber=None,
                        extra={
                            'Código Centro de Costo': codigo_cc_final,
                            'Código Plan de Cuenta': cuentas_globales.get('gasto_default')
                        }
                    )
            elif tipo_doc_str == '34':
                # Tipo 34 (exento) sólo Proveedores + Gastos, y Monto 3 = Monto 1
                monto1 = suma_debe_gastos
                # Para tipo 34 la suma solicitada es copiar Monto 1 (no sumatoria de 1+3 ya que son iguales)
                write_row(
                    descripcion=f'Proveedor Doc {fila_original_idx}',
                    debe=None,
                    haber=monto_total if monto_total is not None else suma_debe_gastos,
                    extra={
                        'Monto 1 Detalle Libro': monto1,
                        'Monto 3 Detalle Libro': monto1,  # se mantiene compatibilidad actual
                        'Monto Suma Detalle Libro': monto1,
                        'Código Plan de Cuenta': cuentas_globales.get('proveedores')
                    }
                )
                for desc_gasto, debe_val, codigo_cc in gastos_rows:
                    codigo_cc_final = mapeo_cc_param.get(codigo_cc, codigo_cc)
                    write_row(
                        descripcion=desc_gasto,
                        debe=debe_val,
                        haber=None,
                        extra={
                            'Código Centro de Costo': codigo_cc_final,
                            'Código Plan de Cuenta': cuentas_globales.get('gasto_default')
                        }
                    )
            elif tipo_doc_str == '61':
                # Tipo 61: espejo de 33 (incluye IVA) invirtiendo Debe/Haber.
                # En 33: IVA (Haber), Proveedor (Haber), Gastos (Debe)
                # En 61: IVA (Debe), Proveedor (Debe), Gastos (Haber)
                # Fila IVA (invertido -> Debe)
                write_row(
                    descripcion=f'IVA Doc {fila_original_idx}',
                    debe=iva_monto,
                    haber=None,
                    extra={
                        'Código Plan de Cuenta': cuentas_globales.get('iva')
                    }
                )
                # Fila Proveedores (invertido -> Debe)
                monto1 = suma_debe_gastos
                monto3 = iva_monto
                write_row(
                    descripcion=f'Proveedor Doc {fila_original_idx}',
                    debe=monto_total if monto_total is not None else (monto1 + monto3),
                    haber=None,
                    extra={
                        'Monto 1 Detalle Libro': monto1,
                        'Monto 3 Detalle Libro': monto3,
                        'Monto Suma Detalle Libro': (monto1 if monto1 is not None else 0) + (monto3 if monto3 is not None else 0),
                        'Código Plan de Cuenta': cuentas_globales.get('proveedores')
                    }
                )
                # Filas Gasto (invertidas -> Haber)
                for desc_gasto, debe_val, codigo_cc in gastos_rows:
                    codigo_cc_final = mapeo_cc_param.get(codigo_cc, codigo_cc)
                    write_row(
                        descripcion=desc_gasto,
                        debe=None,
                        haber=debe_val,
                        extra={
                            'Código Centro de Costo': codigo_cc_final,
                            'Código Plan de Cuenta': cuentas_globales.get('gasto_default')
                        }
                    )
            else:
                # Tipos desconocidos: de momento no se generan movimientos (queda hoja vacía)
                pass

    buffer = BytesIO()
    wb_out.save(buffer)
    excel_content = buffer.getvalue()

    # Guardar Excel en Redis binario
    redis_client_bin = get_redis_client_db1_binary()
    redis_client_bin.setex(f"rg_step1_excel:{usuario_id}:{task_id}", 300, excel_content)

    # Actualizar meta
    metadata.update({
        'estado': 'completado',
        'fin': timezone.now().isoformat(),
        'total_filas': total_filas,
        'grupos': list(sorted(grupos.keys())),
        'archivo_excel_disponible': True,
        'debug_filas': debug_filas[:200]  # limitar tamaño
    })
    redis_client.setex(
        f"rg_step1_meta:{usuario_id}:{task_id}", 300, json.dumps(metadata, ensure_ascii=False)
    )

    return {
        'task_id': task_id,
        'estado': 'completado',
        'total_filas': total_filas,
        'total_grupos': len(grupos),
        'archivo_excel_disponible': True,
    }
