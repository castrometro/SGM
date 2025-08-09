#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Crear un archivo Excel de prueba con caracteres especiales y fechas
para probar la funcionalidad de captura masiva de gastos
"""

import os
import openpyxl
from datetime import datetime

def crear_excel_prueba():
    """
    Crear un Excel con datos de prueba que incluyen caracteres especiales y fechas
    """
    print("📊 Creando archivo Excel de prueba con caracteres especiales...")
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos Gastos"
    
    # Headers - incluyen caracteres especiales
    headers = [
        'Nro', 'Tipo Doc', 'RUT Proveedor', 'Razón Social', 
        'Folio', 'Fecha Docto', 'Monto Total', 'Código cuenta', 
        'Nombre cuenta', 'PyC', 'PS', 'CO', 'Descripción'
    ]
    
    # Escribir headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Datos de prueba con caracteres especiales y fechas
    datos_prueba = [
        [1, 33, '12.345.678-9', 'Peña & Asociados S.A.', 'F001', '30-05-2025', 150000, '1101001', 'Efectivo y equivalentes', 'PC1', 'PS1', 'CO1', 'Provisión de servicios médicos'],
        [2, 33, '98.765.432-1', 'Compañía Española de Seguros Ltda.', 'F002', '31-05-2025', 250000, '1102001', 'Deudores comerciales', 'PC2', 'PS2', 'CO2', 'Consultoría técnica especializada'],
        [3, 34, '11.222.333-4', 'Industrias Metálicas Peñafiel', 'F003', '01-06-2025', 75000, '1201001', 'Existencias materia prima', 'PC3', 'PS3', 'CO3', 'Suministro de materiales metálicos'],
        [4, 34, '55.666.777-8', 'Construcciones & Diseño SpA', 'F004', '02-06-2025', 180000, '1302001', 'Maquinarias y equipos', 'PC4', 'PS4', 'CO4', 'Mantención de equipos industriales'],
        [5, 61, '77.888.999-0', 'Tecnología & Innovación Chile', 'F005', '03-06-2025', 320000, '1401001', 'Software y licencias', 'PC5', 'PS5', 'CO5', 'Implementación sistema gestión'],
        [6, 61, '22.333.444-5', 'Asesorías Contables Peñalolén', 'F006', '04-06-2025', 95000, '2101001', 'Proveedores nacionales', 'PC6', 'PS6', 'CO6', 'Servicios contables mensuales'],
        [7, 33, '44.555.666-7', 'Distribuidora Española Maipú', 'F007', '05-06-2025', 420000, '1103001', 'Otras cuentas por cobrar', 'PC7', 'PS7', 'CO7', 'Venta productos importados'],
        [8, 34, '88.999.111-2', 'Manufacturas & Textiles S.A.', 'F008', '06-06-2025', 280000, '1201002', 'Productos terminados', 'PC8', 'PS8', 'CO8', 'Confección uniformes empresariales'],
        [9, 61, '33.444.555-6', 'Logística & Transporte Región', 'F009', '07-06-2025', 140000, '1501001', 'Vehículos transporte', 'PC9', 'PS9', 'CO9', 'Servicios transporte mercancías'],
        [10, 33, '66.777.888-9', 'Comunicaciones Ñuñoa Ltda.', 'F010', '08-06-2025', 210000, '1601001', 'Equipos comunicación', 'PC10', 'PS10', 'CO10', 'Instalación red telecomunicaciones']
    ]
    
    # Escribir datos
    for row_idx, row_data in enumerate(datos_prueba, 2):
        for col_idx, valor in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=valor)
    
    # Guardar archivo
    archivo_path = '/root/SGM/excel_prueba_caracteres_especiales.xlsx'
    wb.save(archivo_path)
    
    print(f"✅ Archivo creado: {archivo_path}")
    print(f"✅ Contiene {len(datos_prueba)} filas de datos")
    print(f"✅ Grupos por Tipo Doc:")
    print(f"   - Tipo 33: 4 filas")
    print(f"   - Tipo 34: 3 filas") 
    print(f"   - Tipo 61: 3 filas")
    print(f"✅ Incluye caracteres especiales: ñ, tildes, &, etc.")
    print(f"✅ Fechas en formato DD-MM-YYYY")
    
    return archivo_path

if __name__ == "__main__":
    try:
        archivo = crear_excel_prueba()
        print(f"\n🎯 Archivo listo para pruebas: {archivo}")
        print("\n📋 Instrucciones:")
        print("1. Ve a la página de Herramientas → Captura Masiva RindeGastos")
        print("2. Sube este archivo Excel")
        print("3. Verifica que los caracteres especiales se mantengan")
        print("4. Verifica que las fechas salgan en formato DD-MM-YYYY")
        
    except Exception as e:
        print(f"❌ Error creando archivo: {e}")
        import traceback
        traceback.print_exc()
