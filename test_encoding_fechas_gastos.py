#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test para verificar el manejo correcto de encoding UTF-8 y fechas
en la captura masiva de gastos
"""

import openpyxl
from datetime import datetime, date
import json
from io import BytesIO

def serialize_excel_data(data):
    """
    Función actualizada para testing
    """
    if isinstance(data, datetime):
        # Si es datetime, convertir a fecha simple DD-MM-YYYY
        return data.strftime("%d-%m-%Y")
    elif isinstance(data, date):
        # Si es date, mantener formato DD-MM-YYYY  
        return data.strftime("%d-%m-%Y")
    elif isinstance(data, dict):
        return {key: serialize_excel_data(value) for key, value in data.items()}
    elif isinstance(data, list):
        return [serialize_excel_data(item) for item in data]
    elif isinstance(data, str):
        # Asegurar que las strings mantengan encoding UTF-8
        return data
    else:
        return data

def test_caracteres_especiales():
    """
    Test de caracteres especiales (ñ, tildes)
    """
    print("🧪 Probando caracteres especiales...")
    
    test_data = {
        'proveedor': 'Peña & Asociados',
        'razon_social': 'Compañía Española de Seguros', 
        'descripcion': 'Provisión de servicios médicos',
        'cuenta': 'Cuentas por pagar - proveedores extranjeros',
        'fecha': datetime(2025, 5, 30)
    }
    
    # Serializar
    serialized = serialize_excel_data(test_data)
    print(f"✅ Datos serializados: {serialized}")
    
    # Convertir a JSON con UTF-8
    json_str = json.dumps(serialized, ensure_ascii=False)
    print(f"✅ JSON UTF-8: {json_str}")
    
    # Deserializar
    deserialized = json.loads(json_str)
    print(f"✅ Datos deserializados: {deserialized}")
    
    return serialized, json_str, deserialized

def test_fechas():
    """
    Test de formato de fechas
    """
    print("\n📅 Probando formatos de fechas...")
    
    test_dates = [
        datetime(2025, 5, 30, 10, 30, 0),  # datetime completo
        date(2025, 5, 30),                 # solo fecha
        "30-05-2025",                      # string fecha
        "2025-05-30T00:00:00"             # formato ISO
    ]
    
    for test_date in test_dates:
        print(f"Entrada: {test_date} ({type(test_date)})")
        serialized = serialize_excel_data(test_date)
        print(f"Salida: {serialized}")
        print("---")

def test_excel_output():
    """
    Test de escritura a Excel con caracteres especiales y fechas
    """
    print("\n📊 Probando escritura Excel...")
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Headers con caracteres especiales
    headers = ['Nro', 'Tipo Doc', 'RUT Proveedor', 'Razón Social', 
               'Folio', 'Fecha Docto', 'Monto Total', 'Descripción']
    
    # Escribir headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Datos de prueba con caracteres especiales y fechas
    test_rows = [
        [1, 33, '12.345.678-9', 'Peña & Asociados S.A.', 
         'F001', '30-05-2025', 150000, 'Provisión de servicios médicos'],
        [2, 34, '98.765.432-1', 'Compañía Española de Seguros', 
         'F002', '31-05-2025', 250000, 'Consultoría técnica especializada'],
        [3, 61, '11.222.333-4', 'Industrias Metálicas Peñafiel', 
         'F003', '01-06-2025', 75000, 'Suministro de materiales']
    ]
    
    # Escribir datos
    for row_idx, row_data in enumerate(test_rows, 2):
        for col_idx, valor in enumerate(row_data, 1):
            # Procesar fechas especialmente
            if col_idx == 6 and isinstance(valor, str):  # columna fecha
                # Si parece una fecha en formato ISO, convertir a DD-MM-YYYY
                if 'T' in valor and len(valor) > 10:
                    try:
                        dt = datetime.fromisoformat(valor.replace('Z', ''))
                        valor = dt.strftime("%d-%m-%Y")
                    except:
                        pass
            
            ws.cell(row=row_idx, column=col_idx, value=valor)
    
    # Guardar en memory
    buffer = BytesIO()
    wb.save(buffer)
    excel_content = buffer.getvalue()
    
    print(f"✅ Excel creado con {len(excel_content)} bytes")
    print(f"✅ Contiene {len(test_rows)} filas de datos")
    
    # Verificar que se puede leer de vuelta
    wb_read = openpyxl.load_workbook(BytesIO(excel_content))
    ws_read = wb_read.active
    
    print("✅ Verificación de lectura:")
    for row in ws_read.iter_rows(min_row=1, max_row=4, values_only=True):
        print(f"  {row}")
    
    return excel_content

if __name__ == "__main__":
    print("🚀 Iniciando tests de encoding y fechas...")
    
    try:
        # Test 1: Caracteres especiales
        test_caracteres_especiales()
        
        # Test 2: Fechas
        test_fechas()
        
        # Test 3: Excel output
        test_excel_output()
        
        print("\n✅ Todos los tests completados exitosamente!")
        
    except Exception as e:
        print(f"\n❌ Error en tests: {e}")
        import traceback
        traceback.print_exc()
