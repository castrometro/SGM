from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from openpyxl import load_workbook
from io import BytesIO
import unicodedata


def _normalize(text):
    if text is None:
        return ""
    # Normalizar a ASCII simple y lowercase
    if not isinstance(text, str):
        text = str(text)
    text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii')
    return text.strip().lower()


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def leer_headers_excel_rindegastos(request):
    """
    Endpoint exclusivo RindeGastos para leer headers del Excel y detectar
    dinámicamente las columnas de Centros de Costo (CC) como el rango entre
    la última columna que contiene 'Nombre cuenta' y la columna 'Fecha aprobacion'.
    """
    try:
        if 'archivo' not in request.FILES:
            return Response({'error': 'No se encontró archivo en la petición'}, status=400)

        archivo = request.FILES['archivo']

        if not archivo.name.lower().endswith(('.xlsx', '.xls')):
            return Response({'error': 'El archivo debe ser un Excel (.xlsx o .xls)'}, status=400)

        # Leer workbook en memoria
        contenido = archivo.read()
        wb = load_workbook(BytesIO(contenido), read_only=True)
        ws = wb.active

        # Leer primera fila (headers)
        headers = []
        primera_fila = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        for v in primera_fila:
            headers.append(v if v is not None else '')

        # Detectar índice de la última columna que contenga "Nombre cuenta"
        last_nombre_idx = -1
        for i, h in enumerate(headers):
            if 'nombre cuenta' in _normalize(h):
                last_nombre_idx = i

        # Detectar índice de "Fecha aprobacion" (tolerar acentos)
        fecha_ap_idx = None
        for i, h in enumerate(headers):
            hn = _normalize(h)
            if 'fecha' in hn and 'aprobacion' in hn:
                fecha_ap_idx = i
                break

        centros_costo = {}
        # Si encontramos ambos extremos, tomar el rango intermedio como CC
        if last_nombre_idx != -1 and fecha_ap_idx is not None and fecha_ap_idx - last_nombre_idx > 1:
            for pos in range(last_nombre_idx + 1, fecha_ap_idx):
                nombre = headers[pos]
                if nombre and str(nombre).strip() != '':
                    centros_costo[str(nombre)] = {"posicion": pos, "nombre": str(nombre)}

        # Fallback: si no se encontraron por rango, intentar por nombres comunes conocidos
        if not centros_costo:
            conocidos = ['PyC', 'PS', 'EB', 'CO', 'RE', 'TR', 'CF', 'LRC']
            for i, h in enumerate(headers):
                hs = str(h).strip() if h is not None else ''
                if hs in conocidos:
                    centros_costo[hs] = {"posicion": i, "nombre": hs}

        wb.close()

        return Response({
            'headers': [str(h) if h is not None else '' for h in headers],
            'total_columnas': len(headers),
            'centros_costo': centros_costo,
            'mensaje': 'Headers leídos exitosamente (RindeGastos)'
        })

    except Exception as e:
        return Response({'error': f'Error leyendo headers del Excel: {str(e)}'}, status=500)
