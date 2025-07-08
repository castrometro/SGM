#!/usr/bin/env python
"""
Script para diagnosticar qué cuentas ESF se están sumando en el saldo anterior.
Este script simula la lógica del libro mayor para mostrar exactamente qué cuentas
del archivo Excel se clasifican como ESF y cuáles se suman al saldo anterior.
"""

import os
import sys
import logging
from decimal import Decimal
from collections import defaultdict

# Configurar el entorno Django
sys.path.append('/root/SGM')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'sgm.settings')

import django
django.setup()

from django.db import transaction
from contabilidad.models import CuentaContable, AccountClassification, ClasificacionSet, ClasificacionOption
from openpyxl import load_workbook

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def identificar_clasificacion_esf_eri(cuenta_obj):
    """Identifica si una cuenta es ESF o ERI basándose en sus clasificaciones"""
    try:
        # Buscar sets ESF/ERI
        sets_esf_eri = ClasificacionSet.objects.filter(
            cliente=cuenta_obj.cliente,
            nombre__in=['ESF', 'ERI']
        )
        
        clasificaciones = {}
        for set_clas in sets_esf_eri:
            try:
                classification = AccountClassification.objects.get(
                    cuenta=cuenta_obj,
                    set_clas=set_clas
                )
                clasificaciones[set_clas.nombre] = classification.opcion.valor
            except AccountClassification.DoesNotExist:
                clasificaciones[set_clas.nombre] = None
        
        # Lógica de clasificación ESF/ERI
        esf_val = clasificaciones.get('ESF')
        eri_val = clasificaciones.get('ERI')
        
        if esf_val == '1' and eri_val == '0':
            return 'ESF'
        elif esf_val == '0' and eri_val == '1':
            return 'ERI'
        else:
            return None
    except Exception as e:
        logger.error(f"Error identificando clasificación ESF/ERI para cuenta {cuenta_obj.codigo}: {e}")
        return None

def analizar_cuentas_esf_archivo(file_path, cliente_id):
    """Analiza qué cuentas ESF aparecen en el archivo y cuáles se suman"""
    
    print(f"\n{'='*80}")
    print(f"ANÁLISIS DE CUENTAS ESF EN EL ARCHIVO LIBRO MAYOR")
    print(f"{'='*80}")
    print(f"Archivo: {file_path}")
    print(f"Cliente ID: {cliente_id}")
    
    # Cargar el archivo Excel
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    
    # Identificar columnas
    header_row = list(ws.iter_rows(min_row=10, max_row=10, values_only=True))[0]
    
    # Mapear las columnas importantes
    C = None  # Columna de cuenta/descripción
    D = None  # Columna DEBE
    S = None  # Columna SALDO
    
    for i, cell in enumerate(header_row):
        if cell and isinstance(cell, str):
            if cell.strip().upper() == 'CUENTA':
                C = i
            elif cell.strip().upper() == 'DEBE':
                D = i
            elif cell.strip().upper() == 'SALDO':
                S = i
    
    if C is None:
        print("❌ No se encontró la columna CUENTA")
        return
    
    print(f"📊 Columnas identificadas:")
    print(f"  - CUENTA: {C}")
    print(f"  - DEBE: {D}")
    print(f"  - SALDO: {S}")
    
    # Buscar cuentas con saldo anterior
    cuentas_con_saldo_anterior = []
    cuentas_esf_sumadas = []
    cuentas_esf_no_sumadas = []
    total_esf_archivo = Decimal('0')
    
    processed_accounts = {}
    
    for row in ws.iter_rows(min_row=11, values_only=True):
        cell = row[C]
        
        # Buscar filas de SALDO ANTERIOR
        if isinstance(cell, str) and cell.startswith("SALDO ANTERIOR"):
            # Extraer código de cuenta
            texto_completo = cell.split(":", 1)[1].strip()
            partes = texto_completo.split(" ", 1)
            code = partes[0].strip()
            nombre_real = partes[1].strip() if len(partes) > 1 else f"Cuenta {code}"
            
            # Obtener el saldo
            if S and row[S] is not None:
                saldo = Decimal(str(row[S]))
                origen_saldo = f"SALDO={row[S]}"
            else:
                saldo = Decimal(str(row[D] or 0))
                origen_saldo = f"DEBE={row[D]}"
            
            # Buscar la cuenta en la base de datos
            try:
                cuenta_obj = CuentaContable.objects.get(
                    cliente_id=cliente_id,
                    codigo=code
                )
                
                # Verificar clasificación ESF/ERI
                clasificacion = identificar_clasificacion_esf_eri(cuenta_obj)
                
                cuenta_info = {
                    'codigo': code,
                    'nombre': nombre_real,
                    'saldo': saldo,
                    'origen_saldo': origen_saldo,
                    'clasificacion': clasificacion,
                    'es_patrimonio': 'patrimonio' in nombre_real.lower() or 'capital' in nombre_real.lower()
                }
                
                cuentas_con_saldo_anterior.append(cuenta_info)
                
                if clasificacion == 'ESF':
                    cuentas_esf_sumadas.append(cuenta_info)
                    total_esf_archivo += saldo
                
                processed_accounts[code] = cuenta_obj
                
            except CuentaContable.DoesNotExist:
                print(f"⚠️  Cuenta {code} no encontrada en la base de datos")
                continue
    
    # Buscar cuentas ESF en la base de datos que NO aparecen en el archivo
    todas_cuentas_esf = CuentaContable.objects.filter(
        cliente_id=cliente_id,
        accountclassification__set_clas__nombre='ESF',
        accountclassification__opcion__valor='1'
    ).exclude(
        accountclassification__set_clas__nombre='ERI',
        accountclassification__opcion__valor='1'
    )
    
    for cuenta in todas_cuentas_esf:
        if cuenta.codigo not in processed_accounts:
            cuentas_esf_no_sumadas.append({
                'codigo': cuenta.codigo,
                'nombre': cuenta.nombre,
                'saldo': 'N/A (no en archivo)',
                'clasificacion': 'ESF',
                'es_patrimonio': 'patrimonio' in (cuenta.nombre or '').lower() or 'capital' in (cuenta.nombre or '').lower()
            })
    
    # Mostrar resultados
    print(f"\n📋 RESUMEN DE ANÁLISIS:")
    print(f"  - Total cuentas con saldo anterior en archivo: {len(cuentas_con_saldo_anterior)}")
    print(f"  - Cuentas ESF sumadas desde archivo: {len(cuentas_esf_sumadas)}")
    print(f"  - Cuentas ESF en BD que NO aparecen en archivo: {len(cuentas_esf_no_sumadas)}")
    print(f"  - Total ESF acumulado desde archivo: ${total_esf_archivo:,.2f}")
    
    # Mostrar cuentas ESF que SÍ se suman (desde archivo)
    print(f"\n✅ CUENTAS ESF QUE SÍ SE SUMAN (desde archivo):")
    print(f"{'Código':<15} {'Saldo':<15} {'Origen':<12} {'Patrimonio':<10} {'Nombre'}")
    print("-" * 100)
    for cuenta in cuentas_esf_sumadas:
        patrimonio_mark = "SÍ" if cuenta['es_patrimonio'] else "NO"
        print(f"{cuenta['codigo']:<15} ${cuenta['saldo']:>12,.2f} {cuenta['origen_saldo']:<12} {patrimonio_mark:<10} {cuenta['nombre']}")
    
    # Mostrar cuentas ESF que NO se suman (no están en archivo)
    if cuentas_esf_no_sumadas:
        print(f"\n❌ CUENTAS ESF QUE NO SE SUMAN (no están en archivo):")
        print(f"{'Código':<15} {'Patrimonio':<10} {'Nombre'}")
        print("-" * 80)
        for cuenta in cuentas_esf_no_sumadas:
            patrimonio_mark = "SÍ" if cuenta['es_patrimonio'] else "NO"
            print(f"{cuenta['codigo']:<15} {patrimonio_mark:<10} {cuenta['nombre']}")
    
    # Mostrar top 5 cuentas ESF con mayor saldo
    if cuentas_esf_sumadas:
        print(f"\n🔝 TOP 5 CUENTAS ESF CON MAYOR SALDO:")
        cuentas_ordenadas = sorted(cuentas_esf_sumadas, key=lambda x: abs(x['saldo']), reverse=True)
        for i, cuenta in enumerate(cuentas_ordenadas[:5]):
            print(f"{i+1}. {cuenta['codigo']} - ${cuenta['saldo']:,.2f} - {cuenta['nombre']}")
    
    # Análisis especial de cuentas patrimonio
    cuentas_patrimonio = [c for c in cuentas_con_saldo_anterior if c['es_patrimonio']]
    if cuentas_patrimonio:
        print(f"\n🏦 CUENTAS DE PATRIMONIO DETECTADAS:")
        for cuenta in cuentas_patrimonio:
            print(f"  - {cuenta['codigo']}: ${cuenta['saldo']:,.2f} - {cuenta['clasificacion']} - {cuenta['nombre']}")
    
    wb.close()
    return {
        'total_cuentas_archivo': len(cuentas_con_saldo_anterior),
        'cuentas_esf_sumadas': len(cuentas_esf_sumadas),
        'cuentas_esf_no_sumadas': len(cuentas_esf_no_sumadas),
        'total_esf_archivo': total_esf_archivo,
        'cuentas_patrimonio': len(cuentas_patrimonio)
    }

if __name__ == "__main__":
    # Parámetros del análisis
    archivo_ejemplo = "/root/SGM/ejemplo_libro_mayor.xlsx"  # Ajustar path según sea necesario
    cliente_id = 1  # Ajustar según sea necesario
    
    if len(sys.argv) > 1:
        archivo_ejemplo = sys.argv[1]
    if len(sys.argv) > 2:
        cliente_id = int(sys.argv[2])
    
    print(f"Iniciando análisis...")
    print(f"Archivo: {archivo_ejemplo}")
    print(f"Cliente ID: {cliente_id}")
    
    if not os.path.exists(archivo_ejemplo):
        print(f"❌ Archivo no encontrado: {archivo_ejemplo}")
        print("💡 Uso: python debug_cuentas_esf_sumadas.py <path_archivo_excel> [cliente_id]")
        sys.exit(1)
    
    try:
        resultado = analizar_cuentas_esf_archivo(archivo_ejemplo, cliente_id)
        print(f"\n✅ Análisis completado exitosamente")
        print(f"📊 Estadísticas finales: {resultado}")
    except Exception as e:
        print(f"❌ Error durante el análisis: {e}")
        import traceback
        traceback.print_exc()
