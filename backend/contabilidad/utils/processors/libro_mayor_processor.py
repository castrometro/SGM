import hashlib
import os
import re
import datetime
import pandas as pd
from openpyxl import load_workbook
from django.utils import timezone
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from .base_processor import BaseProcessor
from contabilidad.models import (
    UploadLog,
    LibroMayorArchivo,
    CuentaContable,
    AperturaCuenta,
    CentroCosto,
    Auxiliar,
    TipoDocumento,
    NombreIngles,
    ClasificacionCuentaArchivo,
    ClasificacionSet,
    ClasificacionOption,
    AccountClassification,
    ExcepcionValidacion,
    Incidencia,
)
from contabilidad.models_incidencias import IncidenciaResumen
from contabilidad.utils.activity_logger import registrar_actividad_tarjeta


class LibroMayorProcessor(BaseProcessor):
    @staticmethod
    def procesar(upload_log_id):
        """Versión trasladada desde tasks.procesar_libro_mayor"""
        # Copia directa de la lógica antigua
        import hashlib
        processor = LibroMayorProcessor()
        try:
            upload_log = UploadLog.objects.get(id=upload_log_id)
        except UploadLog.DoesNotExist:
            processor.logger.error("UploadLog con id %s no encontrado", upload_log_id)
            return f"Error: UploadLog {upload_log_id} no encontrado"
        upload_log.estado = "procesando"
        upload_log.save(update_fields=["estado"])
        inicio = timezone.now()
        archivo_obj = None
        try:
            es_valido, msg_valid = UploadLog.validar_nombre_archivo(
                upload_log.nombre_archivo_original,
                "LibroMayor",
                upload_log.cliente.rut,
            )
            if not es_valido:
                upload_log.estado = "error"
                upload_log.errores = f"Nombre de archivo inválido: {msg_valid}"
                upload_log.tiempo_procesamiento = timezone.now() - inicio
                upload_log.save()
                return f"Error: {msg_valid}"
            ruta_relativa = upload_log.ruta_archivo
            if not ruta_relativa:
                upload_log.estado = "error"
                upload_log.errores = "No hay ruta de archivo especificada en el upload_log"
                upload_log.tiempo_procesamiento = timezone.now() - inicio
                upload_log.save()
                return "Error: No hay ruta de archivo especificada"
            ruta_completa = default_storage.path(ruta_relativa)
            if not os.path.exists(ruta_completa):
                upload_log.estado = "error"
                upload_log.errores = f"Archivo temporal no encontrado en: {ruta_relativa}"
                upload_log.tiempo_procesamiento = timezone.now() - inicio
                upload_log.save()
                return "Error: Archivo temporal no encontrado"
            with open(ruta_completa, "rb") as f:
                contenido = f.read()
                archivo_hash = hashlib.sha256(contenido).hexdigest()
            upload_log.hash_archivo = archivo_hash
            upload_log.save(update_fields=["hash_archivo"])
            if not upload_log.cierre:
                upload_log.estado = "error"
                upload_log.errores = "LibroMayorUpload has no cierre."
                upload_log.tiempo_procesamiento = timezone.now() - inicio
                upload_log.save()
                return "Error: LibroMayorUpload has no cierre."
            rut_limpio = upload_log.cliente.rut.replace(".", "").replace("-", "") if upload_log.cliente.rut else str(upload_log.cliente.id)
            nombre_sin_ext = re.sub(r"\.(xlsx|xls)$", "", upload_log.nombre_archivo_original, flags=re.IGNORECASE)
            patron_periodo = rf"^{rut_limpio}_(LibroMayor|Mayor)_(\d{{6}})$"
            match = re.match(patron_periodo, nombre_sin_ext)
            periodo = match.group(2) if match else "000000"
            nombre_final = f"libro_mayor_{upload_log.cliente.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            archivo_obj, created = LibroMayorArchivo.objects.get_or_create(
                cliente=upload_log.cliente,
                defaults={
                    "archivo": ContentFile(contenido, name=nombre_final),
                    "estado": "procesando",
                    "procesado": False,
                    "upload_log": upload_log,
                    "periodo": periodo,
                },
            )
            if not created:
                if archivo_obj.archivo:
                    archivo_obj.archivo.delete()
                archivo_obj.archivo.save(nombre_final, ContentFile(contenido))
                archivo_obj.estado = "procesando"
                archivo_obj.procesado = False
                archivo_obj.upload_log = upload_log
                archivo_obj.periodo = periodo
                archivo_obj.save()
            excepciones_por_cuenta = {}
            for exc in ExcepcionValidacion.objects.filter(cliente=upload_log.cliente):
                key = f"{exc.codigo_cuenta}_{exc.tipo_excepcion}"
                excepciones_por_cuenta[key] = exc
            tipos_doc_map = {td.codigo: td for td in TipoDocumento.objects.filter(cliente=upload_log.cliente)}
            nombres_ingles_map = {ni.cuenta_codigo: ni.nombre_ingles for ni in NombreIngles.objects.filter(cliente=upload_log.cliente)}
            clasif_qs = (
                ClasificacionCuentaArchivo.objects.filter(cliente=upload_log.cliente)
                .select_related("upload_log")
                .order_by("-upload_log__fecha_subida", "-upload_log__id")
            )
            clasificaciones_por_cuenta = {}
            for reg in clasif_qs:
                if reg.numero_cuenta not in clasificaciones_por_cuenta:
                    clasificaciones_por_cuenta[reg.numero_cuenta] = reg.clasificaciones
            wb = load_workbook(ruta_completa, read_only=True, data_only=True)
            ws = wb.active
            def clean(h):
                h = h.strip().upper()
                for a, b in [("Á","A"),("É","E"),("Í","I"),("Ó","O"),("Ú","U"),("Ñ","N")]:
                    h = h.replace(a, b)
                return re.sub(r"[^A-Z0-9]", "", h)
            raw = next(ws.iter_rows(min_row=9, max_row=9, values_only=True))
            idx = {clean(h): i for i, h in enumerate(raw) if isinstance(h, str)}
            C = idx["CUENTA"]
            F = idx["FECHA"]
            NC = idx.get("NCOMPROBANTE")
            T = idx.get("TIPO")
            NI = idx.get("NINTERNO")
            CC = idx.get("CENTRODECOSTO")
            AX = idx.get("AUXILIAR")
            TD = idx.get("TIPODOC") or idx.get("TIPODOC.")
            ND = idx.get("NUMERODOC")
            DG = idx.get("DETDEGASTOINSTFINANCIERO")
            D = idx["DEBE"]
            H = idx["HABER"]
            S = idx["SALDO"]
            DS = idx["DESCRIPCION"]
            errores = []
            current_code = None
            fechas_mov = []
            movimientos_creados = 0
            incidencias_creadas = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=11, values_only=True), start=11):
                cell = row[C]
                if isinstance(cell, str) and cell.startswith("SALDO ANTERIOR"):
                    try:
                        resto = cell.split(":", 1)[1].strip()
                        code, name = resto.split(" ", 1)
                        saldo_ant = row[S] or 0
                        cuenta_obj, _ = CuentaContable.objects.get_or_create(
                            cliente=upload_log.cliente,
                            codigo=code,
                            defaults={"nombre": name},
                        )
                        AperturaCuenta.objects.update_or_create(
                            cierre=upload_log.cierre,
                            cuenta=cuenta_obj,
                            defaults={"saldo_anterior": saldo_ant},
                        )
                        current_code = code
                    except Exception as e:
                        errores.append(f"F{row_idx} Apertura: {e}")
                    continue
                if current_code and row[F] is not None:
                    raw_fecha = row[F]
                    try:
                        if isinstance(raw_fecha, str):
                            s = raw_fecha.strip().strip('"').strip("'")
                            fecha = datetime.datetime.strptime(s, "%d/%m/%Y").date()
                        elif isinstance(raw_fecha, datetime.datetime):
                            fecha = raw_fecha.date()
                        elif isinstance(raw_fecha, datetime.date):
                            fecha = raw_fecha
                        else:
                            raise ValueError(f"Formato de fecha no soportado: {raw_fecha!r}")
                        fechas_mov.append(fecha)
                    except Exception as e:
                        errores.append(f"F{row_idx} Fecha: {e}")
                        continue
                    cuenta_obj, _ = CuentaContable.objects.get_or_create(
                        cliente=upload_log.cliente,
                        codigo=current_code,
                    )
                    if not cuenta_obj.nombre_en:
                        nombre_ing = nombres_ingles_map.get(cuenta_obj.codigo)
                        if nombre_ing:
                            cuenta_obj.nombre_en = nombre_ing
                            cuenta_obj.save(update_fields=["nombre_en"])
                    clasif_info = clasificaciones_por_cuenta.get(cuenta_obj.codigo)
                    if clasif_info:
                        for set_nombre, opcion_valor in clasif_info.items():
                            set_obj, _ = ClasificacionSet.objects.get_or_create(
                                cliente=upload_log.cliente, nombre=set_nombre
                            )
                            opcion_obj, _ = ClasificacionOption.objects.get_or_create(
                                set_clas=set_obj, valor=str(opcion_valor).strip()
                            )
                            AccountClassification.objects.update_or_create(
                                cuenta=cuenta_obj,
                                set_clas=set_obj,
                                defaults={"opcion": opcion_obj, "asignado_por": None},
                            )
                    centro_obj = None
                    if CC and row[CC]:
                        centro_obj, _ = CentroCosto.objects.get_or_create(
                            cliente=upload_log.cliente,
                            nombre=str(row[CC]).strip(),
                        )
                    aux_obj = None
                    if AX and row[AX]:
                        aux_obj, _ = Auxiliar.objects.get_or_create(
                            rut_auxiliar=str(row[AX]).strip(),
                            defaults={"nombre": "", "fecha_creacion": timezone.now()},
                        )
                    td_obj = None
                    mov_incompleto = False
                    if TD is not None:
                        codigo_td = str(row[TD] or "").strip()
                        tiene_excepcion_tipodoc = f"{current_code}_tipos_doc_no_reconocidos" in excepciones_por_cuenta or \
                                                f"{current_code}_movimientos_tipodoc_nulo" in excepciones_por_cuenta
                        if codigo_td:
                            td_obj = tipos_doc_map.get(codigo_td)
                            if td_obj is None and not tiene_excepcion_tipodoc:
                                Incidencia.objects.create(
                                    cierre=upload_log.cierre,
                                    tipo="negocio",
                                    descripcion=(
                                        f"Movimiento {row_idx-10}, cuenta {current_code}: "
                                        f"Tipo de documento '{codigo_td}' no encontrado"
                                    ),
                                )
                                incidencias_creadas += 1
                                mov_incompleto = True
                        else:
                            if not tiene_excepcion_tipodoc:
                                Incidencia.objects.create(
                                    cierre=upload_log.cierre,
                                    tipo="negocio",
                                    descripcion=(
                                        f"Movimiento {row_idx-10}, cuenta {current_code}: "
                                        "Tipo de documento vacío"
                                    ),
                                )
                                incidencias_creadas += 1
                                mov_incompleto = True
                    mov = MovimientoContable.objects.create(
                        cierre=upload_log.cierre,
                        cuenta=cuenta_obj,
                        fecha=fecha,
                        tipo_documento=td_obj,
                        tipo_doc_codigo=codigo_td,
                        numero_documento=str(row[ND] or ""),
                        tipo=str(row[T] or ""),
                        numero_comprobante=str(row[NC] or ""),
                        numero_interno=str(row[NI] or ""),
                        centro_costo=centro_obj,
                        auxiliar=aux_obj,
                        detalle_gasto=str(row[DG] or ""),
                        debe=row[D] or 0,
                        haber=row[H] or 0,
                        descripcion=str(row[DS] or ""),
                        flag_incompleto=mov_incompleto,
                    )
                    tiene_excepcion_nombre = f"{current_code}_cuentas_sin_nombre_ingles" in excepciones_por_cuenta
                    tiene_excepcion_clasificacion = f"{current_code}_cuentas_sin_clasificacion" in excepciones_por_cuenta
                    if not cuenta_obj.nombre_en and not tiene_excepcion_nombre:
                        Incidencia.objects.create(
                            cierre=upload_log.cierre,
                            tipo="negocio",
                            descripcion=(
                                f"Movimiento {row_idx-10}, cuenta {current_code}: "
                                "No tiene nombre en inglés"
                            ),
                        )
                        incidencias_creadas += 1
                        mov.flag_incompleto = True
                    if not AccountClassification.objects.filter(cuenta=cuenta_obj).exists() and not tiene_excepcion_clasificacion:
                        Incidencia.objects.create(
                            cierre=upload_log.cierre,
                            tipo="negocio",
                            descripcion=(
                                f"Movimiento {row_idx-10}, cuenta {current_code}: "
                                "Cuenta sin clasificación asignada"
                            ),
                        )
                        incidencias_creadas += 1
                        mov.flag_incompleto = True
                    if mov.flag_incompleto:
                        mov.save(update_fields=["flag_incompleto"])
                    movimientos_creados += 1
            fecha_inicio = min(fechas_mov) if fechas_mov else None
            fecha_fin = max(fechas_mov) if fechas_mov else None
            archivo_obj.estado = "completado" if not errores else "error"
            archivo_obj.procesado = True
            archivo_obj.errores = "\n".join(errores) if errores else ""
            archivo_obj.save()
            upload_log.estado = "completado" if not errores else "error"
            upload_log.errores = "\n".join(errores) if errores else ""
            upload_log.resumen = {
                "movimientos_creados": movimientos_creados,
                "incidencias_creadas": incidencias_creadas,
                "archivo_hash": archivo_hash,
            }
            upload_log.tiempo_procesamiento = timezone.now() - inicio
            upload_log.save()
            try:
                logger = processor.logger
                logger.info(f"Generando incidencias consolidadas para upload_log {upload_log.id}")
                from django.db.models import Q
                raw_counts = {}
                elementos_detallados = {}
                incidencias_tipodoc_vacio = Incidencia.objects.filter(
                    cierre=upload_log.cierre,
                    descripcion__icontains="Tipo de documento vacío",
                )
                if incidencias_tipodoc_vacio.exists():
                    count = incidencias_tipodoc_vacio.count()
                    raw_counts["movimientos_tipodoc_nulo"] = count
                    cuentas_afectadas = []
                    for inc in incidencias_tipodoc_vacio[:50]:
                        match = re.search(r'cuenta\s+([\d\-]+):', inc.descripcion, flags=re.IGNORECASE)
                        if match:
                            codigo_cuenta = match.group(1)
                            if codigo_cuenta not in cuentas_afectadas:
                                cuentas_afectadas.append(codigo_cuenta)
                    elementos_detallados["movimientos_tipodoc_nulo"] = cuentas_afectadas
                incidencias_tipodoc_no_reconocido = Incidencia.objects.filter(
                    cierre=upload_log.cierre,
                    descripcion__icontains="no encontrado",
                ).exclude(
                    descripcion__icontains="vacío",
                )
                if incidencias_tipodoc_no_reconocido.exists():
                    count = incidencias_tipodoc_no_reconocido.count()
                    raw_counts["tipos_doc_no_reconocidos"] = count
                    codigos_no_reconocidos = []
                    for inc in incidencias_tipodoc_no_reconocido[:50]:
                        match = re.search(r"'([^']+)' no encontrado", inc.descripcion)
                        if match:
                            codigo_td = match.group(1)
                            if codigo_td not in codigos_no_reconocidos:
                                codigos_no_reconocidos.append(codigo_td)
                    elementos_detallados["tipos_doc_no_reconocidos"] = codigos_no_reconocidos
                incidencias_sin_clasif = Incidencia.objects.filter(
                    cierre=upload_log.cierre,
                    descripcion__icontains="sin clasificación",
                )
                if incidencias_sin_clasif.exists():
                    count = incidencias_sin_clasif.count()
                    raw_counts["cuentas_sin_clasificacion"] = count
                    cuentas_sin_clasif = []
                    for inc in incidencias_sin_clasif[:50]:
                        match = re.search(r'cuenta\s+([\d\-]+):', inc.descripcion, flags=re.IGNORECASE)
                        if match:
                            codigo_cuenta = match.group(1)
                            if codigo_cuenta not in cuentas_sin_clasif:
                                cuentas_sin_clasif.append(codigo_cuenta)
                    elementos_detallados["cuentas_sin_clasificacion"] = cuentas_sin_clasif
                incidencias_sin_nombre = Incidencia.objects.filter(
                    cierre=upload_log.cierre,
                    descripcion__icontains="nombre en inglés",
                )
                if incidencias_sin_nombre.exists():
                    count = incidencias_sin_nombre.count()
                    raw_counts["cuentas_sin_nombre_ingles"] = count
                    cuentas_sin_nombre = []
                    for inc in incidencias_sin_nombre[:50]:
                        match = re.search(r'cuenta\s+([\d\-]+):', inc.descripcion, flags=re.IGNORECASE)
                        if match:
                            codigo_cuenta = match.group(1)
                            if codigo_cuenta not in cuentas_sin_nombre:
                                cuentas_sin_nombre.append(codigo_cuenta)
                    elementos_detallados["cuentas_sin_nombre_ingles"] = cuentas_sin_nombre
                cuentas_nuevas = upload_log.resumen.get("cuentas_nuevas", 0)
                if cuentas_nuevas > 0:
                    raw_counts["cuentas_nuevas_detectadas"] = cuentas_nuevas
                    cuentas_recientes = list(
                        CuentaContable.objects.filter(cliente=upload_log.cliente).order_by('-id')[:50].values_list('codigo', flat=True)
                    )
                    elementos_detallados["cuentas_nuevas_detectadas"] = cuentas_recientes
                incidencias_consolidadas_creadas = 0
                _MENSAJES = {
                    "movimientos_tipodoc_nulo": "Se encontraron movimientos sin tipo de documento asignado (campo vacío)",
                    "tipos_doc_no_reconocidos": "Se encontraron códigos de tipo de documento que no están registrados en el sistema",
                    "cuentas_sin_clasificacion": "Se detectaron cuentas contables sin clasificación asignada",
                    "cuentas_sin_nombre_ingles": "Se encontraron cuentas sin nombre en inglés configurado",
                    "cuentas_nuevas_detectadas": "Se detectaron nuevas cuentas contables en el libro mayor",
                }
                _ACCIONES = {
                    "movimientos_tipodoc_nulo": "Revisar el archivo fuente y completar los tipos de documento faltantes",
                    "tipos_doc_no_reconocidos": "Cargar archivo de tipos de documento actualizado con los códigos faltantes",
                    "cuentas_sin_clasificacion": "Cargar archivo de clasificaciones o asignar clasificaciones manualmente",
                    "cuentas_sin_nombre_ingles": "Cargar archivo de nombres en inglés actualizado",
                    "cuentas_nuevas_detectadas": "Revisar nuevas cuentas y configurar nombres en inglés y clasificaciones",
                }
                for sub_tipo, count in raw_counts.items():
                    if count > 100:
                        severidad = "critica"
                    elif count > 50:
                        severidad = "alta"
                    elif count > 10:
                        severidad = "media"
                    else:
                        severidad = "baja"
                    elementos_afectados = elementos_detallados.get(sub_tipo, [])
                    IncidenciaResumen.objects.create(
                        upload_log=upload_log,
                        tipo_incidencia=sub_tipo,
                        cantidad_afectada=count,
                        estado="activa",
                        severidad=severidad,
                        mensaje_usuario=_MENSAJES.get(sub_tipo, f"Incidencia detectada: {sub_tipo}"),
                        accion_sugerida=_ACCIONES.get(sub_tipo, "Revisar y corregir manualmente"),
                        codigo_problema=f"{sub_tipo.upper()}_{upload_log.id}",
                        elementos_afectados=elementos_afectados,
                        detalle_muestra={
                            "upload_log_id": upload_log.id,
                            "cantidad": count,
                            "tipo": sub_tipo,
                            "elementos_muestra": elementos_afectados[:10],
                            "total_elementos": len(elementos_afectados),
                        },
                    )
                    incidencias_consolidadas_creadas += 1
                upload_log.resumen.update({
                    "incidencias_consolidadas_creadas": incidencias_consolidadas_creadas,
                    "tipos_incidencia_detectados": list(raw_counts.keys()),
                    "conteos_por_tipo": raw_counts,
                    "sistema_consolidado": True,
                })
                upload_log.save(update_fields=['resumen'])
                logger.info(f"Creadas {incidencias_consolidadas_creadas} incidencias consolidadas por sub-tipo")
            except Exception as e:
                processor.logger.error(f"Error generando incidencias consolidadas: {e}")
                pass
            try:
                os.remove(ruta_completa)
            except OSError:
                pass
            registrar_actividad_tarjeta(
                cliente_id=upload_log.cliente.id,
                periodo=upload_log.cierre.periodo if upload_log.cierre else date.today().strftime("%Y-%m"),
                tarjeta="libro_mayor",
                accion="process_excel",
                descripcion=f"Procesado archivo de libro mayor: {movimientos_creados} movimientos",
                usuario=None,
                detalles={
                    "upload_log_id": upload_log.id,
                    "movimientos_creados": movimientos_creados,
                    "incidencias_creadas": incidencias_creadas,
                },
                resultado="exito" if not errores else "warning",
                ip_address=None,
            )
            return f"Completado: {movimientos_creados} movimientos"
        except Exception as e:
            if archivo_obj:
                archivo_obj.estado = "error"
                archivo_obj.errores = str(e)
                archivo_obj.save()
            upload_log.estado = "error"
            upload_log.errores = str(e)
            upload_log.tiempo_procesamiento = timezone.now() - inicio
            upload_log.save()
            registrar_actividad_tarjeta(
                cliente_id=upload_log.cliente.id,
                periodo=upload_log.cierre.periodo if upload_log.cierre else date.today().strftime("%Y-%m"),
                tarjeta="libro_mayor",
                accion="process_excel",
                descripcion=f"Error al procesar libro mayor: {str(e)}",
                usuario=None,
                detalles={"upload_log_id": upload_log.id},
                resultado="error",
                ip_address=None,
            )
            return f"Error: {str(e)}"
