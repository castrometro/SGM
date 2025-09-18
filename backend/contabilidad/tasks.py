# backend/contabilidad/tasks.py


import datetime
import hashlib
import logging
import os
import re
import time
import json
import redis
from datetime import date
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from celery import shared_task, group, chord
from django.contrib.auth.models import User
from django.core.files.base import ContentFile
from contabilidad.models import (
    AccountClassification,
    # ClasificacionCuentaArchivo,  # OBSOLETO - ELIMINADO EN REDISEÑO
    ClasificacionOption,
    ClasificacionSet,
    CuentaContable,
    CentroCosto,
    Auxiliar,
    LibroMayorArchivo, 
    MovimientoContable,
    AperturaCuenta,
    Incidencia,
    NombresEnInglesUpload,
    TipoDocumento,
    UploadLog,
    ExcepcionValidacion,
    CierreContabilidad,
    ReporteFinanciero,
)
# ✨ NUEVO: Importar modelo de incidencias consolidadas
from contabilidad.models_incidencias import IncidenciaResumen
from contabilidad.utils.parser_libro_mayor import parsear_libro_mayor
from contabilidad.utils.parser_nombre_ingles import procesar_archivo_nombres_ingles
from contabilidad.utils.parser_tipo_documento import parsear_tipo_documento_excel
from contabilidad.utils.activity_logger import registrar_actividad_tarjeta
from django.core.files.storage import default_storage
from django.utils import timezone

logger = logging.getLogger(__name__)


@shared_task
def tarea_de_prueba(nombre):
    logger.info("👋 ¡Hola %s desde Celery!", nombre)
    time.sleep(999)
    logger.info("✅ Tarea completada.")
    return f"Completado por {nombre}"  # esto sale en succeeded


#=======PROCESSORS========



@shared_task
def procesar_libro_mayor(upload_log_id):
    """Procesa archivo de Libro Mayor usando UploadLog"""
    import hashlib
    import os
    import re
    from openpyxl import load_workbook
    import datetime

    logger.info("Iniciando procesamiento de libro mayor para upload_log %s", upload_log_id)

    try:
        upload_log = UploadLog.objects.get(id=upload_log_id)
    except UploadLog.DoesNotExist:
        logger.error("UploadLog con id %s no encontrado", upload_log_id)
        return f"Error: UploadLog {upload_log_id} no encontrado"

    upload_log.estado = "procesando"
    upload_log.save(update_fields=["estado"])
    inicio = timezone.now()

    archivo_obj = None

    try:
        es_valido, msg_valid = UploadLog.validar_nombre_archivo(
            upload_log.nombre_archivo_original,
            "LibroMayor",
            upload_log.cliente.rut,
        )

        if not es_valido:
            upload_log.estado = "error"
            upload_log.errores = f"Nombre de archivo inválido: {msg_valid}"
            upload_log.tiempo_procesamiento = timezone.now() - inicio
            upload_log.save()
            return f"Error: {msg_valid}"

        # Usar la ruta que se guardó en el upload_log
        ruta_relativa = upload_log.ruta_archivo
        if not ruta_relativa:
            upload_log.estado = "error"
            upload_log.errores = "No hay ruta de archivo especificada en el upload_log"
            upload_log.tiempo_procesamiento = timezone.now() - inicio
            upload_log.save()
            return "Error: No hay ruta de archivo especificada"

        ruta_completa = default_storage.path(ruta_relativa)

        if not os.path.exists(ruta_completa):
            upload_log.estado = "error"
            upload_log.errores = f"Archivo temporal no encontrado en: {ruta_relativa}"
            upload_log.tiempo_procesamiento = timezone.now() - inicio
            upload_log.save()
            return "Error: Archivo temporal no encontrado"

        with open(ruta_completa, "rb") as f:
            contenido = f.read()
            archivo_hash = hashlib.sha256(contenido).hexdigest()

        upload_log.hash_archivo = archivo_hash
        upload_log.save(update_fields=["hash_archivo"])

        # Verificar que el upload_log tenga un cierre asignado
        if not upload_log.cierre:
            upload_log.estado = "error"
            upload_log.errores = "LibroMayorUpload has no cierre."
            upload_log.tiempo_procesamiento = timezone.now() - inicio
            upload_log.save()
            return "Error: LibroMayorUpload has no cierre."

        # Extraer período del nombre del archivo
        import re
        rut_limpio = upload_log.cliente.rut.replace(".", "").replace("-", "") if upload_log.cliente.rut else str(upload_log.cliente.id)
        nombre_sin_ext = re.sub(r"\.(xlsx|xls)$", "", upload_log.nombre_archivo_original, flags=re.IGNORECASE)
        patron_periodo = rf"^{rut_limpio}_(LibroMayor|Mayor)_(\d{{6}})$"
        match = re.match(patron_periodo, nombre_sin_ext)
        periodo = match.group(2) if match else "000000"

        nombre_final = f"libro_mayor_{upload_log.cliente.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        archivo_obj, created = LibroMayorArchivo.objects.get_or_create(
            cliente=upload_log.cliente,
            cierre=upload_log.cierre,  # ✅ Agregado campo cierre
            defaults={
                "archivo": ContentFile(contenido, name=nombre_final),
                "estado": "procesando",
                "procesado": False,
                "upload_log": upload_log,
                "periodo": periodo,
            },
        )
        if not created:
            if archivo_obj.archivo:
                try:
                    archivo_obj.archivo.delete()
                except Exception:
                    pass
            archivo_obj.archivo.save(nombre_final, ContentFile(contenido))
            archivo_obj.estado = "procesando"
            archivo_obj.procesado = False
            archivo_obj.upload_log = upload_log
            archivo_obj.periodo = periodo
            archivo_obj.save()

        # Cargar excepciones de validación para este cliente
        excepciones_por_cuenta = {}
        for exc in ExcepcionValidacion.objects.filter(cliente=upload_log.cliente):
            key = f"{exc.codigo_cuenta}_{exc.tipo_excepcion}"
            excepciones_por_cuenta[key] = exc
        
        logger.info(f"Cargadas {len(excepciones_por_cuenta)} excepciones de validación para cliente {upload_log.cliente.id}")

        # Cargar datos auxiliares generados por otras tarjetas para
        # complementar el libro mayor (tipos de documento, nombres en
        # inglés y clasificaciones)
        tipos_doc_map = {
            td.codigo: td
            for td in TipoDocumento.objects.filter(cliente=upload_log.cliente)
        }

        nombres_ingles_map = {
            c.codigo: (c.nombre_en or "")
            for c in CuentaContable.objects.filter(cliente=upload_log.cliente).exclude(nombre_en__isnull=True)
        }

        # Obtener clasificaciones existentes desde AccountClassification (fuente única de verdad)
        clasificaciones_por_cuenta = {}
        for ac in AccountClassification.objects.filter(cliente=upload_log.cliente).select_related('set_clas', 'opcion', 'cuenta'):
            codigo_cuenta = ac.codigo_cuenta_display  # Usa el property que maneja tanto FK como código temporal
            if codigo_cuenta not in clasificaciones_por_cuenta:
                clasificaciones_por_cuenta[codigo_cuenta] = {}
            clasificaciones_por_cuenta[codigo_cuenta][ac.set_clas.nombre] = ac.opcion.valor

        wb = load_workbook(ruta_completa, read_only=True, data_only=True)
        ws = wb.active

        def clean(h):
            h = h.strip().upper()
            for a, b in [("Á","A"),("É","E"),("Í","I"),("Ó","O"),("Ú","U"),("Ñ","N")]:
                h = h.replace(a, b)
            return re.sub(r"[^A-Z0-9]", "", h)

        raw = next(ws.iter_rows(min_row=9, max_row=9, values_only=True))
        idx = {clean(h): i for i, h in enumerate(raw) if isinstance(h, str)}

        C = idx["CUENTA"]
        F = idx["FECHA"]
        NC = idx.get("NCOMPROBANTE")
        T = idx.get("TIPO")
        NI = idx.get("NINTERNO")
        CC = idx.get("CENTRODECOSTO")
        AX = idx.get("AUXILIAR")
        # La cabecera original suele venir como "TIPODOC." con punto al final.
        # Al normalizarla se convierte en "TIPODOC", pero dejamos el alias por
        # claridad.
        TD = idx.get("TIPODOC") or idx.get("TIPODOC.")
        ND = idx.get("NUMERODOC")
        DG = idx.get("DETDEGASTOINSTFINANCIERO")
        D = idx["DEBE"]
        H = idx["HABER"]
        S = idx["SALDO"]
        DS = idx["DESCRIPCION"]

        errores = []
        current_code = None
        fechas_mov = []
        movimientos_creados = 0
        incidencias_creadas = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=11, values_only=True), start=11):
            cell = row[C]

            if isinstance(cell, str) and cell.startswith("SALDO ANTERIOR"):
                try:
                    resto = cell.split(":", 1)[1].strip()
                    code, name = resto.split(" ", 1)
                    saldo_ant = row[S] or 0
                    cuenta_obj, _ = CuentaContable.objects.get_or_create(
                        cliente=upload_log.cliente,
                        codigo=code,
                        defaults={"nombre": name},
                    )
                    AperturaCuenta.objects.update_or_create(
                        cierre=upload_log.cierre,
                        cuenta=cuenta_obj,
                        defaults={"saldo_anterior": saldo_ant},
                    )
                    current_code = code
                except Exception as e:
                    errores.append(f"F{row_idx} Apertura: {e}")
                continue

            if current_code and row[F] is not None: #====es decir, esta fila es un movimiento=====#
                raw_fecha = row[F]
                try:
                    if isinstance(raw_fecha, str):
                        s = raw_fecha.strip().strip('"').strip("'")
                        fecha = datetime.datetime.strptime(s, "%d/%m/%Y").date()
                    elif isinstance(raw_fecha, datetime.datetime):
                        fecha = raw_fecha.date()
                    elif isinstance(raw_fecha, datetime.date):
                        fecha = raw_fecha
                    else:
                        raise ValueError(f"Formato de fecha no soportado: {raw_fecha!r}")
                    fechas_mov.append(fecha)
                except Exception as e:
                    errores.append(f"F{row_idx} Fecha: {e}")
                    continue

                cuenta_obj, _ = CuentaContable.objects.get_or_create(
                    cliente=upload_log.cliente,
                    codigo=current_code,
                )

                # Completar nombre en inglés desde registros previos
                if not cuenta_obj.nombre_en:
                    nombre_ing = nombres_ingles_map.get(cuenta_obj.codigo)
                    if nombre_ing:
                        cuenta_obj.nombre_en = nombre_ing
                        cuenta_obj.save(update_fields=["nombre_en"])

                # Aplicar clasificaciones registradas anteriormente
                clasif_info = clasificaciones_por_cuenta.get(cuenta_obj.codigo)
                if clasif_info:
                    for set_nombre, opcion_valor in clasif_info.items():
                        set_obj, _ = ClasificacionSet.objects.get_or_create(
                            cliente=upload_log.cliente, nombre=set_nombre
                        )
                        opcion_obj, _ = ClasificacionOption.objects.get_or_create(
                            set_clas=set_obj, valor=str(opcion_valor).strip()
                        )
                        AccountClassification.objects.update_or_create(
                            cuenta=cuenta_obj,
                            set_clas=set_obj,
                            defaults={"opcion": opcion_obj, "asignado_por": None},
                        )

                centro_obj = None
                if CC and row[CC]:
                    centro_obj, _ = CentroCosto.objects.get_or_create(
                        cliente=upload_log.cliente,
                        nombre=str(row[CC]).strip(),
                    )

                aux_obj = None
                if AX and row[AX]:
                    aux_obj, _ = Auxiliar.objects.get_or_create(
                        rut_auxiliar=str(row[AX]).strip(),
                        defaults={"nombre": "", "fecha_creacion": timezone.now()},
                    )

                td_obj = None
                mov_incompleto = False
                if TD is not None:
                    codigo_td = str(row[TD] or "").strip()
                    # Verificar si esta cuenta tiene excepción para tipo de documento
                    tiene_excepcion_tipodoc = f"{current_code}_tipos_doc_no_reconocidos" in excepciones_por_cuenta or \
                                            f"{current_code}_movimientos_tipodoc_nulo" in excepciones_por_cuenta
                    
                    if codigo_td:
                        td_obj = tipos_doc_map.get(codigo_td)
                        if td_obj is None and not tiene_excepcion_tipodoc:
                            Incidencia.objects.create(
                                cierre=upload_log.cierre,
                                tipo="negocio",
                                descripcion=(
                                    f"Movimiento {row_idx-10}, cuenta {current_code}: "
                                    f"Tipo de documento '{codigo_td}' no encontrado"
                                ),
                            )
                            incidencias_creadas += 1
                            mov_incompleto = True
                    else:
                        if not tiene_excepcion_tipodoc:
                            Incidencia.objects.create(
                                cierre=upload_log.cierre,
                                tipo="negocio",
                                descripcion=(
                                    f"Movimiento {row_idx-10}, cuenta {current_code}: "
                                    "Tipo de documento vacío"
                                ),
                            )
                            incidencias_creadas += 1
                            mov_incompleto = True
                
                mov = MovimientoContable.objects.create(
                    cierre=upload_log.cierre,
                    cuenta=cuenta_obj,
                    fecha=fecha,
                    tipo_documento=td_obj,
                    tipo_doc_codigo=codigo_td,
                    numero_documento=str(row[ND] or ""),
                    tipo=str(row[T] or ""),
                    numero_comprobante=str(row[NC] or ""),
                    numero_interno=str(row[NI] or ""),
                    centro_costo=centro_obj,
                    auxiliar=aux_obj,
                    detalle_gasto=str(row[DG] or ""),
                    debe=row[D] or 0,
                    haber=row[H] or 0,
                    descripcion=str(row[DS] or ""),
                    flag_incompleto=mov_incompleto,
                )

                # Verificar excepciones para nombres en inglés y clasificaciones
                tiene_excepcion_nombre = f"{current_code}_cuentas_sin_nombre_ingles" in excepciones_por_cuenta
                tiene_excepcion_clasificacion = f"{current_code}_cuentas_sin_clasificacion" in excepciones_por_cuenta

                if not cuenta_obj.nombre_en and not tiene_excepcion_nombre:
                    Incidencia.objects.create(
                        cierre=upload_log.cierre,
                        tipo="negocio",
                        descripcion=(
                            f"Movimiento {row_idx-10}, cuenta {current_code}: "
                            "No tiene nombre en inglés"
                        ),
                    )
                    incidencias_creadas += 1
                    mov.flag_incompleto = True

                if not AccountClassification.objects.filter(cuenta=cuenta_obj).exists() and not tiene_excepcion_clasificacion:
                    Incidencia.objects.create(
                        cierre=upload_log.cierre,
                        tipo="negocio",
                        descripcion=(
                            f"Movimiento {row_idx-10}, cuenta {current_code}: "
                            "Cuenta sin clasificación asignada"
                        ),
                    )
                    incidencias_creadas += 1
                    mov.flag_incompleto = True

                if mov.flag_incompleto:
                    mov.save(update_fields=["flag_incompleto"])
                movimientos_creados += 1

        fecha_inicio = min(fechas_mov) if fechas_mov else None
        fecha_fin = max(fechas_mov) if fechas_mov else None

        archivo_obj.estado = "completado" if not errores else "error"
        archivo_obj.procesado = True
        archivo_obj.errores = "\n".join(errores) if errores else ""
        archivo_obj.save()

        upload_log.estado = "completado" if not errores else "error"
        upload_log.errores = "\n".join(errores) if errores else ""
        upload_log.resumen = {
            "movimientos_creados": movimientos_creados,
            "incidencias_creadas": incidencias_creadas,
            "archivo_hash": archivo_hash,
        }
        upload_log.tiempo_procesamiento = timezone.now() - inicio
        upload_log.save()

        # ✨ NUEVO: Generar incidencias consolidadas por sub-tipo
        try:
            logger.info(f"Generando incidencias consolidadas para upload_log {upload_log.id}")
            
            # Definir mensajes y acciones predefinidas por tipo de incidencia
            _MENSAJES = {
                "movimientos_tipodoc_nulo": "Se encontraron movimientos sin tipo de documento asignado (campo vacío)",
                "tipos_doc_no_reconocidos": "Se encontraron códigos de tipo de documento que no están registrados en el sistema",
                "cuentas_sin_clasificacion": "Se detectaron cuentas contables sin clasificación asignada",
                "cuentas_sin_nombre_ingles": "Se encontraron cuentas sin nombre en inglés configurado",
                "cuentas_nuevas_detectadas": "Se detectaron nuevas cuentas contables en el libro mayor",
            }
            
            _ACCIONES = {
                "movimientos_tipodoc_nulo": "Revisar el archivo fuente y completar los tipos de documento faltantes",
                "tipos_doc_no_reconocidos": "Cargar archivo de tipos de documento actualizado con los códigos faltantes",
                "cuentas_sin_clasificacion": "Cargar archivo de clasificaciones o asignar clasificaciones manualmente", 
                "cuentas_sin_nombre_ingles": "Cargar archivo de nombres en inglés actualizado",
                "cuentas_nuevas_detectadas": "Revisar nuevas cuentas y configurar nombres en inglés y clasificaciones",
            }
            
            # Analizar incidencias del upload_log actual y generar conteos por tipo
            from django.db.models import Q
            import re
            
            raw_counts = {}
            elementos_detallados = {}
            
            # 1. Contar movimientos con tipo de documento vacío y extraer elementos afectados
            incidencias_tipodoc_vacio = Incidencia.objects.filter(
                cierre=upload_log.cierre,
                descripcion__icontains="Tipo de documento vacío"
            )
            if incidencias_tipodoc_vacio.exists():
                count = incidencias_tipodoc_vacio.count()
                raw_counts["movimientos_tipodoc_nulo"] = count
                # Extraer cuentas afectadas de las descripciones
                # Formato: "Movimiento X, cuenta CODIGO: Tipo de documento vacío"
                cuentas_afectadas = []
                for inc in incidencias_tipodoc_vacio[:50]:  # Limitar a primeras 50 para rendimiento
                    match = re.search(r'cuenta\s+([\d\-]+):', inc.descripcion, flags=re.IGNORECASE)
                    if match:
                        codigo_cuenta = match.group(1)
                        if codigo_cuenta not in cuentas_afectadas:
                            cuentas_afectadas.append(codigo_cuenta)
                    else:
                        # Debug: Ver el formato real de la descripción
                        logger.debug(f"Descripción no coincide con regex: {inc.descripcion}")
                elementos_detallados["movimientos_tipodoc_nulo"] = cuentas_afectadas
            
            # 2. Contar incidencias de tipos de documento no encontrados
            incidencias_tipodoc_no_reconocido = Incidencia.objects.filter(
                cierre=upload_log.cierre,
                descripcion__icontains="no encontrado"
            ).exclude(
                descripcion__icontains="vacío"
            )
            if incidencias_tipodoc_no_reconocido.exists():
                count = incidencias_tipodoc_no_reconocido.count()
                raw_counts["tipos_doc_no_reconocidos"] = count
                # Extraer códigos de tipo de documento no reconocidos
                codigos_no_reconocidos = []
                for inc in incidencias_tipodoc_no_reconocido[:50]:
                    match = re.search(r"'([^']+)' no encontrado", inc.descripcion)
                    if match:
                        codigo_td = match.group(1)
                        if codigo_td not in codigos_no_reconocidos:
                            codigos_no_reconocidos.append(codigo_td)
                    else:
                        # Debug: Ver el formato real de la descripción
                        logger.debug(f"Descripción TD no reconocido no coincide: {inc.descripcion}")
                elementos_detallados["tipos_doc_no_reconocidos"] = codigos_no_reconocidos
            
            # 3. Contar cuentas sin clasificación
            incidencias_sin_clasif = Incidencia.objects.filter(
                cierre=upload_log.cierre,
                descripcion__icontains="sin clasificación"
            )
            if incidencias_sin_clasif.exists():
                count = incidencias_sin_clasif.count()
                raw_counts["cuentas_sin_clasificacion"] = count
                # Extraer cuentas sin clasificación
                cuentas_sin_clasif = []
                for inc in incidencias_sin_clasif[:50]:
                    match = re.search(r'cuenta\s+([\d\-]+):', inc.descripcion, flags=re.IGNORECASE)
                    if match:
                        codigo_cuenta = match.group(1)
                        if codigo_cuenta not in cuentas_sin_clasif:
                            cuentas_sin_clasif.append(codigo_cuenta)
                    else:
                        # Debug: Ver el formato real de la descripción
                        logger.debug(f"Descripción sin clasificación no coincide: {inc.descripcion}")
                elementos_detallados["cuentas_sin_clasificacion"] = cuentas_sin_clasif
            
            # 4. Contar cuentas sin nombre en inglés
            incidencias_sin_nombre = Incidencia.objects.filter(
                cierre=upload_log.cierre,
                descripcion__icontains="nombre en inglés"
            )
            if incidencias_sin_nombre.exists():
                count = incidencias_sin_nombre.count()
                raw_counts["cuentas_sin_nombre_ingles"] = count
                # Extraer cuentas sin nombre en inglés
                cuentas_sin_nombre = []
                for inc in incidencias_sin_nombre[:50]:
                    match = re.search(r'cuenta\s+([\d\-]+):', inc.descripcion, flags=re.IGNORECASE)
                    if match:
                        codigo_cuenta = match.group(1)
                        if codigo_cuenta not in cuentas_sin_nombre:
                            cuentas_sin_nombre.append(codigo_cuenta)
                    else:
                        # Debug: Ver el formato real de la descripción
                        logger.debug(f"Descripción sin nombre inglés no coincide: {inc.descripcion}")
                elementos_detallados["cuentas_sin_nombre_ingles"] = cuentas_sin_nombre
            
            # 5. Contar cuentas nuevas detectadas (usando las cuentas creadas durante el procesamiento)
            cuentas_nuevas = upload_log.resumen.get("cuentas_nuevas", 0)
            if cuentas_nuevas > 0:
                raw_counts["cuentas_nuevas_detectadas"] = cuentas_nuevas
                # Para cuentas nuevas, como no tenemos timestamp de creación,
                # obtenemos una muestra de las últimas cuentas del cliente
                cuentas_recientes = list(
                    CuentaContable.objects.filter(
                        cliente=upload_log.cliente
                    ).order_by('-id')[:50].values_list('codigo', flat=True)
                )
                elementos_detallados["cuentas_nuevas_detectadas"] = cuentas_recientes
            
            # Crear registros individuales de IncidenciaResumen por cada sub-tipo
            incidencias_consolidadas_creadas = 0
            for sub_tipo, count in raw_counts.items():
                # Determinar severidad basada en cantidad
                if count > 100:
                    severidad = "critica"
                elif count > 50:
                    severidad = "alta"
                elif count > 10:
                    severidad = "media"
                else:
                    severidad = "baja"
                
                # Obtener elementos afectados para este sub-tipo
                elementos_afectados = elementos_detallados.get(sub_tipo, [])
                
                # Debug: log de elementos extraídos
                logger.info(f"Sub-tipo {sub_tipo}: {count} incidencias, {len(elementos_afectados)} elementos extraídos: {elementos_afectados[:5]}")
                
                # Crear registro consolidado para este sub-tipo
                IncidenciaResumen.objects.create(
                    upload_log=upload_log,
                    tipo_incidencia=sub_tipo,
                    cantidad_afectada=count,
                    estado="activa",
                    severidad=severidad,
                    mensaje_usuario=_MENSAJES.get(sub_tipo, f"Incidencia detectada: {sub_tipo}"),
                    accion_sugerida=_ACCIONES.get(sub_tipo, "Revisar y corregir manualmente"),
                    codigo_problema=f"{sub_tipo.upper()}_{upload_log.id}",
                    elementos_afectados=elementos_afectados,
                    detalle_muestra={
                        "upload_log_id": upload_log.id,
                        "cantidad": count,
                        "tipo": sub_tipo,
                        "elementos_muestra": elementos_afectados[:10],  # Solo primeros 10 para la muestra
                        "total_elementos": len(elementos_afectados)
                    }
                )
                incidencias_consolidadas_creadas += 1
            
            # Actualizar resumen con información consolidada
            upload_log.resumen.update({
                "incidencias_consolidadas_creadas": incidencias_consolidadas_creadas,
                "tipos_incidencia_detectados": list(raw_counts.keys()),
                "conteos_por_tipo": raw_counts,
                "sistema_consolidado": True
            })
            upload_log.save(update_fields=['resumen'])
            
            logger.info(f"Creadas {incidencias_consolidadas_creadas} incidencias consolidadas por sub-tipo")
            
        except Exception as e:
            logger.error(f"Error generando incidencias consolidadas: {e}")
            # No fallar el procesamiento por esto
            pass

        try:
            os.remove(ruta_completa)
        except OSError:
            pass

        registrar_actividad_tarjeta(
            cliente_id=upload_log.cliente.id,
            periodo=upload_log.cierre.periodo if upload_log.cierre else date.today().strftime("%Y-%m"),
            tarjeta="libro_mayor",
            accion="process_excel",
            descripcion=f"Procesado archivo de libro mayor: {movimientos_creados} movimientos",
            usuario=None,
            detalles={
                "upload_log_id": upload_log.id,
                "movimientos_creados": movimientos_creados,
                "incidencias_creadas": incidencias_creadas,
            },
            resultado="exito" if not errores else "warning",
            ip_address=None,
        )

        # FK-only: Ya no hay clasificaciones temporales por código, se omite mapeo
        try:
            logger.info("FK-only activo: se omite mapeo de clasificaciones temporales a FK")
        except Exception:
            pass

        return f"Completado: {movimientos_creados} movimientos"

    except Exception as e:
        if archivo_obj:
            archivo_obj.estado = "error"
            archivo_obj.errores = str(e)
            archivo_obj.save()

        upload_log.estado = "error"
        upload_log.errores = str(e)
        upload_log.tiempo_procesamiento = timezone.now() - inicio
        upload_log.save()

        registrar_actividad_tarjeta(
            cliente_id=upload_log.cliente.id,
            periodo=upload_log.cierre.periodo if upload_log.cierre else date.today().strftime("%Y-%m"),
            tarjeta="libro_mayor",
            accion="process_excel",
            descripcion=f"Error al procesar libro mayor: {str(e)}",
            usuario=None,
            detalles={"upload_log_id": upload_log.id},
            resultado="error",
            ip_address=None,
        )

        return f"Error: {str(e)}"




#=======CLEANING TASKS=======
from celery import shared_task
@shared_task
def limpiar_archivos_temporales_antiguos_task():
    """
    Tarea Celery para limpiar archivos temporales antiguos (>24h)
    """
    from contabilidad.views import limpiar_archivos_temporales_antiguos

    archivos_eliminados = limpiar_archivos_temporales_antiguos()
    logger.info(
        f"🧹 Limpieza automática: {archivos_eliminados} archivos temporales eliminados"
    )
    return f"Eliminados {archivos_eliminados} archivos temporales"

# ===== TAREAS DE REPORTES FINANCIEROS =====

@shared_task
def generar_estado_situacion_financiera(cierre_id, usuario_id=None):
    """
    Genera el Estado de Situación Financiera para un cierre específico.
    
    Args:
        cierre_id (int): ID del cierre para el cual generar el reporte
        usuario_id (int): ID del usuario que solicita el reporte (opcional)
    
    Returns:
        dict: Resultado de la generación del reporte
    """
    from decimal import Decimal
    from django.contrib.auth.models import User
    import json
    
    logger.info(f"🏛️ Iniciando generación de Estado de Situación Financiera para cierre {cierre_id}")
    
    inicio = timezone.now()
    usuario = None
    
    try:
        # Validar que el cierre existe
        try:
            cierre = CierreContabilidad.objects.get(id=cierre_id)
        except CierreContabilidad.DoesNotExist:
            error_msg = f"Cierre con ID {cierre_id} no encontrado"
            logger.error(error_msg)
            return {"error": error_msg}
        
        # Obtener usuario si se proporciona
        if usuario_id:
            try:
                usuario = User.objects.get(id=usuario_id)
            except User.DoesNotExist:
                logger.warning(f"Usuario con ID {usuario_id} no encontrado, continuando sin usuario")
        
        # Verificar si ya existe un reporte para este cierre
        reporte_existente = ReporteFinanciero.objects.filter(
            cierre=cierre,
            tipo_reporte='esf'
        ).first()
        
        if reporte_existente:
            logger.info(f"Ya existe un reporte ESF para el cierre {cierre_id}, actualizando...")
            reporte = reporte_existente
            # IMPORTANTE: Actualizar el usuario generador cuando se regenera
            if usuario:
                reporte.usuario_generador = usuario
        else:
            # Crear nuevo reporte
            reporte = ReporteFinanciero.objects.create(
                cierre=cierre,
                tipo_reporte='esf',
                usuario_generador=usuario,
                estado='generando'
            )
        
        # Actualizar estado a generando
        reporte.estado = 'generando'
        reporte.fecha_generacion = inicio
        reporte.save(update_fields=['estado', 'fecha_generacion', 'usuario_generador'])
        
        # Obtener todas las cuentas con movimientos en el cierre
        movimientos = MovimientoContable.objects.filter(cierre=cierre).select_related('cuenta')
        
        # Obtener saldos de apertura
        saldos_apertura = AperturaCuenta.objects.filter(cierre=cierre).select_related('cuenta')
        
        # Construir diccionario de cuentas únicas
        cuentas_data = {}
        
        # Procesar saldos de apertura
        for apertura in saldos_apertura:
            cuenta_codigo = apertura.cuenta.codigo
            if cuenta_codigo not in cuentas_data:
                cuentas_data[cuenta_codigo] = {
                    'codigo': apertura.cuenta.codigo,
                    'nombre': apertura.cuenta.nombre,
                    'saldo_anterior': float(apertura.saldo_anterior or 0),
                    'debe': 0.0,
                    'haber': 0.0,
                    'saldo_final': float(apertura.saldo_anterior or 0)
                }
            else:
                cuentas_data[cuenta_codigo]['saldo_anterior'] = float(apertura.saldo_anterior or 0)
                cuentas_data[cuenta_codigo]['saldo_final'] = float(apertura.saldo_anterior or 0)
        
        # Procesar movimientos
        for movimiento in movimientos:
            cuenta_codigo = movimiento.cuenta.codigo
            if cuenta_codigo not in cuentas_data:
                cuentas_data[cuenta_codigo] = {
                    'codigo': movimiento.cuenta.codigo,
                    'nombre': movimiento.cuenta.nombre,
                    'saldo_anterior': 0.0,
                    'debe': 0.0,
                    'haber': 0.0,
                    'saldo_final': 0.0
                }
            
            # Acumular debe y haber
            cuentas_data[cuenta_codigo]['debe'] += float(movimiento.debe or 0)
            cuentas_data[cuenta_codigo]['haber'] += float(movimiento.haber or 0)
        
        # Calcular saldos finales
        for cuenta_codigo, data in cuentas_data.items():
            data['saldo_final'] = data['saldo_anterior'] + data['debe'] - data['haber']
        
        # Obtener clasificaciones para estructurar el reporte
        # Buscar el set de clasificación "Estado de Situación Financiera" o similar
        try:
            set_esf = ClasificacionSet.objects.filter(
                nombre__icontains='situación financiera'
            ).first()
            
            if not set_esf:
                # Si no existe, buscar el primer set disponible
                set_esf = ClasificacionSet.objects.first()
                
            if not set_esf:
                raise ValueError("No se encontró ningún set de clasificación disponible")
                
        except Exception as e:
            logger.error(f"Error al obtener set de clasificación: {e}")
            set_esf = None
        
        # Estructura del Estado de Situación Financiera
        estructura_esf = {
            'activos': {
                'corrientes': {},
                'no_corrientes': {},
                'total': 0.0
            },
            'pasivos': {
                'corrientes': {},
                'no_corrientes': {},
                'total': 0.0
            },
            'patrimonio': {
                'capital': {},
                'resultados': {},
                'total': 0.0
            },
            'total_pasivos_patrimonio': 0.0
        }
        
        # Clasificar cuentas según su código (lógica básica)
        for cuenta_codigo, data in cuentas_data.items():
            saldo = data['saldo_final']
            
            # Lógica básica de clasificación por código de cuenta
            if cuenta_codigo.startswith('1'):  # Activos
                if cuenta_codigo.startswith('11'):  # Activos corrientes
                    estructura_esf['activos']['corrientes'][cuenta_codigo] = {
                        'nombre': data['nombre'],
                        'saldo': saldo
                    }
                else:  # Activos no corrientes
                    estructura_esf['activos']['no_corrientes'][cuenta_codigo] = {
                        'nombre': data['nombre'],
                        'saldo': saldo
                    }
                estructura_esf['activos']['total'] += saldo
                
            elif cuenta_codigo.startswith('2'):  # Pasivos
                if cuenta_codigo.startswith('21'):  # Pasivos corrientes
                    estructura_esf['pasivos']['corrientes'][cuenta_codigo] = {
                        'nombre': data['nombre'],
                        'saldo': saldo
                    }
                else:  # Pasivos no corrientes
                    estructura_esf['pasivos']['no_corrientes'][cuenta_codigo] = {
                        'nombre': data['nombre'],
                        'saldo': saldo
                    }
                estructura_esf['pasivos']['total'] += saldo
                
            elif cuenta_codigo.startswith('3'):  # Patrimonio
                if 'capital' in data['nombre'].lower():
                    estructura_esf['patrimonio']['capital'][cuenta_codigo] = {
                        'nombre': data['nombre'],
                        'saldo': saldo
                    }
                else:
                    estructura_esf['patrimonio']['resultados'][cuenta_codigo] = {
                        'nombre': data['nombre'],
                        'saldo': saldo
                    }
                estructura_esf['patrimonio']['total'] += saldo
        
        # Calcular total pasivos + patrimonio
        estructura_esf['total_pasivos_patrimonio'] = (
            estructura_esf['pasivos']['total'] + 
            estructura_esf['patrimonio']['total']
        )
        
        # Metadatos del reporte
        metadatos = {
            'cierre_id': cierre.id,
            'cliente': cierre.cliente.nombre,
            'periodo': cierre.periodo,
            'fecha_generacion': inicio.isoformat(),
            'usuario': usuario.correo_bdo if usuario else 'Sistema',
            'total_cuentas': len(cuentas_data),
            'set_clasificacion_usado': set_esf.nombre if set_esf else 'N/A'
        }
        
        # Preparar datos completos del reporte
        datos_reporte = {
            'metadatos': metadatos,
            'estructura': estructura_esf,
            'detalle_cuentas': cuentas_data
        }
        
        # Guardar el reporte
        reporte.datos_reporte = datos_reporte
        reporte.estado = 'completado'
        reporte.save(update_fields=['datos_reporte', 'estado'])
        
        logger.info(f"✅ Estado de Situación Financiera generado exitosamente para cierre {cierre_id}")
        
        return {
            'success': True,
            'reporte_id': reporte.id,
            'total_cuentas': len(cuentas_data),
            'total_activos': estructura_esf['activos']['total'],
            'total_pasivos_patrimonio': estructura_esf['total_pasivos_patrimonio']
        }
        
    except Exception as e:
        logger.error(f"❌ Error generando Estado de Situación Financiera: {str(e)}")
        
        # Actualizar estado del reporte si existe
        if 'reporte' in locals():
            reporte.estado = 'error'
            reporte.error_mensaje = str(e)
            reporte.save(update_fields=['estado', 'error_mensaje'])
        
        return {
            'error': str(e),
            'cierre_id': cierre_id
        }

@shared_task
def enviar_reporte_a_cache(reporte_id):
    """
    Envía un reporte financiero al caché Redis
    
    Args:
        reporte_id: ID del reporte financiero a enviar al caché
        
    Returns:
        dict: Resultado de la operación
    """
    try:
        from contabilidad.cache_redis import get_cache_system
        
        logger.info(f"Iniciando envío al caché del reporte ID: {reporte_id}")
        
        # Obtener el reporte
        try:
            reporte = ReporteFinanciero.objects.select_related(
                'cierre', 'cierre__cliente'
            ).get(id=reporte_id)
        except ReporteFinanciero.DoesNotExist:
            error_msg = f"Reporte con ID {reporte_id} no encontrado"
            logger.error(error_msg)
            return {'error': error_msg, 'reporte_id': reporte_id}
        
        # Validar que el reporte esté completado
        if reporte.estado != 'completado':
            error_msg = f"Reporte {reporte_id} no está completado. Estado actual: {reporte.estado}"
            logger.error(error_msg)
            return {'error': error_msg, 'reporte_id': reporte_id}
        
        # Validar que tenga datos
        if not reporte.datos_reporte:
            error_msg = f"Reporte {reporte_id} no tiene datos para enviar al caché"
            logger.error(error_msg)
            return {'error': error_msg, 'reporte_id': reporte_id}
        
        # Obtener sistema de caché
        cache_system = get_cache_system()
        
        # Mapear tipo de reporte a clave de caché
        tipo_cache_map = {
            'esf': 'esf',  # Estado de Situación Financiera
            'eri': 'eri',  # Estado de Resultado Integral
            'ecp': 'ecp'   # Estado de Cambios en el Patrimonio
        }
        
        tipo_cache = tipo_cache_map.get(reporte.tipo_reporte)
        if not tipo_cache:
            error_msg = f"Tipo de reporte no soportado para caché: {reporte.tipo_reporte}"
            logger.error(error_msg)
            return {'error': error_msg, 'reporte_id': reporte_id}
        
        # Preparar datos para el caché
        # Obtener información del usuario de forma segura
        usuario_info = None
        if reporte.usuario_generador:
            try:
                # El modelo Usuario usa correo_bdo como USERNAME_FIELD
                if hasattr(reporte.usuario_generador, 'correo_bdo'):
                    usuario_info = reporte.usuario_generador.correo_bdo
                elif hasattr(reporte.usuario_generador, 'nombre'):
                    usuario_info = f"{reporte.usuario_generador.nombre} {getattr(reporte.usuario_generador, 'apellido', '')}"
                else:
                    usuario_info = str(reporte.usuario_generador)
            except Exception as e:
                logger.warning(f"Error obteniendo información del usuario: {e}")
                usuario_info = "Usuario desconocido"
        
        datos_cache = {
            **reporte.datos_reporte,
            '_reporte_metadata': {
                'reporte_id': reporte.id,
                'fecha_generacion': reporte.fecha_generacion.isoformat(),
                'usuario_generador': usuario_info,
                'tipo_reporte': reporte.tipo_reporte,
                'enviado_cache_at': timezone.now().isoformat()
            }
        }
        
        # Enviar al caché con TTL extendido (4 horas para reportes)
        resultado = cache_system.set_estado_financiero(
            cliente_id=reporte.cierre.cliente.id,
            periodo=reporte.cierre.periodo,
            tipo_estado=tipo_cache,
            datos=datos_cache,
            ttl=14400  # 4 horas
        )
        
        if resultado:
            logger.info(
                f"Reporte {reporte_id} ({reporte.get_tipo_reporte_display()}) "
                f"enviado exitosamente al caché. Cliente: {reporte.cierre.cliente.nombre}, "
                f"Período: {reporte.cierre.periodo}"
            )
            
            return {
                'success': True,
                'reporte_id': reporte_id,
                'tipo_reporte': reporte.tipo_reporte,
                'cliente': reporte.cierre.cliente.nombre,
                'periodo': reporte.cierre.periodo,
                'mensaje': 'Reporte enviado al caché exitosamente'
            }
        else:
            error_msg = f"Error enviando reporte {reporte_id} al caché"
            logger.error(error_msg)
            return {'error': error_msg, 'reporte_id': reporte_id}
            
    except Exception as e:
        error_msg = f"Error inesperado enviando reporte {reporte_id} al caché: {str(e)}"
        logger.error(error_msg)
        return {'error': error_msg, 'reporte_id': reporte_id}


@shared_task
def enviar_multiples_reportes_a_cache(reporte_ids):
    """
    Envía múltiples reportes financieros al caché Redis
    
    Args:
        reporte_ids: Lista de IDs de reportes financieros
        
    Returns:
        dict: Resultado consolidado de la operación
    """
    resultados = {
        'exitosos': [],
        'errores': [],
        'total_procesados': len(reporte_ids),
        'total_exitosos': 0,
        'total_errores': 0
    }
    
    logger.info(f"Iniciando envío masivo al caché de {len(reporte_ids)} reportes")
    
    for reporte_id in reporte_ids:
        try:
            resultado = enviar_reporte_a_cache(reporte_id)
            
            if resultado.get('success'):
                resultados['exitosos'].append(resultado)
                resultados['total_exitosos'] += 1
            else:
                resultados['errores'].append(resultado)
                resultados['total_errores'] += 1
                
        except Exception as e:
            error_resultado = {
                'error': f"Error procesando reporte {reporte_id}: {str(e)}",
                'reporte_id': reporte_id
            }
            resultados['errores'].append(error_resultado)
            resultados['total_errores'] += 1
    
    logger.info(
        f"Envío masivo completado. Exitosos: {resultados['total_exitosos']}, "
        f"Errores: {resultados['total_errores']}"
    )
    
    return resultados


# =====================================================
# FUNCIONES AUXILIARES PARA CAPTURA MASIVA DE GASTOS
# =====================================================

def get_redis_client_db1():
    """
    Obtiene cliente Redis para db1 usando la configuración de Django
    Con soporte UTF-8 completo
    """
    redis_password = os.environ.get('REDIS_PASSWORD', '')
    if redis_password:
        return redis.Redis(
            host='redis', 
            port=6379, 
            db=1, 
            password=redis_password, 
            decode_responses=True,
            encoding='utf-8',
            encoding_errors='strict'
        )
    else:
        return redis.Redis(
            host='redis', 
            port=6379, 
            db=1, 
            decode_responses=True,
            encoding='utf-8',
            encoding_errors='strict'
        )

def get_redis_client_db1_binary():
    """
    Obtiene cliente Redis para db1 para datos binarios (sin decode_responses)
    Con soporte UTF-8 para metadatos
    """
    redis_password = os.environ.get('REDIS_PASSWORD', '')
    if redis_password:
        return redis.Redis(
            host='redis', 
            port=6379, 
            db=1, 
            password=redis_password, 
            decode_responses=False,
            encoding='utf-8'
        )
    else:
        return redis.Redis(
            host='redis', 
            port=6379, 
            db=1, 
            decode_responses=False,
            encoding='utf-8'
        )

def serialize_excel_data(data):
    """
    Serializa datos de Excel para que sean compatibles con JSON/Redis
    """
    if isinstance(data, list):
        return [serialize_excel_data(item) for item in data]
    elif isinstance(data, dict):
        return {key: serialize_excel_data(value) for key, value in data.items()}
    elif isinstance(data, (datetime.datetime, datetime.date)):
        return data.isoformat()
    elif isinstance(data, (int, float, str, bool)) or data is None:
        # Asegurar que las strings mantengan encoding UTF-8
        return data
    else:
        return data

def contar_centros_costos(fila, headers, mapeo_cc=None):
    """
    Cuenta la cantidad de centros de costos en una fila de gastos
    """
    if mapeo_cc is None:
        mapeo_cc = {}
    
    count = 0
    # Buscar columnas que contengan información de centros de costos
    for header in headers:
        if header and ('pyc' in header.lower() or 'ps/eb' in header.lower() or 'co' in header.lower()):
            valor = fila.get(header)
            if valor and str(valor).strip():
                count += 1
    
    return count

# =============================================================================
# 🧾 CAPTURA MASIVA RINDE GASTOS - CELERY CHORD
# =============================================================================

import redis
import json
import tempfile
import os
from datetime import datetime, date
from celery import chord
from openpyxl import Workbook, load_workbook
from io import BytesIO
from django.conf import settings

def get_redis_client_db1():
    """
    Obtiene cliente Redis para db1 usando la configuración de Django
    Con soporte UTF-8 completo
    """
    redis_password = os.environ.get('REDIS_PASSWORD', '')
    if redis_password:
        return redis.Redis(
            host='redis', 
            port=6379, 
            db=1, 
            password=redis_password, 
            decode_responses=True,
            encoding='utf-8',
            encoding_errors='strict'
        )
    else:
        return redis.Redis(
            host='redis', 
            port=6379, 
            db=1, 
            decode_responses=True,
            encoding='utf-8',
            encoding_errors='strict'
        )

def get_redis_client_db1_binary():
    """
    Obtiene cliente Redis para db1 para datos binarios (sin decode_responses)
    Con soporte UTF-8 para metadatos
    """
    redis_password = os.environ.get('REDIS_PASSWORD', '')
    if redis_password:
        return redis.Redis(
            host='redis', 
            port=6379, 
            db=1, 
            password=redis_password, 
            decode_responses=False,
            encoding='utf-8'
        )
    else:
        return redis.Redis(
            host='redis', 
            port=6379, 
            db=1, 
            decode_responses=False,
            encoding='utf-8'
        )

def get_headers_salida_contabilidad():
    """
    Retorna los headers que debe tener cada hoja del Excel de salida
    según las especificaciones del sistema contable
    """
    return [
        "Numero",  # Tipo de documento original (33, 34, 61, etc.)
        "Código Plan de Cuenta",
        "Monto al Debe Moneda Base",
        "Monto al Haber Moneda Base",
        "Descripción Movimiento",
        "Equivalencia Moneda",
        "Monto al Debe Moneda Adicional",
        "Monto al Haber Moneda Adicional",
        "Código Condición de Venta",
        "Código Vendedor",
        "Código Ubicación",
        "Código Concepto de Caja",
        "Código Instrumento Financiero",
        "Cantidad Instrumento Financiero",
        "Código Detalle de Gasto",
        "Cantidad Concepto de Gasto",
        "Código Centro de Costo",
        "Tipo Docto. Conciliación",
        "Nro. Docto. Conciliación",
        "Codigo Auxiliar",
        "Tipo Documento",
        "Numero Doc",
        "Fecha Emisión Docto.(DD/MM/AAAA)",
        "Fecha Vencimiento Docto.(DD/MM/AAAA)",
        "Tipo Docto. Referencia",
        "Nro. Docto. Referencia",
        "Nro. Correlativo Interno",
        "Monto 1 Detalle Libro",
        "Monto 2 Detalle Libro",
        "Monto 3 Detalle Libro",
        "Monto 4 Detalle Libro",
        "Monto 5 Detalle Libro",
        "Monto 6 Detalle Libro",
        "Monto 7 Detalle Libro",
        "Monto 8 Detalle Libro",
        "Monto 9 Detalle Libro",
        "Monto Suma Detalle Libro",
        "Número Documento Desde",
        "Número Documento Hasta",
        "Nro. agrupación en igual comprobante",
        "Graba el detalle de libro (S/N) (Opcional, por defecto 'S')",
        "Documento Nulo (S/N) (Opcional, por defecto 'N')",
        "Código flujo efectivo 1 (Opcional)",
        "Monto flujo 1 (Opcional)",
        "Código flujo efectivo 2 (Opcional)",
        "Monto flujo 2 (Opcional)",
        "Código flujo efectivo 3 (Opcional)",
        "Monto flujo 3 (Opcional)",
        "Código flujo efectivo 4 (Opcional)",
        "Monto flujo 4 (Opcional)",
        "Código flujo efectivo 5 (Opcional)",
        "Monto flujo 5 (Opcional)",
        "Código flujo efectivo 6 (Opcional)",
        "Monto flujo 6 (Opcional)",
        "Código flujo efectivo 7 (Opcional)",
        "Monto flujo 7 (Opcional)",
        "Código flujo efectivo 8 (Opcional)",
        "Monto flujo 8 (Opcional)",
        "Código flujo efectivo 9 (Opcional)",
        "Monto flujo 9 (Opcional)",
        "Código flujo efectivo 10 (Opcional)",
        "Monto flujo 10 (Opcional)",
        "Numero Cuota de Pago"
    ]

def aplicar_reglas_tipo_documento(fila, headers_entrada, tipo_doc, cc_count, mapeo_cc, tipos_doc_map):
    """
    Aplica las reglas específicas según el tipo de documento y cantidad de centros de costo
    Transforma una fila del formato de entrada al formato de salida contable
    
    Args:
        fila: diccionario con los datos de la fila de entrada
        headers_entrada: headers del archivo Excel de entrada
        tipo_doc: tipo de documento (33, 34, 61, etc.)
        cc_count: cantidad de centros de costo (1, 2, etc.)
        mapeo_cc: mapeo de códigos de centros de costos
        tipos_doc_map: mapeo de códigos de tipo de documento a nombres
    
    Returns:
        lista de diccionarios con las filas transformadas (puede ser más de una fila por registro)
    """
    headers_salida = get_headers_salida_contabilidad()
    
    if tipo_doc == "33":
        if cc_count == 1:
            return aplicar_reglas_tipo_33_1cc(fila, headers_entrada, mapeo_cc, headers_salida, tipos_doc_map)
        elif cc_count >= 2:
            # La implementación de 3CC itera dinámicamente por todas las posiciones detectadas (soporta >3)
            return aplicar_reglas_tipo_33_3cc(fila, headers_entrada, mapeo_cc, headers_salida, tipos_doc_map)
    elif tipo_doc == "34":
        if cc_count == 1:
            return aplicar_reglas_tipo_34(fila, headers_entrada, mapeo_cc, headers_salida, tipos_doc_map)
        elif cc_count >= 2:
            # Usar la versión 3CC que también soporta N CC dinámicamente
            return aplicar_reglas_tipo_34_3cc(fila, headers_entrada, mapeo_cc, headers_salida, tipos_doc_map)
    elif tipo_doc == "61":
        if cc_count == 1:
            return aplicar_reglas_tipo_61(fila, headers_entrada, mapeo_cc, headers_salida, tipos_doc_map)
        elif cc_count >= 2:
            # Usar la versión 3CC que también soporta N CC dinámicamente
            return aplicar_reglas_tipo_61_3cc(fila, headers_entrada, mapeo_cc, headers_salida, tipos_doc_map)
    else:
        # Para otros tipos de documento, implementar más adelante
        return aplicar_reglas_genericas(fila, headers_entrada, tipo_doc, mapeo_cc, headers_salida, tipos_doc_map)

def aplicar_reglas_tipo_33_1cc(fila, headers_entrada, mapeo_cc, headers_salida, tipos_doc_map):
    """
    Aplica reglas específicas para tipo de documento 33 CON 1 CENTRO DE COSTO
    
    CONTROLADOR EXENTO:
    - Si hay monto en "Monto Exento": Calcula IVA desde "Monto Neto" 
    - Si NO hay exento: Calcula IVA desde "Monto Total" (lógica original)
    
    Genera 3 cuentas: Proveedores (2xxx), Gastos (5xxx), IVA (1xxx)
    
    REGLAS ESPECÍFICAS PARA 1CC:
    - 1 cuenta de Proveedores con Monto 1 y 3 Detalle
    - 1 cuenta de Gastos con el monto neto
    - 1 cuenta de IVA con el monto de IVA
    """
    filas_resultado = []
    
    logger.info(f"🔍 PROCESANDO TIPO 33 CON 1CC")
    logger.info(f"   Fila original: {fila}")
    logger.info(f"   Headers entrada: {headers_entrada}")
    logger.info(f"   Mapeo CC: {mapeo_cc}")
    
    # 🔍 CONTROLADOR: Verificar si hay monto exento
    tiene_exento = verificar_tiene_exento_tipo_33(fila, headers_entrada)
    
    # Extraer datos comunes de la fila original
    datos_comunes = extraer_datos_comunes_tipo_33(fila, headers_entrada)
    
    logger.info(f"🔍 Tipo 33 con 1CC - Datos extraídos: {datos_comunes}")
    
    # Calcular montos con validación
    monto_total_raw = datos_comunes.get('monto_total', 0)
    try:
        monto_total = float(monto_total_raw) if monto_total_raw else 0
    except (ValueError, TypeError):
        logger.warning(f"⚠️ Monto inválido encontrado: {monto_total_raw}, usando 0")
        monto_total = 0
    
    if monto_total <= 0:
        logger.warning(f"⚠️ Monto total es 0 o negativo: {monto_total}")
        return []  # No generar filas si no hay monto válido
    
    # Calcular montos según si hay exento o no
    if tiene_exento:
        # CON EXENTO: Calcular IVA desde "Monto Neto"
        # Obtener monto neto de la columna correspondiente
        mapeo = mapear_columnas_automaticamente(headers_entrada)
        if mapeo.get('monto_neto') is not None:
            monto_neto_raw = fila.get(headers_entrada[mapeo['monto_neto']], 0)
            try:
                monto_neto = float(monto_neto_raw) if monto_neto_raw else 0
            except (ValueError, TypeError):
                logger.warning(f"⚠️ Monto Neto inválido: {monto_neto_raw}, usando cálculo tradicional")
                monto_neto = monto_total / 1.19
        else:
            logger.warning("⚠️ No se encontró columna 'Monto Neto', usando cálculo tradicional")
            monto_neto = monto_total / 1.19
        
        # Obtener monto exento de la columna correspondiente
        if mapeo.get('monto_exento') is not None:
            monto_exento_raw = fila.get(headers_entrada[mapeo['monto_exento']], 0)
            try:
                monto_exento = float(monto_exento_raw) if monto_exento_raw else 0
            except (ValueError, TypeError):
                logger.warning(f"⚠️ Monto Exento inválido: {monto_exento_raw}, usando 0")
                monto_exento = 0
        else:
            monto_exento = 0
        
        # Con exento, el IVA se calcula sobre el monto neto: IVA = Monto Neto * 0.19
        monto_iva = monto_neto * 0.19
        # El monto total de gastos incluye neto + exento
        monto_gastos_total = monto_neto + monto_exento
        logger.info(f"📊 CON EXENTO: Monto Neto = {monto_neto:.2f}, Monto Exento = {monto_exento:.2f}, Monto Gastos Total = {monto_gastos_total:.2f}, Monto IVA (19% del neto) = {monto_iva:.2f}")
    else:
        # SIN EXENTO: Lógica original (IVA desde Monto Total)
        monto_neto = monto_total / 1.19  # Monto 1 Detalle Libro
        monto_iva = monto_total - monto_neto  # Monto 3 Detalle Libro
        monto_exento = 0  # No hay exento
        monto_gastos_total = monto_neto  # Solo el monto neto
        logger.info(f"📊 SIN EXENTO: Monto Neto = {monto_neto:.2f}, Monto IVA = {monto_iva:.2f}")
    
    # Calcular códigos de centros de costos aplicables
    codigos_cc = calcular_codigos_cc_para_fila(fila, headers_entrada, mapeo_cc)
    
    # Obtener código de cuenta original de la columna 8
    codigo_cuenta_original = datos_comunes.get('codigo_cuenta', '')
    
    # Obtener nombre del tipo de documento
    nombre_tipo_documento = tipos_doc_map.get("33", "Factura Electrónica")
    logger.info(f"🏷️ Tipo 33 con 1CC -> Nombre: '{nombre_tipo_documento}'")
    
    # Generar códigos de cuenta basados en el código original
    # Si el código empieza con 2, 5 o 1, usar como base, sino generar códigos genéricos
    if codigo_cuenta_original.startswith('2'):
        codigo_proveedores = codigo_cuenta_original
        codigo_gastos = codigo_cuenta_original.replace('2', '5', 1)
        codigo_iva = codigo_cuenta_original.replace('2', '1', 1)
    elif codigo_cuenta_original.startswith('5'):
        codigo_gastos = codigo_cuenta_original
        codigo_proveedores = codigo_cuenta_original.replace('5', '2', 1)
        codigo_iva = codigo_cuenta_original.replace('5', '1', 1)
    elif codigo_cuenta_original.startswith('1'):
        codigo_iva = codigo_cuenta_original
        codigo_proveedores = codigo_cuenta_original.replace('1', '2', 1)
        codigo_gastos = codigo_cuenta_original.replace('1', '5', 1)
    else:
        # Usar el código original como base y generar variaciones
        codigo_proveedores = f"2{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "2111001"
        codigo_gastos = f"5{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "5111001"
        codigo_iva = f"1{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "1191001"
    
    # 1. Cuenta Proveedores (código empieza con 2)
    fila_proveedores = crear_fila_base_tipo_33(datos_comunes, headers_salida, nombre_tipo_documento)
    fila_proveedores["Código Plan de Cuenta"] = codigo_proveedores
    fila_proveedores["Monto al Haber Moneda Base"] = formatear_monto_clp(round(monto_total, 2))
    # ✨ NUEVO: Agregar campos especiales de detalle en cuenta Proveedores
    fila_proveedores["Monto 1 Detalle Libro"] = formatear_monto_clp(round(monto_neto, 2))
    fila_proveedores["Monto 3 Detalle Libro"] = formatear_monto_clp(round(monto_iva, 2))
    # ✨ NUEVO: Calcular y agregar "Monto Suma Detalle Libro" como suma de montos 1-9
    monto_suma_detalle = round(monto_neto + monto_iva, 2)  # Solo tenemos valores en monto 1 y 3
    fila_proveedores["Monto Suma Detalle Libro"] = formatear_monto_clp(monto_suma_detalle)
    filas_resultado.append(fila_proveedores)
    
    # 2. Cuenta Gastos (código empieza con 5)
    fila_gastos = crear_fila_base_tipo_33(datos_comunes, headers_salida, nombre_tipo_documento)
    fila_gastos["Código Plan de Cuenta"] = codigo_gastos
    fila_gastos["Monto al Debe Moneda Base"] = formatear_monto_clp(round(monto_gastos_total, 2))  # Incluye neto + exento
    # ✨ REMOVIDO: Ya no se asigna "Monto 1 Detalle Libro" aquí
    if codigos_cc:
        fila_gastos["Código Centro de Costo"] = codigos_cc
    filas_resultado.append(fila_gastos)
    
    # 3. Cuenta IVA (código empieza con 1)
    fila_iva = crear_fila_base_tipo_33(datos_comunes, headers_salida, nombre_tipo_documento)
    fila_iva["Código Plan de Cuenta"] = codigo_iva
    fila_iva["Monto al Debe Moneda Base"] = formatear_monto_clp(round(monto_iva, 2))
    # ✨ REMOVIDO: Ya no se asigna "Monto 3 Detalle Libro" aquí
    filas_resultado.append(fila_iva)
    
    return filas_resultado

def aplicar_reglas_tipo_33_2cc(fila, headers_entrada, mapeo_cc, headers_salida, tipos_doc_map):
    """
    Aplica reglas específicas para tipo de documento 33 CON 2 CENTROS DE COSTO
    
    CONTROLADOR EXENTO:
    - Si hay monto en "Monto Exento": Calcula IVA desde "Monto Neto"
    - Si NO hay exento: Calcula IVA desde "Monto Total" (lógica original)
    
    REGLAS ESPECÍFICAS PARA 2CC:
    - 1 cuenta de Proveedores con Monto 1 y 3 Detalle (igual que 1CC)
    - 2 cuentas de Gastos separadas, cada una con su monto ponderado
    - 1 cuenta de IVA (igual que 1CC)
    
    El monto de cada cuenta de gastos = monto_neto * ponderador_del_CC
    Los ponderadores están en las columnas PyC, EB/PS, CO
    """
    filas_resultado = []
    
    # 🔍 CONTROLADOR: Verificar si hay monto exento
    tiene_exento = verificar_tiene_exento_tipo_33(fila, headers_entrada)
    
    # Extraer datos comunes de la fila original
    datos_comunes = extraer_datos_comunes_tipo_33(fila, headers_entrada)
    
    # Calcular montos con validación
    monto_total_raw = datos_comunes.get('monto_total', 0)
    try:
        monto_total = float(monto_total_raw) if monto_total_raw else 0
    except (ValueError, TypeError):
        logger.warning(f"⚠️ Monto inválido encontrado: {monto_total_raw}, usando 0")
        monto_total = 0
    
    if monto_total <= 0:
        logger.warning(f"⚠️ Monto total es 0 o negativo: {monto_total}")
        return []  # No generar filas si no hay monto válido
    
    # Calcular montos según si hay exento o no
    if tiene_exento:
        # CON EXENTO: Calcular IVA desde "Monto Neto"
        # Obtener monto neto de la columna correspondiente
        mapeo = mapear_columnas_automaticamente(headers_entrada)
        if mapeo.get('monto_neto') is not None:
            monto_neto_raw = fila.get(headers_entrada[mapeo['monto_neto']], 0)
            try:
                monto_neto = float(monto_neto_raw) if monto_neto_raw else 0
            except (ValueError, TypeError):
                logger.warning(f"⚠️ Monto Neto inválido: {monto_neto_raw}, usando cálculo tradicional")
                monto_neto = monto_total / 1.19
        else:
            logger.warning("⚠️ No se encontró columna 'Monto Neto', usando cálculo tradicional")
            monto_neto = monto_total / 1.19
        
        # Obtener monto exento de la columna correspondiente
        if mapeo.get('monto_exento') is not None:
            monto_exento_raw = fila.get(headers_entrada[mapeo['monto_exento']], 0)
            try:
                monto_exento = float(monto_exento_raw) if monto_exento_raw else 0
            except (ValueError, TypeError):
                logger.warning(f"⚠️ Monto Exento inválido: {monto_exento_raw}, usando 0")
                monto_exento = 0
        else:
            monto_exento = 0
        
        # Con exento, el IVA se calcula sobre el monto neto: IVA = Monto Neto * 0.19
        monto_iva = monto_neto * 0.19
        # El monto total de gastos incluye neto + exento (para distribución por ponderadores)
        monto_gastos_total = monto_neto + monto_exento
        logger.info(f"📊 CON EXENTO (2CC): Monto Neto = {monto_neto:.2f}, Monto Exento = {monto_exento:.2f}, Monto Gastos Total = {monto_gastos_total:.2f}, Monto IVA (19% del neto) = {monto_iva:.2f}")
    else:
        # SIN EXENTO: Lógica original (IVA desde Monto Total)
        monto_neto = monto_total / 1.19  # Monto 1 Detalle Libro
        monto_iva = monto_total - monto_neto  # Monto 3 Detalle Libro
        monto_exento = 0  # No hay exento
        monto_gastos_total = monto_neto  # Solo el monto neto
        logger.info(f"📊 SIN EXENTO (2CC): Monto Neto = {monto_neto:.2f}, Monto IVA = {monto_iva:.2f}")
    
    # Obtener código de cuenta original
    codigo_cuenta_original = datos_comunes.get('codigo_cuenta', '')
    
    # Obtener nombre del tipo de documento
    nombre_tipo_documento = tipos_doc_map.get("33", "Factura Electrónica")
    logger.info(f"🏷️ Tipo 33 con 2CC -> Nombre: '{nombre_tipo_documento}'")
    
    # Detectar posiciones de centros de costo dinámicamente
    posiciones_cc = detectar_posiciones_centros_costo(headers_entrada)
    logger.info(f"🔍 Posiciones CC detectadas para 2CC: {posiciones_cc}")
    
    # Extraer ponderadores de cada centro de costo
    ponderadores_cc = {}
    total_ponderadores = 0
    
    for nombre_cc, pos in posiciones_cc.items():
        if pos is not None and pos < len(headers_entrada):
            valor_raw = fila.get(headers_entrada[pos], 0)
            try:
                ponderador = float(valor_raw) if valor_raw else 0
                ponderadores_cc[nombre_cc] = ponderador
                total_ponderadores += ponderador
                logger.info(f"  📊 {nombre_cc}: {ponderador}")
            except (ValueError, TypeError):
                logger.warning(f"⚠️ Ponderador inválido en {nombre_cc}: {valor_raw}")
                ponderadores_cc[nombre_cc] = 0
    
    # Validar que los ponderadores sumen 1.0 (o 100% si están en porcentaje)
    if total_ponderadores == 0:
        logger.warning(f"⚠️ Total de ponderadores es 0, usando distribución equitativa")
        # Distribuir equitativamente entre los CC encontrados
        num_cc = len(posiciones_cc)
        for nombre_cc in ponderadores_cc:
            ponderadores_cc[nombre_cc] = 1.0 / num_cc if num_cc > 0 else 0
    elif abs(total_ponderadores - 1.0) > 0.01:  # Si no suman 1.0, normalizar
        logger.info(f"🔧 Normalizando ponderadores (total: {total_ponderadores})")
        for nombre_cc in ponderadores_cc:
            ponderadores_cc[nombre_cc] = ponderadores_cc[nombre_cc] / total_ponderadores
    
    # Generar códigos de cuenta
    if codigo_cuenta_original.startswith('2'):
        codigo_proveedores = codigo_cuenta_original
        codigo_gastos_base = codigo_cuenta_original.replace('2', '5', 1)
        codigo_iva = codigo_cuenta_original.replace('2', '1', 1)
    elif codigo_cuenta_original.startswith('5'):
        codigo_gastos_base = codigo_cuenta_original
        codigo_proveedores = codigo_cuenta_original.replace('5', '2', 1)
        codigo_iva = codigo_cuenta_original.replace('5', '1', 1)
    elif codigo_cuenta_original.startswith('1'):
        codigo_iva = codigo_cuenta_original
        codigo_proveedores = codigo_cuenta_original.replace('1', '2', 1)
        codigo_gastos_base = codigo_cuenta_original.replace('1', '5', 1)
    else:
        codigo_proveedores = f"2{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "2111001"
        codigo_gastos_base = f"5{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "5111001"
        codigo_iva = f"1{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "1191001"
    
    # 1. Cuenta Proveedores (igual que 1CC)
    fila_proveedores = crear_fila_base_tipo_33(datos_comunes, headers_salida, nombre_tipo_documento)
    fila_proveedores["Código Plan de Cuenta"] = codigo_proveedores
    fila_proveedores["Monto al Haber Moneda Base"] = formatear_monto_clp(round(monto_total, 2))
    fila_proveedores["Monto 1 Detalle Libro"] = formatear_monto_clp(round(monto_neto, 2))
    fila_proveedores["Monto 3 Detalle Libro"] = formatear_monto_clp(round(monto_iva, 2))
    monto_suma_detalle = round(monto_neto + monto_iva, 2)
    fila_proveedores["Monto Suma Detalle Libro"] = formatear_monto_clp(monto_suma_detalle)
    filas_resultado.append(fila_proveedores)
    
    # 2. Cuentas de Gastos (una por cada centro de costo)
    contador_gasto = 1
    
    # Mapeo de columnas a keys de mapeo (igual que en calcular_codigos_cc_para_fila)
    # El frontend ahora usa claves lógicas; mantener compatibilidad con claves legacy
    columnas_cc_mapeo = {
        'PyC': 'PyC',
        'PS': 'PS',
        'CO': 'CO',
        'RE': 'RE',
        'TR': 'TR',
        'CF': 'CF',
        'LRC': 'LRC'
    }
    legacy_map = {'PyC': 'col10', 'PS': 'col11', 'CO': 'col12'}
    
    for nombre_cc, ponderador in ponderadores_cc.items():
        if ponderador > 0:  # Solo crear línea si hay ponderador
            # Calcular monto ponderado para este CC (incluye neto + exento)
            monto_gasto_cc = round(monto_gastos_total * ponderador, 2)
            
            # Obtener el código de centro de costo que el usuario asignó en el frontend
            # Usar el mismo mapeo que usa calcular_codigos_cc_para_fila
            mapeo_key = columnas_cc_mapeo.get(nombre_cc, '')
            codigo_cc_usuario = ""
            
            # Buscar primero por clave lógica, luego legacy
            if mapeo_key and mapeo_key in mapeo_cc and mapeo_cc[mapeo_key]:
                codigo_cc_usuario = str(mapeo_cc[mapeo_key]).strip()
            else:
                legacy_key = legacy_map.get(nombre_cc)
                if legacy_key and legacy_key in mapeo_cc and mapeo_cc[legacy_key]:
                    codigo_cc_usuario = str(mapeo_cc[legacy_key]).strip()
            
            if not codigo_cc_usuario:
                logger.warning(f"⚠️ No se encontró código de CC para '{nombre_cc}' (mapeo_key: {mapeo_key}) en mapeo_cc: {mapeo_cc}")
                codigo_cc_usuario = ""  # Dejar vacío si no hay mapeo
            
            # Crear fila de gastos para este CC
            fila_gastos_cc = crear_fila_base_tipo_33(datos_comunes, headers_salida, nombre_tipo_documento)
            fila_gastos_cc["Código Plan de Cuenta"] = codigo_gastos_base  # Sin sufijo _01, _02
            fila_gastos_cc["Monto al Debe Moneda Base"] = formatear_monto_clp(monto_gasto_cc)
            
            # Solo asignar código de centro de costo si existe
            if codigo_cc_usuario:
                fila_gastos_cc["Código Centro de Costo"] = codigo_cc_usuario
            
            logger.info(f"💰 Cuenta Gastos {contador_gasto}: {nombre_cc} -> CC: '{codigo_cc_usuario}' = {formatear_monto_clp(monto_gasto_cc)} (ponderador: {ponderador:.3f})")
            
            filas_resultado.append(fila_gastos_cc)
            contador_gasto += 1
    
    # 3. Cuenta IVA (igual que 1CC)
    fila_iva = crear_fila_base_tipo_33(datos_comunes, headers_salida, nombre_tipo_documento)
    fila_iva["Código Plan de Cuenta"] = codigo_iva
    fila_iva["Monto al Debe Moneda Base"] = formatear_monto_clp(round(monto_iva, 2))
    filas_resultado.append(fila_iva)
    
    logger.info(f"📋 Tipo 33 con 2CC generó {len(filas_resultado)} filas: 1 Proveedores + {contador_gasto-1} Gastos + 1 IVA")
    
    return filas_resultado

def aplicar_reglas_tipo_33_3cc(fila, headers_entrada, mapeo_cc, headers_salida, tipos_doc_map):
    """
    Aplica reglas específicas para tipo de documento 33 CON 3 CENTROS DE COSTO
    
    CONTROLADOR EXENTO:
    - Si hay monto en "Monto Exento": Calcula IVA desde "Monto Neto"
    - Si NO hay exento: Calcula IVA desde "Monto Total" (lógica original)
    
    REGLAS ESPECÍFICAS PARA 3CC:
    - 1 cuenta de Proveedores con Monto 1 y 3 Detalle (igual que 1CC y 2CC)
    - 3 cuentas de Gastos separadas, cada una con su monto ponderado
    - 1 cuenta de IVA (igual que 1CC y 2CC)
    
    El monto de cada cuenta de gastos = monto_neto * ponderador_del_CC
    Los ponderadores están en las columnas PyC, EB/PS, CO
    """
    filas_resultado = []
    
    # 🔍 CONTROLADOR: Verificar si hay monto exento
    tiene_exento = verificar_tiene_exento_tipo_33(fila, headers_entrada)
    
    # Extraer datos comunes de la fila original
    datos_comunes = extraer_datos_comunes_tipo_33(fila, headers_entrada)
    
    # Calcular montos con validación
    monto_total_raw = datos_comunes.get('monto_total', 0)
    try:
        monto_total = float(monto_total_raw) if monto_total_raw else 0
    except (ValueError, TypeError):
        logger.warning(f"⚠️ Monto inválido encontrado: {monto_total_raw}, usando 0")
        monto_total = 0
    
    if monto_total <= 0:
        logger.warning(f"⚠️ Monto total es 0 o negativo: {monto_total}")
        return []  # No generar filas si no hay monto válido
    
    # Calcular montos según si hay exento o no
    if tiene_exento:
        # CON EXENTO: Calcular IVA desde "Monto Neto"
        # Obtener monto neto de la columna correspondiente
        mapeo = mapear_columnas_automaticamente(headers_entrada)
        if mapeo.get('monto_neto') is not None:
            monto_neto_raw = fila.get(headers_entrada[mapeo['monto_neto']], 0)
            try:
                monto_neto = float(monto_neto_raw) if monto_neto_raw else 0
            except (ValueError, TypeError):
                logger.warning(f"⚠️ Monto Neto inválido: {monto_neto_raw}, usando cálculo tradicional")
                monto_neto = monto_total / 1.19
        else:
            logger.warning("⚠️ No se encontró columna 'Monto Neto', usando cálculo tradicional")
            monto_neto = monto_total / 1.19
        
        # Obtener monto exento de la columna correspondiente
        if mapeo.get('monto_exento') is not None:
            monto_exento_raw = fila.get(headers_entrada[mapeo['monto_exento']], 0)
            try:
                monto_exento = float(monto_exento_raw) if monto_exento_raw else 0
            except (ValueError, TypeError):
                logger.warning(f"⚠️ Monto Exento inválido: {monto_exento_raw}, usando 0")
                monto_exento = 0
        else:
            monto_exento = 0
        
        # Con exento, el IVA se calcula sobre el monto neto: IVA = Monto Neto * 0.19
        monto_iva = monto_neto * 0.19
        # El monto total de gastos incluye neto + exento (para distribución por ponderadores)
        monto_gastos_total = monto_neto + monto_exento
        logger.info(f"📊 CON EXENTO (3CC): Monto Neto = {monto_neto:.2f}, Monto Exento = {monto_exento:.2f}, Monto Gastos Total = {monto_gastos_total:.2f}, Monto IVA (19% del neto) = {monto_iva:.2f}")
    else:
        # SIN EXENTO: Lógica original (IVA desde Monto Total)
        monto_neto = monto_total / 1.19  # Monto 1 Detalle Libro
        monto_iva = monto_total - monto_neto  # Monto 3 Detalle Libro
        monto_exento = 0  # No hay exento
        monto_gastos_total = monto_neto  # Solo el monto neto
        logger.info(f"📊 SIN EXENTO (3CC): Monto Neto = {monto_neto:.2f}, Monto IVA = {monto_iva:.2f}")
    
    # Obtener código de cuenta original
    codigo_cuenta_original = datos_comunes.get('codigo_cuenta', '')
    
    # Obtener nombre del tipo de documento
    nombre_tipo_documento = tipos_doc_map.get("33", "Factura Electrónica")
    logger.info(f"🏷️ Tipo 33 con 3CC -> Nombre: '{nombre_tipo_documento}'")
    
    # Detectar posiciones de centros de costo dinámicamente
    posiciones_cc = detectar_posiciones_centros_costo(headers_entrada)
    logger.info(f"🔍 Posiciones CC detectadas para 3CC: {posiciones_cc}")
    
    # Extraer ponderadores de cada centro de costo
    ponderadores_cc = {}
    total_ponderadores = 0
    
    for nombre_cc, pos in posiciones_cc.items():
        if pos is not None and pos < len(headers_entrada):
            valor_raw = fila.get(headers_entrada[pos], 0)
            try:
                ponderador = float(valor_raw) if valor_raw else 0
                ponderadores_cc[nombre_cc] = ponderador
                total_ponderadores += ponderador
                logger.info(f"  📊 {nombre_cc}: {ponderador}")
            except (ValueError, TypeError):
                logger.warning(f"⚠️ Ponderador inválido en {nombre_cc}: {valor_raw}")
                ponderadores_cc[nombre_cc] = 0
    
    # Validar que los ponderadores sumen 1.0 (o 100% si están en porcentaje)
    if total_ponderadores == 0:
        logger.warning(f"⚠️ Total de ponderadores es 0, usando distribución equitativa")
        # Distribuir equitativamente entre los CC encontrados
        num_cc = len(posiciones_cc)
        for nombre_cc in ponderadores_cc:
            ponderadores_cc[nombre_cc] = 1.0 / num_cc if num_cc > 0 else 0
    elif abs(total_ponderadores - 1.0) > 0.01:  # Si no suman 1.0, normalizar
        logger.info(f"🔧 Normalizando ponderadores (total: {total_ponderadores})")
        for nombre_cc in ponderadores_cc:
            ponderadores_cc[nombre_cc] = ponderadores_cc[nombre_cc] / total_ponderadores
    
    # Generar códigos de cuenta
    if codigo_cuenta_original.startswith('2'):
        codigo_proveedores = codigo_cuenta_original
        codigo_gastos_base = codigo_cuenta_original.replace('2', '5', 1)
        codigo_iva = codigo_cuenta_original.replace('2', '1', 1)
    elif codigo_cuenta_original.startswith('5'):
        codigo_gastos_base = codigo_cuenta_original
        codigo_proveedores = codigo_cuenta_original.replace('5', '2', 1)
        codigo_iva = codigo_cuenta_original.replace('5', '1', 1)
    elif codigo_cuenta_original.startswith('1'):
        codigo_iva = codigo_cuenta_original
        codigo_proveedores = codigo_cuenta_original.replace('1', '2', 1)
        codigo_gastos_base = codigo_cuenta_original.replace('1', '5', 1)
    else:
        codigo_proveedores = f"2{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "2111001"
        codigo_gastos_base = f"5{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "5111001"
        codigo_iva = f"1{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "1191001"
    
    # 1. Cuenta Proveedores (igual que 1CC y 2CC)
    fila_proveedores = crear_fila_base_tipo_33(datos_comunes, headers_salida, nombre_tipo_documento)
    fila_proveedores["Código Plan de Cuenta"] = codigo_proveedores
    fila_proveedores["Monto al Haber Moneda Base"] = formatear_monto_clp(round(monto_total, 2))
    fila_proveedores["Monto 1 Detalle Libro"] = formatear_monto_clp(round(monto_neto, 2))
    fila_proveedores["Monto 3 Detalle Libro"] = formatear_monto_clp(round(monto_iva, 2))
    monto_suma_detalle = round(monto_neto + monto_iva, 2)
    fila_proveedores["Monto Suma Detalle Libro"] = formatear_monto_clp(monto_suma_detalle)
    filas_resultado.append(fila_proveedores)
    
    # 2. Cuentas de Gastos (una por cada centro de costo - hasta 3)
    contador_gasto = 1
    
    # Mapeo de columnas a keys de mapeo (igual que en 2CC)
    columnas_cc_mapeo = {
        'PyC': 'PyC',
        'PS': 'PS',
        'CO': 'CO',
        'RE': 'RE',
        'TR': 'TR',
        'CF': 'CF',
        'LRC': 'LRC'
    }
    legacy_map = {'PyC': 'col10', 'PS': 'col11', 'CO': 'col12'}
    
    for nombre_cc, ponderador in ponderadores_cc.items():
        if ponderador > 0:  # Solo crear línea si hay ponderador
            # Calcular monto ponderado para este CC (incluye neto + exento)
            monto_gasto_cc = round(monto_gastos_total * ponderador, 2)
            
            # Obtener el código de centro de costo que el usuario asignó en el frontend
            # Usar el mismo mapeo que usa calcular_codigos_cc_para_fila
            mapeo_key = columnas_cc_mapeo.get(nombre_cc, '')
            codigo_cc_usuario = ""
            
            if mapeo_key and mapeo_key in mapeo_cc and mapeo_cc[mapeo_key]:
                codigo_cc_usuario = str(mapeo_cc[mapeo_key]).strip()
            else:
                legacy_key = legacy_map.get(nombre_cc)
                if legacy_key and legacy_key in mapeo_cc and mapeo_cc[legacy_key]:
                    codigo_cc_usuario = str(mapeo_cc[legacy_key]).strip()
            
            if not codigo_cc_usuario:
                logger.warning(f"⚠️ No se encontró código de CC para '{nombre_cc}' (mapeo_key: {mapeo_key}) en mapeo_cc: {mapeo_cc}")
                codigo_cc_usuario = ""  # Dejar vacío si no hay mapeo
            
            # Crear fila de gastos para este CC
            fila_gastos_cc = crear_fila_base_tipo_33(datos_comunes, headers_salida, nombre_tipo_documento)
            fila_gastos_cc["Código Plan de Cuenta"] = codigo_gastos_base  # Sin sufijo
            fila_gastos_cc["Monto al Debe Moneda Base"] = formatear_monto_clp(monto_gasto_cc)
            
            # Solo asignar código de centro de costo si existe
            if codigo_cc_usuario:
                fila_gastos_cc["Código Centro de Costo"] = codigo_cc_usuario
            
            logger.info(f"💰 Cuenta Gastos {contador_gasto}: {nombre_cc} -> CC: '{codigo_cc_usuario}' = {formatear_monto_clp(monto_gasto_cc)} (ponderador: {ponderador:.3f})")
            
            filas_resultado.append(fila_gastos_cc)
            contador_gasto += 1
    
    # 3. Cuenta IVA (igual que 1CC y 2CC)
    fila_iva = crear_fila_base_tipo_33(datos_comunes, headers_salida, nombre_tipo_documento)
    fila_iva["Código Plan de Cuenta"] = codigo_iva
    fila_iva["Monto al Debe Moneda Base"] = formatear_monto_clp(round(monto_iva, 2))
    filas_resultado.append(fila_iva)
    
    logger.info(f"📋 Tipo 33 con 3CC generó {len(filas_resultado)} filas: 1 Proveedores + {contador_gasto-1} Gastos + 1 IVA")
    
    return filas_resultado


def aplicar_reglas_tipo_34(fila, headers_entrada, mapeo_cc, headers_salida, tipos_doc_map):
    """
    Aplica reglas específicas para tipo de documento 34 con 1CC
    Genera solo 2 cuentas: Proveedores (2xxx) y Gastos (5xxx) - NO incluye IVA
    """
    filas_resultado = []
    
    # Extraer datos comunes de la fila original
    datos_comunes = extraer_datos_comunes_tipo_34(fila, headers_entrada)
    
    # Calcular montos con validación
    monto_total_raw = datos_comunes.get('monto_total', 0)
    try:
        monto_total = float(monto_total_raw) if monto_total_raw else 0
    except (ValueError, TypeError):
        logger.warning(f"⚠️ Monto inválido encontrado en tipo 34: {monto_total_raw}, usando 0")
        monto_total = 0
    
    if monto_total <= 0:
        logger.warning(f"⚠️ Monto total es 0 o negativo en tipo 34: {monto_total}")
        return []  # No generar filas si no hay monto válido
    
    # Para tipo 34, el monto total es el monto del gasto (sin IVA involucrado)
    monto_gasto = monto_total
    
    # Calcular códigos de centros de costos aplicables
    codigos_cc = calcular_codigos_cc_para_fila(fila, headers_entrada, mapeo_cc)
    
    # Obtener código de cuenta original de la columna 8
    codigo_cuenta_original = datos_comunes.get('codigo_cuenta', '')
    
    # Obtener nombre del tipo de documento
    nombre_tipo_documento = tipos_doc_map.get("34", "Factura Exenta Electrónica")
    logger.info(f"🏷️ Tipo 34 -> Nombre: '{nombre_tipo_documento}'")
    
    # Generar códigos de cuenta basados en el código original
    # Si el código empieza con 2 o 5, usar como base, sino generar códigos genéricos
    if codigo_cuenta_original.startswith('2'):
        codigo_proveedores = codigo_cuenta_original
        codigo_gastos = codigo_cuenta_original.replace('2', '5', 1)
    elif codigo_cuenta_original.startswith('5'):
        codigo_gastos = codigo_cuenta_original
        codigo_proveedores = codigo_cuenta_original.replace('5', '2', 1)
    else:
        # Usar el código original como base y generar variaciones
        codigo_proveedores = f"2{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "2111001"
        codigo_gastos = f"5{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "5111001"
    
    # 1. Cuenta Proveedores (código empieza con 2)
    fila_proveedores = crear_fila_base_tipo_34(datos_comunes, headers_salida, nombre_tipo_documento)
    fila_proveedores["Código Plan de Cuenta"] = codigo_proveedores
    fila_proveedores["Monto al Haber Moneda Base"] = formatear_monto_clp(round(monto_total, 2))
    # ✨ CORREGIDO: Usar "Monto 2 Detalle Libro" en cuenta Proveedores para tipo 34
    fila_proveedores["Monto 2 Detalle Libro"] = formatear_monto_clp(round(monto_gasto, 2))
    # ✨ NUEVO: Calcular y agregar "Monto Suma Detalle Libro" (solo monto 2 para tipo 34)
    fila_proveedores["Monto Suma Detalle Libro"] = formatear_monto_clp(round(monto_gasto, 2))
    filas_resultado.append(fila_proveedores)
    
    # 2. Cuenta Gastos (código empieza con 5)
    fila_gastos = crear_fila_base_tipo_34(datos_comunes, headers_salida, nombre_tipo_documento)
    fila_gastos["Código Plan de Cuenta"] = codigo_gastos
    fila_gastos["Monto al Debe Moneda Base"] = formatear_monto_clp(round(monto_gasto, 2))
    # ✨ REMOVIDO: Ya no se asigna "Monto 1 Detalle Libro" aquí, se movió a cuenta Proveedores como "Monto 3"
    if codigos_cc:
        fila_gastos["Código Centro de Costo"] = codigos_cc
    filas_resultado.append(fila_gastos)
    
    # NOTA: Para tipo 34 con 1CC NO se genera cuenta de IVA
    
    return filas_resultado

def aplicar_reglas_tipo_34_2cc(fila, headers_entrada, mapeo_cc, headers_salida, tipos_doc_map):
    """
    Aplica reglas específicas para tipo de documento 34 CON 2 CENTROS DE COSTO
    
    REGLAS ESPECÍFICAS PARA 2CC:
    - 1 cuenta de Proveedores con Monto 2 Detalle (igual que 1CC)
    - 2 cuentas de Gastos separadas, cada una con su monto ponderado
    - NO incluye cuenta de IVA (factura exenta)
    
    El monto de cada cuenta de gastos = monto_total * ponderador_del_CC
    Los ponderadores están en las columnas PyC, EB/PS, CO
    """
    filas_resultado = []
    
    # Extraer datos comunes de la fila original
    datos_comunes = extraer_datos_comunes_tipo_34(fila, headers_entrada)
    
    # Calcular montos con validación
    monto_total_raw = datos_comunes.get('monto_total', 0)
    try:
        monto_total = float(monto_total_raw) if monto_total_raw else 0
    except (ValueError, TypeError):
        logger.warning(f"⚠️ Monto inválido encontrado en tipo 34: {monto_total_raw}, usando 0")
        monto_total = 0
    
    if monto_total <= 0:
        logger.warning(f"⚠️ Monto total es 0 o negativo en tipo 34: {monto_total}")
        return []  # No generar filas si no hay monto válido
    
    # Para tipo 34, el monto total es el monto del gasto (sin IVA involucrado)
    monto_gasto = monto_total
    
    # Obtener código de cuenta original
    codigo_cuenta_original = datos_comunes.get('codigo_cuenta', '')
    
    # Obtener nombre del tipo de documento
    nombre_tipo_documento = tipos_doc_map.get("34", "Factura Exenta Electrónica")
    logger.info(f"🏷️ Tipo 34 con 2CC -> Nombre: '{nombre_tipo_documento}'")
    
    # Detectar posiciones de centros de costo dinámicamente
    posiciones_cc = detectar_posiciones_centros_costo(headers_entrada)
    logger.info(f"🔍 Posiciones CC detectadas para 34 con 2CC: {posiciones_cc}")
    
    # Extraer ponderadores de cada centro de costo
    ponderadores_cc = {}
    total_ponderadores = 0
    
    for nombre_cc, pos in posiciones_cc.items():
        if pos is not None and pos < len(headers_entrada):
            valor_raw = fila.get(headers_entrada[pos], 0)
            try:
                ponderador = float(valor_raw) if valor_raw else 0
                ponderadores_cc[nombre_cc] = ponderador
                total_ponderadores += ponderador
                logger.info(f"  📊 {nombre_cc}: {ponderador}")
            except (ValueError, TypeError):
                logger.warning(f"⚠️ Ponderador inválido en {nombre_cc}: {valor_raw}")
                ponderadores_cc[nombre_cc] = 0
    
    # Validar que los ponderadores sumen 1.0 (o 100% si están en porcentaje)
    if total_ponderadores == 0:
        logger.warning(f"⚠️ Total de ponderadores es 0, usando distribución equitativa")
        # Distribuir equitativamente entre los CC encontrados
        num_cc = len(posiciones_cc)
        for nombre_cc in ponderadores_cc:
            ponderadores_cc[nombre_cc] = 1.0 / num_cc if num_cc > 0 else 0
    elif abs(total_ponderadores - 1.0) > 0.01:  # Si no suman 1.0, normalizar
        logger.info(f"🔧 Normalizando ponderadores (total: {total_ponderadores})")
        for nombre_cc in ponderadores_cc:
            ponderadores_cc[nombre_cc] = ponderadores_cc[nombre_cc] / total_ponderadores
    
    # Generar códigos de cuenta
    if codigo_cuenta_original.startswith('2'):
        codigo_proveedores = codigo_cuenta_original
        codigo_gastos_base = codigo_cuenta_original.replace('2', '5', 1)
    elif codigo_cuenta_original.startswith('5'):
        codigo_gastos_base = codigo_cuenta_original
        codigo_proveedores = codigo_cuenta_original.replace('5', '2', 1)
    else:
        codigo_proveedores = f"2{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "2111001"
        codigo_gastos_base = f"5{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "5111001"
    
    # 1. Cuenta Proveedores (igual que 1CC)
    fila_proveedores = crear_fila_base_tipo_34(datos_comunes, headers_salida, nombre_tipo_documento)
    fila_proveedores["Código Plan de Cuenta"] = codigo_proveedores
    fila_proveedores["Monto al Haber Moneda Base"] = formatear_monto_clp(round(monto_total, 2))
    fila_proveedores["Monto 2 Detalle Libro"] = formatear_monto_clp(round(monto_gasto, 2))
    fila_proveedores["Monto Suma Detalle Libro"] = formatear_monto_clp(round(monto_gasto, 2))
    filas_resultado.append(fila_proveedores)
    
    # 2. Cuentas de Gastos (una por cada centro de costo)
    contador_gasto = 1
    
    # Mapeo de columnas a keys de mapeo (igual que en tipo 33)
    columnas_cc_mapeo = {
        'PyC': 'PyC',
        'PS': 'PS',
        'CO': 'CO',
        'RE': 'RE',
        'TR': 'TR',
        'CF': 'CF',
        'LRC': 'LRC'
    }
    legacy_map = {'PyC': 'col10', 'PS': 'col11', 'CO': 'col12'}
    
    for nombre_cc, ponderador in ponderadores_cc.items():
        if ponderador > 0:  # Solo crear línea si hay ponderador
            # Calcular monto ponderado para este CC
            monto_gasto_cc = round(monto_gasto * ponderador, 2)
            
            # Obtener el código de centro de costo que el usuario asignó en el frontend
            mapeo_key = columnas_cc_mapeo.get(nombre_cc, '')
            codigo_cc_usuario = ""
            
            if mapeo_key and mapeo_key in mapeo_cc and mapeo_cc[mapeo_key]:
                codigo_cc_usuario = str(mapeo_cc[mapeo_key]).strip()
            else:
                legacy_key = legacy_map.get(nombre_cc)
                if legacy_key and legacy_key in mapeo_cc and mapeo_cc[legacy_key]:
                    codigo_cc_usuario = str(mapeo_cc[legacy_key]).strip()
            
            if not codigo_cc_usuario:
                logger.warning(f"⚠️ No se encontró código de CC para '{nombre_cc}' (mapeo_key: {mapeo_key}) en mapeo_cc: {mapeo_cc}")
                codigo_cc_usuario = ""  # Dejar vacío si no hay mapeo
            
            # Crear fila de gastos para este CC
            fila_gastos_cc = crear_fila_base_tipo_34(datos_comunes, headers_salida, nombre_tipo_documento)
            fila_gastos_cc["Código Plan de Cuenta"] = codigo_gastos_base  # Sin sufijo
            fila_gastos_cc["Monto al Debe Moneda Base"] = formatear_monto_clp(monto_gasto_cc)
            
            # Solo asignar código de centro de costo si existe
            if codigo_cc_usuario:
                fila_gastos_cc["Código Centro de Costo"] = codigo_cc_usuario
            
            logger.info(f"💰 Cuenta Gastos {contador_gasto}: {nombre_cc} -> CC: '{codigo_cc_usuario}' = {formatear_monto_clp(monto_gasto_cc)} (ponderador: {ponderador:.3f})")
            
            filas_resultado.append(fila_gastos_cc)
            contador_gasto += 1
    
    # NOTA: Para tipo 34 con 2CC NO se genera cuenta de IVA (factura exenta)
    
    logger.info(f"📋 Tipo 34 con 2CC generó {len(filas_resultado)} filas: 1 Proveedores + {contador_gasto-1} Gastos (sin IVA)")
    
    return filas_resultado

def aplicar_reglas_tipo_34_3cc(fila, headers_entrada, mapeo_cc, headers_salida, tipos_doc_map):
    """
    Aplica reglas específicas para tipo de documento 34 CON 3 CENTROS DE COSTO
    
    REGLAS ESPECÍFICAS PARA 3CC:
    - 1 cuenta de Proveedores con Monto 2 Detalle (igual que 1CC y 2CC)
    - 3 cuentas de Gastos separadas, cada una con su monto ponderado
    - NO incluye cuenta de IVA (factura exenta)
    
    El monto de cada cuenta de gastos = monto_total * ponderador_del_CC
    Los ponderadores están en las columnas PyC, EB/PS, CO
    """
    filas_resultado = []
    
    # Extraer datos comunes de la fila original
    datos_comunes = extraer_datos_comunes_tipo_34(fila, headers_entrada)
    
    # Calcular montos con validación
    monto_total_raw = datos_comunes.get('monto_total', 0)
    try:
        monto_total = float(monto_total_raw) if monto_total_raw else 0
    except (ValueError, TypeError):
        logger.warning(f"⚠️ Monto inválido encontrado en tipo 34: {monto_total_raw}, usando 0")
        monto_total = 0
    
    if monto_total <= 0:
        logger.warning(f"⚠️ Monto total es 0 o negativo en tipo 34: {monto_total}")
        return []  # No generar filas si no hay monto válido
    
    # Para tipo 34, el monto total es el monto del gasto (sin IVA involucrado)
    monto_gasto = monto_total
    
    # Obtener código de cuenta original
    codigo_cuenta_original = datos_comunes.get('codigo_cuenta', '')
    
    # Obtener nombre del tipo de documento
    nombre_tipo_documento = tipos_doc_map.get("34", "Factura Exenta Electrónica")
    logger.info(f"🏷️ Tipo 34 con 3CC -> Nombre: '{nombre_tipo_documento}'")
    
    # Detectar posiciones de centros de costo dinámicamente
    posiciones_cc = detectar_posiciones_centros_costo(headers_entrada)
    logger.info(f"🔍 Posiciones CC detectadas para 34 con 3CC: {posiciones_cc}")
    
    # Extraer ponderadores de cada centro de costo
    ponderadores_cc = {}
    total_ponderadores = 0
    
    for nombre_cc, pos in posiciones_cc.items():
        if pos is not None and pos < len(headers_entrada):
            valor_raw = fila.get(headers_entrada[pos], 0)
            try:
                ponderador = float(valor_raw) if valor_raw else 0
                ponderadores_cc[nombre_cc] = ponderador
                total_ponderadores += ponderador
                logger.info(f"  📊 {nombre_cc}: {ponderador}")
            except (ValueError, TypeError):
                logger.warning(f"⚠️ Ponderador inválido en {nombre_cc}: {valor_raw}")
                ponderadores_cc[nombre_cc] = 0
    
    # Validar que los ponderadores sumen 1.0 (o 100% si están en porcentaje)
    if total_ponderadores == 0:
        logger.warning(f"⚠️ Total de ponderadores es 0, usando distribución equitativa")
        # Distribuir equitativamente entre los CC encontrados
        num_cc = len(posiciones_cc)
        for nombre_cc in ponderadores_cc:
            ponderadores_cc[nombre_cc] = 1.0 / num_cc if num_cc > 0 else 0
    elif abs(total_ponderadores - 1.0) > 0.01:  # Si no suman 1.0, normalizar
        logger.info(f"🔧 Normalizando ponderadores (total: {total_ponderadores})")
        for nombre_cc in ponderadores_cc:
            ponderadores_cc[nombre_cc] = ponderadores_cc[nombre_cc] / total_ponderadores
    
    # Generar códigos de cuenta
    if codigo_cuenta_original.startswith('2'):
        codigo_proveedores = codigo_cuenta_original
        codigo_gastos_base = codigo_cuenta_original.replace('2', '5', 1)
    elif codigo_cuenta_original.startswith('5'):
        codigo_gastos_base = codigo_cuenta_original
        codigo_proveedores = codigo_cuenta_original.replace('5', '2', 1)
    else:
        codigo_proveedores = f"2{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "2111001"
        codigo_gastos_base = f"5{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "5111001"
    
    # 1. Cuenta Proveedores (igual que 1CC y 2CC)
    fila_proveedores = crear_fila_base_tipo_34(datos_comunes, headers_salida, nombre_tipo_documento)
    fila_proveedores["Código Plan de Cuenta"] = codigo_proveedores
    fila_proveedores["Monto al Haber Moneda Base"] = formatear_monto_clp(round(monto_total, 2))
    fila_proveedores["Monto 2 Detalle Libro"] = formatear_monto_clp(round(monto_gasto, 2))
    fila_proveedores["Monto Suma Detalle Libro"] = formatear_monto_clp(round(monto_gasto, 2))
    filas_resultado.append(fila_proveedores)
    
    # 2. Cuentas de Gastos (una por cada centro de costo - hasta 3)
    contador_gasto = 1
    
    # Mapeo de columnas a keys de mapeo (igual que en tipo 33)
    columnas_cc_mapeo = {
        'PyC': 'PyC',
        'PS': 'PS',
        'CO': 'CO',
        'RE': 'RE',
        'TR': 'TR',
        'CF': 'CF',
        'LRC': 'LRC'
    }
    legacy_map = {'PyC': 'col10', 'PS': 'col11', 'CO': 'col12'}
    
    for nombre_cc, ponderador in ponderadores_cc.items():
        if ponderador > 0:  # Solo crear línea si hay ponderador
            # Calcular monto ponderado para este CC
            monto_gasto_cc = round(monto_gasto * ponderador, 2)
            
            # Obtener el código de centro de costo que el usuario asignó en el frontend
            mapeo_key = columnas_cc_mapeo.get(nombre_cc, '')
            codigo_cc_usuario = ""
            
            if mapeo_key and mapeo_key in mapeo_cc and mapeo_cc[mapeo_key]:
                codigo_cc_usuario = str(mapeo_cc[mapeo_key]).strip()
            else:
                legacy_key = legacy_map.get(nombre_cc)
                if legacy_key and legacy_key in mapeo_cc and mapeo_cc[legacy_key]:
                    codigo_cc_usuario = str(mapeo_cc[legacy_key]).strip()
            
            if not codigo_cc_usuario:
                logger.warning(f"⚠️ No se encontró código de CC para '{nombre_cc}' (mapeo_key: {mapeo_key}) en mapeo_cc: {mapeo_cc}")
                codigo_cc_usuario = ""  # Dejar vacío si no hay mapeo
            
            # Crear fila de gastos para este CC
            fila_gastos_cc = crear_fila_base_tipo_34(datos_comunes, headers_salida, nombre_tipo_documento)
            fila_gastos_cc["Código Plan de Cuenta"] = codigo_gastos_base  # Sin sufijo
            fila_gastos_cc["Monto al Debe Moneda Base"] = formatear_monto_clp(monto_gasto_cc)
            
            # Solo asignar código de centro de costo si existe
            if codigo_cc_usuario:
                fila_gastos_cc["Código Centro de Costo"] = codigo_cc_usuario
            
            logger.info(f"💰 Cuenta Gastos {contador_gasto}: {nombre_cc} -> CC: '{codigo_cc_usuario}' = {formatear_monto_clp(monto_gasto_cc)} (ponderador: {ponderador:.3f})")
            
            filas_resultado.append(fila_gastos_cc)
            contador_gasto += 1
    
    # NOTA: Para tipo 34 con 3CC NO se genera cuenta de IVA (factura exenta)
    
    logger.info(f"📋 Tipo 34 con 3CC generó {len(filas_resultado)} filas: 1 Proveedores + {contador_gasto-1} Gastos (sin IVA)")
    
    return filas_resultado

def aplicar_reglas_tipo_61(fila, headers_entrada, mapeo_cc, headers_salida, tipos_doc_map):
    """
    Aplica reglas específicas para tipo de documento 61 (Nota de Crédito) con 1CC
    
    CONTROLADOR IVA:
    - Si hay monto en "Monto IVA": Genera 3 cuentas (Proveedores, Gastos, IVA)
    - Si NO hay IVA: Genera 2 cuentas (Proveedores, Gastos) - monto_neto = monto_total
    
    IGUAL al tipo 33 pero con montos INVERTIDOS (Debe ↔ Haber)
    """
    filas_resultado = []
    
    # 🔍 CONTROLADOR: Verificar si hay IVA
    tiene_iva = verificar_tiene_iva_tipo_61(fila, headers_entrada)
    
    # Extraer datos comunes de la fila original
    datos_comunes = extraer_datos_comunes_tipo_61(fila, headers_entrada)
    
    # Calcular montos con validación
    monto_total_raw = datos_comunes.get('monto_total', 0)
    try:
        monto_total = float(monto_total_raw) if monto_total_raw else 0
    except (ValueError, TypeError):
        logger.warning(f"⚠️ Monto inválido encontrado en tipo 61: {monto_total_raw}, usando 0")
        monto_total = 0
    
    # Para tipo 61 (Nota de Crédito), los montos pueden ser negativos
    # Trabajamos con el valor absoluto para los cálculos
    if monto_total == 0:
        logger.warning(f"⚠️ Monto total es 0 en tipo 61: {monto_total}")
        return []  # No generar filas si no hay monto válido
    
    # Usar valor absoluto para cálculos, pero mantener el signo original en logs
    monto_total_absoluto = abs(monto_total)
    logger.info(f"💰 Tipo 61: Monto original = {monto_total}, trabajando con valor absoluto = {monto_total_absoluto}")
    
    # Calcular montos según si hay IVA o no
    if tiene_iva:
        # CON IVA: Calcular igual que tipo 33
        monto_neto = monto_total_absoluto / 1.19  # Monto 1 Detalle Libro
        monto_iva = monto_total_absoluto - monto_neto  # Monto 3 Detalle Libro
        logger.info(f"📊 CON IVA: Monto Neto = {monto_neto:.2f}, Monto IVA = {monto_iva:.2f}")
    else:
        # SIN IVA: Monto neto = monto total, no hay IVA
        monto_neto = monto_total_absoluto
        monto_iva = 0
        logger.info(f"📊 SIN IVA: Monto Neto = Monto Total = {monto_neto:.2f}")
    
    # Calcular códigos de centros de costos aplicables
    codigos_cc = calcular_codigos_cc_para_fila(fila, headers_entrada, mapeo_cc)
    
    # Obtener código de cuenta original de la columna 8
    codigo_cuenta_original = datos_comunes.get('codigo_cuenta', '')
    
    # Obtener nombre del tipo de documento
    nombre_tipo_documento = tipos_doc_map.get("61", "Nota de Crédito Electrónica")
    logger.info(f"🏷️ Tipo 61 -> Nombre: '{nombre_tipo_documento}'")
    
    # Generar códigos de cuenta basados en el código original (igual que tipo 33)
    if codigo_cuenta_original.startswith('2'):
        codigo_proveedores = codigo_cuenta_original
        codigo_gastos = codigo_cuenta_original.replace('2', '5', 1)
        codigo_iva = codigo_cuenta_original.replace('2', '1', 1)
    elif codigo_cuenta_original.startswith('5'):
        codigo_gastos = codigo_cuenta_original
        codigo_proveedores = codigo_cuenta_original.replace('5', '2', 1)
        codigo_iva = codigo_cuenta_original.replace('5', '1', 1)
    elif codigo_cuenta_original.startswith('1'):
        codigo_iva = codigo_cuenta_original
        codigo_proveedores = codigo_cuenta_original.replace('1', '2', 1)
        codigo_gastos = codigo_cuenta_original.replace('1', '5', 1)
    else:
        # Usar el código original como base y generar variaciones
        codigo_proveedores = f"2{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "2111001"
        codigo_gastos = f"5{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "5111001"
        codigo_iva = f"1{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "1191001"
    
    # 1. Cuenta Proveedores (código empieza con 2) - INVERTIDO: Debe en lugar de Haber
    fila_proveedores = crear_fila_base_tipo_61(datos_comunes, headers_salida, nombre_tipo_documento)
    fila_proveedores["Código Plan de Cuenta"] = codigo_proveedores
    fila_proveedores["Monto al Debe Moneda Base"] = formatear_monto_clp(round(monto_total_absoluto, 2))  # 🔄 INVERTIDO
    
    # Campos especiales: ajustar según si hay IVA o no
    if tiene_iva:
        fila_proveedores["Monto 1 Detalle Libro"] = formatear_monto_clp(round(monto_neto, 2))
        fila_proveedores["Monto 3 Detalle Libro"] = formatear_monto_clp(round(monto_iva, 2))
        monto_suma_detalle = round(monto_neto + monto_iva, 2)
    else:
        # SIN IVA: Solo monto neto, sin IVA
        fila_proveedores["Monto 1 Detalle Libro"] = formatear_monto_clp(round(monto_neto, 2))
        fila_proveedores["Monto 3 Detalle Libro"] = formatear_monto_clp(0)  # Sin IVA
        monto_suma_detalle = round(monto_neto, 2)
    
    fila_proveedores["Monto Suma Detalle Libro"] = formatear_monto_clp(monto_suma_detalle)
    filas_resultado.append(fila_proveedores)
    
    # 2. Cuenta Gastos (código empieza con 5) - INVERTIDO: Haber en lugar de Debe
    fila_gastos = crear_fila_base_tipo_61(datos_comunes, headers_salida, nombre_tipo_documento)
    fila_gastos["Código Plan de Cuenta"] = codigo_gastos
    fila_gastos["Monto al Haber Moneda Base"] = formatear_monto_clp(round(monto_neto, 2))  # 🔄 INVERTIDO
    if codigos_cc:
        fila_gastos["Código Centro de Costo"] = codigos_cc
    filas_resultado.append(fila_gastos)
    
    # 3. Cuenta IVA (código empieza con 1) - SOLO SI HAY IVA - INVERTIDO: Haber en lugar de Debe
    if tiene_iva:
        fila_iva = crear_fila_base_tipo_61(datos_comunes, headers_salida, nombre_tipo_documento)
        fila_iva["Código Plan de Cuenta"] = codigo_iva
        fila_iva["Monto al Haber Moneda Base"] = formatear_monto_clp(round(monto_iva, 2))  # 🔄 INVERTIDO
        filas_resultado.append(fila_iva)
        logger.info(f"📋 Tipo 61 CON IVA generó {len(filas_resultado)} filas: 1 Proveedores + 1 Gastos + 1 IVA (INVERTIDOS)")
    else:
        logger.info(f"📋 Tipo 61 SIN IVA generó {len(filas_resultado)} filas: 1 Proveedores + 1 Gastos (INVERTIDOS)")
    
    return filas_resultado

def aplicar_reglas_tipo_61_2cc(fila, headers_entrada, mapeo_cc, headers_salida, tipos_doc_map):
    """
    Aplica reglas específicas para tipo de documento 61 CON 2 CENTROS DE COSTO
    
    CONTROLADOR IVA:
    - Si hay monto en "Monto IVA": Genera 1 Proveedores + 2 Gastos + 1 IVA
    - Si NO hay IVA: Genera 1 Proveedores + 2 Gastos (sin cuenta IVA) - monto_neto = monto_total
    
    REGLAS ESPECÍFICAS PARA 2CC:
    - 1 cuenta de Proveedores con Monto 1 y 3 Detalle (igual que tipo 33 pero INVERTIDO)
    - 2 cuentas de Gastos separadas, cada una con su monto ponderado (INVERTIDO)
    - 1 cuenta de IVA (INVERTIDO) - SOLO SI HAY IVA
    
    NOTA: Tipo 61 es igual a tipo 33 pero con montos INVERTIDOS (Debe ↔ Haber)
    El monto de cada cuenta de gastos = monto_neto * ponderador_del_CC
    Los ponderadores están en las columnas PyC, EB/PS, CO
    """
    filas_resultado = []
    
    # 🔍 CONTROLADOR: Verificar si hay IVA
    tiene_iva = verificar_tiene_iva_tipo_61(fila, headers_entrada)
    
    # Extraer datos comunes de la fila original
    datos_comunes = extraer_datos_comunes_tipo_61(fila, headers_entrada)
    
    # Calcular montos con validación
    monto_total_raw = datos_comunes.get('monto_total', 0)
    try:
        monto_total = float(monto_total_raw) if monto_total_raw else 0
    except (ValueError, TypeError):
        logger.warning(f"⚠️ Monto inválido encontrado en tipo 61: {monto_total_raw}, usando 0")
        monto_total = 0
    
    # Para tipo 61 (Nota de Crédito), los montos pueden ser negativos
    # Trabajamos con el valor absoluto para los cálculos
    if monto_total == 0:
        logger.warning(f"⚠️ Monto total es 0 en tipo 61 con 2CC: {monto_total}")
        return []  # No generar filas si no hay monto válido
    
    # Usar valor absoluto para cálculos, pero mantener el signo original en logs
    monto_total_absoluto = abs(monto_total)
    logger.info(f"💰 Tipo 61 con 2CC: Monto original = {monto_total}, trabajando con valor absoluto = {monto_total_absoluto}")
    
    # Calcular montos según si hay IVA o no
    if tiene_iva:
        # CON IVA: Calcular igual que tipo 33
        monto_neto = monto_total_absoluto / 1.19  # Monto 1 Detalle Libro
        monto_iva = monto_total_absoluto - monto_neto  # Monto 3 Detalle Libro
        logger.info(f"📊 CON IVA: Monto Neto = {monto_neto:.2f}, Monto IVA = {monto_iva:.2f}")
    else:
        # SIN IVA: Monto neto = monto total, no hay IVA
        monto_neto = monto_total_absoluto
        monto_iva = 0
        logger.info(f"📊 SIN IVA: Monto Neto = Monto Total = {monto_neto:.2f}")
    
    # Obtener código de cuenta original
    codigo_cuenta_original = datos_comunes.get('codigo_cuenta', '')
    
    # Obtener nombre del tipo de documento
    nombre_tipo_documento = tipos_doc_map.get("61", "Nota de Crédito Electrónica")
    logger.info(f"🏷️ Tipo 61 con 2CC -> Nombre: '{nombre_tipo_documento}'")
    
    # Detectar posiciones de centros de costo dinámicamente
    posiciones_cc = detectar_posiciones_centros_costo(headers_entrada)
    logger.info(f"🔍 Posiciones CC detectadas para 61 con 2CC: {posiciones_cc}")
    
    # Extraer ponderadores de cada centro de costo
    ponderadores_cc = {}
    total_ponderadores = 0
    
    for nombre_cc, pos in posiciones_cc.items():
        if pos is not None and pos < len(headers_entrada):
            valor_raw = fila.get(headers_entrada[pos], 0)
            try:
                ponderador = float(valor_raw) if valor_raw else 0
                ponderadores_cc[nombre_cc] = ponderador
                total_ponderadores += ponderador
                logger.info(f"  📊 {nombre_cc}: {ponderador}")
            except (ValueError, TypeError):
                logger.warning(f"⚠️ Ponderador inválido en {nombre_cc}: {valor_raw}")
                ponderadores_cc[nombre_cc] = 0
    
    # Validar que los ponderadores sumen 1.0 (o 100% si están en porcentaje)
    if total_ponderadores == 0:
        logger.warning(f"⚠️ Total de ponderadores es 0, usando distribución equitativa")
        # Distribuir equitativamente entre los CC encontrados
        num_cc = len(posiciones_cc)
        for nombre_cc in ponderadores_cc:
            ponderadores_cc[nombre_cc] = 1.0 / num_cc if num_cc > 0 else 0
    elif abs(total_ponderadores - 1.0) > 0.01:  # Si no suman 1.0, normalizar
        logger.info(f"🔧 Normalizando ponderadores (total: {total_ponderadores})")
        for nombre_cc in ponderadores_cc:
            ponderadores_cc[nombre_cc] = ponderadores_cc[nombre_cc] / total_ponderadores
    
    # Generar códigos de cuenta
    if codigo_cuenta_original.startswith('2'):
        codigo_proveedores = codigo_cuenta_original
        codigo_gastos_base = codigo_cuenta_original.replace('2', '5', 1)
        codigo_iva = codigo_cuenta_original.replace('2', '1', 1)
    elif codigo_cuenta_original.startswith('5'):
        codigo_gastos_base = codigo_cuenta_original
        codigo_proveedores = codigo_cuenta_original.replace('5', '2', 1)
        codigo_iva = codigo_cuenta_original.replace('5', '1', 1)
    elif codigo_cuenta_original.startswith('1'):
        codigo_iva = codigo_cuenta_original
        codigo_proveedores = codigo_cuenta_original.replace('1', '2', 1)
        codigo_gastos_base = codigo_cuenta_original.replace('1', '5', 1)
    else:
        codigo_proveedores = f"2{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "2111001"
        codigo_gastos_base = f"5{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "5111001"
        codigo_iva = f"1{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "1191001"
    
    # 1. Cuenta Proveedores (INVERTIDO: Debe en lugar de Haber)
    fila_proveedores = crear_fila_base_tipo_61(datos_comunes, headers_salida, nombre_tipo_documento)
    fila_proveedores["Código Plan de Cuenta"] = codigo_proveedores
    fila_proveedores["Monto al Debe Moneda Base"] = formatear_monto_clp(round(monto_total_absoluto, 2))  # 🔄 INVERTIDO
    
    # Campos especiales: ajustar según si hay IVA o no
    if tiene_iva:
        fila_proveedores["Monto 1 Detalle Libro"] = formatear_monto_clp(round(monto_neto, 2))
        fila_proveedores["Monto 3 Detalle Libro"] = formatear_monto_clp(round(monto_iva, 2))
        monto_suma_detalle = round(monto_neto + monto_iva, 2)
    else:
        # SIN IVA: Solo monto neto, sin IVA
        fila_proveedores["Monto 1 Detalle Libro"] = formatear_monto_clp(round(monto_neto, 2))
        fila_proveedores["Monto 3 Detalle Libro"] = formatear_monto_clp(0)  # Sin IVA
        monto_suma_detalle = round(monto_neto, 2)
    
    fila_proveedores["Monto Suma Detalle Libro"] = formatear_monto_clp(monto_suma_detalle)
    filas_resultado.append(fila_proveedores)
    
    # 2. Cuentas de Gastos (una por cada centro de costo - INVERTIDO: Haber en lugar de Debe)
    contador_gasto = 1
    
    # Mapeo de columnas a keys de mapeo (igual que en tipo 33)
    columnas_cc_mapeo = {
        'PyC': 'PyC',
        'PS': 'PS',
        'CO': 'CO',
        'RE': 'RE',
        'TR': 'TR',
        'CF': 'CF',
        'LRC': 'LRC'
    }
    legacy_map = {'PyC': 'col10', 'PS': 'col11', 'CO': 'col12'}
    
    for nombre_cc, ponderador in ponderadores_cc.items():
        if ponderador > 0:  # Solo crear línea si hay ponderador
            # Calcular monto ponderado para este CC
            monto_gasto_cc = round(monto_neto * ponderador, 2)
            
            # Obtener el código de centro de costo que el usuario asignó en el frontend
            mapeo_key = columnas_cc_mapeo.get(nombre_cc, '')
            codigo_cc_usuario = ""
            
            if mapeo_key and mapeo_key in mapeo_cc and mapeo_cc[mapeo_key]:
                codigo_cc_usuario = str(mapeo_cc[mapeo_key]).strip()
            else:
                legacy_key = legacy_map.get(nombre_cc)
                if legacy_key and legacy_key in mapeo_cc and mapeo_cc[legacy_key]:
                    codigo_cc_usuario = str(mapeo_cc[legacy_key]).strip()
            
            if not codigo_cc_usuario:
                logger.warning(f"⚠️ No se encontró código de CC para '{nombre_cc}' (mapeo_key: {mapeo_key}) en mapeo_cc: {mapeo_cc}")
                codigo_cc_usuario = ""  # Dejar vacío si no hay mapeo
            
            # Crear fila de gastos para este CC
            fila_gastos_cc = crear_fila_base_tipo_61(datos_comunes, headers_salida, nombre_tipo_documento)
            fila_gastos_cc["Código Plan de Cuenta"] = codigo_gastos_base  # Sin sufijo
            fila_gastos_cc["Monto al Haber Moneda Base"] = formatear_monto_clp(monto_gasto_cc)  # 🔄 INVERTIDO
            
            # Solo asignar código de centro de costo si existe
            if codigo_cc_usuario:
                fila_gastos_cc["Código Centro de Costo"] = codigo_cc_usuario
            
            logger.info(f"💰 Cuenta Gastos {contador_gasto}: {nombre_cc} -> CC: '{codigo_cc_usuario}' = {formatear_monto_clp(monto_gasto_cc)} (ponderador: {ponderador:.3f})")
            
            filas_resultado.append(fila_gastos_cc)
            contador_gasto += 1
    
    # 3. Cuenta IVA (INVERTIDO: Haber en lugar de Debe) - SOLO SI HAY IVA
    if tiene_iva:
        fila_iva = crear_fila_base_tipo_61(datos_comunes, headers_salida, nombre_tipo_documento)
        fila_iva["Código Plan de Cuenta"] = codigo_iva
        fila_iva["Monto al Haber Moneda Base"] = formatear_monto_clp(round(monto_iva, 2))  # 🔄 INVERTIDO
        filas_resultado.append(fila_iva)
        logger.info(f"📋 Tipo 61 con 2CC CON IVA generó {len(filas_resultado)} filas: 1 Proveedores + {contador_gasto-1} Gastos + 1 IVA (INVERTIDOS)")
    else:
        logger.info(f"📋 Tipo 61 con 2CC SIN IVA generó {len(filas_resultado)} filas: 1 Proveedores + {contador_gasto-1} Gastos (INVERTIDOS)")
    
    return filas_resultado

def aplicar_reglas_tipo_61_3cc(fila, headers_entrada, mapeo_cc, headers_salida, tipos_doc_map):
    """
    Aplica reglas específicas para tipo de documento 61 CON 3 CENTROS DE COSTO
    
    CONTROLADOR IVA:
    - Si hay monto en "Monto IVA": Genera 1 Proveedores + 3 Gastos + 1 IVA
    - Si NO hay IVA: Genera 1 Proveedores + 3 Gastos (sin cuenta IVA) - monto_neto = monto_total
    
    REGLAS ESPECÍFICAS PARA 3CC:
    - 1 cuenta de Proveedores con Monto 1 y 3 Detalle (igual que tipo 33 pero INVERTIDO)
    - 3 cuentas de Gastos separadas, cada una con su monto ponderado (INVERTIDO)
    - 1 cuenta de IVA (INVERTIDO) - SOLO SI HAY IVA
    
    NOTA: Tipo 61 es igual a tipo 33 pero con montos INVERTIDOS (Debe ↔ Haber)
    El monto de cada cuenta de gastos = monto_neto * ponderador_del_CC
    Los ponderadores están en las columnas PyC, EB/PS, CO
    """
    filas_resultado = []
    
    # 🔍 CONTROLADOR: Verificar si hay IVA
    tiene_iva = verificar_tiene_iva_tipo_61(fila, headers_entrada)
    
    # Extraer datos comunes de la fila original
    datos_comunes = extraer_datos_comunes_tipo_61(fila, headers_entrada)
    
    # Calcular montos con validación
    monto_total_raw = datos_comunes.get('monto_total', 0)
    try:
        monto_total = float(monto_total_raw) if monto_total_raw else 0
    except (ValueError, TypeError):
        logger.warning(f"⚠️ Monto inválido encontrado en tipo 61: {monto_total_raw}, usando 0")
        monto_total = 0
    
    # Para tipo 61 (Nota de Crédito), los montos pueden ser negativos
    # Trabajamos con el valor absoluto para los cálculos
    if monto_total == 0:
        logger.warning(f"⚠️ Monto total es 0 en tipo 61 con 3CC: {monto_total}")
        return []  # No generar filas si no hay monto válido
    
    # Usar valor absoluto para cálculos, pero mantener el signo original en logs
    monto_total_absoluto = abs(monto_total)
    logger.info(f"💰 Tipo 61 con 3CC: Monto original = {monto_total}, trabajando con valor absoluto = {monto_total_absoluto}")
    
    # Calcular montos según si hay IVA o no
    if tiene_iva:
        # CON IVA: Calcular igual que tipo 33
        monto_neto = monto_total_absoluto / 1.19  # Monto 1 Detalle Libro
        monto_iva = monto_total_absoluto - monto_neto  # Monto 3 Detalle Libro
        logger.info(f"📊 CON IVA: Monto Neto = {monto_neto:.2f}, Monto IVA = {monto_iva:.2f}")
    else:
        # SIN IVA: Monto neto = monto total, no hay IVA
        monto_neto = monto_total_absoluto
        monto_iva = 0
        logger.info(f"📊 SIN IVA: Monto Neto = Monto Total = {monto_neto:.2f}")
    
    # Obtener código de cuenta original
    codigo_cuenta_original = datos_comunes.get('codigo_cuenta', '')
    
    # Obtener nombre del tipo de documento
    nombre_tipo_documento = tipos_doc_map.get("61", "Nota de Crédito Electrónica")
    logger.info(f"🏷️ Tipo 61 con 3CC -> Nombre: '{nombre_tipo_documento}'")
    
    # Detectar posiciones de centros de costo dinámicamente
    posiciones_cc = detectar_posiciones_centros_costo(headers_entrada)
    logger.info(f"🔍 Posiciones CC detectadas para 61 con 3CC: {posiciones_cc}")
    
    # Extraer ponderadores de cada centro de costo
    ponderadores_cc = {}
    total_ponderadores = 0
    
    for nombre_cc, pos in posiciones_cc.items():
        if pos is not None and pos < len(headers_entrada):
            valor_raw = fila.get(headers_entrada[pos], 0)
            try:
                ponderador = float(valor_raw) if valor_raw else 0
                ponderadores_cc[nombre_cc] = ponderador
                total_ponderadores += ponderador
                logger.info(f"  📊 {nombre_cc}: {ponderador}")
            except (ValueError, TypeError):
                logger.warning(f"⚠️ Ponderador inválido en {nombre_cc}: {valor_raw}")
                ponderadores_cc[nombre_cc] = 0
    
    # Validar que los ponderadores sumen 1.0 (o 100% si están en porcentaje)
    if total_ponderadores == 0:
        logger.warning(f"⚠️ Total de ponderadores es 0, usando distribución equitativa")
        # Distribuir equitativamente entre los CC encontrados
        num_cc = len(posiciones_cc)
        for nombre_cc in ponderadores_cc:
            ponderadores_cc[nombre_cc] = 1.0 / num_cc if num_cc > 0 else 0
    elif abs(total_ponderadores - 1.0) > 0.01:  # Si no suman 1.0, normalizar
        logger.info(f"🔧 Normalizando ponderadores (total: {total_ponderadores})")
        for nombre_cc in ponderadores_cc:
            ponderadores_cc[nombre_cc] = ponderadores_cc[nombre_cc] / total_ponderadores
    
    # Generar códigos de cuenta
    if codigo_cuenta_original.startswith('2'):
        codigo_proveedores = codigo_cuenta_original
        codigo_gastos_base = codigo_cuenta_original.replace('2', '5', 1)
        codigo_iva = codigo_cuenta_original.replace('2', '1', 1)
    elif codigo_cuenta_original.startswith('5'):
        codigo_gastos_base = codigo_cuenta_original
        codigo_proveedores = codigo_cuenta_original.replace('5', '2', 1)
        codigo_iva = codigo_cuenta_original.replace('5', '1', 1)
    elif codigo_cuenta_original.startswith('1'):
        codigo_iva = codigo_cuenta_original
        codigo_proveedores = codigo_cuenta_original.replace('1', '2', 1)
        codigo_gastos_base = codigo_cuenta_original.replace('1', '5', 1)
    else:
        codigo_proveedores = f"2{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "2111001"
        codigo_gastos_base = f"5{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "5111001"
        codigo_iva = f"1{codigo_cuenta_original[1:]}" if len(codigo_cuenta_original) > 1 else "1191001"
    
    # 1. Cuenta Proveedores (INVERTIDO: Debe en lugar de Haber)
    fila_proveedores = crear_fila_base_tipo_61(datos_comunes, headers_salida, nombre_tipo_documento)
    fila_proveedores["Código Plan de Cuenta"] = codigo_proveedores
    fila_proveedores["Monto al Debe Moneda Base"] = formatear_monto_clp(round(monto_total_absoluto, 2))  # 🔄 INVERTIDO
    
    # Campos especiales: ajustar según si hay IVA o no
    if tiene_iva:
        fila_proveedores["Monto 1 Detalle Libro"] = formatear_monto_clp(round(monto_neto, 2))
        fila_proveedores["Monto 3 Detalle Libro"] = formatear_monto_clp(round(monto_iva, 2))
        monto_suma_detalle = round(monto_neto + monto_iva, 2)
    else:
        # SIN IVA: Solo monto neto, sin IVA
        fila_proveedores["Monto 1 Detalle Libro"] = formatear_monto_clp(round(monto_neto, 2))
        fila_proveedores["Monto 3 Detalle Libro"] = formatear_monto_clp(0)  # Sin IVA
        monto_suma_detalle = round(monto_neto, 2)
    
    fila_proveedores["Monto Suma Detalle Libro"] = formatear_monto_clp(monto_suma_detalle)
    filas_resultado.append(fila_proveedores)
    
    # 2. Cuentas de Gastos (una por cada centro de costo - hasta 3 - INVERTIDO: Haber en lugar de Debe)
    contador_gasto = 1
    
    # Mapeo de columnas a keys de mapeo (igual que en tipo 33)
    columnas_cc_mapeo = {
        'PyC': 'PyC',
        'PS': 'PS',
        'CO': 'CO',
        'RE': 'RE',
        'TR': 'TR',
        'CF': 'CF',
        'LRC': 'LRC'
    }
    legacy_map = {'PyC': 'col10', 'PS': 'col11', 'CO': 'col12'}
    
    for nombre_cc, ponderador in ponderadores_cc.items():
        if ponderador > 0:  # Solo crear línea si hay ponderador
            # Calcular monto ponderado para este CC
            monto_gasto_cc = round(monto_neto * ponderador, 2)
            
            # Obtener el código de centro de costo que el usuario asignó en el frontend
            mapeo_key = columnas_cc_mapeo.get(nombre_cc, '')
            codigo_cc_usuario = ""
            
            if mapeo_key and mapeo_key in mapeo_cc and mapeo_cc[mapeo_key]:
                codigo_cc_usuario = str(mapeo_cc[mapeo_key]).strip()
            else:
                legacy_key = legacy_map.get(nombre_cc)
                if legacy_key and legacy_key in mapeo_cc and mapeo_cc[legacy_key]:
                    codigo_cc_usuario = str(mapeo_cc[legacy_key]).strip()
            
            if not codigo_cc_usuario:
                logger.warning(f"⚠️ No se encontró código de CC para '{nombre_cc}' (mapeo_key: {mapeo_key}) en mapeo_cc: {mapeo_cc}")
                codigo_cc_usuario = ""  # Dejar vacío si no hay mapeo
            
            # Crear fila de gastos para este CC
            fila_gastos_cc = crear_fila_base_tipo_61(datos_comunes, headers_salida, nombre_tipo_documento)
            fila_gastos_cc["Código Plan de Cuenta"] = codigo_gastos_base  # Sin sufijo
            fila_gastos_cc["Monto al Haber Moneda Base"] = formatear_monto_clp(monto_gasto_cc)  # 🔄 INVERTIDO
            
            # Solo asignar código de centro de costo si existe
            if codigo_cc_usuario:
                fila_gastos_cc["Código Centro de Costo"] = codigo_cc_usuario
            
            logger.info(f"💰 Cuenta Gastos {contador_gasto}: {nombre_cc} -> CC: '{codigo_cc_usuario}' = {formatear_monto_clp(monto_gasto_cc)} (ponderador: {ponderador:.3f})")
            
            filas_resultado.append(fila_gastos_cc)
            contador_gasto += 1
    
    # 3. Cuenta IVA (INVERTIDO: Haber en lugar de Debe) - SOLO SI HAY IVA
    if tiene_iva:
        fila_iva = crear_fila_base_tipo_61(datos_comunes, headers_salida, nombre_tipo_documento)
        fila_iva["Código Plan de Cuenta"] = codigo_iva
        fila_iva["Monto al Haber Moneda Base"] = formatear_monto_clp(round(monto_iva, 2))  # 🔄 INVERTIDO
        filas_resultado.append(fila_iva)
        logger.info(f"📋 Tipo 61 con 3CC CON IVA generó {len(filas_resultado)} filas: 1 Proveedores + {contador_gasto-1} Gastos + 1 IVA (INVERTIDOS)")
    else:
        logger.info(f"📋 Tipo 61 con 3CC SIN IVA generó {len(filas_resultado)} filas: 1 Proveedores + {contador_gasto-1} Gastos (INVERTIDOS)")
    
    return filas_resultado

def extraer_datos_comunes_tipo_33(fila, headers_entrada):
    """
    Extrae los datos comunes necesarios para tipo 33 desde la fila original
    Busca columnas dinámicamente para mayor flexibilidad con diferentes layouts de Excel
    """
    datos = {}
    
    # Mapear columnas automáticamente
    mapeo = mapear_columnas_automaticamente(headers_entrada)
    
    # Extraer datos usando el mapeo dinámico
    datos['nro'] = fila.get(headers_entrada[mapeo.get('nro', 0)], "") if mapeo.get('nro') is not None else ""
    datos['tipo_doc'] = fila.get(headers_entrada[mapeo.get('tipo_doc', 1)], "") if mapeo.get('tipo_doc') is not None else ""
    datos['rut_proveedor'] = fila.get(headers_entrada[mapeo.get('rut_proveedor', 2)], "") if mapeo.get('rut_proveedor') is not None else ""
    datos['razon_social'] = fila.get(headers_entrada[mapeo.get('razon_social', 3)], "") if mapeo.get('razon_social') is not None else ""
    datos['folio'] = fila.get(headers_entrada[mapeo.get('folio', 4)], "") if mapeo.get('folio') is not None else ""
    datos['fecha'] = fila.get(headers_entrada[mapeo.get('fecha', 5)], "") if mapeo.get('fecha') is not None else ""
    
    # Monto Total - crítico para el procesamiento
    if mapeo.get('monto_total') is not None:
        monto_raw = fila.get(headers_entrada[mapeo['monto_total']], 0)
        try:
            datos['monto_total'] = float(monto_raw) if monto_raw else 0
        except (ValueError, TypeError):
            datos['monto_total'] = 0
    else:
        datos['monto_total'] = 0
    
    # Código y nombre de cuenta
    datos['codigo_cuenta'] = fila.get(headers_entrada[mapeo.get('codigo_cuenta', 7)], "") if mapeo.get('codigo_cuenta') is not None else ""
    datos['nombre_cuenta'] = fila.get(headers_entrada[mapeo.get('nombre_cuenta', 8)], "") if mapeo.get('nombre_cuenta') is not None else ""
    
    # Extraer RUT sin dígito verificador
    rut_completo = datos.get('rut_proveedor', '')
    if rut_completo and '-' in rut_completo:
        datos['rut_sin_dv'] = rut_completo.split('-')[0].replace('.', '')
    else:
        datos['rut_sin_dv'] = rut_completo.replace('.', '') if rut_completo else "11111111"
    
    return datos

def extraer_datos_comunes_tipo_61(fila, headers_entrada):
    """
    Extrae los datos comunes necesarios para tipo 61 desde la fila original
    Busca columnas dinámicamente para mayor flexibilidad con diferentes layouts de Excel
    """
    datos = {}
    
    # Mapear columnas automáticamente
    mapeo = mapear_columnas_automaticamente(headers_entrada)
    
    # Extraer datos usando el mapeo dinámico (misma lógica que tipo 33)
    datos['nro'] = fila.get(headers_entrada[mapeo.get('nro', 0)], "") if mapeo.get('nro') is not None else ""
    datos['tipo_doc'] = fila.get(headers_entrada[mapeo.get('tipo_doc', 1)], "") if mapeo.get('tipo_doc') is not None else ""
    datos['rut_proveedor'] = fila.get(headers_entrada[mapeo.get('rut_proveedor', 2)], "") if mapeo.get('rut_proveedor') is not None else ""
    datos['razon_social'] = fila.get(headers_entrada[mapeo.get('razon_social', 3)], "") if mapeo.get('razon_social') is not None else ""
    datos['folio'] = fila.get(headers_entrada[mapeo.get('folio', 4)], "") if mapeo.get('folio') is not None else ""
    datos['fecha'] = fila.get(headers_entrada[mapeo.get('fecha', 5)], "") if mapeo.get('fecha') is not None else ""
    
    # Monto Total - crítico para el procesamiento
    if mapeo.get('monto_total') is not None:
        monto_raw = fila.get(headers_entrada[mapeo['monto_total']], 0)
        try:
            datos['monto_total'] = float(monto_raw) if monto_raw else 0
        except (ValueError, TypeError):
            datos['monto_total'] = 0
    else:
        datos['monto_total'] = 0
    
    # Código y nombre de cuenta
    datos['codigo_cuenta'] = fila.get(headers_entrada[mapeo.get('codigo_cuenta', 7)], "") if mapeo.get('codigo_cuenta') is not None else ""
    datos['nombre_cuenta'] = fila.get(headers_entrada[mapeo.get('nombre_cuenta', 8)], "") if mapeo.get('nombre_cuenta') is not None else ""
    
    # Extraer RUT sin dígito verificador
    rut_completo = datos.get('rut_proveedor', '')
    if rut_completo and '-' in rut_completo:
        datos['rut_sin_dv'] = rut_completo.split('-')[0].replace('.', '')
    else:
        datos['rut_sin_dv'] = rut_completo.replace('.', '') if rut_completo else "11111111"
    
    return datos

def crear_fila_base_tipo_61(datos_comunes, headers_salida, nombre_tipo_documento="Nota de Crédito"):
    """
    Crea una fila base con los datos comunes para tipo 61
    """
    fila_base = {header: "" for header in headers_salida}
    
    # Llenar campos comunes usando los datos reales del Excel
    fila_base["Numero"] = "61"  # Tipo de documento original
    fila_base["Tipo Documento"] = "Nota de Crédito"  # Nombre para tipo 61
    fila_base["Codigo Auxiliar"] = datos_comunes.get('rut_sin_dv', "")
    fila_base["Numero Doc"] = datos_comunes.get('folio', "")
    # ✨ Usar limpieza UTF-8 para descripción
    fila_base["Descripción Movimiento"] = limpiar_texto_utf8(datos_comunes.get('razon_social', ""))
    
    # ✨ Configurar moneda en CLP
    fila_base["Equivalencia Moneda"] = "CLP"
    
    # Procesar fecha
    fecha_original = datos_comunes.get('fecha', "")
    if fecha_original:
        try:
            if 'T' in str(fecha_original):
                dt = datetime.fromisoformat(str(fecha_original).replace('Z', ''))
                fila_base["Fecha Emisión Docto.(DD/MM/AAAA)"] = dt.strftime("%d/%m/%Y")
            else:
                # Asumir que ya viene en formato correcto o convertir
                fila_base["Fecha Emisión Docto.(DD/MM/AAAA)"] = str(fecha_original)
        except:
            fila_base["Fecha Emisión Docto.(DD/MM/AAAA)"] = str(fecha_original)
    
    return fila_base

def crear_fila_base_tipo_33(datos_comunes, headers_salida, nombre_tipo_documento="Factura"):
    """
    Crea una fila base con los datos comunes para tipo 33
    """
    fila_base = {header: "" for header in headers_salida}
    
    # Log de debug del nombre del tipo de documento
    logger.info(f"🏷️ Creando fila base tipo 33 con nombre: '{nombre_tipo_documento}'")
    
    # Llenar campos comunes usando los datos reales del Excel
    fila_base["Numero"] = "33"  # Tipo de documento original
    fila_base["Tipo Documento"] = "Factura"  # Nombre para tipo 33
    fila_base["Codigo Auxiliar"] = datos_comunes.get('rut_sin_dv', "")
    fila_base["Numero Doc"] = datos_comunes.get('folio', "")
    # ✨ NUEVO: Usar limpieza UTF-8 para descripción
    fila_base["Descripción Movimiento"] = limpiar_texto_utf8(datos_comunes.get('razon_social', ""))
    
    # ✨ NUEVO: Configurar moneda en CLP
    fila_base["Equivalencia Moneda"] = "CLP"
    
    # Procesar fecha
    fecha_original = datos_comunes.get('fecha', "")
    if fecha_original:
        try:
            if 'T' in str(fecha_original):
                dt = datetime.fromisoformat(str(fecha_original).replace('Z', ''))
                fila_base["Fecha Emisión Docto.(DD/MM/AAAA)"] = dt.strftime("%d/%m/%Y")
            else:
                # Asumir que ya viene en formato correcto o convertir
                fila_base["Fecha Emisión Docto.(DD/MM/AAAA)"] = str(fecha_original)
        except:
            fila_base["Fecha Emisión Docto.(DD/MM/AAAA)"] = str(fecha_original)
    
    return fila_base

def extraer_datos_comunes_tipo_34(fila, headers_entrada):
    """
    Extrae los datos comunes necesarios para tipo 34 desde la fila original
    Busca columnas dinámicamente para mayor flexibilidad con diferentes layouts de Excel
    """
    datos = {}
    
    # Mapear columnas automáticamente
    mapeo = mapear_columnas_automaticamente(headers_entrada)
    
    # Extraer datos usando el mapeo dinámico (misma lógica que tipo 33)
    datos['nro'] = fila.get(headers_entrada[mapeo.get('nro', 0)], "") if mapeo.get('nro') is not None else ""
    datos['tipo_doc'] = fila.get(headers_entrada[mapeo.get('tipo_doc', 1)], "") if mapeo.get('tipo_doc') is not None else ""
    datos['rut_proveedor'] = fila.get(headers_entrada[mapeo.get('rut_proveedor', 2)], "") if mapeo.get('rut_proveedor') is not None else ""
    datos['razon_social'] = fila.get(headers_entrada[mapeo.get('razon_social', 3)], "") if mapeo.get('razon_social') is not None else ""
    datos['folio'] = fila.get(headers_entrada[mapeo.get('folio', 4)], "") if mapeo.get('folio') is not None else ""
    datos['fecha'] = fila.get(headers_entrada[mapeo.get('fecha', 5)], "") if mapeo.get('fecha') is not None else ""
    
    # Monto Total - crítico para el procesamiento
    if mapeo.get('monto_total') is not None:
        monto_raw = fila.get(headers_entrada[mapeo['monto_total']], 0)
        try:
            datos['monto_total'] = float(monto_raw) if monto_raw else 0
        except (ValueError, TypeError):
            datos['monto_total'] = 0
    else:
        datos['monto_total'] = 0
    
    # Código y nombre de cuenta
    datos['codigo_cuenta'] = fila.get(headers_entrada[mapeo.get('codigo_cuenta', 7)], "") if mapeo.get('codigo_cuenta') is not None else ""
    datos['nombre_cuenta'] = fila.get(headers_entrada[mapeo.get('nombre_cuenta', 8)], "") if mapeo.get('nombre_cuenta') is not None else ""
    
    # Extraer RUT sin dígito verificador
    rut_completo = datos.get('rut_proveedor', '')
    if rut_completo and '-' in rut_completo:
        datos['rut_sin_dv'] = rut_completo.split('-')[0].replace('.', '')
    else:
        datos['rut_sin_dv'] = rut_completo.replace('.', '') if rut_completo else "11111111"
    
    return datos

def crear_fila_base_tipo_34(datos_comunes, headers_salida, nombre_tipo_documento="Factura Excenta"):
    """
    Crea una fila base con los datos comunes para tipo 34
    """
    fila_base = {header: "" for header in headers_salida}
    
    # Llenar campos comunes usando los datos reales del Excel
    fila_base["Numero"] = "34"  # Tipo de documento original
    fila_base["Tipo Documento"] = "Factura Excenta"  # Nombre para tipo 34
    fila_base["Codigo Auxiliar"] = datos_comunes.get('rut_sin_dv', "")
    fila_base["Numero Doc"] = datos_comunes.get('folio', "")
    # ✨ NUEVO: Usar limpieza UTF-8 para descripción
    fila_base["Descripción Movimiento"] = limpiar_texto_utf8(datos_comunes.get('razon_social', ""))
    
    # ✨ NUEVO: Configurar moneda en CLP
    fila_base["Equivalencia Moneda"] = "CLP"
    
    # Procesar fecha
    fecha_original = datos_comunes.get('fecha', "")
    if fecha_original:
        try:
            if 'T' in str(fecha_original):
                dt = datetime.fromisoformat(str(fecha_original).replace('Z', ''))
                fila_base["Fecha Emisión Docto.(DD/MM/AAAA)"] = dt.strftime("%d/%m/%Y")
            else:
                # Asumir que ya viene en formato correcto o convertir
                fila_base["Fecha Emisión Docto.(DD/MM/AAAA)"] = str(fecha_original)
        except:
            fila_base["Fecha Emisión Docto.(DD/MM/AAAA)"] = str(fecha_original)
    
    return fila_base

def aplicar_reglas_genericas(fila, headers_entrada, tipo_doc, mapeo_cc, headers_salida, tipos_doc_map):
    """
    Aplica reglas genéricas para tipos de documento no implementados específicamente
    """
    # Inicializar fila de salida con valores vacíos
    fila_salida = {header: "" for header in headers_salida}
    
    # Obtener nombre del tipo de documento
    nombre_tipo_documento = tipos_doc_map.get(tipo_doc, f"Tipo {tipo_doc}")
    logger.info(f"🏷️ Tipo {tipo_doc} -> Nombre: '{nombre_tipo_documento}'")
    
    # Campos básicos que se pueden mapear directamente
    fila_salida["Numero"] = str(tipo_doc)  # Tipo de documento original
    # ✨ NUEVO: Usar limpieza UTF-8 para descripción
    descripcion_raw = fila.get(headers_entrada[7] if len(headers_entrada) > 7 else "", "")
    fila_salida["Descripción Movimiento"] = limpiar_texto_utf8(descripcion_raw)
    fila_salida["Tipo Documento"] = tipo_doc
    
    # Calcular códigos de centros de costos aplicables
    codigos_cc = calcular_codigos_cc_para_fila(fila, headers_entrada, mapeo_cc)
    if codigos_cc:
        fila_salida["Código Centro de Costo"] = codigos_cc
    
    # Fecha de emisión
    fecha_original = fila.get(headers_entrada[4] if len(headers_entrada) > 4 else "", "")
    if fecha_original:
        try:
            if 'T' in str(fecha_original):
                dt = datetime.fromisoformat(str(fecha_original).replace('Z', ''))
                fila_salida["Fecha Emisión Docto.(DD/MM/AAAA)"] = dt.strftime("%d/%m/%Y")
            else:
                fila_salida["Fecha Emisión Docto.(DD/MM/AAAA)"] = str(fecha_original)
        except:
            fila_salida["Fecha Emisión Docto.(DD/MM/AAAA)"] = str(fecha_original)
    
    # Monto
    monto_original = fila.get(headers_entrada[5] if len(headers_entrada) > 5 else "", "")
    if monto_original:
        try:
            monto_formateado = formatear_monto_clp(float(monto_original))
            fila_salida["Monto al Debe Moneda Base"] = monto_formateado
        except (ValueError, TypeError):
            fila_salida["Monto al Debe Moneda Base"] = str(monto_original)
    
    return [fila_salida]  # Retornar como lista para consistencia

def limpiar_texto_utf8(texto):
    """
    Limpia y normaliza texto para asegurar correcta codificación UTF-8
    Preserva caracteres especiales como Ñ, ñ, tildes y otros acentos
    """
    if not texto:
        return ""
    
    # Convertir a string si no lo es
    texto_str = str(texto)
    
    # Asegurar que esté en UTF-8 y eliminar caracteres de control
    try:
        # Normalizar el texto para manejar caracteres compuestos
        import unicodedata
        texto_normalizado = unicodedata.normalize('NFC', texto_str)
        
        # Eliminar caracteres de control pero preservar caracteres especiales válidos
        texto_limpio = ''.join(char for char in texto_normalizado 
                              if not unicodedata.category(char).startswith('C') or char in '\n\r\t')
        
        # Limpiar espacios extra
        texto_final = ' '.join(texto_limpio.split())
        
        return texto_final
        
    except Exception:
        # Fallback: solo limpiar espacios
        return ' '.join(texto_str.split())

def formatear_monto_clp(monto):
    """
    Formatea un monto en formato pesos chilenos con separadores de miles
    Aplica redondeo específico: ≥0.5 sube al entero siguiente, <0.5 trunca al entero actual
    Ejemplo: 1000000.56 -> "1.000.001" (≥0.5, sube a 1000001)
             1000000.44 -> "1.000.000" (<0.5, queda en 1000000)
             1000000.50 -> "1.000.001" (=0.5, sube a 1000001)
    """
    if monto == 0 or monto is None:
        return "0"
    
    # Convertir a float si no lo es
    monto_float = float(monto)
    
    # Obtener parte entera y decimal
    parte_entera = int(monto_float)
    parte_decimal = monto_float - parte_entera
    
    # Aplicar lógica de redondeo específica
    if parte_decimal >= 0.5:
        # Si decimal es ≥0.5, subir al entero siguiente
        monto_final = parte_entera + 1
    else:
        # Si decimal es <0.5, usar solo la parte entera
        monto_final = parte_entera
    
    # Formatear con separadores de miles (punto como separador)
    monto_formateado = f"{monto_final:,}".replace(',', '.')
    
    return monto_formateado

def detectar_posiciones_centros_costo(headers):
    """
    Detecta automáticamente las posiciones de las columnas de centros de costo
    Busca PyC, PS/EB, CO por nombre en lugar de posición fija
    Retorna un diccionario con el mapeo dinámico
    """
    mapeo_dinamico = {}
    
    for i, header in enumerate(headers):
        if header is None:
            continue
        h = str(header).strip()
        if h == 'PyC':
            mapeo_dinamico['PyC'] = i
        elif h in ['PS', 'EB']:
            mapeo_dinamico['PS'] = i
        elif h == 'CO':
            mapeo_dinamico['CO'] = i
        elif h == 'RE':
            mapeo_dinamico['RE'] = i
        elif h == 'TR':
            mapeo_dinamico['TR'] = i
        elif h == 'CF':
            mapeo_dinamico['CF'] = i
        elif h == 'LRC':
            mapeo_dinamico['LRC'] = i
    
    return mapeo_dinamico

def buscar_columna_codigo_cuenta(headers):
    """
    Busca la primera columna que contenga "Codigo cuenta" (con variaciones como 1, 2, etc.)
    Retorna el índice de la columna o None si no se encuentra
    """
    for i, header in enumerate(headers):
        if header and 'Codigo cuenta' in str(header):
            return i
    return None

def buscar_columna_nombre_cuenta(headers):
    """
    Busca la primera columna que contenga "Nombre cuenta" (con variaciones como 1, 2, etc.)
    Retorna el índice de la columna o None si no se encuentra
    """
    for i, header in enumerate(headers):
        if header and 'Nombre cuenta' in str(header):
            return i
    return None

def buscar_columna_monto_total(headers):
    """
    Busca la columna "Monto Total" 
    Retorna el índice de la columna o None si no se encuentra
    """
    for i, header in enumerate(headers):
        if header and str(header).strip() == 'Monto Total':
            return i
    return None

def mapear_columnas_automaticamente(headers):
    """
    Mapea automáticamente las columnas importantes basándose en sus nombres
    Retorna un diccionario con las posiciones encontradas
    """
    mapeo = {}
    
    for i, header in enumerate(headers):
        if not header:
            continue
            
        header_clean = str(header).strip()
        
        # Mapeo de columnas conocidas
        if header_clean == 'Nro':
            mapeo['nro'] = i
        elif header_clean == 'Tipo Doc':
            mapeo['tipo_doc'] = i  
        elif header_clean == 'RUT Proveedor':
            mapeo['rut_proveedor'] = i
        elif header_clean == 'Razon Social':
            mapeo['razon_social'] = i
        elif header_clean == 'Folio':
            mapeo['folio'] = i
        elif header_clean == 'Fecha Docto':
            mapeo['fecha'] = i
        elif header_clean == 'Monto Total':
            mapeo['monto_total'] = i
        elif header_clean == 'Monto IVA Recuperable':
            mapeo['monto_iva'] = i
        elif header_clean == 'Monto Exento':
            mapeo['monto_exento'] = i
        elif header_clean == 'Monto Neto':
            mapeo['monto_neto'] = i
        elif 'Codigo cuenta' in header_clean and 'codigo_cuenta' not in mapeo:
            mapeo['codigo_cuenta'] = i  # Tomar el primero
        elif 'Nombre cuenta' in header_clean and 'nombre_cuenta' not in mapeo:
            mapeo['nombre_cuenta'] = i  # Tomar el primero
    
    return mapeo

def verificar_tiene_iva_tipo_61(fila, headers_entrada):
    """
    Verifica si la fila tiene IVA revisando la columna "Monto IVA Recuperable"
    Retorna True si hay IVA, False si no lo hay
    """
    # Mapear columnas automáticamente
    mapeo = mapear_columnas_automaticamente(headers_entrada)
    
    # Verificar si existe la columna "Monto IVA Recuperable"
    if mapeo.get('monto_iva') is None:
        logger.info("🔍 No se encontró columna 'Monto IVA Recuperable', asumiendo que SÍ hay IVA (lógica por defecto)")
        return True
    
    # Obtener el valor de la columna "Monto IVA Recuperable"
    monto_iva_raw = fila.get(headers_entrada[mapeo['monto_iva']], 0)
    
    try:
        monto_iva = float(monto_iva_raw) if monto_iva_raw else 0
    except (ValueError, TypeError):
        logger.warning(f"⚠️ Valor inválido en columna 'Monto IVA Recuperable': {monto_iva_raw}, asumiendo 0")
        monto_iva = 0
    
    tiene_iva = abs(monto_iva) > 0.01  # Considerar que hay IVA si es mayor a 1 centavo
    
    logger.info(f"🔍 Verificación IVA: Monto IVA Recuperable = {monto_iva} -> {'SÍ' if tiene_iva else 'NO'} hay IVA")
    
    return tiene_iva

def verificar_tiene_exento_tipo_33(fila, headers_entrada):
    """
    Verifica si la fila tiene monto exento revisando la columna "Monto Exento"
    Retorna True si hay monto exento, False si no lo hay
    Si hay exento, se debe calcular el IVA desde "Monto Neto" en lugar del "Monto Total"
    """
    # Mapear columnas automáticamente
    mapeo = mapear_columnas_automaticamente(headers_entrada)
    
    # Verificar si existe la columna "Monto Exento"
    if mapeo.get('monto_exento') is None:
        logger.info("🔍 No se encontró columna 'Monto Exento', usando lógica estándar (desde Monto Total)")
        return False
    
    # Obtener el valor de la columna "Monto Exento"
    monto_exento_raw = fila.get(headers_entrada[mapeo['monto_exento']], 0)
    
    try:
        monto_exento = float(monto_exento_raw) if monto_exento_raw else 0
    except (ValueError, TypeError):
        logger.warning(f"⚠️ Valor inválido en columna 'Monto Exento': {monto_exento_raw}, asumiendo 0")
        monto_exento = 0
    
    tiene_exento = abs(monto_exento) > 0.01  # Considerar que hay exento si es mayor a 1 centavo
    
    logger.info(f"🔍 Verificación Exento: Monto Exento = {monto_exento} -> {'SÍ' if tiene_exento else 'NO'} hay exento (IVA desde {'Monto Neto' if tiene_exento else 'Monto Total'})")
    
    return tiene_exento

def calcular_codigos_cc_para_fila(fila, headers, mapeo_cc):
    """
    Calcula qué códigos de centros de costos se aplican a una fila específica
    Retorna una string con los códigos separados por comas
    Busca las columnas PyC, PS/EB, CO por nombre (EB es equivalente a PS)
    """
    # Detectar posiciones dinámicamente
    posiciones_cc = detectar_posiciones_centros_costo(headers)
    
    # El frontend ahora envía mapeo por tipo lógico (PyC, PS, CO, RE, TR, CF, LRC)
    # Mantener compatibilidad si llegan claves antiguas col10/col11/col12
    legacy_map = {'PyC': 'col10', 'PS': 'col11', 'CO': 'col12'}
    
    codigos_aplicables = []
    
    # Buscar cada tipo de centro de costo
    for tipo_cc, posicion in posiciones_cc.items():
        # Encuentra código desde mapeo lógico o claves legacy
        codigo_cc = None
        if tipo_cc in mapeo_cc and mapeo_cc[tipo_cc]:
            codigo_cc = mapeo_cc[tipo_cc]
        else:
            legacy_key = legacy_map.get(tipo_cc)
            if legacy_key and legacy_key in mapeo_cc:
                codigo_cc = mapeo_cc.get(legacy_key)

        if codigo_cc is None:
            continue

        header_name = headers[posicion]
        valor = fila.get(header_name)
        
        # Un centro de costo es válido si NO es: None, "-", 0, "0" y tiene código mapeado
        if (valor is not None and 
            valor != "-" and 
            valor != 0 and 
            valor != "0" and 
            str(valor).strip() != "" and
            str(codigo_cc).strip() != ""):
            
            codigo_limpio = str(codigo_cc).strip()
            if codigo_limpio not in codigos_aplicables:
                codigos_aplicables.append(codigo_limpio)
    
    # Retornar códigos separados por comas o string vacía si no hay ninguno
    return ", ".join(codigos_aplicables) if codigos_aplicables else ""


def contar_centros_costos(fila, headers, mapeo_cc=None):
    """
    Cuenta la cantidad de centros de costos válidos en las columnas PyC, PS/EB, CO
    Un centro de costo es válido si NO es: None, "-", 0
    
    Nota: PS y EB son equivalentes (mismo campo, diferentes nombres en archivos)
    """
    # Detectar posiciones dinámicamente
    posiciones_cc = detectar_posiciones_centros_costo(headers)
    count = 0
    
    logger.info(f"🔍 Contando CC para fila con posiciones detectadas: {posiciones_cc}")
    
    # Contar centros de costo válidos
    for tipo_cc, posicion in posiciones_cc.items():
        if posicion < len(headers):
            header_name = headers[posicion]
            valor = fila.get(header_name)
            
            # Un centro de costo es válido si NO es: None, "-", 0, "0" o string vacío
            es_valido = (valor is not None and 
                        valor != "-" and 
                        valor != 0 and 
                        valor != "0" and 
                        str(valor).strip() != "")
            
            if es_valido:
                count += 1
                logger.info(f"  ✅ {tipo_cc} (pos {posicion}): '{valor}' = VÁLIDO")
            else:
                logger.info(f"  ❌ {tipo_cc} (pos {posicion}): '{valor}' = INVÁLIDO")
    
    logger.info(f"📊 Total centros de costo válidos: {count}")
    return count

def serialize_excel_data(data):
    """
    Convierte objetos datetime y otros tipos no serializables a strings
    Mantiene el formato de fecha original para Excel
    """
    if isinstance(data, datetime):
        # Si es datetime, convertir a fecha simple DD-MM-YYYY
        return data.strftime("%d-%m-%Y")
    elif isinstance(data, date):
        # Si es date, mantener formato DD-MM-YYYY  
        return data.strftime("%d-%m-%Y")
    elif isinstance(data, dict):
        return {key: serialize_excel_data(value) for key, value in data.items()}
    elif isinstance(data, list):
        return [serialize_excel_data(item) for item in data]
    elif isinstance(data, str):
        # Asegurar que las strings mantengan encoding UTF-8
        return data
    else:
        return data

@shared_task(bind=True)
def procesar_captura_masiva_gastos_task(self, archivo_content, archivo_nombre, usuario_id, mapeo_cc=None):
    """
    Tarea principal que recibe el archivo Excel y dispara el procesamiento con Chord
    """
    logger.info(f"🧾 Iniciando captura masiva de gastos para archivo: {archivo_nombre}")
    
    # Si no hay mapeo, usar diccionario vacío
    if mapeo_cc is None:
        mapeo_cc = {}
    
    try:
        # Configurar Redis db1 para resultados temporales - usar configuración de Django
        redis_client = get_redis_client_db1()
        
        # Cargar el archivo Excel desde el contenido
        wb = load_workbook(BytesIO(archivo_content))
        ws = wb.active
        
        # Obtener headers de la primera fila
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        logger.info(f"📋 Headers encontrados: {headers}")
        
        # Procesar filas desde la fila 2 (saltar header)
        filas_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):  # Solo procesar filas que no estén completamente vacías
                fila_dict = dict(zip(headers, row))
                # Serializar datos para evitar problemas con datetime
                fila_dict = serialize_excel_data(fila_dict)
                filas_data.append(fila_dict)
        
        logger.info(f"📊 Total filas a procesar: {len(filas_data)}")
        
        # Agrupar por "Tipo Doc" (columna 2) + cantidad de Centros de Costos (PyC, PS/EB, CO)
        tipo_doc_column = headers[1] if len(headers) > 1 else None
        if not tipo_doc_column:
            raise ValueError("No se encontró la columna de Tipo Doc (columna 2)")
        
        grupos_por_tipo_cc = {}
        for i, fila in enumerate(filas_data):
            tipo_doc = fila.get(tipo_doc_column, 'Sin Tipo')
            cc_count = contar_centros_costos(fila, headers, mapeo_cc)
            
            # Crear clave de grupo: "TipoDoc con XCC"
            clave_grupo = f"{tipo_doc} con {cc_count}CC"
            
            if clave_grupo not in grupos_por_tipo_cc:
                grupos_por_tipo_cc[clave_grupo] = []
            grupos_por_tipo_cc[clave_grupo].append(fila)
        
        logger.info(f"🗂️ Grupos creados por Tipo Doc + CC: {list(grupos_por_tipo_cc.keys())}")
        
        # Generar task_id único para coordinar resultados
        task_id = self.request.id
        
        # Almacenar metadatos en Redis con TTL de 5 minutos - incluir usuario_id en la clave
        metadata = {
            'task_id': task_id,
            'archivo_nombre': archivo_nombre,
            'usuario_id': usuario_id,
            'total_filas': len(filas_data),
            'grupos': list(grupos_por_tipo_cc.keys()),
            'headers': headers,
            'mapeo_cc': mapeo_cc,
            'inicio': timezone.now().isoformat(),
            'estado': 'procesando'
        }
        redis_client.setex(
            f"captura_gastos_meta:{usuario_id}:{task_id}", 
            300, 
            json.dumps(metadata, ensure_ascii=False)
        )
        
        # Crear las tareas del chord - una por cada grupo de tipo de documento + CC
        chord_tasks = []
        for clave_grupo, filas_grupo in grupos_por_tipo_cc.items():
            # Serializar las filas antes de pasarlas a la tarea
            filas_serializadas = serialize_excel_data(filas_grupo)
            chord_tasks.append(
                procesar_grupo_tipo_doc_task.s(
                    task_id, usuario_id, clave_grupo, filas_serializadas, headers, mapeo_cc
                )
            )
        
        # Ejecutar Chord con concurrencia 2 (requisito para contabilidad)
        chord_job = chord(chord_tasks)(
            consolidar_resultados_captura_gastos_task.s(task_id, usuario_id)
        )
        
        logger.info(f"🎼 Chord iniciado con {len(chord_tasks)} tareas paralelas")
        
        return {
            'task_id': task_id,
            'estado': 'procesando',
            'total_grupos': len(grupos_por_tipo_cc),
            'total_filas': len(filas_data)
        }
        
    except Exception as e:
        logger.error(f"❌ Error procesando captura masiva de gastos: {str(e)}")
        raise


@shared_task(bind=True)
def procesar_grupo_tipo_doc_task(self, task_id, usuario_id, clave_grupo, filas_grupo, headers, mapeo_cc=None):
    """
    Procesa un grupo de filas del mismo tipo de documento y cantidad de centros de costos
    """
    logger.info(f"⚙️ Procesando grupo '{clave_grupo}' con {len(filas_grupo)} filas para usuario {usuario_id}")
    
    # Log detallado de las filas en este grupo
    tipo_doc_column = headers[1] if len(headers) > 1 else None
    if tipo_doc_column:
        for i, fila in enumerate(filas_grupo[:3]):  # Solo las primeras 3 filas para no saturar logs
            tipo_doc_fila = fila.get(tipo_doc_column, 'Sin Tipo')
            cc_count = contar_centros_costos(fila, headers, mapeo_cc)
            logger.info(f"    Fila {i+1} en grupo '{clave_grupo}': Tipo Doc='{tipo_doc_fila}', CC Count={cc_count}")
    
    # Si no hay mapeo, usar diccionario vacío
    if mapeo_cc is None:
        mapeo_cc = {}
    
    try:
        redis_client = get_redis_client_db1()
        
        # Simular procesamiento del grupo
        time.sleep(1)  # Simular trabajo real
        
        # Preparar resultado del grupo
        resultado_grupo = {
            'clave_grupo': clave_grupo,
            'filas_procesadas': len(filas_grupo),
            'filas_data': serialize_excel_data(filas_grupo),  # Serializar antes de guardar
            'headers': headers,
            'procesado_en': timezone.now().isoformat(),
            'estado': 'completado'
        }
        
        # Almacenar resultado parcial en Redis - incluir usuario_id en la clave
        # Usar clave_grupo pero reemplazar espacios para evitar problemas en Redis
        clave_redis = clave_grupo.replace(" ", "_")
        redis_client.setex(
            f"captura_gastos_grupo:{usuario_id}:{task_id}:{clave_redis}", 
            300, 
            json.dumps(resultado_grupo, ensure_ascii=False)
        )
        
        logger.info(f"✅ Grupo '{clave_grupo}' procesado exitosamente")
        
        return {
            'clave_grupo': clave_grupo,
            'filas_procesadas': len(filas_grupo),
            'estado': 'completado'
        }
        
    except Exception as e:
        logger.error(f"❌ Error procesando grupo '{clave_grupo}': {str(e)}")
        raise


@shared_task(bind=True)
def consolidar_resultados_captura_gastos_task(self, resultados_grupos, task_id, usuario_id):
    """
    Consolida los resultados de todos los grupos y genera el Excel final
    """
    logger.info(f"🔄 Consolidando resultados para task_id: {task_id}, usuario: {usuario_id}")
    
    try:
        redis_client = get_redis_client_db1()
        
        # Recuperar metadata - incluir usuario_id en la clave
        metadata_raw = redis_client.get(f"captura_gastos_meta:{usuario_id}:{task_id}")
        if not metadata_raw:
            raise ValueError(f"No se encontraron metadatos para usuario {usuario_id}, task_id: {task_id}")
        
        metadata = json.loads(metadata_raw)
        headers = metadata['headers']
        
        # Mapeo estático de tipos de documento chilenos comunes
        tipos_doc_map = {
            "33": "Factura Electrónica",
            "34": "Factura Exenta Electrónica", 
            "39": "Boleta Electrónica",
            "41": "Boleta Exenta Electrónica",
            "43": "Liquidación Factura Electrónica",
            "46": "Factura de Compra Electrónica",
            "52": "Guía de Despacho Electrónica",
            "56": "Nota de Débito Electrónica",
            "61": "Nota de Crédito Electrónica",
            "110": "Factura de Exportación Electrónica",
            "111": "Nota de Débito de Exportación Electrónica",
            "112": "Nota de Crédito de Exportación Electrónica"
        }
        logger.info(f"📋 Usando mapeo estático de {len(tipos_doc_map)} tipos de documento chilenos")
        
        # Crear un nuevo workbook con pestañas por tipo de documento
        wb = Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto
        
        total_filas_procesadas = 0
        
        # Recuperar resultados de cada grupo y crear pestañas
        for resultado in resultados_grupos:
            clave_grupo = resultado['clave_grupo']
            
            # Recuperar datos detallados del grupo desde Redis - incluir usuario_id en la clave
            # Usar clave_grupo pero reemplazar espacios para Redis
            clave_redis = clave_grupo.replace(" ", "_")
            grupo_data_raw = redis_client.get(f"captura_gastos_grupo:{usuario_id}:{task_id}:{clave_redis}")
            if grupo_data_raw:
                grupo_data = json.loads(grupo_data_raw)
                filas_data = grupo_data['filas_data']
                
                # Crear hoja para este grupo - convertir a string y limpiar para Excel
                nombre_hoja = str(clave_grupo)[:31]  # Excel limita a 31 caracteres
                # Reemplazar caracteres no válidos para nombres de hoja en Excel
                nombre_hoja = nombre_hoja.replace(":", "-").replace("/", "-").replace("\\", "-")
                ws = wb.create_sheet(title=nombre_hoja)
                
                # Obtener mapeo de centros de costos desde metadatos
                mapeo_cc = metadata.get('mapeo_cc', {})
                
                # Headers de salida según especificaciones contables
                headers_salida = get_headers_salida_contabilidad()
                
                # Escribir headers de salida
                for col, header in enumerate(headers_salida, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # Escribir datos transformados
                row_idx = 2  # Comenzar después del header
                
                for fila in filas_data:
                    # Extraer tipo de documento y cantidad de CC del nombre del grupo (ej: "33 con 2CC" -> "33", 2)
                    partes_grupo = clave_grupo.split(' ')
                    tipo_doc = partes_grupo[0]
                    cc_count = int(partes_grupo[2].replace('CC', '')) if len(partes_grupo) >= 3 else 1
                    
                    # Aplicar reglas de transformación según tipo de documento y cantidad de CC
                    # Esto puede retornar múltiples filas (ej: para tipo 33 retorna 3 filas: Proveedores, Gastos, IVA)
                    filas_transformadas = aplicar_reglas_tipo_documento(fila, headers, tipo_doc, cc_count, mapeo_cc, tipos_doc_map)
                    
                    # Escribir cada fila transformada
                    for fila_transformada in filas_transformadas:
                        for col_idx, header in enumerate(headers_salida, 1):
                            valor = fila_transformada.get(header, "")
                            ws.cell(row=row_idx, column=col_idx, value=valor)
                        row_idx += 1
                
                # Calcular total de filas de salida para este grupo
                filas_salida_grupo = 0
                for fila in filas_data:
                    partes_grupo = clave_grupo.split(' ')
                    tipo_doc = partes_grupo[0]
                    cc_count = int(partes_grupo[2].replace('CC', '')) if len(partes_grupo) >= 3 else 1
                    filas_transformadas = aplicar_reglas_tipo_documento(fila, headers, tipo_doc, cc_count, mapeo_cc, tipos_doc_map)
                    filas_salida_grupo += len(filas_transformadas)
                
                total_filas_procesadas += filas_salida_grupo
                logger.info(f"📊 Hoja '{clave_grupo}' creada con {len(filas_data)} filas de entrada → {filas_salida_grupo} filas de salida")
        
        # Convertir workbook a bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_content = excel_buffer.getvalue()
        
        # Almacenar el archivo Excel en Redis con TTL de 5 minutos - incluir usuario_id en la clave
        redis_client_binary = get_redis_client_db1_binary()
        redis_client_binary.setex(
            f"captura_gastos_excel:{usuario_id}:{task_id}", 
            300, 
            excel_content
        )
        
        # Actualizar metadata con resultado final - incluir usuario_id en la clave
        metadata.update({
            'estado': 'completado',
            'fin': timezone.now().isoformat(),
            'total_filas_procesadas': total_filas_procesadas,
            'archivo_excel_disponible': True
        })
        redis_client.setex(
            f"captura_gastos_meta:{usuario_id}:{task_id}", 
            300, 
            json.dumps(metadata, ensure_ascii=False)
        )
        
        logger.info(f"✅ Consolidación completada para task_id: {task_id}")
        
        return {
            'task_id': task_id,
            'estado': 'completado',
            'total_filas_procesadas': total_filas_procesadas,
            'total_grupos': len(resultados_grupos),
            'archivo_excel_disponible': True
        }
        
    except Exception as e:
        logger.error(f"❌ Error consolidando resultados: {str(e)}")
        
        # Marcar como error en Redis - incluir usuario_id en la clave
        try:
            redis_client = get_redis_client_db1()
            error_metadata = {
                'task_id': task_id,
                'usuario_id': usuario_id,
                'estado': 'error',
                'error': str(e),
                'fin': timezone.now().isoformat()
            }
            redis_client.setex(
                f"captura_gastos_meta:{usuario_id}:{task_id}", 
                300, 
                json.dumps(error_metadata, ensure_ascii=False)
            )
        except:
            pass
        
        raise



